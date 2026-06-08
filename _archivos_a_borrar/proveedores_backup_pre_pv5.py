# ============================================================
# MÓDULO PROVEEDORES - SCHIRO / LK-D Mayorista
# Pegar en app.py ANTES de la línea: app.run(debug=True, ...)
# ============================================================
# v2 — Ampliado con Ciclo de Compras Módulo 4 (Cta Cte Proveedores):
#   - FacturaCompra: + fecha_vencimiento, saldo_pendiente, estado, cae, fecha_vto_cae
#   - PagoProveedor: + numero_recibo, punto_venta_recibo, estado
#   - NUEVOS: PagoProveedorNumerador, PagoProveedorDetalle, PagoProveedorMedio,
#             ProveedorMovimiento, ChequePropio, ChequeTercero
#   Ninguna ruta existente fue modificada — todo lo nuevo se construye
#   en las fases siguientes del ciclo de compras.
# ============================================================

# ──────────────────────────────────────────────────────────────
# MODELOS
# ──────────────────────────────────────────────────────────────
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, session, send_file
from extensions import db
from sqlalchemy import Numeric, or_
from datetime import datetime
from decimal import Decimal
from io import BytesIO

proveedores_bp = Blueprint('proveedores', __name__)

class Proveedor(db.Model):
    __tablename__ = 'proveedor'

    id             = db.Column(db.Integer, primary_key=True)
    razon_social   = db.Column(db.String(150), nullable=False)
    cuit           = db.Column(db.String(20))
    condicion_iva  = db.Column(db.String(50), nullable=False, default='Responsable Inscripto')
    direccion      = db.Column(db.String(200))
    telefono       = db.Column(db.String(30))
    email          = db.Column(db.String(100))
    saldo          = db.Column(Numeric(12, 2), nullable=False, default=0.00)  # positivo = le debo
    activo         = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)

    facturas_compra = db.relationship('FacturaCompra', backref='proveedor', lazy='dynamic')
    pagos           = db.relationship('PagoProveedor',  backref='proveedor', lazy='dynamic')
    movimientos     = db.relationship('ProveedorMovimiento', backref='proveedor', lazy='dynamic')

    def __repr__(self):
        return f'<Proveedor {self.razon_social}>'

    def to_dict(self):
        return {
            'id':            self.id,
            'razon_social':  self.razon_social,
            'cuit':          self.cuit,
            'condicion_iva': self.condicion_iva,
            'direccion':     self.direccion,
            'telefono':      self.telefono,
            'email':         self.email,
            'saldo':         float(self.saldo),
            'activo':        self.activo,
        }


class FacturaCompra(db.Model):
    __tablename__ = 'factura_compra'

    id                = db.Column(db.Integer, primary_key=True)
    proveedor_id      = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=False)
    fecha             = db.Column(db.Date, nullable=False)
    fecha_vencimiento = db.Column(db.Date)                                          # NUEVO
    tipo_comprobante  = db.Column(db.String(5), nullable=False, default='A')        # A, B, C
    punto_venta       = db.Column(db.String(5), nullable=False)
    numero            = db.Column(db.String(10), nullable=False)
    neto_gravado_21   = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    iva_21            = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    neto_gravado_105  = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    iva_105           = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    neto_no_gravado   = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    otros_impuestos   = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    total             = db.Column(Numeric(12, 2), nullable=False)
    saldo_pendiente   = db.Column(Numeric(12, 2), nullable=False, default=0.00)     # NUEVO
    estado            = db.Column(db.String(15), nullable=False, default='pendiente') # NUEVO: pendiente/parcial/pagada/anulada
    cae               = db.Column(db.String(20))                                    # NUEVO (del OCR)
    fecha_vto_cae     = db.Column(db.Date)                                          # NUEVO
    observaciones     = db.Column(db.Text)
    usuario_id        = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    fecha_creacion    = db.Column(db.DateTime, default=datetime.now)

    usuario = db.relationship('Usuario', backref='facturas_compra')

    # Relaciones con el módulo Cta Cte
    imputaciones = db.relationship('PagoProveedorDetalle', backref='factura', lazy='dynamic')

    def numero_completo(self):
        return f"{self.tipo_comprobante} {self.punto_venta.zfill(4)}-{self.numero.zfill(8)}"

    def to_dict(self):
        return {
            'id':                self.id,
            'proveedor_id':      self.proveedor_id,
            'proveedor':         self.proveedor.razon_social if self.proveedor else '',
            'fecha':             self.fecha.strftime('%d/%m/%Y') if self.fecha else '',
            'fecha_vencimiento': self.fecha_vencimiento.strftime('%d/%m/%Y') if self.fecha_vencimiento else '',
            'tipo_comprobante':  self.tipo_comprobante,
            'punto_venta':       self.punto_venta,
            'numero':            self.numero,
            'numero_completo':   self.numero_completo(),
            'neto_gravado_21':   float(self.neto_gravado_21),
            'iva_21':            float(self.iva_21),
            'neto_gravado_105':  float(self.neto_gravado_105),
            'iva_105':           float(self.iva_105),
            'neto_no_gravado':   float(self.neto_no_gravado),
            'otros_impuestos':   float(self.otros_impuestos),
            'total':             float(self.total),
            'saldo_pendiente':   float(self.saldo_pendiente or 0),
            'estado':            self.estado or 'pendiente',
            'cae':               self.cae or '',
            'fecha_vto_cae':     self.fecha_vto_cae.strftime('%d/%m/%Y') if self.fecha_vto_cae else '',
            'observaciones':     self.observaciones,
        }


class PagoProveedor(db.Model):
    __tablename__ = 'pago_proveedor'

    id                 = db.Column(db.Integer, primary_key=True)
    proveedor_id       = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=False)
    fecha              = db.Column(db.Date, nullable=False)
    importe            = db.Column(Numeric(12, 2), nullable=False)
    numero_recibo      = db.Column(db.Integer, nullable=False, default=0)             # NUEVO
    punto_venta_recibo = db.Column(db.Integer, nullable=False, default=1)             # NUEVO
    estado             = db.Column(db.String(15), nullable=False, default='activo')   # NUEVO: activo/anulado
    forma_pago         = db.Column(db.String(30), nullable=False, default='efectivo') # histórico (pagos viejos)
    referencia         = db.Column(db.String(100))                                    # histórico (pagos viejos)
    observaciones      = db.Column(db.Text)
    usuario_id         = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    fecha_creacion     = db.Column(db.DateTime, default=datetime.now)

    usuario = db.relationship('Usuario', backref='pagos_proveedor')

    # Relaciones con el módulo Cta Cte
    detalles = db.relationship('PagoProveedorDetalle', backref='pago', lazy='dynamic',
                                cascade='all, delete-orphan')
    medios   = db.relationship('PagoProveedorMedio',   backref='pago', lazy='dynamic',
                                cascade='all, delete-orphan')

    def numero_recibo_completo(self):
        """Devuelve 'R 0001-00000012' — solo útil para pagos nuevos (con recibo)."""
        if self.numero_recibo:
            return f"R {str(self.punto_venta_recibo).zfill(4)}-{str(self.numero_recibo).zfill(8)}"
        return ''

    def to_dict(self):
        return {
            'id':                 self.id,
            'proveedor_id':       self.proveedor_id,
            'proveedor':          self.proveedor.razon_social if self.proveedor else '',
            'fecha':              self.fecha.strftime('%d/%m/%Y') if self.fecha else '',
            'importe':            float(self.importe),
            'numero_recibo':      self.numero_recibo or 0,
            'punto_venta_recibo': self.punto_venta_recibo or 1,
            'numero_recibo_completo': self.numero_recibo_completo(),
            'estado':             self.estado or 'activo',
            'forma_pago':         self.forma_pago,
            'referencia':         self.referencia,
            'observaciones':      self.observaciones,
        }


# ══════════════════════════════════════════════════════════════
# MODELOS NUEVOS — Ciclo de Compras Módulo 4 (Cta Cte Proveedores)
# ══════════════════════════════════════════════════════════════

class PagoProveedorNumerador(db.Model):
    """Correlativo de recibos/OP por punto de venta."""
    __tablename__ = 'pago_proveedor_numerador'

    id             = db.Column(db.Integer, primary_key=True)
    punto_venta    = db.Column(db.Integer, nullable=False, default=1, unique=True)
    ultimo_numero  = db.Column(db.Integer, nullable=False, default=0)

    @staticmethod
    def siguiente_numero(punto_venta=1):
        """
        Toma y reserva el siguiente número de recibo de forma atómica.
        Uso dentro de una transacción:
            nro = PagoProveedorNumerador.siguiente_numero(1)
            db.session.commit()
        """
        num = PagoProveedorNumerador.query.filter_by(punto_venta=punto_venta)\
                .with_for_update().first()
        if not num:
            num = PagoProveedorNumerador(punto_venta=punto_venta, ultimo_numero=0)
            db.session.add(num)
            db.session.flush()
        num.ultimo_numero = (num.ultimo_numero or 0) + 1
        return num.ultimo_numero


class PagoProveedorDetalle(db.Model):
    """Imputación del pago a facturas específicas (1 pago → N facturas)."""
    __tablename__ = 'pago_proveedor_detalle'

    id             = db.Column(db.Integer, primary_key=True)
    pago_id        = db.Column(db.Integer, db.ForeignKey('pago_proveedor.id'), nullable=False)
    factura_id     = db.Column(db.Integer, db.ForeignKey('factura_compra.id'), nullable=False)
    monto_imputado = db.Column(Numeric(12, 2), nullable=False, default=0.00)

    def to_dict(self):
        return {
            'id':             self.id,
            'pago_id':        self.pago_id,
            'factura_id':     self.factura_id,
            'monto_imputado': float(self.monto_imputado),
            'factura_numero': self.factura.numero_completo() if self.factura else '',
            'factura_total':  float(self.factura.total) if self.factura else 0,
        }


class PagoProveedorMedio(db.Model):
    """Medios de pago usados en un pago (1 pago → N medios)."""
    __tablename__ = 'pago_proveedor_medio'

    id                = db.Column(db.Integer, primary_key=True)
    pago_id           = db.Column(db.Integer, db.ForeignKey('pago_proveedor.id'), nullable=False)
    medio             = db.Column(db.String(20), nullable=False)
    # medio: efectivo / transferencia / cheque_propio / cheque_tercero / otro
    monto             = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    cheque_propio_id  = db.Column(db.Integer, db.ForeignKey('cheque_propio.id'))
    cheque_tercero_id = db.Column(db.Integer, db.ForeignKey('cheque_tercero.id'))
    banco_destino     = db.Column(db.String(50))
    cbu_destino       = db.Column(db.String(22))
    observaciones     = db.Column(db.String(200))

    cheque_propio  = db.relationship('ChequePropio',  foreign_keys=[cheque_propio_id])
    cheque_tercero = db.relationship('ChequeTercero', foreign_keys=[cheque_tercero_id])

    def to_dict(self):
        d = {
            'id':            self.id,
            'pago_id':       self.pago_id,
            'medio':         self.medio,
            'monto':         float(self.monto),
            'banco_destino': self.banco_destino,
            'cbu_destino':   self.cbu_destino,
            'observaciones': self.observaciones,
        }
        if self.cheque_propio:
            d['cheque_propio'] = {
                'id':     self.cheque_propio.id,
                'banco':  self.cheque_propio.banco,
                'numero': self.cheque_propio.numero,
                'vto':    self.cheque_propio.fecha_vencimiento.strftime('%d/%m/%Y') if self.cheque_propio.fecha_vencimiento else '',
            }
        if self.cheque_tercero:
            d['cheque_tercero'] = {
                'id':       self.cheque_tercero.id,
                'banco':    self.cheque_tercero.banco,
                'numero':   self.cheque_tercero.numero,
                'librador': self.cheque_tercero.librador,
                'vto':      self.cheque_tercero.fecha_vencimiento.strftime('%d/%m/%Y') if self.cheque_tercero.fecha_vencimiento else '',
            }
        return d


class ProveedorMovimiento(db.Model):
    """
    Libro mayor del proveedor — fuente de verdad para la Cta Cte.
    Se alimenta al cargar facturas, registrar pagos, cargar NC o en la migración inicial.
    """
    __tablename__ = 'proveedor_movimiento'

    id              = db.Column(db.Integer, primary_key=True)
    proveedor_id    = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=False)
    fecha           = db.Column(db.Date, nullable=False)
    tipo            = db.Column(db.String(20), nullable=False)
    # tipo: factura / pago / nota_credito / saldo_inicial / ajuste
    referencia_id   = db.Column(db.Integer)   # id de factura_compra, pago_proveedor, etc.
    descripcion     = db.Column(db.String(200))
    debe            = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    haber           = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    saldo_acumulado = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    fecha_carga     = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id':              self.id,
            'proveedor_id':    self.proveedor_id,
            'fecha':           self.fecha.strftime('%d/%m/%Y') if self.fecha else '',
            'tipo':            self.tipo,
            'referencia_id':   self.referencia_id,
            'descripcion':     self.descripcion,
            'debe':            float(self.debe or 0),
            'haber':           float(self.haber or 0),
            'saldo_acumulado': float(self.saldo_acumulado or 0),
        }


class ChequePropio(db.Model):
    """Chequera propia — ciclo de vida completo."""
    __tablename__ = 'cheque_propio'

    id                = db.Column(db.Integer, primary_key=True)
    banco             = db.Column(db.String(50), nullable=False)
    cuenta            = db.Column(db.String(30))
    numero            = db.Column(db.String(20), nullable=False)
    fecha_emision     = db.Column(db.Date, nullable=False)
    fecha_vencimiento = db.Column(db.Date)
    monto             = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    estado            = db.Column(db.String(15), nullable=False, default='pendiente')
    # estado: pendiente / entregado / cobrado / rechazado / anulado
    proveedor_id      = db.Column(db.Integer, db.ForeignKey('proveedor.id'))
    pago_id           = db.Column(db.Integer, db.ForeignKey('pago_proveedor.id'))
    fecha_entrega     = db.Column(db.Date)
    fecha_cobro       = db.Column(db.Date)
    observaciones     = db.Column(db.Text)
    fecha_carga       = db.Column(db.DateTime, default=datetime.now)

    proveedor = db.relationship('Proveedor',      foreign_keys=[proveedor_id])
    pago      = db.relationship('PagoProveedor',  foreign_keys=[pago_id])

    def to_dict(self):
        return {
            'id':                self.id,
            'banco':             self.banco,
            'cuenta':            self.cuenta,
            'numero':            self.numero,
            'fecha_emision':     self.fecha_emision.strftime('%d/%m/%Y') if self.fecha_emision else '',
            'fecha_vencimiento': self.fecha_vencimiento.strftime('%d/%m/%Y') if self.fecha_vencimiento else '',
            'monto':             float(self.monto),
            'estado':            self.estado,
            'proveedor_id':      self.proveedor_id,
            'proveedor':         self.proveedor.razon_social if self.proveedor else '',
            'pago_id':           self.pago_id,
            'fecha_entrega':     self.fecha_entrega.strftime('%d/%m/%Y') if self.fecha_entrega else '',
            'fecha_cobro':       self.fecha_cobro.strftime('%d/%m/%Y') if self.fecha_cobro else '',
            'observaciones':     self.observaciones,
        }


class ChequeTercero(db.Model):
    """Cheques recibidos de clientes — pueden endosarse a proveedores."""
    __tablename__ = 'cheque_tercero'

    id                   = db.Column(db.Integer, primary_key=True)
    banco                = db.Column(db.String(50), nullable=False)
    numero               = db.Column(db.String(20), nullable=False)
    librador             = db.Column(db.String(100))
    cuit_librador        = db.Column(db.String(13))
    fecha_emision        = db.Column(db.Date)
    fecha_vencimiento    = db.Column(db.Date)
    monto                = db.Column(Numeric(12, 2), nullable=False, default=0.00)
    estado               = db.Column(db.String(15), nullable=False, default='en_cartera')
    # estado: en_cartera / depositado / endosado / cobrado / rechazado
    cliente_origen_id    = db.Column(db.Integer)   # sin FK (se agrega cuando confirmemos tabla clientes)
    proveedor_destino_id = db.Column(db.Integer, db.ForeignKey('proveedor.id'))
    pago_id              = db.Column(db.Integer, db.ForeignKey('pago_proveedor.id'))
    fecha_recepcion      = db.Column(db.Date)
    fecha_endoso         = db.Column(db.Date)
    observaciones        = db.Column(db.Text)
    fecha_carga          = db.Column(db.DateTime, default=datetime.now)

    proveedor_destino = db.relationship('Proveedor',      foreign_keys=[proveedor_destino_id])
    pago              = db.relationship('PagoProveedor',  foreign_keys=[pago_id])

    def to_dict(self):
        return {
            'id':                   self.id,
            'banco':                self.banco,
            'numero':               self.numero,
            'librador':             self.librador,
            'cuit_librador':        self.cuit_librador,
            'fecha_emision':        self.fecha_emision.strftime('%d/%m/%Y') if self.fecha_emision else '',
            'fecha_vencimiento':    self.fecha_vencimiento.strftime('%d/%m/%Y') if self.fecha_vencimiento else '',
            'monto':                float(self.monto),
            'estado':               self.estado,
            'cliente_origen_id':    self.cliente_origen_id,
            'proveedor_destino_id': self.proveedor_destino_id,
            'proveedor_destino':    self.proveedor_destino.razon_social if self.proveedor_destino else '',
            'pago_id':              self.pago_id,
            'fecha_recepcion':      self.fecha_recepcion.strftime('%d/%m/%Y') if self.fecha_recepcion else '',
            'fecha_endoso':         self.fecha_endoso.strftime('%d/%m/%Y') if self.fecha_endoso else '',
            'observaciones':        self.observaciones,
        }


# ──────────────────────────────────────────────────────────────
# RUTAS – ABM PROVEEDORES
# ──────────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores')
def proveedores():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    lista = Proveedor.query.filter_by(activo=True).order_by(Proveedor.razon_social).all()
    return render_template('proveedores.html', proveedores=lista)


@proveedores_bp.route('/api/proveedores', methods=['GET'])
def api_proveedores():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        q = request.args.get('q', '').strip()
        query = Proveedor.query.filter_by(activo=True)
        if q:
            query = query.filter(
                or_(
                    Proveedor.razon_social.ilike(f'%{q}%'),
                    Proveedor.cuit.ilike(f'%{q}%')
                )
            )
        provs = query.order_by(Proveedor.razon_social).all()
        return jsonify({'success': True, 'proveedores': [p.to_dict() for p in provs]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/proveedor/nuevo', methods=['POST'])
def api_proveedor_nuevo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data = request.json
        prov = Proveedor(
            razon_social  = data['razon_social'].strip(),
            cuit          = data.get('cuit', '').strip() or None,
            condicion_iva = data.get('condicion_iva', 'Responsable Inscripto'),
            direccion     = data.get('direccion', '').strip() or None,
            telefono      = data.get('telefono', '').strip() or None,
            email         = data.get('email', '').strip() or None,
        )
        db.session.add(prov)
        db.session.commit()
        return jsonify({'success': True, 'proveedor': prov.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/proveedor/<int:prov_id>', methods=['GET'])
def api_proveedor_get(prov_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    prov = Proveedor.query.get_or_404(prov_id)
    return jsonify({'success': True, 'proveedor': prov.to_dict()})


@proveedores_bp.route('/api/proveedor/<int:prov_id>/editar', methods=['POST'])
def api_proveedor_editar(prov_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov = Proveedor.query.get_or_404(prov_id)
        data = request.json
        prov.razon_social  = data['razon_social'].strip()
        prov.cuit          = data.get('cuit', '').strip() or None
        prov.condicion_iva = data.get('condicion_iva', prov.condicion_iva)
        prov.direccion     = data.get('direccion', '').strip() or None
        prov.telefono      = data.get('telefono', '').strip() or None
        prov.email         = data.get('email', '').strip() or None
        db.session.commit()
        return jsonify({'success': True, 'proveedor': prov.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/proveedor/<int:prov_id>/eliminar', methods=['POST'])
def api_proveedor_eliminar(prov_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov = Proveedor.query.get_or_404(prov_id)
        prov.activo = False
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────
# RUTAS – FACTURAS DE COMPRA
# ──────────────────────────────────────────────────────────────

@proveedores_bp.route('/facturas_compra')
def facturas_compra():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.razon_social).all()
    return render_template('facturas_compra.html', proveedores=proveedores)


@proveedores_bp.route('/api/factura_compra/nueva', methods=['POST'])
def api_factura_compra_nueva():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data = request.json

        # Validar unicidad de comprobante por proveedor
        existente = FacturaCompra.query.filter_by(
            proveedor_id     = data['proveedor_id'],
            tipo_comprobante = data['tipo_comprobante'],
            punto_venta      = data['punto_venta'].zfill(4),
            numero           = data['numero'].zfill(8),
        ).first()
        if existente:
            return jsonify({'success': False, 'error': 'Ya existe esa factura para ese proveedor'}), 400

        neto21  = Decimal(str(data.get('neto_gravado_21',  0)))
        neto105 = Decimal(str(data.get('neto_gravado_105', 0)))
        iva21   = Decimal(str(data.get('iva_21',  0)))
        iva105  = Decimal(str(data.get('iva_105', 0)))
        neto_ng = Decimal(str(data.get('neto_no_gravado',  0)))
        otros   = Decimal(str(data.get('otros_impuestos',  0)))
        total   = neto21 + iva21 + neto105 + iva105 + neto_ng + otros

        fc = FacturaCompra(
            proveedor_id      = data['proveedor_id'],
            fecha             = datetime.strptime(data['fecha'], '%Y-%m-%d').date(),
            tipo_comprobante  = data['tipo_comprobante'],
            punto_venta       = data['punto_venta'].zfill(4),
            numero            = data['numero'].zfill(8),
            neto_gravado_21   = neto21,
            iva_21            = iva21,
            neto_gravado_105  = neto105,
            iva_105           = iva105,
            neto_no_gravado   = neto_ng,
            otros_impuestos   = otros,
            total             = total,
            observaciones     = data.get('observaciones', '').strip() or None,
            usuario_id        = session['user_id'],
        )
        db.session.add(fc)

        # Actualizar saldo del proveedor (le debemos más)
        prov = Proveedor.query.get(data['proveedor_id'])
        prov.saldo = Decimal(str(prov.saldo)) + total

        db.session.commit()
        return jsonify({'success': True, 'factura': fc.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/factura_compra/listar', methods=['GET'])
def api_factura_compra_listar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov_id = request.args.get('proveedor_id')
        desde   = request.args.get('desde')
        hasta   = request.args.get('hasta')

        query = FacturaCompra.query
        if prov_id:
            query = query.filter_by(proveedor_id=prov_id)
        if desde:
            query = query.filter(FacturaCompra.fecha >= datetime.strptime(desde, '%Y-%m-%d').date())
        if hasta:
            query = query.filter(FacturaCompra.fecha <= datetime.strptime(hasta, '%Y-%m-%d').date())

        facturas = query.order_by(FacturaCompra.fecha.desc()).all()
        return jsonify({'success': True, 'facturas': [f.to_dict() for f in facturas]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/factura_compra/<int:fc_id>/eliminar', methods=['POST'])
def api_factura_compra_eliminar(fc_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        fc = FacturaCompra.query.get_or_404(fc_id)
        # Revertir saldo
        prov = fc.proveedor
        prov.saldo = Decimal(str(prov.saldo)) - Decimal(str(fc.total))
        db.session.delete(fc)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────
# RUTAS – PAGOS A PROVEEDORES
# ──────────────────────────────────────────────────────────────

@proveedores_bp.route('/api/pago_proveedor/nuevo', methods=['POST'])
def api_pago_proveedor_nuevo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data   = request.json
        importe = Decimal(str(data['importe']))

        pago = PagoProveedor(
            proveedor_id  = data['proveedor_id'],
            fecha         = datetime.strptime(data['fecha'], '%Y-%m-%d').date(),
            importe       = importe,
            forma_pago    = data.get('forma_pago', 'efectivo'),
            referencia    = data.get('referencia', '').strip() or None,
            observaciones = data.get('observaciones', '').strip() or None,
            usuario_id    = session['user_id'],
        )
        db.session.add(pago)

        # Descontar del saldo del proveedor
        prov = Proveedor.query.get(data['proveedor_id'])
        prov.saldo = Decimal(str(prov.saldo)) - importe

        db.session.commit()
        return jsonify({'success': True, 'pago': pago.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/pago_proveedor/listar', methods=['GET'])
def api_pago_proveedor_listar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov_id = request.args.get('proveedor_id')
        query = PagoProveedor.query
        if prov_id:
            query = query.filter_by(proveedor_id=prov_id)
        pagos = query.order_by(PagoProveedor.fecha.desc()).all()
        return jsonify({'success': True, 'pagos': [p.to_dict() for p in pagos]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/pago_proveedor/<int:pago_id>/eliminar', methods=['POST'])
def api_pago_proveedor_eliminar(pago_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        pago = PagoProveedor.query.get_or_404(pago_id)
        prov = pago.proveedor
        # Revertir: si borro el pago, vuelvo a deberle
        prov.saldo = Decimal(str(prov.saldo)) + Decimal(str(pago.importe))
        db.session.delete(pago)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────
# RUTA – CUENTA CORRIENTE DE UN PROVEEDOR
# ──────────────────────────────────────────────────────────────

@proveedores_bp.route('/cta_cte_proveedor/<int:prov_id>')
def cta_cte_proveedor(prov_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    prov = Proveedor.query.get_or_404(prov_id)
    return render_template('cta_cte_proveedor.html', proveedor=prov)


@proveedores_bp.route('/api/cta_cte_proveedor/<int:prov_id>')
def api_cta_cte_proveedor(prov_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov = Proveedor.query.get_or_404(prov_id)

        facturas = FacturaCompra.query.filter_by(proveedor_id=prov_id)\
                        .order_by(FacturaCompra.fecha.asc()).all()
        pagos    = PagoProveedor.query.filter_by(proveedor_id=prov_id)\
                        .order_by(PagoProveedor.fecha.asc()).all()

        # Combinar y ordenar por fecha para armar el movimiento cronológico
        movimientos = []
        for f in facturas:
            movimientos.append({
                'tipo':        'factura',
                'fecha':       f.fecha,
                'descripcion': f'Factura {f.numero_completo()}',
                'debe':        float(f.total),
                'haber':       0.0,
                'id':          f.id,
            })
        for p in pagos:
            movimientos.append({
                'tipo':        'pago',
                'fecha':       p.fecha,
                'descripcion': f'Pago {p.forma_pago.capitalize()}' + (f' [{p.referencia}]' if p.referencia else ''),
                'debe':        0.0,
                'haber':       float(p.importe),
                'id':          p.id,
            })

        movimientos.sort(key=lambda x: x['fecha'])

        # Calcular saldo acumulado por movimiento
        saldo_acum = 0.0
        for m in movimientos:
            saldo_acum += m['debe'] - m['haber']
            m['saldo']  = round(saldo_acum, 2)
            m['fecha']  = m['fecha'].strftime('%d/%m/%Y')

        return jsonify({
            'success':       True,
            'proveedor':     prov.to_dict(),
            'movimientos':   movimientos,
            'saldo_actual':  float(prov.saldo),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────
# RUTA – IVA COMPRAS (listado + exportar PDF)
# ──────────────────────────────────────────────────────────────

@proveedores_bp.route('/iva_compras')
def iva_compras():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.razon_social).all()
    return render_template('iva_compras.html', proveedores=proveedores)


@proveedores_bp.route('/api/iva_compras', methods=['GET'])
def api_iva_compras():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        desde   = request.args.get('desde')
        hasta   = request.args.get('hasta')
        prov_id = request.args.get('proveedor_id')
        tipo    = request.args.get('tipo_comprobante')  # 'A', 'B', 'C' o vacío = todos

        if not desde or not hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'}), 400

        # Usamos SQL crudo para traer también los campos de percepciones/flete/descuento
        # que no están mapeados en el modelo SQLAlchemy de FacturaCompra.
        from sqlalchemy import text
        params = {
            'desde': datetime.strptime(desde, '%Y-%m-%d').date(),
            'hasta': datetime.strptime(hasta, '%Y-%m-%d').date(),
        }
        sql_where = "WHERE fc.fecha BETWEEN :desde AND :hasta"
        if prov_id:
            sql_where += " AND fc.proveedor_id = :prov"
            params['prov'] = prov_id
        if tipo and tipo.upper() in ('A', 'B', 'C'):
            sql_where += " AND UPPER(fc.tipo_comprobante) = :tipo"
            params['tipo'] = tipo.upper()

        sql = f"""
            SELECT fc.id, fc.fecha, fc.tipo_comprobante, fc.punto_venta, fc.numero,
                   fc.neto_gravado_21, fc.iva_21, fc.neto_gravado_105, fc.iva_105,
                   fc.neto_no_gravado, fc.otros_impuestos, fc.total,
                   COALESCE(fc.descuento, 0)             AS descuento,
                   COALESCE(fc.flete, 0)                  AS flete,
                   COALESCE(fc.percepcion_iva, 0)         AS percepcion_iva,
                   COALESCE(fc.percepcion_iibb, 0)        AS percepcion_iibb,
                   COALESCE(fc.percepcion_ganancias, 0)   AS percepcion_ganancias,
                   fc.cae,
                   p.razon_social AS proveedor_razon_social,
                   p.cuit         AS proveedor_cuit,
                   p.condicion_iva AS proveedor_cond_iva
              FROM factura_compra fc
              JOIN proveedor p ON p.id = fc.proveedor_id
              {sql_where}
             ORDER BY fc.fecha ASC, fc.id ASC
        """
        rows = db.session.execute(text(sql), params).mappings().all()

        def f(v):
            return float(v or 0)

        facturas = []
        t_neto21 = t_iva21 = t_neto105 = t_iva105 = 0.0
        t_neto_ng = t_otros = t_flete = t_desc = 0.0
        t_piva = t_piibb = t_pgan = 0.0
        t_general = 0.0

        # Subtotales por tipo de comprobante (A/B/C/otros)
        por_tipo = {}

        for r in rows:
            neto21  = f(r['neto_gravado_21'])
            iva21   = f(r['iva_21'])
            neto105 = f(r['neto_gravado_105'])
            iva105  = f(r['iva_105'])
            neto_ng = f(r['neto_no_gravado'])
            otros   = f(r['otros_impuestos'])
            flete   = f(r['flete'])
            desc    = f(r['descuento'])
            piva    = f(r['percepcion_iva'])
            piibb   = f(r['percepcion_iibb'])
            pgan    = f(r['percepcion_ganancias'])
            total   = f(r['total'])

            t_neto21 += neto21; t_iva21 += iva21
            t_neto105 += neto105; t_iva105 += iva105
            t_neto_ng += neto_ng; t_otros += otros
            t_flete += flete; t_desc += desc
            t_piva += piva; t_piibb += piibb; t_pgan += pgan
            t_general += total

            t_comp = (r['tipo_comprobante'] or '?').upper()
            por_tipo.setdefault(t_comp, {'cant': 0, 'total': 0.0, 'iva': 0.0})
            por_tipo[t_comp]['cant']  += 1
            por_tipo[t_comp]['total'] += total
            por_tipo[t_comp]['iva']   += iva21 + iva105

            pv = str(r['punto_venta'] or '').zfill(4)
            nro = str(r['numero'] or '').zfill(8)
            facturas.append({
                'id': r['id'],
                'fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
                'fecha_iso': r['fecha'].isoformat() if r['fecha'] else '',
                'tipo_comprobante': r['tipo_comprobante'] or '',
                'punto_venta': pv,
                'numero': nro,
                'numero_completo': f"{r['tipo_comprobante']} {pv}-{nro}",
                'proveedor': r['proveedor_razon_social'] or '',
                'proveedor_cuit': r['proveedor_cuit'] or '',
                'proveedor_cond_iva': r['proveedor_cond_iva'] or '',
                'neto_gravado_21': round(neto21, 2),
                'iva_21': round(iva21, 2),
                'neto_gravado_105': round(neto105, 2),
                'iva_105': round(iva105, 2),
                'neto_no_gravado': round(neto_ng, 2),
                'otros_impuestos': round(otros, 2),
                'flete': round(flete, 2),
                'descuento': round(desc, 2),
                'percepcion_iva': round(piva, 2),
                'percepcion_iibb': round(piibb, 2),
                'percepcion_ganancias': round(pgan, 2),
                'total': round(total, 2),
                'cae': r['cae'] or '',
            })

        return jsonify({
            'success':  True,
            'facturas': facturas,
            'totales': {
                'cantidad':         len(facturas),
                'neto_gravado_21':  round(t_neto21, 2),
                'iva_21':           round(t_iva21, 2),
                'neto_gravado_105': round(t_neto105, 2),
                'iva_105':          round(t_iva105, 2),
                'neto_no_gravado':  round(t_neto_ng, 2),
                'otros_impuestos':  round(t_otros, 2),
                'flete':            round(t_flete, 2),
                'descuento':        round(t_desc, 2),
                'percepcion_iva':   round(t_piva, 2),
                'percepcion_iibb':  round(t_piibb, 2),
                'percepcion_ganancias': round(t_pgan, 2),
                'iva_total':        round(t_iva21 + t_iva105, 2),
                'percepciones_total': round(t_piva + t_piibb + t_pgan, 2),
                'total':            round(t_general, 2),
            },
            'por_tipo': {k: {'cant': v['cant'],
                             'total': round(v['total'], 2),
                             'iva':   round(v['iva'], 2)}
                         for k, v in por_tipo.items()},
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/iva_compras/exportar_excel')
def exportar_excel_iva_compras():
    """
    Exporta el Libro IVA Compras a Excel (.xlsx) con formato profesional:
    - Encabezado con período, totales y subtotales por tipo de comprobante
    - Tabla detallada con todas las columnas (incluidas percepciones)
    - Formato de moneda argentina en columnas numéricas
    """
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        from sqlalchemy import text

        desde   = request.args.get('desde')
        hasta   = request.args.get('hasta')
        prov_id = request.args.get('proveedor_id')
        tipo    = request.args.get('tipo_comprobante')

        if not desde or not hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'}), 400

        params = {
            'desde': datetime.strptime(desde, '%Y-%m-%d').date(),
            'hasta': datetime.strptime(hasta, '%Y-%m-%d').date(),
        }
        sql_where = "WHERE fc.fecha BETWEEN :desde AND :hasta"
        if prov_id:
            sql_where += " AND fc.proveedor_id = :prov"
            params['prov'] = prov_id
        if tipo and tipo.upper() in ('A', 'B', 'C'):
            sql_where += " AND UPPER(fc.tipo_comprobante) = :tipo"
            params['tipo'] = tipo.upper()

        sql = f"""
            SELECT fc.fecha, fc.tipo_comprobante, fc.punto_venta, fc.numero,
                   fc.neto_gravado_21, fc.iva_21, fc.neto_gravado_105, fc.iva_105,
                   fc.neto_no_gravado, fc.otros_impuestos, fc.total,
                   COALESCE(fc.percepcion_iva, 0)         AS percepcion_iva,
                   COALESCE(fc.percepcion_iibb, 0)        AS percepcion_iibb,
                   COALESCE(fc.percepcion_ganancias, 0)   AS percepcion_ganancias,
                   fc.cae,
                   p.razon_social AS proveedor_razon_social,
                   p.cuit         AS proveedor_cuit,
                   p.condicion_iva AS proveedor_cond_iva
              FROM factura_compra fc
              JOIN proveedor p ON p.id = fc.proveedor_id
              {sql_where}
             ORDER BY fc.fecha ASC, fc.id ASC
        """
        rows = db.session.execute(text(sql), params).mappings().all()

        # ─── Workbook ───
        wb = Workbook()
        ws = wb.active
        ws.title = "IVA Compras"

        # Estilos
        header_fill  = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
        header_font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        title_font   = Font(name='Calibri', size=14, bold=True, color='0D6EFD')
        subtit_font  = Font(name='Calibri', size=10, bold=True, color='6C757D')
        total_fill   = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        total_font   = Font(name='Calibri', size=10, bold=True)
        border_thin  = Border(left=Side(style='thin', color='CCCCCC'),
                              right=Side(style='thin', color='CCCCCC'),
                              top=Side(style='thin', color='CCCCCC'),
                              bottom=Side(style='thin', color='CCCCCC'))
        money_fmt    = '"$"#,##0.00'

        # ─── Título y período ───
        ws.merge_cells('A1:R1')
        ws['A1'] = 'LIBRO IVA COMPRAS'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:R2')
        ws['A2'] = f"Período: {desde}  a  {hasta}"
        ws['A2'].font = subtit_font
        ws['A2'].alignment = Alignment(horizontal='center')

        # ─── Encabezados ───
        headers = [
            'Fecha', 'Tipo', 'PV', 'Número', 'Proveedor', 'CUIT', 'Cond. IVA',
            'Neto 21%', 'IVA 21%', 'Neto 10,5%', 'IVA 10,5%',
            'No Gravado', 'Otros Imp.',
            'Perc. IVA', 'Perc. IIBB', 'Perc. Gcias.',
            'Total', 'CAE'
        ]
        row_header = 4
        for col_i, h in enumerate(headers, 1):
            c = ws.cell(row=row_header, column=col_i, value=h)
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border_thin

        # ─── Filas ───
        money_cols = {8, 9, 10, 11, 12, 13, 14, 15, 16, 17}  # columnas de moneda

        t_neto21 = t_iva21 = t_neto105 = t_iva105 = 0.0
        t_ng = t_otros = t_piva = t_piibb = t_pgan = t_total = 0.0

        for i, r in enumerate(rows, start=row_header + 1):
            neto21  = float(r['neto_gravado_21']  or 0)
            iva21   = float(r['iva_21']           or 0)
            neto105 = float(r['neto_gravado_105'] or 0)
            iva105  = float(r['iva_105']          or 0)
            ng      = float(r['neto_no_gravado']  or 0)
            otros   = float(r['otros_impuestos']  or 0)
            piva    = float(r['percepcion_iva']   or 0)
            piibb   = float(r['percepcion_iibb']  or 0)
            pgan    = float(r['percepcion_ganancias'] or 0)
            total   = float(r['total']            or 0)

            t_neto21 += neto21; t_iva21 += iva21
            t_neto105 += neto105; t_iva105 += iva105
            t_ng += ng; t_otros += otros
            t_piva += piva; t_piibb += piibb; t_pgan += pgan
            t_total += total

            valores = [
                r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '',
                r['tipo_comprobante'] or '',
                str(r['punto_venta'] or '').zfill(4),
                str(r['numero'] or '').zfill(8),
                r['proveedor_razon_social'] or '',
                r['proveedor_cuit'] or '',
                r['proveedor_cond_iva'] or '',
                neto21, iva21, neto105, iva105, ng, otros,
                piva, piibb, pgan, total,
                r['cae'] or ''
            ]
            for col_i, v in enumerate(valores, 1):
                c = ws.cell(row=i, column=col_i, value=v)
                c.border = border_thin
                if col_i in money_cols:
                    c.number_format = money_fmt
                if i % 2 == 0:
                    c.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

        # ─── Fila de totales ───
        row_total = row_header + 1 + len(rows)
        ws.cell(row=row_total, column=1, value='TOTALES').font = total_font
        tot_row_vals = {
            8: t_neto21, 9: t_iva21, 10: t_neto105, 11: t_iva105,
            12: t_ng, 13: t_otros,
            14: t_piva, 15: t_piibb, 16: t_pgan, 17: t_total
        }
        for col_i, val in tot_row_vals.items():
            c = ws.cell(row=row_total, column=col_i, value=val)
            c.font = total_font; c.fill = total_fill
            c.number_format = money_fmt; c.border = border_thin

        # ─── Hoja 2: Resumen por tipo de comprobante ───
        ws2 = wb.create_sheet('Resumen')
        ws2['A1'] = 'RESUMEN POR TIPO DE COMPROBANTE'
        ws2['A1'].font = title_font
        ws2.merge_cells('A1:D1')
        ws2.cell(row=3, column=1, value='Tipo').font = header_font
        ws2.cell(row=3, column=2, value='Cantidad').font = header_font
        ws2.cell(row=3, column=3, value='IVA Total').font = header_font
        ws2.cell(row=3, column=4, value='Total').font = header_font
        for c in range(1, 5):
            ws2.cell(row=3, column=c).fill = header_fill

        por_tipo = {}
        for r in rows:
            t = (r['tipo_comprobante'] or '?').upper()
            por_tipo.setdefault(t, {'cant': 0, 'total': 0.0, 'iva': 0.0})
            por_tipo[t]['cant']  += 1
            por_tipo[t]['total'] += float(r['total'] or 0)
            por_tipo[t]['iva']   += float(r['iva_21'] or 0) + float(r['iva_105'] or 0)

        r_i = 4
        for k, v in sorted(por_tipo.items()):
            ws2.cell(row=r_i, column=1, value=f'Factura {k}')
            ws2.cell(row=r_i, column=2, value=v['cant'])
            ws2.cell(row=r_i, column=3, value=v['iva']).number_format  = money_fmt
            ws2.cell(row=r_i, column=4, value=v['total']).number_format = money_fmt
            r_i += 1

        # ─── Anchos de columna ───
        widths = [11, 6, 7, 11, 30, 14, 14, 13, 13, 13, 13, 13, 12, 12, 12, 12, 14, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 16
        ws2.column_dimensions['D'].width = 16

        # Freeze panes (mantener encabezado visible al scrollear)
        ws.freeze_panes = f'A{row_header + 1}'

        # ─── Enviar archivo ───
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"iva_compras_{desde}_{hasta}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/iva_compras/exportar_pdf')
def exportar_pdf_iva_compras():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        from reporte_iva_compras_pdf import generar_pdf_iva_compras

        desde   = request.args.get('desde')
        hasta   = request.args.get('hasta')
        prov_id = request.args.get('proveedor_id')

        if not desde or not hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'}), 400

        query = FacturaCompra.query.filter(
            FacturaCompra.fecha >= datetime.strptime(desde, '%Y-%m-%d').date(),
            FacturaCompra.fecha <= datetime.strptime(hasta, '%Y-%m-%d').date(),
        )
        if prov_id:
            query = query.filter_by(proveedor_id=prov_id)

        facturas = query.order_by(FacturaCompra.fecha.asc()).all()

        datos = []
        for f in facturas:
            datos.append({
                'fecha':            f.fecha.strftime('%d/%m/%Y'),
                'proveedor':        f.proveedor.razon_social,
                'cuit':             f.proveedor.cuit or '',
                'tipo':             f.tipo_comprobante,
                'numero':           f.numero_completo(),
                'neto_gravado_21':  float(f.neto_gravado_21),
                'iva_21':           float(f.iva_21),
                'neto_gravado_105': float(f.neto_gravado_105),
                'iva_105':          float(f.iva_105),
                'neto_no_gravado':  float(f.neto_no_gravado),
                'otros_impuestos':  float(f.otros_impuestos),
                'total':            float(f.total),
            })

        pdf_bytes = generar_pdf_iva_compras(datos, desde, hasta)
        nombre    = f'IVA_Compras_{desde}_{hasta}.pdf'

        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre,
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════
# RUTAS — CHEQUES PROPIOS (Fase 2a — Ciclo de Compras)
# ══════════════════════════════════════════════════════════════
# Agregar este bloque AL FINAL de proveedores.py, después de la ruta
# de exportar_pdf_iva_compras y antes de cualquier otra cosa.
# ══════════════════════════════════════════════════════════════

@proveedores_bp.route('/cheques_propios')
def cheques_propios():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.razon_social).all()
    return render_template('cheques_propios.html', proveedores=proveedores)


@proveedores_bp.route('/api/cheques_propios/listar', methods=['GET'])
def api_cheques_propios_listar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        estado     = request.args.get('estado', '').strip()
        vto_desde  = request.args.get('vto_desde', '').strip()
        vto_hasta  = request.args.get('vto_hasta', '').strip()
        q          = request.args.get('q', '').strip()

        query = ChequePropio.query

        if estado:
            query = query.filter(ChequePropio.estado == estado)
        if vto_desde:
            query = query.filter(ChequePropio.fecha_vencimiento >=
                                 datetime.strptime(vto_desde, '%Y-%m-%d').date())
        if vto_hasta:
            query = query.filter(ChequePropio.fecha_vencimiento <=
                                 datetime.strptime(vto_hasta, '%Y-%m-%d').date())
        if q:
            query = query.filter(
                or_(
                    ChequePropio.banco.ilike(f'%{q}%'),
                    ChequePropio.numero.ilike(f'%{q}%')
                )
            )

        cheques = query.order_by(
            ChequePropio.fecha_vencimiento.is_(None).asc(),
            ChequePropio.fecha_vencimiento.asc(),
            ChequePropio.id.desc()
        ).all()

        return jsonify({'success': True, 'cheques': [c.to_dict() for c in cheques]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_propio/<int:ch_id>', methods=['GET'])
def api_cheque_propio_get(ch_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    ch = ChequePropio.query.get_or_404(ch_id)
    return jsonify({'success': True, 'cheque': ch.to_dict()})


@proveedores_bp.route('/api/cheque_propio/nuevo', methods=['POST'])
def api_cheque_propio_nuevo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data = request.json

        # Validaciones básicas
        if not data.get('banco', '').strip():
            return jsonify({'success': False, 'error': 'El banco es obligatorio'}), 400
        if not data.get('numero', '').strip():
            return jsonify({'success': False, 'error': 'El número de cheque es obligatorio'}), 400
        if not data.get('fecha_emision'):
            return jsonify({'success': False, 'error': 'La fecha de emisión es obligatoria'}), 400

        banco  = data['banco'].strip()
        numero = data['numero'].strip()

        # Unicidad banco+número
        existente = ChequePropio.query.filter_by(banco=banco, numero=numero).first()
        if existente:
            return jsonify({
                'success': False,
                'error': f'Ya existe un cheque con ese número para {banco}'
            }), 400

        # Estado (default pendiente)
        estado_valido = ['pendiente', 'entregado', 'cobrado', 'rechazado', 'anulado']
        estado = data.get('estado', 'pendiente')
        if estado not in estado_valido:
            return jsonify({'success': False, 'error': f'Estado inválido: {estado}'}), 400

        # Si se crea directamente como "entregado" a mano, requerimos proveedor
        proveedor_id = data.get('proveedor_id') or None
        if estado == 'entregado' and not proveedor_id:
            return jsonify({
                'success': False,
                'error': 'Si el estado es "entregado" hay que indicar a qué proveedor se entregó'
            }), 400

        ch = ChequePropio(
            banco             = banco,
            cuenta            = data.get('cuenta', '').strip() or None,
            numero            = numero,
            fecha_emision     = datetime.strptime(data['fecha_emision'], '%Y-%m-%d').date(),
            fecha_vencimiento = datetime.strptime(data['fecha_vencimiento'], '%Y-%m-%d').date()
                                if data.get('fecha_vencimiento') else None,
            monto             = Decimal(str(data.get('monto', 0))),
            estado            = estado,
            proveedor_id      = proveedor_id,
            pago_id           = None,  # Al crear a mano NUNCA se vincula a un pago
            fecha_entrega     = datetime.strptime(data['fecha_entrega'], '%Y-%m-%d').date()
                                if data.get('fecha_entrega') else None,
            fecha_cobro       = datetime.strptime(data['fecha_cobro'], '%Y-%m-%d').date()
                                if data.get('fecha_cobro') else None,
            observaciones     = data.get('observaciones', '').strip() or None,
        )
        db.session.add(ch)
        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_propio/<int:ch_id>/editar', methods=['POST'])
def api_cheque_propio_editar(ch_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequePropio.query.get_or_404(ch_id)
        data = request.json

        # Si el cheque tiene pago_id (fue usado en la pantalla de pago),
        # solo se pueden editar fechas de cobro/entrega y observaciones.
        if ch.pago_id:
            ch.fecha_cobro = datetime.strptime(data['fecha_cobro'], '%Y-%m-%d').date() \
                             if data.get('fecha_cobro') else ch.fecha_cobro
            ch.observaciones = data.get('observaciones', '').strip() or None
            db.session.commit()
            return jsonify({
                'success': True,
                'cheque': ch.to_dict(),
                'warning': 'Cheque vinculado a un pago: solo se editaron fecha de cobro y observaciones.'
            })

        # Cheque sin pago: se puede editar todo
        banco  = data.get('banco', ch.banco).strip()
        numero = data.get('numero', ch.numero).strip()

        # Unicidad banco+número (excluyéndose a sí mismo)
        existente = ChequePropio.query.filter(
            ChequePropio.banco == banco,
            ChequePropio.numero == numero,
            ChequePropio.id != ch_id
        ).first()
        if existente:
            return jsonify({
                'success': False,
                'error': f'Ya existe otro cheque con ese número para {banco}'
            }), 400

        ch.banco             = banco
        ch.cuenta            = data.get('cuenta', '').strip() or None
        ch.numero            = numero
        ch.fecha_emision     = datetime.strptime(data['fecha_emision'], '%Y-%m-%d').date() \
                               if data.get('fecha_emision') else ch.fecha_emision
        ch.fecha_vencimiento = datetime.strptime(data['fecha_vencimiento'], '%Y-%m-%d').date() \
                               if data.get('fecha_vencimiento') else None
        ch.monto             = Decimal(str(data.get('monto', ch.monto)))
        ch.proveedor_id      = data.get('proveedor_id') or None
        ch.fecha_entrega     = datetime.strptime(data['fecha_entrega'], '%Y-%m-%d').date() \
                               if data.get('fecha_entrega') else None
        ch.fecha_cobro       = datetime.strptime(data['fecha_cobro'], '%Y-%m-%d').date() \
                               if data.get('fecha_cobro') else None
        ch.observaciones     = data.get('observaciones', '').strip() or None

        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_propio/<int:ch_id>/estado', methods=['POST'])
def api_cheque_propio_estado(ch_id):
    """
    Cambio rápido de estado del cheque. Valida transiciones razonables.
    Body: { "estado": "cobrado"/"rechazado"/"anulado"/"entregado",
            "fecha": "YYYY-MM-DD" (opcional, para cobro/entrega),
            "proveedor_id": int (obligatorio si estado=entregado) }
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequePropio.query.get_or_404(ch_id)
        data = request.json
        nuevo = data.get('estado', '')

        validos = ['pendiente', 'entregado', 'cobrado', 'rechazado', 'anulado']
        if nuevo not in validos:
            return jsonify({'success': False, 'error': f'Estado inválido: {nuevo}'}), 400

        # Transiciones permitidas
        transiciones = {
            'pendiente': ['entregado', 'anulado'],
            'entregado': ['cobrado', 'rechazado'],
            'cobrado':   [],   # estado final
            'rechazado': ['entregado'],  # rehacer si se vuelve a entregar el mismo físicamente (raro)
            'anulado':   [],   # estado final
        }
        if nuevo not in transiciones.get(ch.estado, []):
            return jsonify({
                'success': False,
                'error': f'No se puede pasar de "{ch.estado}" a "{nuevo}".'
            }), 400

        # Validaciones específicas
        if nuevo == 'entregado':
            proveedor_id = data.get('proveedor_id') or None
            if not proveedor_id:
                return jsonify({
                    'success': False,
                    'error': 'Para marcar "entregado" hay que indicar a qué proveedor se entregó.'
                }), 400
            ch.proveedor_id  = proveedor_id
            ch.fecha_entrega = datetime.strptime(data['fecha'], '%Y-%m-%d').date() \
                               if data.get('fecha') else datetime.now().date()

        if nuevo == 'cobrado':
            ch.fecha_cobro = datetime.strptime(data['fecha'], '%Y-%m-%d').date() \
                             if data.get('fecha') else datetime.now().date()

        ch.estado = nuevo
        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_propio/<int:ch_id>/eliminar', methods=['POST'])
def api_cheque_propio_eliminar(ch_id):
    """
    Eliminar cheque — solo se permite si estado='pendiente' o 'anulado'
    y no tiene pago_id. Cheques entregados/cobrados/rechazados NO se borran
    (se marcan anulados si hiciera falta).
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequePropio.query.get_or_404(ch_id)

        if ch.pago_id:
            return jsonify({
                'success': False,
                'error': 'No se puede eliminar: el cheque está vinculado a un pago.'
            }), 400

        if ch.estado not in ('pendiente', 'anulado'):
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar un cheque en estado "{ch.estado}". '
                         'Solo se pueden eliminar cheques pendientes o anulados.'
            }), 400

        db.session.delete(ch)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════
# RUTAS — CHEQUES DE TERCEROS (Fase 2b — Ciclo de Compras)
# ══════════════════════════════════════════════════════════════
# Agregar este bloque AL FINAL de proveedores.py, después del bloque
# de cheques_propios.
#
# Pre-requisito: haber corrido schiro_ctacte_proveedores_2b_fk.sql
# (que agrega la FK cheque_tercero.cliente_origen_id → cliente.id)
# ══════════════════════════════════════════════════════════════

@proveedores_bp.route('/cheques_terceros')
def cheques_terceros():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.razon_social).all()
    return render_template('cheques_terceros.html', proveedores=proveedores)


@proveedores_bp.route('/api/cheques_terceros/listar', methods=['GET'])
def api_cheques_terceros_listar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        from sqlalchemy import text

        estado     = request.args.get('estado', '').strip()
        vto_desde  = request.args.get('vto_desde', '').strip()
        vto_hasta  = request.args.get('vto_hasta', '').strip()
        q          = request.args.get('q', '').strip()

        query = ChequeTercero.query

        if estado:
            query = query.filter(ChequeTercero.estado == estado)
        if vto_desde:
            query = query.filter(ChequeTercero.fecha_vencimiento >=
                                 datetime.strptime(vto_desde, '%Y-%m-%d').date())
        if vto_hasta:
            query = query.filter(ChequeTercero.fecha_vencimiento <=
                                 datetime.strptime(vto_hasta, '%Y-%m-%d').date())
        if q:
            query = query.filter(
                or_(
                    ChequeTercero.banco.ilike(f'%{q}%'),
                    ChequeTercero.numero.ilike(f'%{q}%'),
                    ChequeTercero.librador.ilike(f'%{q}%')
                )
            )

        cheques = query.order_by(
            ChequeTercero.fecha_vencimiento.is_(None).asc(),
            ChequeTercero.fecha_vencimiento.asc(),
            ChequeTercero.id.desc()
        ).all()

        # Enriquecer cada cheque con el nombre del cliente origen
        # (no hay relationship backref — hacemos lookup puntual vía SQL)
        result = []
        for c in cheques:
            d = c.to_dict()
            if c.cliente_origen_id:
                row = db.session.execute(
                    text("SELECT nombre FROM cliente WHERE id = :id"),
                    {"id": c.cliente_origen_id}
                ).fetchone()
                d['cliente_origen_nombre'] = row[0] if row else ''
            else:
                d['cliente_origen_nombre'] = ''
            result.append(d)

        return jsonify({'success': True, 'cheques': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_tercero/<int:ch_id>', methods=['GET'])
def api_cheque_tercero_get(ch_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from sqlalchemy import text
    ch = ChequeTercero.query.get_or_404(ch_id)
    d = ch.to_dict()
    if ch.cliente_origen_id:
        row = db.session.execute(
            text("SELECT nombre FROM cliente WHERE id = :id"),
            {"id": ch.cliente_origen_id}
        ).fetchone()
        d['cliente_origen_nombre'] = row[0] if row else ''
    else:
        d['cliente_origen_nombre'] = ''
    return jsonify({'success': True, 'cheque': d})


@proveedores_bp.route('/api/cheque_tercero/nuevo', methods=['POST'])
def api_cheque_tercero_nuevo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data = request.json

        if not data.get('banco', '').strip():
            return jsonify({'success': False, 'error': 'El banco es obligatorio'}), 400
        if not data.get('numero', '').strip():
            return jsonify({'success': False, 'error': 'El número de cheque es obligatorio'}), 400
        if not data.get('monto') or Decimal(str(data['monto'])) <= 0:
            return jsonify({'success': False, 'error': 'El monto debe ser mayor a 0'}), 400

        banco  = data['banco'].strip()
        numero = data['numero'].strip()

        # Unicidad banco+número
        existente = ChequeTercero.query.filter_by(banco=banco, numero=numero).first()
        if existente:
            return jsonify({
                'success': False,
                'error': f'Ya existe un cheque de terceros con número {numero} del banco {banco}'
            }), 400

        estados_validos = ['en_cartera', 'depositado', 'endosado', 'cobrado', 'rechazado']
        estado = data.get('estado', 'en_cartera')
        if estado not in estados_validos:
            return jsonify({'success': False, 'error': f'Estado inválido: {estado}'}), 400

        # Endoso manual: requiere proveedor destino
        proveedor_destino_id = data.get('proveedor_destino_id') or None
        if estado == 'endosado' and not proveedor_destino_id:
            return jsonify({
                'success': False,
                'error': 'Si el estado es "endosado" hay que indicar a qué proveedor se endosó.'
            }), 400

        ch = ChequeTercero(
            banco                = banco,
            numero               = numero,
            librador             = data.get('librador', '').strip() or None,
            cuit_librador        = data.get('cuit_librador', '').strip() or None,
            fecha_emision        = datetime.strptime(data['fecha_emision'], '%Y-%m-%d').date()
                                   if data.get('fecha_emision') else None,
            fecha_vencimiento    = datetime.strptime(data['fecha_vencimiento'], '%Y-%m-%d').date()
                                   if data.get('fecha_vencimiento') else None,
            monto                = Decimal(str(data['monto'])),
            estado               = estado,
            cliente_origen_id    = data.get('cliente_origen_id') or None,
            proveedor_destino_id = proveedor_destino_id,
            pago_id              = None,   # Alta manual NUNCA se vincula a un pago
            fecha_recepcion      = datetime.strptime(data['fecha_recepcion'], '%Y-%m-%d').date()
                                   if data.get('fecha_recepcion') else datetime.now().date(),
            fecha_endoso         = datetime.strptime(data['fecha_endoso'], '%Y-%m-%d').date()
                                   if data.get('fecha_endoso') else None,
            observaciones        = data.get('observaciones', '').strip() or None,
        )
        db.session.add(ch)
        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_tercero/<int:ch_id>/editar', methods=['POST'])
def api_cheque_tercero_editar(ch_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequeTercero.query.get_or_404(ch_id)
        data = request.json

        # Si el cheque fue usado en un pago, solo permitir editar observaciones
        if ch.pago_id:
            ch.observaciones = data.get('observaciones', '').strip() or None
            db.session.commit()
            return jsonify({
                'success': True,
                'cheque': ch.to_dict(),
                'warning': 'Cheque vinculado a un pago: solo se editaron observaciones.'
            })

        banco  = data.get('banco',  ch.banco).strip()
        numero = data.get('numero', ch.numero).strip()

        # Unicidad banco+número (excluyéndose a sí mismo)
        existente = ChequeTercero.query.filter(
            ChequeTercero.banco == banco,
            ChequeTercero.numero == numero,
            ChequeTercero.id != ch_id
        ).first()
        if existente:
            return jsonify({
                'success': False,
                'error': f'Ya existe otro cheque de terceros con número {numero} del banco {banco}'
            }), 400

        ch.banco             = banco
        ch.numero            = numero
        ch.librador          = data.get('librador', '').strip() or None
        ch.cuit_librador     = data.get('cuit_librador', '').strip() or None
        ch.fecha_emision     = datetime.strptime(data['fecha_emision'], '%Y-%m-%d').date() \
                               if data.get('fecha_emision') else None
        ch.fecha_vencimiento = datetime.strptime(data['fecha_vencimiento'], '%Y-%m-%d').date() \
                               if data.get('fecha_vencimiento') else None
        ch.monto             = Decimal(str(data.get('monto', ch.monto)))
        ch.cliente_origen_id = data.get('cliente_origen_id') or None
        ch.fecha_recepcion   = datetime.strptime(data['fecha_recepcion'], '%Y-%m-%d').date() \
                               if data.get('fecha_recepcion') else ch.fecha_recepcion
        ch.observaciones     = data.get('observaciones', '').strip() or None

        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_tercero/<int:ch_id>/estado', methods=['POST'])
def api_cheque_tercero_estado(ch_id):
    """
    Cambio de estado del cheque de tercero. Valida transiciones.
    Body según acción:
      { "estado": "depositado", "fecha": "YYYY-MM-DD",
        "banco_deposito": "Galicia", "cuenta_deposito": "12345" }
      { "estado": "endosado", "fecha": "YYYY-MM-DD", "proveedor_destino_id": 5 }
      { "estado": "cobrado",   "fecha": "YYYY-MM-DD" }
      { "estado": "rechazado", "fecha": "YYYY-MM-DD" }
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequeTercero.query.get_or_404(ch_id)
        data = request.json
        nuevo = data.get('estado', '')

        validos = ['en_cartera', 'depositado', 'endosado', 'cobrado', 'rechazado']
        if nuevo not in validos:
            return jsonify({'success': False, 'error': f'Estado inválido: {nuevo}'}), 400

        # Transiciones permitidas
        transiciones = {
            'en_cartera': ['depositado', 'endosado', 'rechazado'],
            'depositado': ['cobrado', 'rechazado'],
            'endosado':   ['rechazado'],
            'cobrado':    [],
            'rechazado':  ['en_cartera'],
        }
        if nuevo not in transiciones.get(ch.estado, []):
            return jsonify({
                'success': False,
                'error': f'No se puede pasar de "{ch.estado}" a "{nuevo}".'
            }), 400

        if nuevo == 'depositado':
            banco_dep  = (data.get('banco_deposito')  or '').strip()
            cuenta_dep = (data.get('cuenta_deposito') or '').strip()
            if not banco_dep or not cuenta_dep:
                return jsonify({
                    'success': False,
                    'error': 'Para depositar hay que indicar banco y cuenta de destino.'
                }), 400
            obs_dep = f"DEPOSITADO en {banco_dep} - Cta {cuenta_dep}"
            ch.observaciones = ((ch.observaciones + ' | ') if ch.observaciones else '') + obs_dep

        if nuevo == 'endosado':
            proveedor_destino_id = data.get('proveedor_destino_id') or None
            if not proveedor_destino_id:
                return jsonify({
                    'success': False,
                    'error': 'Para endosar hay que indicar el proveedor destino.'
                }), 400
            ch.proveedor_destino_id = proveedor_destino_id
            ch.fecha_endoso = datetime.strptime(data['fecha'], '%Y-%m-%d').date() \
                              if data.get('fecha') else datetime.now().date()

        ch.estado = nuevo
        db.session.commit()
        return jsonify({'success': True, 'cheque': ch.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheque_tercero/<int:ch_id>/eliminar', methods=['POST'])
def api_cheque_tercero_eliminar(ch_id):
    """
    Eliminar cheque de tercero — solo si está en_cartera y no tiene pago_id.
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        ch = ChequeTercero.query.get_or_404(ch_id)

        if ch.pago_id:
            return jsonify({
                'success': False,
                'error': 'No se puede eliminar: el cheque está vinculado a un pago.'
            }), 400

        if ch.estado != 'en_cartera':
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar un cheque en estado "{ch.estado}". '
                         'Solo se pueden eliminar cheques en cartera sin movimientos posteriores.'
            }), 400

        db.session.delete(ch)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500      


# ══════════════════════════════════════════════════════════════
# APIs AUXILIARES — Pago a Proveedor (Fase 3a — Ciclo de Compras)
# ══════════════════════════════════════════════════════════════
# Agregar este bloque AL FINAL de proveedores.py, después de los
# bloques de cheques_propios y cheques_terceros.
#
# Estas APIs son las que consumirá la pantalla /pagar_proveedor
# para armar el formulario de pago.
# ══════════════════════════════════════════════════════════════

@proveedores_bp.route('/api/proveedor/<int:prov_id>/facturas_pendientes', methods=['GET'])
def api_proveedor_facturas_pendientes(prov_id):
    """
    Devuelve facturas de compra del proveedor con saldo pendiente > 0.
    Ordenadas por fecha de vencimiento (las más urgentes primero), y las
    que no tienen vto al final.
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        prov = Proveedor.query.get_or_404(prov_id)

        facturas = FacturaCompra.query.filter(
            FacturaCompra.proveedor_id == prov_id,
            FacturaCompra.saldo_pendiente > 0,
            FacturaCompra.estado != 'anulada'
        ).order_by(
            FacturaCompra.fecha_vencimiento.is_(None).asc(),
            FacturaCompra.fecha_vencimiento.asc(),
            FacturaCompra.fecha.asc()
        ).all()

        # Enriquecemos con dias_para_vencer para que el front pueda colorear
        hoy = datetime.now().date()
        result = []
        for f in facturas:
            d = f.to_dict()
            if f.fecha_vencimiento:
                d['dias_para_vencer'] = (f.fecha_vencimiento - hoy).days
            else:
                d['dias_para_vencer'] = None
            result.append(d)

        return jsonify({
            'success': True,
            'proveedor': prov.to_dict(),
            'facturas': result,
            'total_pendiente': float(sum(f.saldo_pendiente for f in facturas)),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheques_propios/disponibles', methods=['GET'])
def api_cheques_propios_disponibles():
    """
    Devuelve cheques propios en estado 'pendiente' (emitidos pero no entregados).
    Estos son los que se pueden USAR en un pago a proveedor.
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        cheques = ChequePropio.query.filter_by(estado='pendiente')\
            .order_by(ChequePropio.fecha_vencimiento.is_(None).asc(),
                      ChequePropio.fecha_vencimiento.asc(),
                      ChequePropio.id.asc()).all()
        return jsonify({'success': True, 'cheques': [c.to_dict() for c in cheques]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@proveedores_bp.route('/api/cheques_terceros/disponibles', methods=['GET'])
def api_cheques_terceros_disponibles():
    """
    Devuelve cheques de terceros en estado 'en_cartera' (disponibles para endosar).
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        from sqlalchemy import text

        cheques = ChequeTercero.query.filter_by(estado='en_cartera')\
            .order_by(ChequeTercero.fecha_vencimiento.is_(None).asc(),
                      ChequeTercero.fecha_vencimiento.asc(),
                      ChequeTercero.id.asc()).all()

        # Enriquecer con nombre del cliente origen
        result = []
        for c in cheques:
            d = c.to_dict()
            if c.cliente_origen_id:
                row = db.session.execute(
                    text("SELECT nombre FROM cliente WHERE id = :id"),
                    {"id": c.cliente_origen_id}
                ).fetchone()
                d['cliente_origen_nombre'] = row[0] if row else ''
            else:
                d['cliente_origen_nombre'] = ''
            result.append(d)

        return jsonify({'success': True, 'cheques': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500# ══════════════════════════════════════════════════════════════════════════
# RUTAS — Módulo PAGAR PROVEEDOR (Ciclo de Compras Fase 3)
# ══════════════════════════════════════════════════════════════════════════
# PEGAR ESTE BLOQUE COMPLETO AL FINAL DE proveedores.py
#
# Usa los modelos que ya están declarados al principio del archivo:
#   Proveedor, FacturaCompra, PagoProveedor, PagoProveedorNumerador,
#   PagoProveedorDetalle, PagoProveedorMedio, ProveedorMovimiento,
#   ChequePropio, ChequeTercero
#
# Y el blueprint proveedores_bp que ya existe.
# ══════════════════════════════════════════════════════════════════════════

from flask import send_file
from io import BytesIO


# --------------------------------------------------------------------------
# Helpers internos (privados al módulo)
# --------------------------------------------------------------------------
def _tipo_medio_normalizado(tipo):
    """
    El modelo PagoProveedorMedio.medio admite:
      efectivo / transferencia / cheque_propio / cheque_tercero / otro
    El front manda 'cheque_propio_nuevo' y 'cheque_propio_existente' por
    separado — acá los unificamos en 'cheque_propio'.
    """
    if tipo in ('cheque_propio_nuevo', 'cheque_propio_existente'):
        return 'cheque_propio'
    if tipo == 'cheque_tercero_existente':
        return 'cheque_tercero'
    return tipo or 'otro'


def _resumen_medios(medios):
    """String corto para forma_pago. Ej: 'Efectivo + Cheque Propio'."""
    nombres_map = {
        'efectivo':       'Efectivo',
        'transferencia':  'Transferencia',
        'cheque_propio':  'Cheque Propio',
        'cheque_tercero': 'Cheque Tercero',
    }
    tipos = []
    for m in medios:
        t = _tipo_medio_normalizado(m.get('tipo'))
        nombre = nombres_map.get(t, t.replace('_', ' ').title())
        if nombre not in tipos:
            tipos.append(nombre)
    return ' + '.join(tipos) if tipos else 'efectivo'


def _resumen_medios_obj(medios_list):
    """Versión para objetos PagoProveedorMedio (del listado)."""
    nombres_map = {
        'efectivo':       'Efectivo',
        'transferencia':  'Transferencia',
        'cheque_propio':  'Cheque Propio',
        'cheque_tercero': 'Cheque Tercero',
    }
    tipos = []
    for m in medios_list:
        nombre = nombres_map.get(m.medio, (m.medio or '').replace('_', ' ').title())
        if nombre not in tipos:
            tipos.append(nombre)
    return ' + '.join(tipos) if tipos else 'efectivo'


# ══════════════════════════════════════════════════════════════════════════
# 1. GET /pagar_proveedor  — renderiza la pantalla
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/pagar_proveedor')
def pagar_proveedor_view():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('pagar_proveedor.html')


# ══════════════════════════════════════════════════════════════════════════
# 2. GET /pagos_proveedores  — renderiza el listado
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/pagos_proveedores')
def pagos_proveedores_view():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('pagos_proveedores.html')


# ══════════════════════════════════════════════════════════════════════════
# 3. POST /api/pagar_proveedor  — procesa el pago completo ⭐
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/api/pagar_proveedor', methods=['POST'])
def api_pagar_proveedor():
    """
    Payload (coincide con pagar_proveedor.html):
    {
        "proveedor_id": 5,
        "fecha": "2026-04-22",
        "punto_venta_recibo": 1,
        "observaciones": "...",
        "imputaciones": [{"factura_id": 10, "monto": 5000.00}, ...],
        "medios": [
          {"tipo": "efectivo", "monto": 2000.00},
          {"tipo": "transferencia", "monto": 3000.00, "banco_destino": "...", "cbu_destino": "..."},
          {"tipo": "cheque_propio_nuevo", "monto": 3000, "banco": "Galicia",
           "cuenta": "123", "numero": "00000123",
           "fecha_emision": "2026-04-22", "fecha_vencimiento": "2026-05-22"},
          {"tipo": "cheque_propio_existente", "monto": 1000, "cheque_id": 5},
          {"tipo": "cheque_tercero_existente", "monto": 2000, "cheque_id": 8}
        ]
    }
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json() or {}

        # ─── Validaciones básicas ────────────────────────────────────
        proveedor_id = data.get('proveedor_id')
        if not proveedor_id:
            return jsonify({'success': False, 'error': 'Proveedor obligatorio'}), 400

        fecha_str = data.get('fecha')
        if not fecha_str:
            return jsonify({'success': False, 'error': 'Fecha obligatoria'}), 400
        try:
            fecha_pago = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': 'Formato de fecha inválido (YYYY-MM-DD)'}), 400

        punto_venta_recibo = int(data.get('punto_venta_recibo') or 1)
        observaciones      = (data.get('observaciones') or '').strip() or None
        imputaciones       = data.get('imputaciones') or []
        medios             = data.get('medios') or []

        if not imputaciones:
            return jsonify({'success': False, 'error': 'Tenés que imputar al menos una factura'}), 400
        if not medios:
            return jsonify({'success': False, 'error': 'Tenés que cargar al menos un medio de pago'}), 400

        prov = Proveedor.query.get(proveedor_id)
        if not prov:
            return jsonify({'success': False, 'error': 'Proveedor no encontrado'}), 404

        # ─── Calcular totales y que cuadren ──────────────────────────
        total_imputado = Decimal('0')
        for imp in imputaciones:
            monto = Decimal(str(imp.get('monto') or 0))
            if monto <= 0:
                return jsonify({'success': False,
                                'error': 'Todas las imputaciones deben tener monto > 0'}), 400
            total_imputado += monto

        total_medios = Decimal('0')
        for m in medios:
            mo = Decimal(str(m.get('monto') or 0))
            if mo <= 0:
                return jsonify({'success': False,
                                'error': 'Todos los medios deben tener monto > 0'}), 400
            total_medios += mo

        # Tolerancia de 1 centavo por redondeos
        if abs(total_medios - total_imputado) > Decimal('0.01'):
            return jsonify({
                'success': False,
                'error': f'Los medios de pago (${total_medios}) no coinciden con lo imputado (${total_imputado})'
            }), 400

        # ─── Traer facturas y validar saldos ─────────────────────────
        facturas_ids = [int(imp['factura_id']) for imp in imputaciones]
        facturas = FacturaCompra.query.filter(
            FacturaCompra.id.in_(facturas_ids),
            FacturaCompra.proveedor_id == proveedor_id
        ).all()
        fac_by_id = {f.id: f for f in facturas}

        for imp in imputaciones:
            fid   = int(imp['factura_id'])
            monto = Decimal(str(imp['monto']))
            f     = fac_by_id.get(fid)
            if not f:
                return jsonify({'success': False,
                                'error': f'Factura {fid} no existe o no pertenece al proveedor'}), 400
            if f.estado == 'anulada':
                return jsonify({'success': False,
                                'error': f'Factura {f.numero_completo()} está anulada'}), 400
            saldo_disp = Decimal(str(f.saldo_pendiente or 0))
            if monto > saldo_disp:
                return jsonify({
                    'success': False,
                    'error': f'Querés imputar ${monto} a la factura {f.numero_completo()} pero solo debe ${saldo_disp}'
                }), 400

        # ─── Número de recibo atómico ────────────────────────────────
        numero_recibo = PagoProveedorNumerador.siguiente_numero(punto_venta_recibo)

        # ─── Pago cabecera ───────────────────────────────────────────
        pago = PagoProveedor(
            proveedor_id       = proveedor_id,
            fecha              = fecha_pago,
            importe            = total_imputado,
            numero_recibo      = numero_recibo,
            punto_venta_recibo = punto_venta_recibo,
            estado             = 'activo',
            forma_pago         = _resumen_medios(medios),
            referencia         = None,
            observaciones      = observaciones,
            usuario_id         = session['user_id'],
        )
        db.session.add(pago)
        db.session.flush()  # necesitamos pago.id

        # ─── Imputaciones a facturas ─────────────────────────────────
        for imp in imputaciones:
            fid   = int(imp['factura_id'])
            monto = Decimal(str(imp['monto']))
            f     = fac_by_id[fid]

            det = PagoProveedorDetalle(
                pago_id        = pago.id,
                factura_id     = fid,
                monto_imputado = monto,
            )
            db.session.add(det)

            # Actualizar saldo y estado de la factura
            nuevo_saldo = Decimal(str(f.saldo_pendiente or 0)) - monto
            if nuevo_saldo <= Decimal('0.01'):
                f.saldo_pendiente = Decimal('0')
                f.estado = 'pagada'
            else:
                f.saldo_pendiente = nuevo_saldo
                f.estado = 'parcial'

        # ─── Medios de pago ──────────────────────────────────────────
        for m in medios:
            tipo  = m.get('tipo')
            monto = Decimal(str(m.get('monto') or 0))

            medio_row = PagoProveedorMedio(
                pago_id = pago.id,
                medio   = _tipo_medio_normalizado(tipo),
                monto   = monto,
            )

            if tipo == 'transferencia':
                medio_row.banco_destino = (m.get('banco_destino') or '').strip() or None
                medio_row.cbu_destino   = (m.get('cbu_destino') or '').strip() or None

            elif tipo == 'cheque_propio_nuevo':
                banco = (m.get('banco') or '').strip()
                numero = (m.get('numero') or '').strip()
                fecha_em_str = m.get('fecha_emision')
                if not banco or not numero or not fecha_em_str:
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': 'Cheque propio nuevo: banco, número y fecha de emisión son obligatorios'}), 400
                try:
                    fecha_em = datetime.strptime(fecha_em_str, '%Y-%m-%d').date()
                except ValueError:
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': 'Formato de fecha de emisión inválido'}), 400

                fecha_vto = None
                if m.get('fecha_vencimiento'):
                    try:
                        fecha_vto = datetime.strptime(m['fecha_vencimiento'], '%Y-%m-%d').date()
                    except ValueError:
                        pass

                cheque_nuevo = ChequePropio(
                    banco             = banco,
                    cuenta            = (m.get('cuenta') or '').strip() or None,
                    numero            = numero,
                    fecha_emision     = fecha_em,
                    fecha_vencimiento = fecha_vto,
                    monto             = monto,
                    estado            = 'entregado',
                    proveedor_id      = proveedor_id,
                    pago_id           = pago.id,
                    fecha_entrega     = fecha_pago,
                )
                db.session.add(cheque_nuevo)
                db.session.flush()
                medio_row.cheque_propio_id = cheque_nuevo.id

            elif tipo == 'cheque_propio_existente':
                cheque_id = m.get('cheque_id')
                if not cheque_id:
                    db.session.rollback()
                    return jsonify({'success': False, 'error': 'cheque_propio_existente sin cheque_id'}), 400
                cheque = ChequePropio.query.get(cheque_id)
                if not cheque:
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': f'Cheque propio {cheque_id} no encontrado'}), 404
                if cheque.estado != 'pendiente':
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': f'Cheque propio {cheque.numero} no está disponible (estado: {cheque.estado})'}), 400
                cheque.estado        = 'entregado'
                cheque.proveedor_id  = proveedor_id
                cheque.pago_id       = pago.id
                cheque.fecha_entrega = fecha_pago
                medio_row.cheque_propio_id = cheque.id

            elif tipo == 'cheque_tercero_existente':
                cheque_id = m.get('cheque_id')
                if not cheque_id:
                    db.session.rollback()
                    return jsonify({'success': False, 'error': 'cheque_tercero_existente sin cheque_id'}), 400
                cheque = ChequeTercero.query.get(cheque_id)
                if not cheque:
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': f'Cheque tercero {cheque_id} no encontrado'}), 404
                if cheque.estado != 'en_cartera':
                    db.session.rollback()
                    return jsonify({'success': False,
                                    'error': f'Cheque tercero {cheque.numero} no está en cartera (estado: {cheque.estado})'}), 400
                cheque.estado               = 'endosado'
                cheque.proveedor_destino_id = proveedor_id
                cheque.pago_id              = pago.id
                cheque.fecha_endoso         = fecha_pago
                medio_row.cheque_tercero_id = cheque.id

            db.session.add(medio_row)

        # ─── Actualizar saldo del proveedor ──────────────────────────
        prov.saldo = Decimal(str(prov.saldo or 0)) - total_imputado

        # ─── Asiento en el libro mayor ───────────────────────────────
        mov = ProveedorMovimiento(
            proveedor_id    = proveedor_id,
            fecha           = fecha_pago,
            tipo            = 'pago',
            referencia_id   = pago.id,
            descripcion     = f'Pago OP {pago.numero_recibo_completo()}',
            debe            = Decimal('0'),
            haber           = total_imputado,
            saldo_acumulado = Decimal(str(prov.saldo or 0)),
        )
        db.session.add(mov)

        db.session.commit()

        return jsonify({
            'success':                       True,
            'pago_id':                       pago.id,
            'numero_recibo':                 pago.numero_recibo,
            'punto_venta_recibo':            pago.punto_venta_recibo,
            'numero_recibo_completo':        pago.numero_recibo_completo(),
            'total':                         float(total_imputado),
            'saldo_proveedor_actualizado':   float(prov.saldo),
        })

    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# 4. GET /api/pagos_proveedores/listar  — listado con filtros
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/api/pagos_proveedores/listar')
def api_pagos_proveedores_listar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        prov_id   = request.args.get('proveedor_id', '').strip()
        desde_str = request.args.get('desde', '').strip()
        hasta_str = request.args.get('hasta', '').strip()
        estado    = request.args.get('estado', '').strip()

        q = PagoProveedor.query
        if prov_id:
            q = q.filter(PagoProveedor.proveedor_id == int(prov_id))
        if desde_str:
            try:
                q = q.filter(PagoProveedor.fecha >= datetime.strptime(desde_str, '%Y-%m-%d').date())
            except ValueError:
                pass
        if hasta_str:
            try:
                q = q.filter(PagoProveedor.fecha <= datetime.strptime(hasta_str, '%Y-%m-%d').date())
            except ValueError:
                pass
        if estado and estado != 'todos':
            q = q.filter(PagoProveedor.estado == estado)

        pagos = q.order_by(PagoProveedor.fecha.desc(), PagoProveedor.id.desc()).limit(500).all()

        out = []
        for p in pagos:
            d = p.to_dict()
            d['cant_facturas']  = p.detalles.count()
            d['resumen_medios'] = _resumen_medios_obj(list(p.medios))
            out.append(d)

        return jsonify({'success': True, 'pagos': out})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# 5. POST /api/pago_proveedor/<id>/anular  — anular pago (revierte todo)
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/api/pago_proveedor/<int:pago_id>/anular', methods=['POST'])
def api_pago_proveedor_anular_v2(pago_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json(silent=True) or {}
        motivo = (data.get('motivo') or '').strip()

        pago = PagoProveedor.query.get(pago_id)
        if not pago:
            return jsonify({'success': False, 'error': 'Pago no encontrado'}), 404
        if pago.estado == 'anulado':
            return jsonify({'success': False, 'error': 'El pago ya estaba anulado'}), 400

        # Revertir imputaciones
        for det in pago.detalles:
            f = det.factura
            if not f:
                continue
            f.saldo_pendiente = Decimal(str(f.saldo_pendiente or 0)) + Decimal(str(det.monto_imputado))
            if f.saldo_pendiente >= f.total:
                f.estado = 'pendiente'
                f.saldo_pendiente = f.total
            else:
                f.estado = 'parcial'

        # Revertir cheques
        for medio in pago.medios:
            if medio.cheque_propio_id and medio.cheque_propio and medio.cheque_propio.pago_id == pago.id:
                ch = medio.cheque_propio
                ch.estado        = 'pendiente'
                ch.proveedor_id  = None
                ch.pago_id       = None
                ch.fecha_entrega = None
            if medio.cheque_tercero_id and medio.cheque_tercero and medio.cheque_tercero.pago_id == pago.id:
                ch = medio.cheque_tercero
                ch.estado               = 'en_cartera'
                ch.proveedor_destino_id = None
                ch.pago_id              = None
                ch.fecha_endoso         = None

        # Sumar al saldo del proveedor
        prov = pago.proveedor
        prov.saldo = Decimal(str(prov.saldo or 0)) + Decimal(str(pago.importe))

        # Construir descripción: con motivo si lo mandaron
        desc_base = f'Anulación OP {pago.numero_recibo_completo()}'
        desc = f'{desc_base} — {motivo}' if motivo else desc_base

        # Asiento de ajuste en el libro mayor
        mov = ProveedorMovimiento(
            proveedor_id    = prov.id,
            fecha           = datetime.now().date(),
            tipo            = 'ajuste',
            referencia_id   = pago.id,
            descripcion     = desc,
            debe            = Decimal(str(pago.importe)),
            haber           = Decimal('0'),
            saldo_acumulado = Decimal(str(prov.saldo or 0)),
        )
        db.session.add(mov)

        # Agregar el motivo a las observaciones del pago para trazabilidad
        if motivo:
            obs_anterior = pago.observaciones or ''
            sep = ' · ' if obs_anterior else ''
            pago.observaciones = f'{obs_anterior}{sep}[ANULADO] {motivo}'

        pago.estado = 'anulado'

        db.session.commit()
        return jsonify({'success': True, 'saldo_proveedor_actualizado': float(prov.saldo)})

    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# 6. GET /api/pago_proveedor/<id>/detalle  — Detalle completo del pago
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/api/pago_proveedor/<int:pago_id>/detalle')
def api_pago_proveedor_detalle(pago_id):
    """Devuelve pago + imputaciones + medios (con info de cheque si corresponde)
    para el modal 'Ver detalle' del listado de pagos."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        pago = PagoProveedor.query.get(pago_id)
        if not pago:
            return jsonify({'success': False, 'error': 'Pago no encontrado'}), 404

        # Pago (to_dict ya incluye numero_recibo_completo, proveedor, importe, estado, observaciones)
        pago_dict = pago.to_dict()

        # Imputaciones (to_dict trae factura_numero y factura_total)
        detalles = [d.to_dict() for d in pago.detalles]

        # Medios (to_dict trae cheque_propio y cheque_tercero con banco/numero/vto/librador)
        medios = [m.to_dict() for m in pago.medios]

        return jsonify({
            'success':  True,
            'pago':     pago_dict,
            'detalles': detalles,
            'medios':   medios,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════
# 7. GET /api/pago_proveedor/<id>/recibo_pdf  — PDF del recibo
# ══════════════════════════════════════════════════════════════════════════
@proveedores_bp.route('/api/pago_proveedor/<int:pago_id>/recibo_pdf')
def api_pago_proveedor_recibo_pdf(pago_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)

        pago = PagoProveedor.query.get(pago_id)
        if not pago:
            return jsonify({'error': 'Pago no encontrado'}), 404

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)

        story = []
        st_titulo = ParagraphStyle('titulo', fontName='Helvetica-Bold',
                                   fontSize=18, alignment=TA_CENTER, spaceAfter=4)
        st_sub    = ParagraphStyle('sub', fontName='Helvetica', fontSize=10,
                                   alignment=TA_CENTER,
                                   textColor=colors.HexColor('#6c757d'), spaceAfter=15)
        st_seccion = ParagraphStyle('sec', fontName='Helvetica-Bold',
                                    fontSize=11, spaceBefore=10, spaceAfter=4,
                                    textColor=colors.HexColor('#0a58ca'))
        st_normal = ParagraphStyle('n', fontName='Helvetica', fontSize=10, leading=13)

        # Encabezado
        story.append(Paragraph('RECIBO DE PAGO', st_titulo))
        fecha_fmt = pago.fecha.strftime('%d/%m/%Y') if pago.fecha else ''
        story.append(Paragraph(f'Nº <b>{pago.numero_recibo_completo()}</b>  ·  Fecha: {fecha_fmt}', st_sub))

        # Proveedor
        story.append(Paragraph('Proveedor', st_seccion))
        prov = pago.proveedor
        data_prov = [
            ['Razón social:', prov.razon_social if prov else ''],
            ['CUIT:',         (prov.cuit or '') if prov else ''],
            ['Dirección:',    (prov.direccion or '') if prov else ''],
        ]
        t_prov = Table(data_prov, colWidths=[35*mm, 120*mm])
        t_prov.setStyle(TableStyle([
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(t_prov)

        # Imputaciones
        story.append(Paragraph('Facturas imputadas', st_seccion))
        data_imp = [['Factura', 'Total factura', 'Imputado']]
        for det in pago.detalles:
            f = det.factura
            data_imp.append([
                f.numero_completo() if f else f'ID {det.factura_id}',
                f'${float(f.total):,.2f}' if f else '',
                f'${float(det.monto_imputado):,.2f}'
            ])
        t_imp = Table(data_imp, colWidths=[70*mm, 45*mm, 40*mm])
        t_imp.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e8f1ff')),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_imp)

        # Medios de pago
        story.append(Paragraph('Medios de pago', st_seccion))
        data_med = [['Medio', 'Detalle', 'Monto']]
        for m in pago.medios:
            if m.medio == 'efectivo':
                detalle = '—'
            elif m.medio == 'transferencia':
                detalle = m.banco_destino or ''
            elif m.medio == 'cheque_propio' and m.cheque_propio:
                c = m.cheque_propio
                detalle = f'{c.banco} Nº {c.numero}'
            elif m.medio == 'cheque_tercero' and m.cheque_tercero:
                c = m.cheque_tercero
                detalle = f'{c.banco} Nº {c.numero} (librador: {c.librador or "-"})'
            else:
                detalle = m.observaciones or '—'

            nombre_medio = {
                'efectivo':       'Efectivo',
                'transferencia':  'Transferencia',
                'cheque_propio':  'Cheque propio',
                'cheque_tercero': 'Cheque de tercero',
            }.get(m.medio, m.medio)
            data_med.append([nombre_medio, detalle, f'${float(m.monto):,.2f}'])

        t_med = Table(data_med, colWidths=[40*mm, 75*mm, 40*mm])
        t_med.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e8f1ff')),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t_med)

        # Total
        story.append(Spacer(1, 10))
        data_total = [['TOTAL', f'${float(pago.importe):,.2f}']]
        t_total = Table(data_total, colWidths=[115*mm, 40*mm])
        t_total.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#0a58ca')),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 12),
            ('RIGHTPADDING', (0,0), (-1,-1), 12),
        ]))
        story.append(t_total)

        if pago.observaciones:
            story.append(Paragraph('Observaciones', st_seccion))
            story.append(Paragraph(pago.observaciones, st_normal))

        # Firmas
        story.append(Spacer(1, 40))
        data_firmas = [['_______________________________', '', '_______________________________'],
                       ['Recibí conforme', '', 'Firma autorizada']]
        t_firmas = Table(data_firmas, colWidths=[65*mm, 25*mm, 65*mm])
        t_firmas.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor('#6c757d')),
        ]))
        story.append(t_firmas)

        doc.build(story)

        filename = f"recibo_{pago.numero_recibo_completo().replace(' ', '_')}.pdf"
        return send_file(
            BytesIO(buf.getvalue()),
            mimetype='application/pdf',
            as_attachment=False,
            download_name=filename,
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════
# LIBRO IVA DIGITAL (RG 4597) — EXPORTACIÓN DE COMPRAS
# ═══════════════════════════════════════════════════════════════════════

@proveedores_bp.route('/api/libro_iva/exportar_compras_zip')
def api_libro_iva_exportar_compras_zip():
    """
    Genera un ZIP con los archivos del Libro IVA Digital de Compras (RG 4597):
        - LIBRO_IVA_DIGITAL_COMPRAS_CBTE.txt
        - LIBRO_IVA_DIGITAL_COMPRAS_ALICUOTAS.txt
        - LEEME.txt

    Parámetros GET:
        ?desde=YYYY-MM-DD
        ?hasta=YYYY-MM-DD   (debe ser del MISMO mes que 'desde')
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        from sqlalchemy import text
        from exportador_libro_iva import (
            generar_compras_cbte,
            generar_compras_alicuotas,
            armar_zip_libro_iva_compras,
        )
        try:
            from config_cliente import CUIT as CUIT_EMISOR
        except Exception:
            CUIT_EMISOR = ''

        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'}), 400

        desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()

        # Validación: mismo mes calendario (requisito AFIP)
        if desde_dt.year != hasta_dt.year or desde_dt.month != hasta_dt.month:
            return jsonify({
                'success': False,
                'error': 'El período debe ser de un único mes calendario. '
                         'AFIP no permite mezclar meses en un mismo Libro IVA.'
            }), 400

        # ────────── QUERY FACTURAS DE COMPRA + PERCEPCIONES ──────────
        # Usamos SQL crudo porque las percepciones no están mapeadas
        # en el modelo SQLAlchemy (mismo enfoque que /api/iva_compras)
        sql = """
            SELECT fc.id, fc.fecha, fc.tipo_comprobante, fc.punto_venta, fc.numero,
                   fc.neto_gravado_21, fc.iva_21, fc.neto_gravado_105, fc.iva_105,
                   fc.neto_no_gravado, fc.otros_impuestos, fc.total,
                   COALESCE(fc.percepcion_iva, 0)       AS percepcion_iva,
                   COALESCE(fc.percepcion_iibb, 0)      AS percepcion_iibb,
                   COALESCE(fc.percepcion_ganancias, 0) AS percepcion_ganancias,
                   p.razon_social AS prov_nombre,
                   p.cuit         AS prov_cuit
              FROM factura_compra fc
              JOIN proveedor p ON p.id = fc.proveedor_id
             WHERE fc.fecha BETWEEN :desde AND :hasta
               AND (fc.estado IS NULL OR fc.estado != 'anulada')
             ORDER BY fc.fecha ASC, fc.id ASC
        """
        rows = db.session.execute(text(sql), {
            'desde': desde_dt, 'hasta': hasta_dt
        }).mappings().all()

        if not rows:
            return jsonify({
                'success': False,
                'error': 'No hay comprobantes de compra en el período seleccionado.'
            }), 400

        # ────────── ARMAR DATOS PARA EL EXPORTADOR ──────────
        cbte_rows = []
        alic_rows = []

        for r in rows:
            tipo = (r['tipo_comprobante'] or 'A').strip().upper()
            # En compras, tipos B y C no discriminan IVA
            no_discrim = tipo in ('B', 'C', 'NB', 'NC', 'NDB', 'NDC',
                                  '006', '007', '008', '011', '012', '013')

            neto21  = float(r['neto_gravado_21']  or 0)
            iva21   = float(r['iva_21']           or 0)
            neto105 = float(r['neto_gravado_105'] or 0)
            iva105  = float(r['iva_105']          or 0)

            # Alícuotas presentes en esta factura
            alicuotas_esta_compra = []
            if not no_discrim:
                if neto21 > 0 or iva21 > 0:
                    alicuotas_esta_compra.append({
                        'porcentaje': 21, 'neto': neto21, 'iva': iva21,
                    })
                if neto105 > 0 or iva105 > 0:
                    alicuotas_esta_compra.append({
                        'porcentaje': 10.5, 'neto': neto105, 'iva': iva105,
                    })

            # Si es A pero no tiene ningún neto/iva cargado, dejar cant_alic=1
            # para que AFIP no rechace (caso raro, se podría dar con factura a 0)
            cant_alic = 0 if no_discrim else (len(alicuotas_esta_compra) or 1)

            cbte_rows.append({
                'fecha':                r['fecha'],
                'tipo_comprobante':     tipo,
                'punto_venta':          r['punto_venta'],
                'numero':               r['numero'],
                'prov_doc_numero':      r['prov_cuit'] or '0',
                'prov_nombre':          r['prov_nombre'] or 'SIN NOMBRE',
                'total':                float(r['total'] or 0),
                'neto_gravado':         neto21 + neto105,
                'iva':                  iva21 + iva105,
                'neto_no_gravado':      float(r['neto_no_gravado'] or 0),
                'exento':               0,
                'percepcion_iva':       float(r['percepcion_iva'] or 0),
                'percepcion_otros_nac': float(r['percepcion_ganancias'] or 0),
                'percepcion_iibb':      float(r['percepcion_iibb'] or 0),
                'percepcion_municipal': 0,
                'impuestos_internos':   0,
                'otros_tributos':       float(r['otros_impuestos'] or 0),
                'credito_fiscal_computable': 0 if no_discrim else (iva21 + iva105),
                'cant_alicuotas':       cant_alic,
            })

            # Alícuotas (solo si discrimina IVA)
            for a in alicuotas_esta_compra:
                alic_rows.append({
                    'tipo_comprobante': tipo,
                    'punto_venta':      r['punto_venta'],
                    'numero':           r['numero'],
                    'prov_doc_numero':  r['prov_cuit'] or '0',
                    'neto_gravado':     a['neto'],
                    'porcentaje_iva':   a['porcentaje'],
                    'iva_liquidado':    a['iva'],
                })

        # ────────── GENERAR ARCHIVOS Y ZIP ──────────
        cbte_txt = generar_compras_cbte(cbte_rows)
        alic_txt = generar_compras_alicuotas(alic_rows)

        periodo_aaaamm = desde_dt.strftime('%Y%m')
        zip_buf = armar_zip_libro_iva_compras(
            compras_cbte_txt=cbte_txt,
            compras_alic_txt=alic_txt,
            cuit_informante=CUIT_EMISOR,
            periodo_aaaamm=periodo_aaaamm,
        )

        cuit_str = CUIT_EMISOR or 'SINCUIT'
        nombre_zip = f"Libro_IVA_Digital_Compras_{cuit_str}_{periodo_aaaamm}.zip"
        return send_file(
            zip_buf,
            as_attachment=True,
            download_name=nombre_zip,
            mimetype='application/zip'
        )

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500