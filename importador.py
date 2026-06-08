"""
═══════════════════════════════════════════════════════════════════════════════
IMPORTADOR - Carga masiva desde Excel/CSV con detección automática de columnas
═══════════════════════════════════════════════════════════════════════════════

Módulo independiente — no contamina app.py. Solo expone:
  - Blueprint con todas las rutas bajo /importador/
  - Función init_importador(app, db, **modelos) para registrarlo

Uso desde app.py:
    from importador import init_importador
    init_importador(app, db, Cliente=Cliente, Producto=Producto)

Soporta archivos .xlsx, .xls y .csv. Detección automática de columnas.
Implementa Clientes y Productos. Proveedores y Saldos quedan placeholder.
═══════════════════════════════════════════════════════════════════════════════
"""

from flask import Blueprint, render_template, request, jsonify, redirect, session, send_file
from io import BytesIO, StringIO
from decimal import Decimal, InvalidOperation
import re
import csv

importador_bp = Blueprint('importador', __name__, url_prefix='/importador')

# ════════════════════════════════════════════════════════════════════════════
# Modelos inyectados desde app.py
# ════════════════════════════════════════════════════════════════════════════
_db = None
_models = {}


# ════════════════════════════════════════════════════════════════════════════
# DEFINICIÓN DE CAMPOS POR ENTIDAD
# ════════════════════════════════════════════════════════════════════════════
CAMPOS_CLIENTE = [
    {'key': 'nombre',         'label': 'Nombre / Razón Social', 'required': True,  'type': 'str',
     'aliases': ['nombre', 'razon social', 'razón social', 'cliente', 'razonsocial', 'apellido y nombre', 'denominacion', 'denominación']},
    {'key': 'documento',      'label': 'CUIT / DNI',            'required': False, 'type': 'doc',
     'aliases': ['cuit', 'c.u.i.t.', 'c.u.i.t', 'dni', 'documento', 'doc', 'nro documento', 'numero documento', 'cuit/dni']},
    {'key': 'tipo_documento', 'label': 'Tipo Doc (CUIT/DNI)',   'required': False, 'type': 'str',
     'aliases': ['tipo doc', 'tipo documento', 'tipo de documento', 'tipodoc']},
    {'key': 'email',          'label': 'Email',                 'required': False, 'type': 'email',
     'aliases': ['email', 'e-mail', 'mail', 'correo', 'correo electronico', 'correo electrónico']},
    {'key': 'telefono',       'label': 'Teléfono',              'required': False, 'type': 'str',
     'aliases': ['telefono', 'teléfono', 'tel', 'tel.', 'celular', 'cel', 'movil', 'móvil', 'whatsapp']},
    {'key': 'direccion',      'label': 'Dirección',             'required': False, 'type': 'str',
     'aliases': ['direccion', 'dirección', 'domicilio', 'calle', 'address']},
    {'key': 'condicion_iva',  'label': 'Condición IVA',         'required': False, 'type': 'iva',
     'aliases': ['condicion iva', 'condición iva', 'cond iva', 'cond. iva', 'iva', 'condicion frente al iva', 'situacion iva', 'situación iva']},
    {'key': 'lista_precio',   'label': 'Lista de Precios (1-5)','required': False, 'type': 'int',
     'aliases': ['lista', 'lista precio', 'lista de precios', 'lista_precio', 'precio lista', 'tipo lista']},
]

CAMPOS_PRODUCTO = [
    {'key': 'codigo',         'label': 'Código',                'required': True,  'type': 'str',
     'aliases': ['codigo', 'código', 'code', 'cod', 'cod.', 'sku', 'codigo articulo', 'codigo producto']},
    {'key': 'nombre',         'label': 'Nombre / Descripción',  'required': True,  'type': 'str',
     'aliases': ['nombre', 'descripcion', 'descripción', 'producto', 'articulo', 'artículo', 'detalle', 'denominacion']},
    {'key': 'precio',         'label': 'Precio (lista 1)',      'required': True,  'type': 'money',
     'aliases': ['precio', 'precio venta', 'p venta', 'pventa', 'precio 1', 'precio_1', 'precio lista 1', 'precio_lista_1', 'precio publico']},
    {'key': 'precio2',        'label': 'Precio Lista 2',        'required': False, 'type': 'money',
     'aliases': ['precio 2', 'precio_2', 'precio lista 2', 'precio_lista_2', 'p2', 'lista 2']},
    {'key': 'precio3',        'label': 'Precio Lista 3',        'required': False, 'type': 'money',
     'aliases': ['precio 3', 'precio_3', 'precio lista 3', 'precio_lista_3', 'p3', 'lista 3']},
    {'key': 'precio4',        'label': 'Precio Lista 4',        'required': False, 'type': 'money',
     'aliases': ['precio 4', 'precio_4', 'precio lista 4', 'precio_lista_4', 'p4', 'lista 4']},
    {'key': 'precio5',        'label': 'Precio Lista 5',        'required': False, 'type': 'money',
     'aliases': ['precio 5', 'precio_5', 'precio lista 5', 'precio_lista_5', 'p5', 'lista 5']},
    {'key': 'costo',          'label': 'Costo (sin IVA)',      'required': False, 'type': 'money',
     'aliases': ['costo', 'precio costo', 'p costo', 'pcosto', 'costo sin iva', 'costo neto', 'precio_costo']},
    {'key': 'stock',          'label': 'Stock inicial',         'required': False, 'type': 'money',
     'aliases': ['stock', 'cantidad', 'existencia', 'inventario', 'stock inicial', 'stock actual']},
    {'key': 'iva',            'label': 'Alícuota IVA (%)',      'required': False, 'type': 'money',
     'aliases': ['iva', 'alicuota iva', 'alícuota iva', 'alicuota', 'alícuota', 'iva %', '% iva']},
    {'key': 'categoria',      'label': 'Categoría',             'required': False, 'type': 'str',
     'aliases': ['categoria', 'categoría', 'rubro', 'familia', 'grupo', 'tipo']},
    {'key': 'codigo_barras',  'label': 'Código de Barras',      'required': False, 'type': 'str',
     'aliases': ['codigo barras', 'código de barras', 'cod barras', 'ean', 'codigo_barras', 'barcode', 'cb']},
    {'key': 'stock_minimo',   'label': 'Stock Mínimo',          'required': False, 'type': 'money',
     'aliases': ['stock minimo', 'stock mínimo', 'min', 'minimo', 'mínimo', 'stock_minimo']},
    {'key': 'es_pesable',     'label': 'Es Pesable (S/N)',      'required': False, 'type': 'bool',
     'aliases': ['pesable', 'es pesable', 'balanza', 'por peso']},
]

CAMPOS_PROVEEDOR = [
    {'key': 'razon_social',   'label': 'Razón Social',          'required': True,  'type': 'str',
     'aliases': ['razon social', 'razón social', 'razonsocial', 'nombre', 'proveedor', 'denominacion', 'denominación', 'apellido y nombre']},
    {'key': 'cuit',           'label': 'CUIT',                  'required': False, 'type': 'doc',
     'aliases': ['cuit', 'c.u.i.t.', 'c.u.i.t', 'documento', 'doc', 'nro documento', 'numero documento']},
    {'key': 'condicion_iva',  'label': 'Condición IVA',         'required': False, 'type': 'iva',
     'aliases': ['condicion iva', 'condición iva', 'cond iva', 'cond. iva', 'iva', 'condicion frente al iva', 'situacion iva']},
    {'key': 'direccion',      'label': 'Dirección',             'required': False, 'type': 'str',
     'aliases': ['direccion', 'dirección', 'domicilio', 'calle', 'address']},
    {'key': 'telefono',       'label': 'Teléfono',              'required': False, 'type': 'str',
     'aliases': ['telefono', 'teléfono', 'tel', 'tel.', 'celular', 'cel', 'movil', 'móvil', 'whatsapp']},
    {'key': 'email',          'label': 'Email',                 'required': False, 'type': 'email',
     'aliases': ['email', 'e-mail', 'mail', 'correo', 'correo electronico']},
]

CAMPOS_SALDO = [
    {'key': 'cuit',           'label': 'CUIT (para identificar)','required': True,  'type': 'doc',
     'aliases': ['cuit', 'c.u.i.t.', 'c.u.i.t', 'cuit/dni', 'documento', 'doc', 'nro documento', 'numero documento']},
    {'key': 'monto',          'label': 'Monto del saldo',       'required': True,  'type': 'money',
     'aliases': ['monto', 'saldo', 'importe', 'deuda', 'debe', 'total', 'saldo deudor', 'saldo acreedor']},
    {'key': 'razon_social',   'label': 'Razón Social (referencia)','required': False,'type': 'str',
     'aliases': ['razon social', 'razón social', 'nombre', 'cliente', 'proveedor', 'denominacion']},
]

COND_IVA_MAP = {
    'responsable inscripto': 'Responsable Inscripto',
    'resp inscripto': 'Responsable Inscripto',
    'ri': 'Responsable Inscripto',
    'iva responsable inscripto': 'Responsable Inscripto',
    'monotributo': 'Monotributista',
    'monotributista': 'Monotributista',
    'mt': 'Monotributista',
    'exento': 'Exento',
    'iva exento': 'Exento',
    'no responsable': 'Exento',
    'consumidor final': 'Consumidor Final',
    'cf': 'Consumidor Final',
    'final': 'Consumidor Final',
    'no categorizado': 'No Categorizado',
    'no inscripto': 'No Categorizado',
}


# ════════════════════════════════════════════════════════════════════════════
# HELPERS DE NORMALIZACIÓN
# ════════════════════════════════════════════════════════════════════════════
def _norm(s):
    if s is None:
        return ''
    s = str(s).strip().lower()
    repl = {'á':'a', 'é':'e', 'í':'i', 'ó':'o', 'ú':'u', 'ñ':'n'}
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r'[._\-]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _solo_digitos(s):
    if s is None:
        return ''
    return re.sub(r'\D', '', str(s))


def validar_cuit(cuit):
    """Valida CUIT con algoritmo de dígito verificador."""
    digitos = _solo_digitos(cuit)
    if len(digitos) != 11:
        return False, digitos
    multiplicadores = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    try:
        suma = sum(int(digitos[i]) * multiplicadores[i] for i in range(10))
    except (ValueError, IndexError):
        return False, digitos
    resto = suma % 11
    dv_calc = 11 - resto
    if dv_calc == 11:
        dv_calc = 0
    elif dv_calc == 10:
        return False, digitos
    return dv_calc == int(digitos[10]), digitos


def to_decimal(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace('$', '').replace(' ', '')
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_bool(v):
    if v is None or v == '':
        return False
    s = str(v).strip().lower()
    return s in ('s', 'si', 'sí', 'y', 'yes', 'true', '1', 'verdadero', 'x')


def normalizar_iva(valor):
    if not valor:
        return None
    n = _norm(valor)
    return COND_IVA_MAP.get(n, str(valor).strip())


# ════════════════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ════════════════════════════════════════════════════════════════════════════
def leer_archivo(file_storage):
    """Lee un Excel (xlsx/xls) o CSV. Devuelve (headers, rows) o (None, error_msg)."""
    nombre = (file_storage.filename or '').lower()
    contenido = file_storage.read()
    try:
        if nombre.endswith('.xlsx'):
            return _leer_xlsx(contenido)
        elif nombre.endswith('.xls'):
            return _leer_xls(contenido)
        elif nombre.endswith('.csv'):
            return _leer_csv(contenido)
        else:
            return None, 'Formato no soportado. Usá .xlsx, .xls o .csv'
    except Exception as e:
        return None, f'Error leyendo archivo: {str(e)}'


def _leer_xlsx(contenido):
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    return _procesar_filas(rows_raw)


def _leer_xls(contenido):
    try:
        import xlrd
    except ImportError:
        return None, 'Para leer archivos .xls instalá: pip install xlrd==1.2.0'
    wb = xlrd.open_workbook(file_contents=contenido)
    sh = wb.sheet_by_index(0)
    rows_raw = []
    for r in range(sh.nrows):
        rows_raw.append(tuple(sh.cell(r, c).value for c in range(sh.ncols)))
    return _procesar_filas(rows_raw)


def _leer_csv(contenido):
    txt = None
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            txt = contenido.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if txt is None:
        return None, 'No se pudo decodificar el CSV (probá guardarlo como UTF-8)'
    sample = txt[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        sep = dialect.delimiter
    except Exception:
        sep = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.reader(StringIO(txt), delimiter=sep)
    rows_raw = list(reader)
    return _procesar_filas(rows_raw)


def _procesar_filas(rows_raw):
    if not rows_raw:
        return None, 'Archivo vacío'
    idx = 0
    while idx < len(rows_raw) and all((c is None or str(c).strip() == '') for c in rows_raw[idx]):
        idx += 1
    if idx >= len(rows_raw):
        return None, 'No hay datos en el archivo'

    headers_raw = rows_raw[idx]
    headers = []
    for i, c in enumerate(headers_raw):
        h = str(c).strip() if c is not None and str(c).strip() else f'Columna_{i+1}'
        headers.append(h)

    rows = []
    for r in rows_raw[idx+1:]:
        if all((c is None or str(c).strip() == '') for c in r):
            continue
        row = {i: r[i] if i < len(r) else None for i in range(len(headers))}
        rows.append(row)
    return (headers, rows), None


# ════════════════════════════════════════════════════════════════════════════
# DETECCIÓN AUTOMÁTICA DE COLUMNAS
# ════════════════════════════════════════════════════════════════════════════
def detectar_mapeo(headers, rows, campos):
    """Para cada campo del modelo, busca la columna del Excel que mejor matchea.
    Devuelve {key: columna_index} o {key: None} si no encontró."""
    mapeo = {}
    headers_norm = [_norm(h) for h in headers]
    usadas = set()

    # 1. Match exacto por alias
    for campo in campos:
        if campo['key'] in mapeo:
            continue
        for alias in campo['aliases']:
            alias_n = _norm(alias)
            for i, h in enumerate(headers_norm):
                if i in usadas:
                    continue
                if h == alias_n:
                    mapeo[campo['key']] = i
                    usadas.add(i)
                    break
            if campo['key'] in mapeo:
                break

    # 2. Match parcial (contiene)
    for campo in campos:
        if campo['key'] in mapeo:
            continue
        for alias in campo['aliases']:
            alias_n = _norm(alias)
            for i, h in enumerate(headers_norm):
                if i in usadas:
                    continue
                if alias_n in h or h in alias_n:
                    mapeo[campo['key']] = i
                    usadas.add(i)
                    break
            if campo['key'] in mapeo:
                break

    # 3. Heurísticas por contenido (si todavía falta documento, email)
    if rows:
        sample = rows[:min(5, len(rows))]
        for campo in campos:
            if campo['key'] in mapeo:
                continue
            for i in range(len(headers)):
                if i in usadas:
                    continue
                valores = [str(r.get(i, '') or '').strip() for r in sample]
                if campo['type'] == 'email':
                    if sum(1 for v in valores if '@' in v) >= max(1, len(valores) // 2):
                        mapeo[campo['key']] = i
                        usadas.add(i)
                        break
                elif campo['type'] == 'doc':
                    digitos = [_solo_digitos(v) for v in valores]
                    if sum(1 for d in digitos if len(d) in (7, 8, 11)) >= max(1, len(valores) // 2):
                        mapeo[campo['key']] = i
                        usadas.add(i)
                        break

    # Completar con None lo que falte
    for campo in campos:
        if campo['key'] not in mapeo:
            mapeo[campo['key']] = None
    return mapeo


# ════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN Y NORMALIZACIÓN DE FILAS
# ════════════════════════════════════════════════════════════════════════════
def procesar_fila_cliente(row, mapeo):
    """Aplica el mapeo a una fila y devuelve (datos_normalizados, errores, advertencias)."""
    datos = {}
    errores = []
    advertencias = []

    def get(key):
        idx = mapeo.get(key)
        if idx is None:
            return None
        v = row.get(idx)
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # Nombre (obligatorio)
    nombre = get('nombre')
    if not nombre:
        errores.append('Falta nombre')
        return None, errores, advertencias
    datos['nombre'] = nombre[:100]

    # Documento
    doc_raw = get('documento')
    tipo_doc = get('tipo_documento')
    if doc_raw:
        digitos = _solo_digitos(doc_raw)
        if len(digitos) == 11:
            valido, _ = validar_cuit(digitos)
            if valido:
                datos['documento'] = digitos
                datos['tipo_documento'] = 'CUIT'
            else:
                advertencias.append(f'CUIT {digitos} con dígito verificador inválido (se carga igual)')
                datos['documento'] = digitos
                datos['tipo_documento'] = 'CUIT'
        elif len(digitos) in (7, 8):
            datos['documento'] = digitos
            datos['tipo_documento'] = 'DNI'
        else:
            datos['documento'] = digitos[:20]
            datos['tipo_documento'] = (tipo_doc or 'OTRO')[:10]

    # Email
    email = get('email')
    if email:
        if '@' in email and '.' in email.split('@')[-1]:
            datos['email'] = email[:100]
        else:
            advertencias.append(f'Email "{email}" parece inválido (se carga igual)')
            datos['email'] = email[:100]

    # Teléfono / Dirección
    tel = get('telefono')
    if tel:
        datos['telefono'] = tel[:20]
    direc = get('direccion')
    if direc:
        datos['direccion'] = direc

    # Condición IVA
    iva = get('condicion_iva')
    if iva:
        datos['condicion_iva'] = normalizar_iva(iva)
    else:
        datos['condicion_iva'] = 'Consumidor Final'

    # Lista de precios
    lista = get('lista_precio')
    if lista:
        try:
            n = int(_solo_digitos(lista) or '1')
            if 1 <= n <= 5:
                datos['lista_precio'] = n
            else:
                advertencias.append(f'Lista de precios {n} fuera de rango (1-5), se asigna lista 1')
                datos['lista_precio'] = 1
        except ValueError:
            datos['lista_precio'] = 1
    else:
        datos['lista_precio'] = 1

    return datos, errores, advertencias


def procesar_fila_proveedor(row, mapeo):
    """Aplica el mapeo a una fila de proveedor."""
    datos = {}
    errores = []
    advertencias = []

    def get(key):
        idx = mapeo.get(key)
        if idx is None:
            return None
        v = row.get(idx)
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # Razón social (obligatorio)
    razon = get('razon_social')
    if not razon:
        errores.append('Falta razón social')
        return None, errores, advertencias
    datos['razon_social'] = razon[:150]

    # CUIT
    cuit_raw = get('cuit')
    if cuit_raw:
        digitos = _solo_digitos(cuit_raw)
        if len(digitos) == 11:
            valido, _ = validar_cuit(digitos)
            if valido:
                datos['cuit'] = digitos
            else:
                advertencias.append(f'CUIT {digitos} con dígito verificador inválido (se carga igual)')
                datos['cuit'] = digitos
        else:
            datos['cuit'] = digitos[:20]

    # Condición IVA (default Responsable Inscripto - es lo más común para proveedores)
    iva = get('condicion_iva')
    if iva:
        datos['condicion_iva'] = normalizar_iva(iva)
    else:
        datos['condicion_iva'] = 'Responsable Inscripto'

    # Email
    email = get('email')
    if email:
        if '@' in email and '.' in email.split('@')[-1]:
            datos['email'] = email[:100]
        else:
            advertencias.append(f'Email "{email}" parece inválido (se carga igual)')
            datos['email'] = email[:100]

    # Teléfono / Dirección
    tel = get('telefono')
    if tel:
        datos['telefono'] = tel[:30]
    direc = get('direccion')
    if direc:
        datos['direccion'] = direc[:200]

    return datos, errores, advertencias


def procesar_fila_saldo(row, mapeo):
    """Procesa fila de saldo - solo extrae cuit y monto."""
    datos = {}
    errores = []
    advertencias = []

    def get(key):
        idx = mapeo.get(key)
        if idx is None:
            return None
        v = row.get(idx)
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # CUIT (obligatorio)
    cuit_raw = get('cuit')
    if not cuit_raw:
        errores.append('Falta CUIT')
        return None, errores, advertencias
    digitos = _solo_digitos(cuit_raw)
    if len(digitos) < 7:
        errores.append(f'CUIT/DNI inválido: "{cuit_raw}"')
        return None, errores, advertencias
    datos['cuit'] = digitos

    # Monto (obligatorio)
    monto_raw = get('monto')
    if monto_raw is None or monto_raw == '':
        errores.append('Falta monto')
        return None, errores, advertencias
    monto = to_decimal(monto_raw)
    if monto is None:
        errores.append(f'Monto inválido: "{monto_raw}"')
        return None, errores, advertencias
    datos['monto'] = monto

    # Razón social (referencia, no obligatoria)
    razon = get('razon_social')
    if razon:
        datos['razon_social'] = razon[:150]

    return datos, errores, advertencias


def procesar_fila_producto(row, mapeo, margen_default=Decimal('30')):
    """Aplica el mapeo a una fila de producto."""
    datos = {}
    errores = []
    advertencias = []

    def get(key):
        idx = mapeo.get(key)
        if idx is None:
            return None
        v = row.get(idx)
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # Código (obligatorio)
    codigo = get('codigo')
    if not codigo:
        errores.append('Falta código')
        return None, errores, advertencias
    datos['codigo'] = codigo[:50]

    # Nombre (obligatorio)
    nombre = get('nombre')
    if not nombre:
        errores.append('Falta nombre')
        return None, errores, advertencias
    datos['nombre'] = nombre[:200]

    # Precio (obligatorio)
    precio = to_decimal(get('precio'))
    if precio is None or precio < 0:
        errores.append('Precio inválido o falta')
        return None, errores, advertencias
    datos['precio'] = precio

    # Precios 2-5
    for i in range(2, 6):
        p = to_decimal(get(f'precio{i}'))
        if p is not None and p >= 0:
            datos[f'precio{i}'] = p

    # IVA
    iva = to_decimal(get('iva'))
    if iva is not None and 0 <= iva <= 50:
        datos['iva'] = iva
    else:
        datos['iva'] = Decimal('21')

    # Costo y margen
    costo_raw = to_decimal(get('costo'))
    if costo_raw is not None and costo_raw > 0:
        datos['costo'] = costo_raw
        # Calcular margen real basado en precio_lista_1 vs costo
        try:
            margen_real = ((precio - costo_raw) / costo_raw) * Decimal('100')
            datos['margen'] = margen_real.quantize(Decimal('0.01'))
        except (InvalidOperation, ZeroDivisionError):
            datos['margen'] = margen_default
    else:
        # Sin costo en el Excel → calculamos costo = precio / (1 + margen/100)
        try:
            datos['costo'] = (precio / (Decimal('1') + margen_default / Decimal('100'))).quantize(Decimal('0.01'))
            datos['margen'] = margen_default
        except (InvalidOperation, ZeroDivisionError):
            datos['costo'] = Decimal('0')
            datos['margen'] = margen_default

    # Stock
    stock = to_decimal(get('stock'))
    datos['stock'] = stock if stock is not None and stock >= 0 else Decimal('0')

    # Stock mínimo
    sm = to_decimal(get('stock_minimo'))
    if sm is not None and sm >= 0:
        datos['stock_minimo'] = sm

    # Categoría
    cat = get('categoria')
    if cat:
        datos['categoria'] = cat[:100]

    # Código de barras
    cb = get('codigo_barras')
    if cb:
        datos['codigo_barras'] = cb[:50]

    # Pesable
    pesable_raw = get('es_pesable')
    if pesable_raw is not None:
        datos['es_pesable'] = parse_bool(pesable_raw)
    else:
        datos['es_pesable'] = False

    return datos, errores, advertencias


# ════════════════════════════════════════════════════════════════════════════
# RUTAS
# ════════════════════════════════════════════════════════════════════════════
@importador_bp.route('/')
def inicio():
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('importador_inicio.html')


@importador_bp.route('/clientes')
def pantalla_clientes():
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('importador_clientes.html', campos=CAMPOS_CLIENTE)


@importador_bp.route('/productos')
def pantalla_productos():
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('importador_productos.html', campos=CAMPOS_PRODUCTO)


@importador_bp.route('/proveedores')
def pantalla_proveedores():
    if 'user_id' not in session:
        return redirect('/login')
    if 'Proveedor' not in _models:
        return render_template('importador_proveedores.html', campos=CAMPOS_PROVEEDOR, modulo_disponible=False)
    return render_template('importador_proveedores.html', campos=CAMPOS_PROVEEDOR, modulo_disponible=True)


@importador_bp.route('/saldos')
def pantalla_saldos():
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('importador_saldos.html', campos=CAMPOS_SALDO,
                           proveedor_disponible='Proveedor' in _models)


@importador_bp.route('/<entidad>/preview', methods=['POST'])
def preview(entidad):
    """Recibe el archivo, devuelve mapeo detectado + primeras 10 filas."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    # Para saldos, la entidad real puede ser 'saldos_clientes' o 'saldos_proveedores'
    entidades_validas = ('clientes', 'productos', 'proveedores', 'saldos_clientes', 'saldos_proveedores')
    if entidad not in entidades_validas:
        return jsonify({'success': False, 'error': 'Entidad no soportada'}), 400

    if entidad == 'proveedores' and 'Proveedor' not in _models:
        return jsonify({'success': False, 'error': 'Módulo de proveedores no disponible en este sistema'}), 400

    if entidad.startswith('saldos_proveedores') and 'Proveedor' not in _models:
        return jsonify({'success': False, 'error': 'Módulo de proveedores no disponible'}), 400

    if 'archivo' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibió archivo'}), 400

    archivo = request.files['archivo']
    resultado, error = leer_archivo(archivo)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    headers, rows = resultado
    if not rows:
        return jsonify({'success': False, 'error': 'El archivo no tiene filas de datos'}), 400

    if entidad == 'clientes':
        campos = CAMPOS_CLIENTE
    elif entidad == 'productos':
        campos = CAMPOS_PRODUCTO
    elif entidad == 'proveedores':
        campos = CAMPOS_PROVEEDOR
    elif entidad in ('saldos_clientes', 'saldos_proveedores'):
        campos = CAMPOS_SALDO
    else:
        campos = CAMPOS_CLIENTE
    mapeo = detectar_mapeo(headers, rows, campos)

    sample_rows = []
    for i, r in enumerate(rows[:10]):
        sample_rows.append({
            'numero': i + 1,
            'valores': [str(r.get(j, '') or '') for j in range(len(headers))]
        })

    return jsonify({
        'success': True,
        'headers': headers,
        'mapeo': mapeo,
        'total_filas': len(rows),
        'sample': sample_rows,
        'archivo_token': _guardar_archivo_temp(archivo, resultado)
    })


# Cache simple de archivos en memoria (por user_id)
_archivos_temp = {}


def _guardar_archivo_temp(archivo, resultado):
    import time
    user = session.get('user_id', 'anon')
    token = f"{user}_{int(time.time() * 1000)}"
    _archivos_temp[token] = resultado
    # Limpiar entradas viejas (>1h)
    cutoff = time.time() - 3600
    for k in list(_archivos_temp.keys()):
        try:
            ts = int(k.split('_')[-1]) / 1000
            if ts < cutoff:
                del _archivos_temp[k]
        except (ValueError, IndexError):
            pass
    return token


@importador_bp.route('/<entidad>/confirmar', methods=['POST'])
def confirmar(entidad):
    """Ejecuta la importación con el mapeo confirmado."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    entidades_validas = ('clientes', 'productos', 'proveedores', 'saldos_clientes', 'saldos_proveedores')
    if entidad not in entidades_validas:
        return jsonify({'success': False, 'error': 'Entidad no soportada'}), 400

    if entidad == 'proveedores' and 'Proveedor' not in _models:
        return jsonify({'success': False, 'error': 'Módulo de proveedores no disponible'}), 400

    if entidad == 'saldos_proveedores' and 'Proveedor' not in _models:
        return jsonify({'success': False, 'error': 'Módulo de proveedores no disponible'}), 400

    data = request.get_json() or {}
    token = data.get('archivo_token')
    mapeo_raw = data.get('mapeo', {})
    margen_default = data.get('margen_default', 30)

    if not token or token not in _archivos_temp:
        return jsonify({'success': False, 'error': 'Token de archivo expirado, subí el archivo de nuevo'}), 400

    headers, rows = _archivos_temp[token]

    # Mapeo: convertir índices a int (vienen como str del JSON)
    mapeo = {}
    for k, v in mapeo_raw.items():
        if v is None or v == '' or v == 'null':
            mapeo[k] = None
        else:
            try:
                mapeo[k] = int(v)
            except (ValueError, TypeError):
                mapeo[k] = None

    if entidad == 'clientes':
        return _confirmar_clientes(rows, mapeo)
    elif entidad == 'proveedores':
        return _confirmar_proveedores(rows, mapeo)
    elif entidad == 'saldos_clientes':
        return _confirmar_saldos(rows, mapeo, 'cliente')
    elif entidad == 'saldos_proveedores':
        return _confirmar_saldos(rows, mapeo, 'proveedor')
    else:  # productos
        try:
            margen_dec = Decimal(str(margen_default))
        except InvalidOperation:
            margen_dec = Decimal('30')
        return _confirmar_productos(rows, mapeo, margen_dec)


def _confirmar_clientes(rows, mapeo):
    Cliente = _models['Cliente']

    # Cargar CUITs/DNIs existentes para detectar duplicados
    docs_existentes = set()
    nombres_existentes = set()
    for c in Cliente.query.with_entities(Cliente.documento, Cliente.nombre).all():
        if c.documento:
            docs_existentes.add(_solo_digitos(c.documento))
        if c.nombre:
            nombres_existentes.add(c.nombre.strip().lower())

    importados = 0
    duplicados = []
    errores = []
    advertencias_global = []

    for i, row in enumerate(rows):
        numero_fila = i + 2  # +2 porque la fila 1 es el header
        datos, errs, advs = procesar_fila_cliente(row, mapeo)

        if errs:
            errores.append({'fila': numero_fila, 'errores': errs})
            continue

        # Saltar duplicados por documento
        if datos.get('documento') and datos['documento'] in docs_existentes:
            duplicados.append({'fila': numero_fila, 'nombre': datos['nombre'], 'doc': datos['documento']})
            continue

        # Saltar duplicados por nombre exacto si no hay documento
        if not datos.get('documento') and datos['nombre'].strip().lower() in nombres_existentes:
            duplicados.append({'fila': numero_fila, 'nombre': datos['nombre'], 'doc': ''})
            continue

        # Crear cliente
        try:
            c = Cliente(**{k: v for k, v in datos.items() if hasattr(Cliente, k)})
            _db.session.add(c)
            importados += 1
            if datos.get('documento'):
                docs_existentes.add(datos['documento'])
            nombres_existentes.add(datos['nombre'].strip().lower())
            if advs:
                advertencias_global.append({'fila': numero_fila, 'nombre': datos['nombre'], 'advertencias': advs})
        except Exception as e:
            errores.append({'fila': numero_fila, 'errores': [f'Error al crear: {str(e)}']})

    try:
        _db.session.commit()
    except Exception as e:
        _db.session.rollback()
        return jsonify({'success': False, 'error': f'Error al guardar en BD: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'importados': importados,
        'duplicados': duplicados,
        'errores': errores,
        'advertencias': advertencias_global,
        'total': len(rows)
    })


def _confirmar_productos(rows, mapeo, margen_default):
    Producto = _models['Producto']

    codigos_existentes = set()
    for p in Producto.query.with_entities(Producto.codigo).all():
        if p.codigo:
            codigos_existentes.add(p.codigo.strip())

    importados = 0
    duplicados = []
    errores = []
    advertencias_global = []

    for i, row in enumerate(rows):
        numero_fila = i + 2
        datos, errs, advs = procesar_fila_producto(row, mapeo, margen_default)

        if errs:
            errores.append({'fila': numero_fila, 'errores': errs})
            continue

        if datos['codigo'] in codigos_existentes:
            duplicados.append({'fila': numero_fila, 'codigo': datos['codigo'], 'nombre': datos['nombre']})
            continue

        try:
            datos_filtrados = {k: v for k, v in datos.items() if hasattr(Producto, k)}
            p = Producto(**datos_filtrados)
            _db.session.add(p)
            importados += 1
            codigos_existentes.add(datos['codigo'])
            if advs:
                advertencias_global.append({'fila': numero_fila, 'codigo': datos['codigo'], 'advertencias': advs})
        except Exception as e:
            errores.append({'fila': numero_fila, 'errores': [f'Error al crear: {str(e)}']})

    try:
        _db.session.commit()
    except Exception as e:
        _db.session.rollback()
        return jsonify({'success': False, 'error': f'Error al guardar en BD: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'importados': importados,
        'duplicados': duplicados,
        'errores': errores,
        'advertencias': advertencias_global,
        'total': len(rows)
    })


def _confirmar_proveedores(rows, mapeo):
    Proveedor = _models['Proveedor']

    # Cargar CUITs y razones sociales existentes
    cuits_existentes = set()
    razones_existentes = set()
    for p in Proveedor.query.with_entities(Proveedor.cuit, Proveedor.razon_social).all():
        if p.cuit:
            cuits_existentes.add(_solo_digitos(p.cuit))
        if p.razon_social:
            razones_existentes.add(p.razon_social.strip().lower())

    importados = 0
    duplicados = []
    errores = []
    advertencias_global = []

    for i, row in enumerate(rows):
        numero_fila = i + 2
        datos, errs, advs = procesar_fila_proveedor(row, mapeo)

        if errs:
            errores.append({'fila': numero_fila, 'errores': errs})
            continue

        # Saltar duplicados por CUIT
        if datos.get('cuit') and datos['cuit'] in cuits_existentes:
            duplicados.append({'fila': numero_fila, 'nombre': datos['razon_social'], 'doc': datos['cuit']})
            continue

        # Si no tiene CUIT, comparar por razón social exacta
        if not datos.get('cuit') and datos['razon_social'].strip().lower() in razones_existentes:
            duplicados.append({'fila': numero_fila, 'nombre': datos['razon_social'], 'doc': ''})
            continue

        try:
            p = Proveedor(**{k: v for k, v in datos.items() if hasattr(Proveedor, k)})
            _db.session.add(p)
            importados += 1
            if datos.get('cuit'):
                cuits_existentes.add(datos['cuit'])
            razones_existentes.add(datos['razon_social'].strip().lower())
            if advs:
                advertencias_global.append({'fila': numero_fila, 'nombre': datos['razon_social'], 'advertencias': advs})
        except Exception as e:
            errores.append({'fila': numero_fila, 'errores': [f'Error al crear: {str(e)}']})

    try:
        _db.session.commit()
    except Exception as e:
        _db.session.rollback()
        return jsonify({'success': False, 'error': f'Error al guardar en BD: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'importados': importados,
        'duplicados': duplicados,
        'errores': errores,
        'advertencias': advertencias_global,
        'total': len(rows)
    })


def _confirmar_saldos(rows, mapeo, tipo):
    """Importa saldos. tipo = 'cliente' o 'proveedor'.
    REEMPLAZA el saldo existente (no suma)."""
    if tipo == 'cliente':
        Modelo = _models['Cliente']
        campo_doc = 'documento'
        campo_nombre = 'nombre'
    else:  # proveedor
        Modelo = _models['Proveedor']
        campo_doc = 'cuit'
        campo_nombre = 'razon_social'

    # Cargar registros existentes indexados por CUIT/documento limpio
    registros_por_cuit = {}
    for r in Modelo.query.all():
        doc = getattr(r, campo_doc, None)
        if doc:
            digitos = _solo_digitos(doc)
            if digitos:
                registros_por_cuit[digitos] = r

    actualizados = 0
    no_encontrados = []
    errores = []
    advertencias_global = []

    for i, row in enumerate(rows):
        numero_fila = i + 2
        datos, errs, advs = procesar_fila_saldo(row, mapeo)

        if errs:
            errores.append({'fila': numero_fila, 'errores': errs})
            continue

        cuit_clean = datos['cuit']
        registro = registros_por_cuit.get(cuit_clean)

        if not registro:
            ref = datos.get('razon_social', '(sin razón social)')
            no_encontrados.append({'fila': numero_fila, 'cuit': cuit_clean, 'nombre': ref})
            continue

        try:
            registro.saldo = datos['monto']
            actualizados += 1
            nombre_referencia = getattr(registro, campo_nombre, '?')
            if advs:
                advertencias_global.append({'fila': numero_fila, 'nombre': nombre_referencia, 'advertencias': advs})
        except Exception as e:
            errores.append({'fila': numero_fila, 'errores': [f'Error al actualizar: {str(e)}']})

    try:
        _db.session.commit()
    except Exception as e:
        _db.session.rollback()
        return jsonify({'success': False, 'error': f'Error al guardar en BD: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'importados': actualizados,
        'duplicados': no_encontrados,  # reusamos el campo "duplicados" para "no encontrados"
        'errores': errores,
        'advertencias': advertencias_global,
        'total': len(rows),
        'modo_saldos': True  # flag para que el frontend cambie textos
    })


# ════════════════════════════════════════════════════════════════════════════
# PLANTILLAS DESCARGABLES
# ════════════════════════════════════════════════════════════════════════════
@importador_bp.route('/plantilla/<entidad>')
def descargar_plantilla(entidad):
    if 'user_id' not in session:
        return redirect('/login')
    if entidad not in ('clientes', 'productos', 'proveedores', 'saldos_clientes', 'saldos_proveedores'):
        return 'Entidad no soportada', 400

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entidad.replace('_', ' ').capitalize()

    if entidad == 'clientes':
        campos = CAMPOS_CLIENTE
        ejemplo = ['Juan Pérez S.A.', '20-12345678-9', 'CUIT', 'juan@ejemplo.com',
                   '11-4444-5555', 'Av. Corrientes 1234', 'Responsable Inscripto', '1']
    elif entidad == 'productos':
        campos = CAMPOS_PRODUCTO
        ejemplo = ['PROD001', 'Producto de ejemplo', '1500.00', '1600.00', '1700.00',
                   '1800.00', '1900.00', '1000.00', '50', '21', 'Categoría A',
                   '7790000000001', '5', 'N']
    elif entidad == 'proveedores':
        campos = CAMPOS_PROVEEDOR
        ejemplo = ['Proveedor Ejemplo S.A.', '30-12345678-9', 'Responsable Inscripto',
                   'Av. Industrial 500', '11-3333-4444', 'ventas@proveedor.com']
    elif entidad in ('saldos_clientes', 'saldos_proveedores'):
        campos = CAMPOS_SALDO
        if entidad == 'saldos_clientes':
            ejemplo = ['20-12345678-9', '5000.00', 'Juan Pérez S.A.']
        else:
            ejemplo = ['30-12345678-9', '15000.00', 'Proveedor Ejemplo S.A.']
    else:
        campos = CAMPOS_CLIENTE
        ejemplo = []

    # Header
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col, campo in enumerate(campos, 1):
        c = ws.cell(row=1, column=col, value=campo['label'] + ('*' if campo['required'] else ''))
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = max(18, len(campo['label']) + 4)

    # Ejemplo
    for col, val in enumerate(ejemplo[:len(campos)], 1):
        ws.cell(row=2, column=col, value=val)

    # Nota
    ws.cell(row=4, column=1, value='Los campos con * son obligatorios. Borrá la fila de ejemplo antes de cargar tus datos.')

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'plantilla_{entidad}.xlsx')


# ════════════════════════════════════════════════════════════════════════════
# INIT
# ════════════════════════════════════════════════════════════════════════════
def init_importador(app, db, **modelos):
    """Registra el blueprint y guarda los modelos.
    Modelos esperados: Cliente, Producto, Proveedor (opcional).
    Si falta Proveedor, las pantallas de Proveedores y Saldos de Proveedores
    quedan en modo 'no disponible' (la pantalla informa al usuario).
    Si falta Cliente o Producto, esas entidades fallan."""
    global _db, _models
    _db = db
    _models = modelos
    app.register_blueprint(importador_bp)
