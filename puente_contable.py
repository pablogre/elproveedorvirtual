"""
═══════════════════════════════════════════════════════════════════════════════
PUENTE CONTABLE - Exportación de asientos para Estudios Contables
═══════════════════════════════════════════════════════════════════════════════

Genera un Excel con asientos contables listos para importar en Tango Gestión
(o cualquier otro sistema mediante reorganización en una planilla).

Arquitectura:
  - El blueprint NO redefine modelos. Recibe las clases por inyección desde app.py
    (mismo patrón que se usa en init_caja_system).
  - Esto evita el error "Table is already defined for this MetaData instance"
    que tendríamos al re-importar Factura/Cliente/etc.

Uso desde app.py:
    from puente_contable import init_puente_contable
    init_puente_contable(
        app, db,
        Factura=Factura, DetalleFactura=DetalleFactura, MedioPago=MedioPago,
        NotaCredito=NotaCredito, DetalleNotaCredito=DetalleNotaCredito,
        Cliente=Cliente, Usuario=Usuario, Gasto=Gasto,
        FacturaCompra=FacturaCompra, PagoProveedor=PagoProveedor,
        PagoProveedorMedio=PagoProveedorMedio, Proveedor=Proveedor,
        CajaAperturaModel=CajaAperturaModel, MovimientoCajaModel=MovimientoCajaModel
    )
═══════════════════════════════════════════════════════════════════════════════
"""

from flask import Blueprint, render_template, request, send_file, redirect, session, jsonify
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

# openpyxl ya viene con pandas, igual lo importamos directo para tener control
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

puente_contable_bp = Blueprint('puente_contable', __name__, url_prefix='/puente_contable')

# ════════════════════════════════════════════════════════════════════════════
# Modelos inyectados desde app.py — ver init_puente_contable()
# ════════════════════════════════════════════════════════════════════════════
_db = None
_models = {}


# ════════════════════════════════════════════════════════════════════════════
# PLAN DE CUENTAS ESTÁNDAR (FACPCE-compatible)
# ════════════════════════════════════════════════════════════════════════════
PLAN_CUENTAS = {
    # ── ACTIVO ──────────────────────────────────────────────────────────────
    '1.1.1.01': 'Caja',
    '1.1.1.02': 'Banco Cuenta Corriente',
    '1.1.1.03': 'Mercado Pago',
    '1.1.2.01': 'Deudores por Ventas',
    '1.1.2.02': 'Cheques de Terceros en Cartera',
    '1.1.2.05': 'Tarjetas de Crédito a Cobrar',
    '1.1.2.06': 'Tarjetas de Débito a Cobrar',
    '1.1.3.01': 'Mercaderías',
    '1.1.4.01': 'IVA Crédito Fiscal 21%',
    '1.1.4.02': 'IVA Crédito Fiscal 10.5%',
    # ── PASIVO ──────────────────────────────────────────────────────────────
    '2.1.1.01': 'Proveedores',
    '2.1.2.01': 'Cheques Propios Emitidos',
    '2.1.3.01': 'IVA Débito Fiscal 21%',
    '2.1.3.02': 'IVA Débito Fiscal 10.5%',
    # ── PATRIMONIO NETO ─────────────────────────────────────────────────────
    '3.1.1.01': 'Capital',
    '3.2.1.01': 'Resultados no Asignados',
    # ── INGRESOS ────────────────────────────────────────────────────────────
    '4.1.1.01': 'Ventas',
    '4.2.9.01': 'Sobrantes de Caja',
    # ── EGRESOS / GASTOS ────────────────────────────────────────────────────
    '5.1.1.01': 'Compras / Mercaderías',
    '5.2.1.01': 'Gastos Generales',
    '5.2.1.02': 'Insumos y Materiales',
    '5.2.1.03': 'Servicios Públicos',
    '5.2.1.04': 'Transporte y Combustible',
    '5.2.1.05': 'Sueldos y Cargas Sociales',
    '5.2.1.06': 'Mantenimiento',
    '5.2.1.07': 'Impuestos y Tasas',
    '5.2.1.08': 'Otros Gastos',
    '5.2.9.01': 'Faltantes de Caja',
    '5.9.9.99': 'Movimientos de Caja a Clasificar',
}

# Mapeo medio de pago de venta (MedioPago.medio_pago) → cuenta contable
CUENTAS_COBRO = {
    'efectivo':     '1.1.1.01',
    'credito':      '1.1.2.05',
    'debito':       '1.1.2.06',
    'mercado_pago': '1.1.1.03',
    # fallback
    'transferencia': '1.1.1.02',
}

# Mapeo medio de pago de proveedor (PagoProveedorMedio.medio) → cuenta contable
CUENTAS_PAGO_PROV = {
    'efectivo':       '1.1.1.01',
    'transferencia':  '1.1.1.02',
    'cheque_propio':  '2.1.2.01',
    'cheque_tercero': '1.1.2.02',
    'otro':           '5.9.9.99',
}

# Mapeo método de pago de Gasto → cuenta contable
CUENTAS_GASTO_MP = {
    'efectivo':      '1.1.1.01',
    'transferencia': '1.1.1.02',
    'debito':        '1.1.2.06',
    'credito':       '1.1.2.05',
}

# Mapeo categoría de Gasto → cuenta contable
CUENTAS_GASTO_CAT = {
    'general':       '5.2.1.01',
    'insumos':       '5.2.1.02',
    'servicios':     '5.2.1.03',
    'transporte':    '5.2.1.04',
    'personal':      '5.2.1.05',
    'mantenimiento': '5.2.1.06',
    'impuestos':     '5.2.1.07',
    'otros':         '5.2.1.08',
}


# ════════════════════════════════════════════════════════════════════════════
# HELPERS DE ASIENTO
# ════════════════════════════════════════════════════════════════════════════
def _q(x):
    """Cuantizar a 2 decimales como Decimal."""
    if x is None:
        return Decimal('0.00')
    return Decimal(str(x)).quantize(Decimal('0.01'))


def _get_cuenta(codigo):
    """Devuelve descripción de cuenta o el mismo código si no existe."""
    return PLAN_CUENTAS.get(codigo, codigo)


# ════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE ASIENTOS POR ORIGEN
# ════════════════════════════════════════════════════════════════════════════
def asientos_de_ventas(fecha_desde, fecha_hasta):
    """Asientos por cada Factura autorizada en el rango."""
    Factura = _models['Factura']
    DetalleFactura = _models['DetalleFactura']
    MedioPago = _models['MedioPago']

    asientos = []
    facturas = (Factura.query
                .filter(Factura.fecha >= fecha_desde)
                .filter(Factura.fecha <= fecha_hasta)
                .filter(Factura.estado == 'autorizada')
                .order_by(Factura.fecha, Factura.id)
                .all())

    for f in facturas:
        cliente_nombre = f.cliente.nombre if f.cliente else 'Consumidor Final'
        cliente_doc = (f.cliente.documento or '') if f.cliente else ''
        concepto = f"Venta {f.tipo_comprobante or ''} {f.numero or ''} - {cliente_nombre}".strip()

        # Discriminar IVA por alícuota desde DetalleFactura
        iva_21 = Decimal('0')
        iva_105 = Decimal('0')
        for d in f.detalles:
            ali = float(d.porcentaje_iva or 21)
            imp_iva = _q(d.importe_iva or 0)
            if abs(ali - 21.0) < 0.5:
                iva_21 += imp_iva
            elif abs(ali - 10.5) < 0.5:
                iva_105 += imp_iva
            else:
                iva_21 += imp_iva  # fallback al 21

        subtotal = _q(f.subtotal)
        total = _q(f.total)
        fecha_a = f.fecha.date() if hasattr(f.fecha, 'date') else f.fecha

        # ─── ASIENTO DE VENTA ──────────────────────────────────────────────
        # Si la factura tiene cobros registrados (MedioPago) → la asentamos
        # como ventas a cuentas de cobro directamente.
        # Si NO tiene cobros (es a cuenta corriente) → contra Deudores.
        medios = list(MedioPago.query.filter_by(factura_id=f.id).all())

        lineas = []  # (cuenta, debe, haber, concepto)

        if medios:
            # Cobro directo: cada medio a su cuenta
            for m in medios:
                cta = CUENTAS_COBRO.get(m.medio_pago, '1.1.1.01')
                lineas.append((cta, _q(m.importe), Decimal('0'), concepto))
        else:
            # Cta Cte
            lineas.append(('1.1.2.01', total, Decimal('0'), concepto + ' (cta cte)'))

        # Haberes
        lineas.append(('4.1.1.01', Decimal('0'), subtotal, concepto))
        if iva_21 > 0:
            lineas.append(('2.1.3.01', Decimal('0'), iva_21, f'IVA Débito 21% - {concepto}'))
        if iva_105 > 0:
            lineas.append(('2.1.3.02', Decimal('0'), iva_105, f'IVA Débito 10.5% - {concepto}'))

        asientos.append({
            'fecha': fecha_a,
            'tipo': 'VTA',
            'comprobante': f.numero or f.id,
            'concepto': concepto,
            'auxiliar': cliente_doc,
            'lineas': lineas,
        })

    return asientos


def asientos_de_notas_credito(fecha_desde, fecha_hasta):
    """Asientos por cada Nota de Crédito autorizada (asiento inverso)."""
    if 'NotaCredito' not in _models:
        return []
    NC = _models['NotaCredito']
    asientos = []
    ncs = (NC.query
           .filter(NC.fecha >= fecha_desde)
           .filter(NC.fecha <= fecha_hasta)
           .filter(NC.estado == 'autorizada')
           .order_by(NC.fecha, NC.id)
           .all())

    for nc in ncs:
        cliente_nombre = nc.cliente.nombre if getattr(nc, 'cliente', None) else 'Consumidor Final'
        cliente_doc = (nc.cliente.documento or '') if getattr(nc, 'cliente', None) else ''
        concepto = f"NC {nc.tipo_comprobante or ''} {nc.numero or ''} - {cliente_nombre}".strip()

        # Reusamos el detalle si está; si no, total contra IVA 21
        iva_21 = Decimal('0')
        iva_105 = Decimal('0')
        subtotal = _q(getattr(nc, 'subtotal', None) or 0)
        total = _q(getattr(nc, 'total', None) or 0)

        detalles = list(getattr(nc, 'detalles', []) or [])
        for d in detalles:
            ali = float(getattr(d, 'porcentaje_iva', 21) or 21)
            imp_iva = _q(getattr(d, 'importe_iva', 0) or 0)
            if abs(ali - 21.0) < 0.5:
                iva_21 += imp_iva
            elif abs(ali - 10.5) < 0.5:
                iva_105 += imp_iva

        if not detalles and total > 0:
            iva_21 = (total - subtotal) if (total > subtotal) else Decimal('0')

        fecha_a = nc.fecha.date() if hasattr(nc.fecha, 'date') else nc.fecha

        # Asiento inverso
        lineas = []
        lineas.append(('4.1.1.01', subtotal, Decimal('0'), f'Anula venta - {concepto}'))
        if iva_21 > 0:
            lineas.append(('2.1.3.01', iva_21, Decimal('0'), f'Anula IVA Débito 21% - {concepto}'))
        if iva_105 > 0:
            lineas.append(('2.1.3.02', iva_105, Decimal('0'), f'Anula IVA Débito 10.5% - {concepto}'))
        # Contra Deudores (la NC genera saldo a favor del cliente)
        lineas.append(('1.1.2.01', Decimal('0'), total, concepto))

        asientos.append({
            'fecha': fecha_a,
            'tipo': 'NCV',
            'comprobante': nc.numero or nc.id,
            'concepto': concepto,
            'auxiliar': cliente_doc,
            'lineas': lineas,
        })

    return asientos


def asientos_de_compras(fecha_desde, fecha_hasta):
    """Asientos por cada Factura de Compra registrada."""
    if 'FacturaCompra' not in _models:
        return []
    FC = _models['FacturaCompra']

    asientos = []
    facturas = (FC.query
                .filter(FC.fecha >= fecha_desde)
                .filter(FC.fecha <= fecha_hasta)
                .filter(FC.estado != 'anulada')
                .order_by(FC.fecha, FC.id)
                .all())

    for f in facturas:
        prov_nombre = f.proveedor.razon_social if f.proveedor else 'Proveedor'
        prov_cuit = (f.proveedor.cuit or '') if f.proveedor else ''
        concepto = f"Compra {f.tipo_comprobante} {f.punto_venta}-{f.numero} - {prov_nombre}"

        neto21 = _q(f.neto_gravado_21)
        neto105 = _q(f.neto_gravado_105)
        nogravado = _q(f.neto_no_gravado)
        iva21 = _q(f.iva_21)
        iva105 = _q(f.iva_105)
        otros = _q(f.otros_impuestos)
        total = _q(f.total)

        lineas = []
        # Compras (DEBE)
        compras_total = neto21 + neto105 + nogravado
        if compras_total > 0:
            lineas.append(('5.1.1.01', compras_total, Decimal('0'), concepto))
        # IVA Crédito (DEBE)
        if iva21 > 0:
            lineas.append(('1.1.4.01', iva21, Decimal('0'), f'IVA CF 21% - {concepto}'))
        if iva105 > 0:
            lineas.append(('1.1.4.02', iva105, Decimal('0'), f'IVA CF 10.5% - {concepto}'))
        if otros > 0:
            lineas.append(('5.2.1.07', otros, Decimal('0'), f'Otros impuestos - {concepto}'))
        # Proveedor (HABER)
        lineas.append(('2.1.1.01', Decimal('0'), total, concepto))

        asientos.append({
            'fecha': f.fecha,
            'tipo': 'CPA',
            'comprobante': f"{f.punto_venta}-{f.numero}",
            'concepto': concepto,
            'auxiliar': prov_cuit,
            'lineas': lineas,
        })

    return asientos


def asientos_de_pagos_proveedor(fecha_desde, fecha_hasta):
    """Asientos por cada PagoProveedor (con sus PagoProveedorMedio)."""
    if 'PagoProveedor' not in _models:
        return []
    PP = _models['PagoProveedor']
    PPM = _models['PagoProveedorMedio']

    asientos = []
    pagos = (PP.query
             .filter(PP.fecha >= fecha_desde)
             .filter(PP.fecha <= fecha_hasta)
             .filter(PP.estado == 'activo')
             .order_by(PP.fecha, PP.id)
             .all())

    for p in pagos:
        prov_nombre = p.proveedor.razon_social if p.proveedor else 'Proveedor'
        prov_cuit = (p.proveedor.cuit or '') if p.proveedor else ''
        nro_recibo = p.numero_recibo_completo() if hasattr(p, 'numero_recibo_completo') else ''
        concepto = f"Pago a {prov_nombre} {nro_recibo}".strip()

        importe_total = _q(p.importe)

        lineas = []
        # Proveedor (DEBE) — saldamos la deuda
        lineas.append(('2.1.1.01', importe_total, Decimal('0'), concepto))

        # Medios de pago (HABER)
        medios = list(PPM.query.filter_by(pago_id=p.id).all())
        if medios:
            for m in medios:
                cta = CUENTAS_PAGO_PROV.get(m.medio, '1.1.1.01')
                detalle_medio = ''
                if m.medio == 'cheque_propio' and m.cheque_propio:
                    detalle_medio = f" Cheque {m.cheque_propio.banco} Nº{m.cheque_propio.numero}"
                elif m.medio == 'cheque_tercero' and m.cheque_tercero:
                    detalle_medio = f" Cheque 3ros {m.cheque_tercero.banco} Nº{m.cheque_tercero.numero}"
                elif m.medio == 'transferencia' and m.banco_destino:
                    detalle_medio = f" Transferencia {m.banco_destino}"
                lineas.append((cta, Decimal('0'), _q(m.monto), concepto + detalle_medio))
        else:
            # Pagos viejos sin detalle de medios — usamos forma_pago
            cta = CUENTAS_PAGO_PROV.get(p.forma_pago or 'efectivo', '1.1.1.01')
            lineas.append((cta, Decimal('0'), importe_total, concepto))

        asientos.append({
            'fecha': p.fecha,
            'tipo': 'PAG',
            'comprobante': nro_recibo or p.id,
            'concepto': concepto,
            'auxiliar': prov_cuit,
            'lineas': lineas,
        })

    return asientos


def asientos_de_gastos(fecha_desde, fecha_hasta):
    """Asientos por cada Gasto activo en el rango."""
    if 'Gasto' not in _models:
        return []
    G = _models['Gasto']

    asientos = []
    gastos = (G.query
              .filter(G.fecha >= fecha_desde)
              .filter(G.fecha <= fecha_hasta)
              .filter(G.activo == True)
              .order_by(G.fecha, G.id)
              .all())

    for g in gastos:
        cta_gasto = CUENTAS_GASTO_CAT.get(g.categoria, '5.2.1.01')
        cta_pago = CUENTAS_GASTO_MP.get(g.metodo_pago, '1.1.1.01')
        monto = _q(g.monto)
        concepto = f"Gasto: {g.descripcion[:80]}"

        lineas = [
            (cta_gasto, monto, Decimal('0'), concepto),
            (cta_pago, Decimal('0'), monto, concepto),
        ]

        asientos.append({
            'fecha': g.fecha,
            'tipo': 'GAS',
            'comprobante': g.id,
            'concepto': concepto,
            'auxiliar': '',
            'lineas': lineas,
        })

    return asientos


def asientos_de_movimientos_caja(fecha_desde, fecha_hasta):
    """Asientos por cada MovimientoCaja manual (ingreso/egreso)."""
    if 'MovimientoCajaModel' not in _models:
        return []
    MC = _models['MovimientoCajaModel']

    asientos = []
    movs = (MC.query
            .filter(MC.fecha >= fecha_desde)
            .filter(MC.fecha <= fecha_hasta)
            .order_by(MC.fecha, MC.id)
            .all())

    for m in movs:
        monto = _q(m.monto)
        concepto = f"Mov.Caja: {m.descripcion[:80]}"
        fecha_a = m.fecha.date() if hasattr(m.fecha, 'date') else m.fecha

        if m.tipo == 'ingreso':
            # Caja DEBE / A clasificar HABER
            lineas = [
                ('1.1.1.01', monto, Decimal('0'), concepto),
                ('5.9.9.99', Decimal('0'), monto, concepto),
            ]
        else:  # egreso
            # A clasificar DEBE / Caja HABER
            lineas = [
                ('5.9.9.99', monto, Decimal('0'), concepto),
                ('1.1.1.01', Decimal('0'), monto, concepto),
            ]

        asientos.append({
            'fecha': fecha_a,
            'tipo': 'MCA',
            'comprobante': m.id,
            'concepto': concepto,
            'auxiliar': '',
            'lineas': lineas,
        })

    return asientos


def asientos_de_diferencias_caja(fecha_desde, fecha_hasta):
    """Asientos por las diferencias de cierre de caja (sobrantes/faltantes)."""
    if 'CajaAperturaModel' not in _models:
        return []
    CA = _models['CajaAperturaModel']

    asientos = []
    cierres = (CA.query
               .filter(CA.estado == 'cerrada')
               .filter(CA.fecha_cierre >= fecha_desde)
               .filter(CA.fecha_cierre <= fecha_hasta)
               .all())

    for c in cierres:
        if not c.diferencia or _q(c.diferencia) == Decimal('0'):
            continue
        dif = _q(c.diferencia)
        fecha_a = c.fecha_cierre.date() if hasattr(c.fecha_cierre, 'date') else c.fecha_cierre
        concepto = f"Diferencia cierre Caja #{c.id}"

        if dif > 0:
            # Sobrante: Caja DEBE / Sobrantes HABER
            lineas = [
                ('1.1.1.01', dif, Decimal('0'), concepto),
                ('4.2.9.01', Decimal('0'), dif, concepto),
            ]
        else:
            # Faltante: Faltantes DEBE / Caja HABER
            dif_abs = abs(dif)
            lineas = [
                ('5.2.9.01', dif_abs, Decimal('0'), concepto),
                ('1.1.1.01', Decimal('0'), dif_abs, concepto),
            ]

        asientos.append({
            'fecha': fecha_a,
            'tipo': 'DCA',
            'comprobante': c.id,
            'concepto': concepto,
            'auxiliar': '',
            'lineas': lineas,
        })

    return asientos


# ════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DEL EXCEL
# ════════════════════════════════════════════════════════════════════════════
def generar_excel(asientos, fecha_desde, fecha_hasta):
    """Genera el .xlsx con 5 hojas."""
    wb = openpyxl.Workbook()

    # Estilos
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    money_fmt = '#,##0.00'

    # ── HOJA 1: ASIENTOS ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Asientos'
    headers = ['Fecha', 'Tipo Asiento', 'Nro Asiento', 'Cuenta', 'Descripción Cuenta',
               'Debe', 'Haber', 'Concepto', 'Cód. Auxiliar']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_thin

    row = 2
    nro_asiento = 0
    for a in asientos:
        nro_asiento += 1
        for linea in a['lineas']:
            cuenta, debe, haber, concepto = linea
            ws.cell(row=row, column=1, value=a['fecha']).number_format = 'DD/MM/YYYY'
            ws.cell(row=row, column=2, value=a['tipo'])
            ws.cell(row=row, column=3, value=nro_asiento)
            ws.cell(row=row, column=4, value=cuenta)
            ws.cell(row=row, column=5, value=_get_cuenta(cuenta))
            ws.cell(row=row, column=6, value=float(debe) if debe else None).number_format = money_fmt
            ws.cell(row=row, column=7, value=float(haber) if haber else None).number_format = money_fmt
            ws.cell(row=row, column=8, value=concepto)
            ws.cell(row=row, column=9, value=str(a['auxiliar']))
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border_thin
            row += 1

    # Anchos
    widths = [12, 13, 11, 11, 30, 14, 14, 50, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ── HOJA 2: RESUMEN (Sumas y Saldos) ──────────────────────────────────
    ws2 = wb.create_sheet('Resumen')
    ws2.cell(row=1, column=1, value=f'Sumas y Saldos del período {fecha_desde.strftime("%d/%m/%Y")} al {fecha_hasta.strftime("%d/%m/%Y")}').font = Font(bold=True, size=13)
    ws2.cell(row=1, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
    ws2.merge_cells('A1:E1')

    h2 = ['Cuenta', 'Descripción', 'Total Debe', 'Total Haber', 'Saldo']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = border_thin

    # Acumular por cuenta
    saldos = {}
    for a in asientos:
        for cuenta, debe, haber, _c in a['lineas']:
            if cuenta not in saldos:
                saldos[cuenta] = {'debe': Decimal('0'), 'haber': Decimal('0')}
            saldos[cuenta]['debe'] += debe
            saldos[cuenta]['haber'] += haber

    row = 4
    total_d = Decimal('0')
    total_h = Decimal('0')
    for cta in sorted(saldos.keys()):
        d = saldos[cta]['debe']
        h = saldos[cta]['haber']
        saldo = d - h
        ws2.cell(row=row, column=1, value=cta)
        ws2.cell(row=row, column=2, value=_get_cuenta(cta))
        ws2.cell(row=row, column=3, value=float(d)).number_format = money_fmt
        ws2.cell(row=row, column=4, value=float(h)).number_format = money_fmt
        ws2.cell(row=row, column=5, value=float(saldo)).number_format = money_fmt
        for col in range(1, 6):
            ws2.cell(row=row, column=col).border = border_thin
        total_d += d
        total_h += h
        row += 1

    # Totales
    ws2.cell(row=row, column=2, value='TOTALES').font = Font(bold=True)
    ws2.cell(row=row, column=3, value=float(total_d)).number_format = money_fmt
    ws2.cell(row=row, column=3).font = Font(bold=True)
    ws2.cell(row=row, column=4, value=float(total_h)).number_format = money_fmt
    ws2.cell(row=row, column=4).font = Font(bold=True)
    ws2.cell(row=row, column=5, value=float(total_d - total_h)).number_format = money_fmt
    ws2.cell(row=row, column=5).font = Font(bold=True)
    diff_row = row
    if total_d == total_h:
        ws2.cell(row=row+2, column=1, value='✓ Asientos balanceados (Debe = Haber)').font = Font(bold=True, color='008000')
    else:
        ws2.cell(row=row+2, column=1, value=f'⚠ Diferencia detectada: {float(total_d - total_h)}').font = Font(bold=True, color='CC0000')

    for i, w in enumerate([11, 32, 16, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── HOJA 3: PLAN DE CUENTAS ──────────────────────────────────────────
    ws3 = wb.create_sheet('Plan de Cuentas')
    ws3.cell(row=1, column=1, value='Plan de cuentas usado por el sistema').font = Font(bold=True, size=13)
    ws3.cell(row=1, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
    ws3.merge_cells('A1:B1')

    for col, h in enumerate(['Código', 'Descripción'], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = border_thin

    row = 4
    for cta in sorted(PLAN_CUENTAS.keys()):
        ws3.cell(row=row, column=1, value=cta)
        ws3.cell(row=row, column=2, value=PLAN_CUENTAS[cta])
        ws3.cell(row=row, column=1).border = border_thin
        ws3.cell(row=row, column=2).border = border_thin
        row += 1

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 42

    # ── HOJA 4: EQUIVALENCIAS ────────────────────────────────────────────
    ws4 = wb.create_sheet('Equivalencias')
    ws4.cell(row=1, column=1, value='Mapeo a su Plan de Cuentas (a completar por el contador)').font = Font(bold=True, size=13)
    ws4.cell(row=1, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
    ws4.merge_cells('A1:D1')
    ws4.cell(row=2, column=1, value='Si su plan de cuentas usa códigos distintos, complete la columna "Mi Cuenta" y haga Buscar/Reemplazar en la hoja Asientos antes de importar.').font = Font(italic=True, color='606060')
    ws4.merge_cells('A2:D2')

    for col, h in enumerate(['Código sistema', 'Descripción sistema', 'Mi Cuenta', 'Mi Descripción'], 1):
        c = ws4.cell(row=4, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = border_thin

    row = 5
    for cta in sorted(PLAN_CUENTAS.keys()):
        ws4.cell(row=row, column=1, value=cta)
        ws4.cell(row=row, column=2, value=PLAN_CUENTAS[cta])
        ws4.cell(row=row, column=3, value='')
        ws4.cell(row=row, column=4, value='')
        for col in range(1, 5):
            ws4.cell(row=row, column=col).border = border_thin
        row += 1

    for i, w in enumerate([14, 32, 14, 32], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── HOJA 5: INSTRUCCIONES ────────────────────────────────────────────
    ws5 = wb.create_sheet('Instrucciones')
    ws5.column_dimensions['A'].width = 100

    instrucciones = [
        ('PUENTE CONTABLE - Instrucciones de uso', True, 16, 'DDEBF7'),
        ('', False, 11, None),
        (f'Período exportado: {fecha_desde.strftime("%d/%m/%Y")} al {fecha_hasta.strftime("%d/%m/%Y")}', True, 11, None),
        (f'Cantidad de asientos: {len(asientos)}', False, 11, None),
        ('', False, 11, None),
        ('═══ HOJA 1: ASIENTOS ═══', True, 12, 'F2F2F2'),
        ('Contiene todos los asientos contables generados por el sistema en el período.', False, 11, None),
        ('Cada asiento puede tener varias líneas (un Debe y uno o varios Haber, o viceversa).', False, 11, None),
        ('La columna "Nro Asiento" agrupa las líneas que pertenecen al mismo asiento.', False, 11, None),
        ('', False, 11, None),
        ('Tipos de asiento:', True, 11, None),
        ('  VTA = Venta              CPA = Compra', False, 11, None),
        ('  NCV = Nota Crédito Venta GAS = Gasto', False, 11, None),
        ('  PAG = Pago Proveedor     MCA = Mov. Caja manual', False, 11, None),
        ('  DCA = Diferencia Caja', False, 11, None),
        ('', False, 11, None),
        ('═══ HOJA 2: RESUMEN ═══', True, 12, 'F2F2F2'),
        ('Sumas y saldos del período, agrupado por cuenta. Sirve de control rápido.', False, 11, None),
        ('La fila final muestra Total Debe = Total Haber. Si no cuadra, hay un error en algún asiento.', False, 11, None),
        ('', False, 11, None),
        ('═══ HOJA 3: PLAN DE CUENTAS ═══', True, 12, 'F2F2F2'),
        ('Plan de cuentas estándar usado por el sistema (compatible con FACPCE).', False, 11, None),
        ('', False, 11, None),
        ('═══ HOJA 4: EQUIVALENCIAS ═══', True, 12, 'F2F2F2'),
        ('Si su plan de cuentas usa códigos distintos a los del sistema:', False, 11, None),
        ('  1) Complete la columna "Mi Cuenta" con sus códigos.', False, 11, None),
        ('  2) En la hoja Asientos, use Ctrl+H (Buscar y Reemplazar) para cambiar cada código.', False, 11, None),
        ('  3) Importe la hoja Asientos a su sistema.', False, 11, None),
        ('', False, 11, None),
        ('═══ IMPORTACIÓN EN TANGO GESTIÓN CONTABILIDAD ═══', True, 12, 'FFF2CC'),
        ('  1) En Tango → Procesos Periódicos → Importación de Asientos Masivos.', False, 11, None),
        ('  2) Configure el "Modelo de importación" con las columnas en este orden:', False, 11, None),
        ('     Fecha | Tipo | Nro Asiento | Cuenta | Debe | Haber | Concepto | Auxiliar', False, 11, None),
        ('  3) Seleccione la hoja "Asientos" de este Excel y ejecute la importación.', False, 11, None),
        ('  4) Tango le mostrará un previsualizador. Revise que los asientos cuadren.', False, 11, None),
        ('  5) Confirme la importación.', False, 11, None),
        ('', False, 11, None),
        ('═══ NOTAS IMPORTANTES ═══', True, 12, 'FFF2CC'),
        ('• Los asientos de "Mov.Caja a Clasificar" (cuenta 5.9.9.99) son retiros, refuerzos', False, 11, None),
        ('  y movimientos manuales que el sistema no puede categorizar automáticamente.', False, 11, None),
        ('  Reclasificar manualmente según corresponda (Cuenta Particular, Depósitos, etc.).', False, 11, None),
        ('• Las facturas a Cuenta Corriente (sin cobro registrado) generan asiento contra', False, 11, None),
        ('  Deudores por Ventas (1.1.2.01). Los cobros posteriores aparecen en otro asiento.', False, 11, None),
        ('• Las diferencias de cierre de caja se registran como Sobrantes (4.2.9.01) o', False, 11, None),
        ('  Faltantes (5.2.9.01) según corresponda.', False, 11, None),
        ('', False, 11, None),
        ('═══ SOPORTE ═══', True, 12, 'F2F2F2'),
        ('Pablo Ré - Desarrollador', True, 11, None),
        ('Email: pablogustavore@gmail.com', False, 11, None),
        ('WhatsApp: +54 9 3364 537093', False, 11, None),
        ('Web: pablore.com.ar', False, 11, None),
    ]
    for i, (texto, bold, sz, fill) in enumerate(instrucciones, 1):
        c = ws5.cell(row=i, column=1, value=texto)
        c.font = Font(bold=bold, size=sz)
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        c.alignment = Alignment(wrap_text=True, vertical='center')

    # Salvar a buffer
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ════════════════════════════════════════════════════════════════════════════
# RUTAS
# ════════════════════════════════════════════════════════════════════════════
@puente_contable_bp.route('/')
def pantalla():
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('puente_contable.html')


@puente_contable_bp.route('/exportar', methods=['POST'])
def exportar():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    try:
        desde_str = request.form.get('desde')
        hasta_str = request.form.get('hasta')
        if not desde_str or not hasta_str:
            return jsonify({'success': False, 'error': 'Falta fecha desde/hasta'}), 400

        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()

        if fecha_hasta < fecha_desde:
            return jsonify({'success': False, 'error': 'Fecha hasta no puede ser anterior a fecha desde'}), 400

        # Filtros opcionales (qué incluir)
        incluir = {
            'ventas':    request.form.get('inc_ventas', '1') == '1',
            'nc':        request.form.get('inc_nc', '1') == '1',
            'compras':   request.form.get('inc_compras', '1') == '1',
            'pagos':     request.form.get('inc_pagos', '1') == '1',
            'gastos':    request.form.get('inc_gastos', '1') == '1',
            'mov_caja':  request.form.get('inc_movcaja', '1') == '1',
            'dif_caja':  request.form.get('inc_difcaja', '1') == '1',
        }

        asientos = []
        # Convertir a datetime para queries que comparan contra DateTime
        f_desde_dt = datetime.combine(fecha_desde, datetime.min.time())
        f_hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

        if incluir['ventas']:
            asientos += asientos_de_ventas(f_desde_dt, f_hasta_dt)
        if incluir['nc']:
            asientos += asientos_de_notas_credito(f_desde_dt, f_hasta_dt)
        if incluir['compras']:
            asientos += asientos_de_compras(fecha_desde, fecha_hasta)
        if incluir['pagos']:
            asientos += asientos_de_pagos_proveedor(fecha_desde, fecha_hasta)
        if incluir['gastos']:
            asientos += asientos_de_gastos(fecha_desde, fecha_hasta)
        if incluir['mov_caja']:
            asientos += asientos_de_movimientos_caja(f_desde_dt, f_hasta_dt)
        if incluir['dif_caja']:
            asientos += asientos_de_diferencias_caja(f_desde_dt, f_hasta_dt)

        # Ordenar por fecha
        asientos.sort(key=lambda x: (x['fecha'], x['tipo']))

        if not asientos:
            return jsonify({'success': False, 'error': 'No hay movimientos en el período seleccionado'}), 404

        excel_buffer = generar_excel(asientos, fecha_desde, fecha_hasta)
        nombre = f"Puente_Contable_{fecha_desde.strftime('%Y%m%d')}_a_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@puente_contable_bp.route('/preview', methods=['POST'])
def preview():
    """Devuelve la cantidad de asientos que se generarán, sin armar Excel."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    import traceback as _tb

    def _safe_count(label, fn, *args):
        """Llama una funcion generadora y devuelve la cantidad. Si falla, loguea y devuelve 0."""
        try:
            return len(fn(*args))
        except Exception as exc:
            print(f"[Puente Contable] ERROR generando '{label}': {exc}")
            _tb.print_exc()
            return -1  # marca de error para mostrar en UI

    try:
        desde_str = request.form.get('desde') or (request.json or {}).get('desde')
        hasta_str = request.form.get('hasta') or (request.json or {}).get('hasta')
        if not desde_str or not hasta_str:
            return jsonify({'success': False, 'error': 'Falta fecha desde/hasta'}), 400

        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        f_desde_dt = datetime.combine(fecha_desde, datetime.min.time())
        f_hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

        print(f"[Puente Contable] Preview {fecha_desde} a {fecha_hasta}")

        c_ventas  = _safe_count('ventas',         asientos_de_ventas,            f_desde_dt, f_hasta_dt)
        c_nc      = _safe_count('notas_credito',  asientos_de_notas_credito,     f_desde_dt, f_hasta_dt)
        c_compras = _safe_count('compras',        asientos_de_compras,           fecha_desde, fecha_hasta)
        c_pagos   = _safe_count('pagos_prov',     asientos_de_pagos_proveedor,   fecha_desde, fecha_hasta)
        c_gastos  = _safe_count('gastos',         asientos_de_gastos,            fecha_desde, fecha_hasta)
        c_mov     = _safe_count('mov_caja',       asientos_de_movimientos_caja,  f_desde_dt, f_hasta_dt)
        c_dif     = _safe_count('dif_caja',       asientos_de_diferencias_caja,  f_desde_dt, f_hasta_dt)

        # Si todos dieron -1 (todos rompieron) devolvemos error
        valores = [c_ventas, c_nc, c_compras, c_pagos, c_gastos, c_mov, c_dif]
        if all(v == -1 for v in valores):
            return jsonify({'success': False, 'error': 'Error generando todos los asientos. Revisar consola Flask.'}), 500

        # Sumamos solo los que no dieron error
        total = sum(v for v in valores if v >= 0)

        return jsonify({
            'success': True,
            'ventas':   c_ventas if c_ventas  >= 0 else 'ERROR',
            'nc':       c_nc      if c_nc      >= 0 else 'ERROR',
            'compras':  c_compras if c_compras >= 0 else 'ERROR',
            'pagos':    c_pagos   if c_pagos   >= 0 else 'ERROR',
            'gastos':   c_gastos  if c_gastos  >= 0 else 'ERROR',
            'mov_caja': c_mov     if c_mov     >= 0 else 'ERROR',
            'dif_caja': c_dif     if c_dif     >= 0 else 'ERROR',
            'total': total,
        })
    except Exception as e:
        _tb.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════
# INIT (inyección de modelos desde app.py)
# ════════════════════════════════════════════════════════════════════════════
def init_puente_contable(app, db, **modelos):
    """Registra el blueprint y guarda los modelos para uso interno.

    Modelos esperados (todos opcionales — si falta alguno, esa parte del
    export simplemente devuelve []):
      Factura, DetalleFactura, MedioPago, NotaCredito, Cliente, Usuario,
      Gasto, FacturaCompra, PagoProveedor, PagoProveedorMedio, Proveedor,
      CajaAperturaModel, MovimientoCajaModel
    """
    global _db, _models
    _db = db
    _models = modelos
    app.register_blueprint(puente_contable_bp)