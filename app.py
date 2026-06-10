# ================ SISTEMA DE COMPRAS ================
from compras import compras_bp # app.py - Sistema de Punto de Venta Argentina con Flask, MySQL, ARCA e Impresión Térmica

#http://localhost:5080/api/comparar_stocks muestra stock dnamico en combos y el stock de prord. base

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from functools import wraps
import sys 
from extensions import db
from sqlalchemy import Numeric, or_, and_, func, desc, asc, case, text  
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from qr_afip import crear_generador_qr
from io import BytesIO
import os
import requests
import xml.etree.ElementTree as ET
from zeep import Client
from zeep.wsse import BinarySignature
import base64
import hashlib
import csv
import io
from flask import make_response
# Importar la función del PDF
from reporte_ventas_pdf import generar_pdf_reporte_ventas
from cryptography import x509
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import padding
import json
import subprocess
import pymysql
pymysql.install_as_MySQLdb()
import MySQLdb.cursors
from estadisticas import init_estadisticas
from caja import init_caja_system, CajaAperturaModel
from notas_credito import notas_credito_bp
from cta_cte import init_cta_cte
from stock_audit import init_stock_audit, registrar_movimiento_stock
from pedidos import init_pedidos
from cta_cte import (
    guardar_venta_fiada,
    obtener_productos_pendientes,
    marcar_productos_como_pagados,
    obtener_saldo_cliente,
    registrar_comprobante_en_cta_cte
)
from reporte_ctacte_pdf import generar_pdf_cuentas_corrientes

# ================ SISTEMA DE VERIFICACIÓN DE LICENCIAS (WEB) ================
from verificador_licencias_web import verificar_licencia
from functools import wraps

# ================ MODULO DE PRESUPUESTO ================
from presupuestos import init_presupuestos

# ================ MODULO PUENTE CONTABLE (Exportacion para Estudios Contables) ================
from puente_contable import init_puente_contable

# ================ MODULO DE IMPORTACION ================
from importador import init_importador
from mobile import init_mobile
from mercadopago_qr import init_mp_qr

# ================ FIX SSL COMPATIBLE PARA AFIP ================
import ssl
import urllib3
from urllib3.util import ssl_
from requests.adapters import HTTPAdapter
from requests import Session




def configurar_ssl_afip():
    """Configuración SSL compatible para todas las versiones de Python"""
    
    # Variables de entorno
    os.environ['PYTHONHTTPSVERIFY'] = '0'
    os.environ['CURL_CA_BUNDLE'] = ''
    os.environ['REQUESTS_CA_BUNDLE'] = ''
    
    def create_afip_ssl_context():
        """Crear contexto SSL para AFIP"""
        try:
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        except AttributeError:
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS)
        
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        # Ciphers más permisivos - probar niveles de seguridad
        try:
            ctx.set_ciphers('ALL:@SECLEVEL=0')
        except ssl.SSLError:
            try:
                ctx.set_ciphers('ALL:@SECLEVEL=1')
            except ssl.SSLError:
                ctx.set_ciphers('ALL')
        
        # Aplicar opciones SSL disponibles
        for opcion in ['OP_LEGACY_SERVER_CONNECT', 'OP_ALLOW_UNSAFE_LEGACY_RENEGOTIATION', 'OP_ALL']:
            if hasattr(ssl, opcion):
                try:
                    ctx.options |= getattr(ssl, opcion)
                except:
                    pass
        
        return ctx
    
    # Aplicar configuración
    ssl._create_default_https_context = create_afip_ssl_context
    ssl_.create_urllib3_context = create_afip_ssl_context
    urllib3.disable_warnings()
    
    print("✅ Configuración SSL para AFIP aplicada")

# Aplicar configuración SSL
configurar_ssl_afip()

def crear_session_afip():
    """Crear sesión HTTP personalizada para AFIP"""
    
    class AFIPAdapter(HTTPAdapter):
        def init_poolmanager(self, *args, **kwargs):
            try:
                ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
            except AttributeError:
                ctx = ssl.SSLContext(ssl.PROTOCOL_TLS)
            
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            try:
                ctx.set_ciphers('ALL:@SECLEVEL=0')
            except ssl.SSLError:
                try:
                    ctx.set_ciphers('ALL:@SECLEVEL=1')
                except ssl.SSLError:
                    ctx.set_ciphers('ALL')
            
            # Aplicar opciones disponibles
            for opcion in ['OP_LEGACY_SERVER_CONNECT', 'OP_ALLOW_UNSAFE_LEGACY_RENEGOTIATION', 'OP_ALL']:
                if hasattr(ssl, opcion):
                    try:
                        ctx.options |= getattr(ssl, opcion)
                    except:
                        pass
            
            kwargs['ssl_context'] = ctx
            return super().init_poolmanager(*args, **kwargs)
    
    session = Session()
    session.mount('https://', AFIPAdapter())
    session.verify = False
    return session

app = Flask(__name__)

app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui'

app.jinja_env.globals['hasattr'] = hasattr

# Inicializar módulo de cuenta corriente y los demas 
init_cta_cte(app)
init_stock_audit(app)
init_pedidos(app, db)


# Intentar cargar configuración local, si no existe usar por defecto
try:
    from config_cliente import Config, ARCAConfig  # ← Configuración centralizada
    app.config.from_object(Config)
    ARCA_CONFIG = ARCAConfig()
except ImportError:
    # Configuración por defecto si no existe config_cliente.py
    app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://pos_user:pos_password@localhost/carnave'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    
    class DefaultARCAConfig:
        CUIT = '27333429433'
        PUNTO_VENTA = 1
        CERT_PATH = 'certificados/certificado.crt'
        KEY_PATH = 'certificados/private.key'
        USE_HOMOLOGACION = False  # PRODUCCIÓN
        
        @property
        def WSAA_URL(self):
            return 'https://wsaahomo.afip.gov.ar/ws/services/LoginCms' if self.USE_HOMOLOGACION else 'https://wsaa.afip.gov.ar/ws/services/LoginCms'
        
        @property
        def WSFEv1_URL(self):
            return 'https://wswhomo.afip.gov.ar/wsfev1/service.asmx?WSDL' if self.USE_HOMOLOGACION else 'https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL'
        
        TOKEN_CACHE_FILE = 'cache/token_arca.json'
    
    ARCA_CONFIG = DefaultARCAConfig()

# ================ VERIFICACIÓN DE LICENCIA AL INICIO ================
print(f"\n{'='*60}")
print(f"🔐 VERIFICANDO LICENCIA DEL SISTEMA")
print(f"{'='*60}")

# Obtener CUIT desde la configuración
CUIT_SISTEMA = ARCA_CONFIG.CUIT
print(f"CUIT: {CUIT_SISTEMA}")

# Verificar licencia
resultado_licencia = verificar_licencia(CUIT_SISTEMA)


# Guardar info de licencia para usar en templates
app.config['LICENCIA_INFO'] = resultado_licencia

# ================ DECORADOR PARA PROTEGER RUTAS ================
def requiere_licencia_activa(f):
    """
    Decorador que verifica la licencia antes de permitir acceso
    Redirige a página de bloqueo si está bloqueado
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        licencia_info = _licencia_actual()
        tipo_bloqueo = licencia_info.get('tipo_bloqueo', 'error')
        
        # Si está bloqueado, redirigir a página de bloqueo
        if tipo_bloqueo in ['bloqueado', 'no_encontrada', 'error']:
            return redirect(url_for('licencia_bloqueada'))
        
        # Si está en mora o sin bloqueo, permitir acceso
        return f(*args, **kwargs)
    
    return decorated_function


def _licencia_actual():
    """Devuelve el estado de licencia. Si el estado GUARDADO es un fallo
    recuperable (error / offline_expirado), REINTENTA verificar — por si
    ya volvió internet o ya se generó el cache local. Si el estado es
    válido (sin_bloqueo / mora), usa lo guardado SIN reintentar (rápido,
    no agrega latencia al uso normal).

    Esto evita que el sistema quede bloqueado para siempre solo porque
    al arrancar (python app.py) no había internet en ese instante.
    Antes la licencia se verificaba UNA sola vez al inicio y el
    resultado quedaba congelado hasta reiniciar el sistema.
    """
    info = app.config.get('LICENCIA_INFO', {})
    tipo = info.get('tipo_bloqueo', 'error')
    if tipo in ['error', 'offline_expirado']:
        try:
            nuevo = verificar_licencia(CUIT_SISTEMA)
            app.config['LICENCIA_INFO'] = nuevo
            return nuevo
        except Exception as e:
            print(f"WARN: reintento de verificación de licencia falló: {e}")
            return info
    return info


db.init_app(app)

# ================ MODULO DE REMITOS  ================
from remitos import remitos_bp
app.register_blueprint(remitos_bp)

# ================ Blueprint Zonas de Reparto ================
from zonas import zonas_bp
app.register_blueprint(zonas_bp)

# ================ Blueprint de Reparto ================
from reparto import reparto_bp
app.register_blueprint(reparto_bp)


# ================ SISTEMA DE PROVEEDORES ================
from proveedores import proveedores_bp
app.register_blueprint(proveedores_bp)

# ================ VENDEDORES (módulo inline, no requiere vendedores.py) ================
# ABM de vendedores: modelo + rutas. Baja lógica (activo=0).
# Requiere: tabla `vendedor` en la BD (crear_tabla_vendedor.sql) y
# templates/vendedores.html

from flask import Blueprint

vendedores_bp = Blueprint('vendedores', __name__)


class Vendedor(db.Model):
    __tablename__ = 'vendedor'

    id             = db.Column(db.Integer, primary_key=True)
    nombre         = db.Column(db.String(150), nullable=False)   # Nombre y Apellido
    tipo_documento = db.Column(db.String(20), nullable=False, default='DNI')
    documento      = db.Column(db.String(20))
    direccion      = db.Column(db.String(200))
    telefono       = db.Column(db.String(30))
    notas          = db.Column(db.Text)
    activo         = db.Column(db.Boolean, nullable=False, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)

    def __repr__(self):
        return f'<Vendedor {self.nombre}>'

    def to_dict(self):
        return {
            'id':             self.id,
            'nombre':         self.nombre,
            'tipo_documento': self.tipo_documento or 'DNI',
            'documento':      self.documento or '',
            'direccion':      self.direccion or '',
            'telefono':       self.telefono or '',
            'notas':          self.notas or '',
            'activo':         bool(self.activo),
        }


@vendedores_bp.route('/vendedores')
def vendedores():
    """Pantalla ABM de vendedores."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    lista = Vendedor.query.filter_by(activo=True).order_by(Vendedor.nombre).all()
    return render_template('vendedores.html', vendedores=lista)


@vendedores_bp.route('/api/vendedores', methods=['GET'])
def api_vendedores():
    """Lista de vendedores activos. Filtro opcional ?q= por nombre o documento."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        q = request.args.get('q', '').strip()
        query = Vendedor.query.filter_by(activo=True)
        if q:
            query = query.filter(or_(
                Vendedor.nombre.ilike(f'%{q}%'),
                Vendedor.documento.ilike(f'%{q}%')
            ))
        lista = query.order_by(Vendedor.nombre).all()
        return jsonify({'success': True, 'vendedores': [v.to_dict() for v in lista]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@vendedores_bp.route('/api/vendedor/<int:vend_id>', methods=['GET'])
def api_vendedor_get(vend_id):
    """Datos de un vendedor (para cargar el modal de edición)."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    v = Vendedor.query.get_or_404(vend_id)
    return jsonify(v.to_dict())


@vendedores_bp.route('/api/vendedor/nuevo', methods=['POST'])
def api_vendedor_nuevo():
    """Alta de vendedor."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data = request.json or {}
        nombre = (data.get('nombre') or '').strip()
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre es obligatorio'}), 400

        documento = (data.get('documento') or '').strip()
        if documento:
            existe = Vendedor.query.filter_by(documento=documento, activo=True).first()
            if existe:
                return jsonify({'success': False,
                                'error': f'Ya existe un vendedor activo con documento {documento}: {existe.nombre}'}), 400

        v = Vendedor(
            nombre         = nombre,
            tipo_documento = (data.get('tipo_documento') or 'DNI').strip(),
            documento      = documento or None,
            direccion      = (data.get('direccion') or '').strip() or None,
            telefono       = (data.get('telefono') or '').strip() or None,
            notas          = (data.get('notas') or '').strip() or None,
            activo         = True,
        )
        db.session.add(v)
        db.session.commit()
        return jsonify({'success': True, 'id': v.id,
                        'message': f'Vendedor "{v.nombre}" creado correctamente'})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@vendedores_bp.route('/api/vendedor/<int:vend_id>/editar', methods=['POST'])
def api_vendedor_editar(vend_id):
    """Modificación de vendedor."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        v = Vendedor.query.get_or_404(vend_id)
        data = request.json or {}
        nombre = (data.get('nombre') or '').strip()
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre es obligatorio'}), 400

        documento = (data.get('documento') or '').strip()
        if documento:
            existe = Vendedor.query.filter(
                Vendedor.documento == documento,
                Vendedor.activo == True,
                Vendedor.id != v.id
            ).first()
            if existe:
                return jsonify({'success': False,
                                'error': f'Ya existe otro vendedor activo con documento {documento}: {existe.nombre}'}), 400

        v.nombre         = nombre
        v.tipo_documento = (data.get('tipo_documento') or 'DNI').strip()
        v.documento      = documento or None
        v.direccion      = (data.get('direccion') or '').strip() or None
        v.telefono       = (data.get('telefono') or '').strip() or None
        v.notas          = (data.get('notas') or '').strip() or None
        db.session.commit()
        return jsonify({'success': True,
                        'message': f'Vendedor "{v.nombre}" actualizado correctamente'})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@vendedores_bp.route('/api/vendedor/<int:vend_id>/eliminar', methods=['POST'])
def api_vendedor_eliminar(vend_id):
    """Baja LÓGICA del vendedor (activo=0). No se borra el registro para no
    romper referencias futuras (ventas, comisiones, etc.)."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        v = Vendedor.query.get_or_404(vend_id)
        v.activo = False
        db.session.commit()
        return jsonify({'success': True,
                        'message': f'Vendedor "{v.nombre}" eliminado correctamente'})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


app.register_blueprint(vendedores_bp)

# ================ ORDENES DE COMPRA (MÓDULO 5) ================
from ordenes_compra import ordenes_compra_bp
app.register_blueprint(ordenes_compra_bp)

# ================ categorías ================
from categorias import categorias_bp
app.register_blueprint(categorias_bp)

# ================ sugerencias compras ================
from sugerencia_compra import sugerencia_compra_bp     # ← NUEVO
app.register_blueprint(sugerencia_compra_bp)           # ← NUEVO

# ================ SISTEMA DE COMPRAS CON DETALLE ================
from compras import compras_bp
app.register_blueprint(compras_bp)

# ================ INTEGRACIÓN MERCADO PAGO QR ================
init_mp_qr(app)

# ================ SISTEMA DE IMPRESIÓN TÉRMICA ================
import tempfile
try:
    import win32print
    import win32api
    IMPRESION_DISPONIBLE = True
    print("✅ Sistema de impresión disponible")
except ImportError:
    IMPRESION_DISPONIBLE = False
    print("⚠️ Sistema de impresión no disponible (instalar: pip install pywin32)")

# IMPORTAR LA IMPRESORA DESDE EL ARCHIVO SEPARADO
from impresora_termica import impresora_termica

# Modelos de Base de Datos
class Configuracion(db.Model):
    __tablename__ = 'configuracion'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    clave = db.Column(db.String(100), unique=True, nullable=False)
    valor = db.Column(db.Text)
    descripcion = db.Column(db.String(255))
    fecha_modificacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


# ═══════════════════════════════════════════════════════════
# REDONDEO DE PRECIOS — Configurable por instalación
# Si configuracion.redondear_precios = 'true':  10.58 → 11
# Si configuracion.redondear_precios = 'false': 10.58 → 10.58 (default)
# ═══════════════════════════════════════════════════════════
import time as _time_redondeo
_cache_redondeo = {'valor': None, 'timestamp': 0}

def debe_redondear_precios():
    """Lee la config con cache de 60 seg para no consultar BD en cada cálculo"""
    ahora = _time_redondeo.time()
    if _cache_redondeo['valor'] is None or (ahora - _cache_redondeo['timestamp']) > 60:
        try:
            config = Configuracion.query.filter_by(clave='redondear_precios').first()
            _cache_redondeo['valor'] = (config and config.valor and 
                                        config.valor.strip().lower() == 'true')
            _cache_redondeo['timestamp'] = ahora
        except Exception as e:
            print(f"⚠️ Error leyendo config redondear_precios: {e}")
            _cache_redondeo['valor'] = False
    return _cache_redondeo['valor']


def aplicar_redondeo_precio(precio):
    """
    Aplica redondeo a un precio según la flag de configuración.
    - Si la flag está activa: 10.58 → Decimal('11'), 10.25 → Decimal('10')
    - Si está desactivada: 10.58 → Decimal('10.58') (comportamiento original)
    Devuelve None si el precio es inválido.
    """
    if precio is None:
        return None
    try:
        precio_float = float(precio)
        if debe_redondear_precios():
            return Decimal(str(round(precio_float)))
        else:
            return Decimal(str(round(precio_float, 2)))
    except (ValueError, TypeError):
        return None


def invalidar_cache_redondeo():
    """Llamar cuando se cambia la config desde la UI para que aplique al instante"""
    _cache_redondeo['valor'] = None
    _cache_redondeo['timestamp'] = 0


@app.route('/api/config/redondeo')
def api_config_redondeo():
    """Endpoint para que el frontend lea si tiene que redondear precios"""
    return jsonify({'redondear_precios': debe_redondear_precios()})


class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    rol = db.Column(db.String(50), default='vendedor')
    activo = db.Column(db.Boolean, default=True)
    punto_venta = db.Column(db.Integer, nullable=False, default=1)   # PV asignado al usuario (1..N)
    puede_liquidar = db.Column(db.Boolean, default=False, nullable=False)  # liquidación a intermediarios

class Zona(db.Model):
    __tablename__ = 'zona'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(60), nullable=False, unique=True)
    descripcion = db.Column(db.String(255))
    color = db.Column(db.String(7), default='#0d6efd')
    orden_reparto = db.Column(db.Integer, default=0)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    documento = db.Column(db.String(20))
    tipo_documento = db.Column(db.String(10))  # DNI, CUIT, etc.
    email = db.Column(db.String(100))
    telefono = db.Column(db.String(20))
    direccion = db.Column(db.Text)
    condicion_iva = db.Column(db.String(50))  # Responsable Inscripto, Monotributista, etc.
    tipo_precio = db.Column(db.String(10), default='venta')
    lista_precio = db.Column(db.Integer, default=1)
    saldo = db.Column(db.Numeric(12, 2), default=0.00)
    zona_id = db.Column(db.Integer, db.ForeignKey('zona.id'), nullable=True)  # zona de reparto
    es_intermediario = db.Column(db.Boolean, default=False, nullable=False)  # consignatario / intermediario
    vendedor_id = db.Column(db.Integer, nullable=True)  # vendedor asignado (FK a vendedor.id a nivel BD)
    dias_vencimiento_factura = db.Column(db.Integer, nullable=True)  # la factura vence a N días desde la emisión

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(Numeric(10, 2), nullable=False)
    stock = db.Column(Numeric(10, 3), nullable=False, default=0.000)
    categoria = db.Column(db.String(100))
    iva = db.Column(Numeric(5, 2), default=21.00)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    fecha_actualizacion_precio = db.Column(db.DateTime, nullable=True)
    costo = db.Column(Numeric(10, 2), default=0.00)
    margen = db.Column(Numeric(5, 2), default=30.00)
    stock_minimo = db.Column(Numeric(10, 3), default=0)
    
    # Campos de combo
    es_combo = db.Column(db.Boolean, default=False)
    producto_base_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=True)
    cantidad_combo = db.Column(Numeric(8, 3), default=1.000)
    precio_unitario_base = db.Column(Numeric(10, 2), nullable=True)
    descuento_porcentaje = db.Column(Numeric(5, 2), default=0.00)
    acceso_rapido = db.Column(db.Boolean, default=False)
    orden_acceso_rapido = db.Column(db.Integer, default=0)
    producto_base_2_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=True)
    cantidad_combo_2 = db.Column(Numeric(8, 3), default=0.000)
    producto_base_3_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=True)
    cantidad_combo_3 = db.Column(Numeric(8, 3), default=0.000)
    codigo_barras = db.Column(db.String(50), nullable=True, default=None)
    margen2 = db.Column(Numeric(5, 2), nullable=True)
    margen3 = db.Column(Numeric(5, 2), nullable=True)
    margen4 = db.Column(Numeric(5, 2), nullable=True)
    margen5 = db.Column(Numeric(5, 2), nullable=True)
    precio2 = db.Column(Numeric(10, 2), nullable=True)
    precio3 = db.Column(Numeric(10, 2), nullable=True)
    precio4 = db.Column(Numeric(10, 2), nullable=True)
    precio5 = db.Column(Numeric(10, 2), nullable=True)
    es_pesable = db.Column(db.Boolean, default=False)
    imagen_url = db.Column(db.String(500), nullable=True)

    # Relaciones
    producto_base = db.relationship('Producto', 
                                    foreign_keys=[producto_base_id], 
                                    remote_side=[id], 
                                    backref='combos_derivados')

    producto_base_2 = db.relationship('Producto', 
                                    foreign_keys=[producto_base_2_id], 
                                    remote_side=[id])

    producto_base_3 = db.relationship('Producto', 
                                    foreign_keys=[producto_base_3_id], 
                                    remote_side=[id])
        
    def __repr__(self):
        return f'<Producto {self.codigo}: {self.nombre}>'
    
    def to_dict(self):
        """Convertir producto a diccionario"""
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'precio': float(self.precio),
            'costo': float(self.costo) if self.costo else 0.0,
            'margen': float(self.margen) if self.margen else 0.0,
            'stock': self.stock_dinamico,
            'categoria': self.categoria,
            'iva': float(self.iva),
            'activo': self.activo,
            'fecha_creacion': self.fecha_creacion.isoformat() if self.fecha_creacion else None,
            'fecha_modificacion': self.fecha_modificacion.isoformat() if self.fecha_modificacion else None,
            'es_combo': self.es_combo,
            'producto_base_id': self.producto_base_id,
            'cantidad_combo': float(self.cantidad_combo) if self.cantidad_combo else 1.0,
            'precio_unitario_base': float(self.precio_unitario_base) if self.precio_unitario_base else float(self.precio),
            'descuento_porcentaje': float(self.descuento_porcentaje) if self.descuento_porcentaje else 0.0,
            'ahorro_combo': self.calcular_ahorro_combo(),
            'precio_normal': self.calcular_precio_normal(),
            'producto_base_nombre': self.producto_base.nombre if self.producto_base else None
        }
    
    @property
    def precio_calculado(self):
        """Calcular precio basado en costo y margen"""
        if self.costo and self.margen is not None:
            return float(self.costo) * (1 + (float(self.margen) / 100))
        return float(self.precio)
    
    def actualizar_precio_desde_costo_margen(self):
        """Actualizar el precio basado en costo y margen"""
        if self.costo and self.margen is not None:
            self.precio = Decimal(str(self.precio_calculado))
            self.fecha_modificacion = datetime.now()
    
    @staticmethod
    def calcular_precio_venta(costo, margen):
        """Método estático para calcular precio de venta"""
        if not costo or margen is None:
            return 0.0
        return float(costo) * (1 + (float(margen) / 100))
    
    def calcular_precio_normal(self):
        """Calcular precio normal sin descuento"""
        if self.es_combo and self.precio_unitario_base and self.cantidad_combo:
            return float(self.precio_unitario_base) * float(self.cantidad_combo)
        return float(self.precio)
    
    def calcular_ahorro_combo(self):
        """Calcular cuánto se ahorra con el combo"""
        if self.es_combo:
            precio_normal = self.calcular_precio_normal()
            precio_combo = float(self.precio)
            return precio_normal - precio_combo
        return 0.0
    
    def obtener_descripcion_completa(self):
        """Obtener descripción que incluye información del combo"""
        if self.es_combo:
            ahorro = self.calcular_ahorro_combo()
            cantidad_str = f"{self.cantidad_combo:g}"
            return f"{self.nombre} - {cantidad_str} unidades (Ahorro: ${ahorro:.0f})"
        return self.nombre
    
    def obtener_precio_con_oferta(self, cantidad):
        """Obtener precio considerando ofertas por volumen"""
        try:
            cantidad_decimal = float(cantidad)
            
            oferta = OfertaVolumen.query.filter(
                and_(
                    OfertaVolumen.producto_id == self.id,
                    OfertaVolumen.cantidad_minima <= cantidad_decimal,
                    OfertaVolumen.activo == True
                )
            ).order_by(OfertaVolumen.cantidad_minima.desc()).first()
            
            if oferta:
                return float(oferta.precio_oferta)
            else:
                return float(self.precio)
                
        except Exception as e:
            print(f"Error calculando precio con oferta: {e}")
            return float(self.precio)

    def obtener_info_oferta(self, cantidad):
        """Obtener información detallada de la oferta aplicada"""
        try:
            cantidad_decimal = float(cantidad)
            precio_normal = float(self.precio)
            precio_con_oferta = self.obtener_precio_con_oferta(cantidad_decimal)
            
            oferta = OfertaVolumen.query.filter(
                and_(
                    OfertaVolumen.producto_id == self.id,
                    OfertaVolumen.cantidad_minima <= cantidad_decimal,
                    OfertaVolumen.activo == True
                )
            ).order_by(OfertaVolumen.cantidad_minima.desc()).first()
            
            if oferta and precio_con_oferta < precio_normal:
                ahorro_unitario = precio_normal - precio_con_oferta
                ahorro_total = ahorro_unitario * cantidad_decimal
                
                return {
                    'tiene_oferta': True,
                    'precio_normal': precio_normal,
                    'precio_oferta': precio_con_oferta,
                    'ahorro_unitario': round(ahorro_unitario, 2),
                    'ahorro_total': round(ahorro_total, 2),
                    'cantidad_minima': float(oferta.cantidad_minima),
                    'descripcion_oferta': oferta.descripcion or f"Oferta por volumen desde {oferta.cantidad_minima} unidades"
                }
            
            return {
                'tiene_oferta': False,
                'precio_normal': precio_normal,
                'precio_oferta': precio_normal
            }
            
        except Exception as e:
            print(f"Error obteniendo info de oferta: {e}")
            return {
                'tiene_oferta': False,
                'precio_normal': float(self.precio),
                'precio_oferta': float(self.precio)
            }

    def tiene_ofertas_volumen(self):
        """Verificar si el producto tiene ofertas por volumen activas"""
        return OfertaVolumen.query.filter_by(
            producto_id=self.id,
            activo=True
        ).count() > 0

    def calcular_stock_disponible_combo(self):
        """Calcular stock disponible para combos basado en productos base"""
        if not self.es_combo:
            return self.stock
        
        try:
            stocks_disponibles = []
            
            # Producto base 1 (obligatorio)
            if self.producto_base_id and self.cantidad_combo and float(self.cantidad_combo) > 0:
                producto_base = Producto.query.get(self.producto_base_id)
                if producto_base and producto_base.activo:
                    cantidad_necesaria = float(self.cantidad_combo)
                    stock_posible = int(float(producto_base.stock) / cantidad_necesaria) if cantidad_necesaria > 0 else 0
                    stocks_disponibles.append(stock_posible)
            
            # Producto base 2 (opcional)
            if self.producto_base_2_id and self.cantidad_combo_2 and float(self.cantidad_combo_2) > 0:
                producto_base_2 = Producto.query.get(self.producto_base_2_id)
                if producto_base_2 and producto_base_2.activo:
                    cantidad_necesaria = float(self.cantidad_combo_2)
                    stock_posible = int(float(producto_base_2.stock) / cantidad_necesaria) if cantidad_necesaria > 0 else 0
                    stocks_disponibles.append(stock_posible)
            
            # Producto base 3 (opcional)
            if self.producto_base_3_id and self.cantidad_combo_3 and float(self.cantidad_combo_3) > 0:
                producto_base_3 = Producto.query.get(self.producto_base_3_id)
                if producto_base_3 and producto_base_3.activo:
                    cantidad_necesaria = float(self.cantidad_combo_3)
                    stock_posible = int(float(producto_base_3.stock) / cantidad_necesaria) if cantidad_necesaria > 0 else 0
                    stocks_disponibles.append(stock_posible)
            
            return min(stocks_disponibles) if stocks_disponibles else 0
            
        except Exception as e:
            print(f"Error calculando stock de combo {self.codigo}: {e}")
            return 0

    @property
    def stock_dinamico(self):
        """Propiedad que devuelve stock dinámico para combos, stock normal para productos base"""
        if self.es_combo:
            return self.calcular_stock_disponible_combo()
        else:
            return self.stock

    def debug_stock_combo(self):
        """Función de debug para ver cálculo de stock paso a paso"""
        if not self.es_combo:
            return f"Producto base {self.codigo}: stock normal = {self.stock}"
        
        debug_info = [f"DEBUG COMBO {self.codigo}:"]
        
        if self.producto_base_id and self.cantidad_combo:
            producto_base = Producto.query.get(self.producto_base_id)
            if producto_base:
                debug_info.append(f"  Base 1: {producto_base.codigo} stock={producto_base.stock}, necesita={self.cantidad_combo}")
        
        if self.producto_base_2_id and self.cantidad_combo_2:
            producto_base_2 = Producto.query.get(self.producto_base_2_id)
            if producto_base_2:
                debug_info.append(f"  Base 2: {producto_base_2.codigo} stock={producto_base_2.stock}, necesita={self.cantidad_combo_2}")
        
        if self.producto_base_3_id and self.cantidad_combo_3:
            producto_base_3 = Producto.query.get(self.producto_base_3_id)
            if producto_base_3:
                debug_info.append(f"  Base 3: {producto_base_3.codigo} stock={producto_base_3.stock}, necesita={self.cantidad_combo_3}")
        
        debug_info.append(f"  Stock dinámico resultante: {self.stock_dinamico}")
        return "\n".join(debug_info)

    @staticmethod
    def obtener_productos_con_ofertas():
        """Obtener productos base con sus ofertas"""
        productos_base = Producto.query.filter_by(es_combo=False, activo=True).all()
        
        resultado = []
        for producto_base in productos_base:
            item_base = producto_base.to_dict()
            item_base['tipo'] = 'BASE'
            resultado.append(item_base)
            
            combos = Producto.query.filter_by(
                producto_base_id=producto_base.id, 
                es_combo=True, 
                activo=True
            ).order_by(Producto.precio).all()
            
            for combo in combos:
                item_combo = combo.to_dict()
                item_combo['tipo'] = 'COMBO'
                resultado.append(item_combo)
        
        return resultado

    @staticmethod
    def obtener_con_ofertas():
        """Obtener productos que tienen ofertas por volumen"""
        return db.session.query(Producto).join(OfertaVolumen).filter(
            and_(
                Producto.activo == True,
                OfertaVolumen.activo == True
            )
        ).distinct().all()



class OfertaVolumen(db.Model):
    """Modelo para ofertas por volumen de productos"""
    __tablename__ = 'ofertas_volumen'
    
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad_minima = db.Column(Numeric(10, 3), nullable=False)
    precio_oferta = db.Column(Numeric(10, 2), nullable=False)
    descripcion = db.Column(db.String(200))
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    # Relación con Producto
    producto = db.relationship('Producto', backref=db.backref('ofertas_volumen', lazy=True))
    
    def __repr__(self):
        return f'<OfertaVolumen {self.producto.codigo if self.producto else "SIN_PRODUCTO"}: {self.cantidad_minima}+ = ${self.precio_oferta}>'
    
    def to_dict(self):
        """Convertir a diccionario"""
        return {
            'id': self.id,
            'producto_id': self.producto_id,
            'cantidad_minima': float(self.cantidad_minima),
            'precio_oferta': float(self.precio_oferta),
            'descripcion': self.descripcion,
            'activo': self.activo,
            'fecha_creacion': self.fecha_creacion.isoformat() if self.fecha_creacion else None,
            'producto': {
                'codigo': self.producto.codigo,
                'nombre': self.producto.nombre
            } if self.producto else None
        }


    

class Factura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(50), unique=True)
    tipo_comprobante = db.Column(db.String(10))  # FA, FB, FC, etc.
    punto_venta = db.Column(db.Integer)
    fecha = db.Column(db.DateTime, default=datetime.now)  # ← Cambiar de utcnow a now
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    subtotal = db.Column(Numeric(10, 2))
    iva = db.Column(Numeric(10, 2))
    total = db.Column(Numeric(10, 2))
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, autorizada, anulada
    cae = db.Column(db.String(50))  # Código de Autorización Electrónico
    vto_cae = db.Column(db.Date)
    interno_origen_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=True, unique=True)  # consignación: factura que proviene de un Interno
    
    cliente = db.relationship('Cliente', backref='facturas')
    usuario = db.relationship('Usuario', backref='facturas')

class DetalleFactura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(Numeric(10, 2), nullable=False)
    subtotal = db.Column(Numeric(10, 2), nullable=False)
    porcentaje_iva = db.Column(Numeric(5, 2), nullable=False, default=21.00)  # ← NUEVO CAMPO
    importe_iva = db.Column(Numeric(10, 2), nullable=False, default=0.00)    # ← NUEVO CAMPO
    costo_unitario = db.Column(Numeric(10, 2), nullable=True, default=0.00)  # ← costo al momento de la venta
    
    factura = db.relationship('Factura', backref='detalles')
    producto = db.relationship('Producto', backref='detalles_factura')


# ═══════════════════════════════════════════════════════════════════════
# LIQUIDACIONES A INTERMEDIARIOS (consignación FASE 2)
# ═══════════════════════════════════════════════════════════════════════
class LiquidacionIntermediario(db.Model):
    __tablename__ = 'liquidacion_intermediario'
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(50), unique=True, nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.now)
    intermediario_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    total_vendido = db.Column(Numeric(12, 2), default=0)
    total_costo = db.Column(Numeric(12, 2), default=0)
    total_liquidar = db.Column(Numeric(12, 2), default=0)
    base_calculo = db.Column(db.String(20), default='con_iva')  # 'con_iva' o 'sin_iva'
    estado = db.Column(db.String(20), default='pendiente_pago')  # pendiente_pago / pagada / anulada
    medio_pago = db.Column(db.String(50), nullable=True)
    fecha_pago = db.Column(db.DateTime, nullable=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    motivo = db.Column(db.Text, nullable=True)

    intermediario = db.relationship('Cliente', backref='liquidaciones')
    usuario = db.relationship('Usuario', backref='liquidaciones_generadas')


class LiquidacionDetalle(db.Model):
    __tablename__ = 'liquidacion_detalle'
    id = db.Column(db.Integer, primary_key=True)
    liquidacion_id = db.Column(db.Integer, db.ForeignKey('liquidacion_intermediario.id', ondelete='CASCADE'), nullable=False)
    interno_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=False)
    factura_derivada_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=True)
    monto_interno = db.Column(Numeric(12, 2))
    monto_factura = db.Column(Numeric(12, 2))
    monto_diferencia = db.Column(Numeric(12, 2))

    liquidacion = db.relationship('LiquidacionIntermediario', backref=db.backref('detalles', cascade='all, delete-orphan'))
    interno = db.relationship('Factura', foreign_keys=[interno_id], backref='liquidacion_detalles_como_interno')
    factura_derivada = db.relationship('Factura', foreign_keys=[factura_derivada_id], backref='liquidacion_detalles_como_factura')


class DescuentoFactura(db.Model):
    """Registro de descuentos aplicados a facturas - tabla independiente"""
    __tablename__ = 'descuentos_factura'
    
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'), unique=True, nullable=False)
    porcentaje_descuento = db.Column(Numeric(5, 2), nullable=False)
    monto_descuento = db.Column(Numeric(10, 2), nullable=False)
    total_original = db.Column(Numeric(10, 2), nullable=False)  # total antes del descuento
    fecha_aplicacion = db.Column(db.DateTime, default=datetime.now)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relaciones
    factura = db.relationship('Factura', backref=db.backref('descuento_aplicado', uselist=False))
    usuario = db.relationship('Usuario', backref='descuentos_aplicados')
    
    def __repr__(self):
        return f'<DescuentoFactura {self.factura_id}: {self.porcentaje_descuento}% = ${self.monto_descuento}>'
    
    def to_dict(self):
        return {
            'id': self.id,
            'factura_id': self.factura_id,
            'porcentaje_descuento': float(self.porcentaje_descuento),
            'monto_descuento': float(self.monto_descuento),
            'total_original': float(self.total_original),
            'fecha_aplicacion': self.fecha_aplicacion.isoformat(),
            'usuario': self.usuario.nombre if self.usuario else None
        }

#######################################################################################################
# ============================================================================
# PASO 1: MODELO DE NOTA DE CRÉDITO
# Agregar en app.py después del modelo Factura (después de línea 603)
# ============================================================================

class NotaCredito(db.Model):
    """Notas de Crédito electrónicas para anular/corregir facturas"""
    __tablename__ = 'notas_credito'
    
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(50), unique=True)  # 0001-00000001
    tipo_comprobante = db.Column(db.String(10))  # 03, 08, 13 (NC A, B, C)
    punto_venta = db.Column(db.Integer)
    fecha = db.Column(db.DateTime, default=datetime.now)
    
    # Relación con la factura que anula
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=False)
    factura_numero = db.Column(db.String(50))  # Para referencia rápida
    
    # Cliente (mismo que la factura)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    
    # Importes (normalmente iguales a la factura para NC total)
    subtotal = db.Column(Numeric(10, 2))
    iva = db.Column(Numeric(10, 2))
    total = db.Column(Numeric(10, 2))
    
    # Estado y autorización AFIP
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, autorizada, error_afip
    cae = db.Column(db.String(50))  # CAE de la NC
    vto_cae = db.Column(db.Date)
    error_afip = db.Column(db.Text, nullable=True)

    # Motivo de la NC
    motivo = db.Column(db.String(500))
    
    # Timestamps
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_autorizacion = db.Column(db.DateTime, nullable=True)
    
    # Relaciones
    factura = db.relationship('Factura', backref='notas_credito', foreign_keys=[factura_id])
    cliente = db.relationship('Cliente', backref='notas_credito')
    usuario = db.relationship('Usuario', backref='notas_credito')
    
    def __repr__(self):
        return f'<NotaCredito {self.numero} - Factura {self.factura_numero}>'
    
    def to_dict(self):
        return {
            'id': self.id,
            'numero': self.numero,
            'tipo_comprobante': self.tipo_comprobante,
            'fecha': self.fecha.isoformat() if self.fecha else None,
            'factura_numero': self.factura_numero,
            'cliente': self.cliente.nombre if self.cliente else None,
            'subtotal': float(self.subtotal) if self.subtotal else 0,
            'iva': float(self.iva) if self.iva else 0,
            'total': float(self.total) if self.total else 0,
            'estado': self.estado,
            'cae': self.cae,
            'vto_cae': self.vto_cae.isoformat() if self.vto_cae else None,
            'motivo': self.motivo
        }


class DetalleNotaCredito(db.Model):
    """Detalle de items de la Nota de Crédito"""
    __tablename__ = 'detalle_nota_credito'
    
    id = db.Column(db.Integer, primary_key=True)
    nota_credito_id = db.Column(db.Integer, db.ForeignKey('notas_credito.id'))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    cantidad = db.Column(Numeric(10, 3), nullable=False)  # Permitir decimales
    precio_unitario = db.Column(Numeric(10, 2), nullable=False)
    subtotal = db.Column(Numeric(10, 2), nullable=False)
    porcentaje_iva = db.Column(Numeric(5, 2), nullable=False, default=21.00)
    importe_iva = db.Column(Numeric(10, 2), nullable=False, default=0.00)
    
    # Relaciones
    nota_credito = db.relationship('NotaCredito', backref='detalles')
    producto = db.relationship('Producto', backref='detalles_nc')
    
    def __repr__(self):
        return f'<DetalleNC {self.nota_credito_id} - {self.cantidad}x Producto {self.producto_id}>'





########################################################################################################
class MedioPago(db.Model):
    """Tabla para registrar los medios de pago de cada factura"""
    __tablename__ = 'medios_pago'
    
    id = db.Column(db.Integer, primary_key=True)
    factura_id = db.Column(db.Integer, db.ForeignKey('factura.id'), nullable=False)
    medio_pago = db.Column(db.String(20), nullable=False)  # efectivo, credito, debito, mercado_pago
    importe = db.Column(Numeric(10, 2), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.now)
    
    # Relación con Factura
    factura = db.relationship('Factura', backref=db.backref('medios_pago', lazy=True))
    
    def __repr__(self):
        return f'<MedioPago {self.medio_pago}: ${self.importe}>'
    
    def to_dict(self):
        """Convertir a diccionario para JSON"""
        return {
            'id': self.id,
            'factura_id': self.factura_id,
            'medio_pago': self.medio_pago,
            'importe': float(self.importe),
            'fecha_registro': self.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
        }
    
    @staticmethod
    def obtener_medios_disponibles():
        """Retorna los medios de pago disponibles"""
        return [
            {'codigo': 'efectivo', 'nombre': 'Efectivo', 'icono': 'fas fa-money-bill-wave'},
            {'codigo': 'credito', 'nombre': 'Tarjeta de Crédito', 'icono': 'fas fa-credit-card'},
            {'codigo': 'debito', 'nombre': 'Tarjeta de Débito', 'icono': 'fas fa-credit-card'},
            {'codigo': 'mercado_pago', 'nombre': 'Mercado Pago', 'icono': 'fas fa-mobile-alt'}
        ]
    
    @staticmethod
    def calcular_recaudacion_por_fecha(fecha_desde, fecha_hasta):
        """Calcular recaudación por medio de pago en un rango de fechas"""
        try:
            from sqlalchemy import func, and_
            
            resultado = db.session.query(
                MedioPago.medio_pago,
                func.sum(MedioPago.importe).label('total'),
                func.count(MedioPago.id).label('cantidad_operaciones')
            ).filter(
                and_(
                    MedioPago.fecha_registro >= fecha_desde,
                    MedioPago.fecha_registro <= fecha_hasta
                )
            ).group_by(MedioPago.medio_pago).all()
            
            # Convertir a diccionario
            recaudacion = {}
            total_general = 0
            
            for medio, total, cantidad in resultado:
                recaudacion[medio] = {
                    'total': float(total),
                    'cantidad_operaciones': cantidad
                }
                total_general += float(total)
            
            return {
                'recaudacion_por_medio': recaudacion,
                'total_general': total_general,
                'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
                'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            print(f"Error calculando recaudación: {e}")
            return None

# Agregar este modelo después de la clase MedioPago en tu app.py

class Gasto(db.Model):
    """Modelo para registrar gastos y egresos del negocio"""
    __tablename__ = 'gastos'
    
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    monto = db.Column(Numeric(10, 2), nullable=False)
    categoria = db.Column(db.String(50), nullable=False, default='general')
    metodo_pago = db.Column(db.String(30), nullable=False, default='efectivo')
    notas = db.Column(db.Text)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    activo = db.Column(db.Boolean, default=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))  # Usuario que registró el gasto
    caja_id = db.Column(db.Integer, db.ForeignKey('cajas.id'), nullable=True) 

    # Relación con Usuario
    usuario = db.relationship('Usuario', backref=db.backref('gastos', lazy=True))
    
    def __repr__(self):
        return f'<Gasto {self.descripcion}: ${self.monto}>'
    
    def to_dict(self):
        """Convertir gasto a diccionario para JSON"""
        return {
            'id': self.id,
            'fecha': self.fecha.strftime('%Y-%m-%d') if self.fecha else None,
            'descripcion': self.descripcion,
            'monto': float(self.monto),
            'categoria': self.categoria,
            'metodo_pago': self.metodo_pago,
            'notas': self.notas,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None,
            'fecha_modificacion': self.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_modificacion else None,
            'activo': self.activo,
            'usuario': self.usuario.nombre if self.usuario else None
        }
    
    @staticmethod
    def obtener_categorias_disponibles():
        """Retorna las categorías de gastos disponibles"""
        return [
            {'codigo': 'general', 'nombre': 'General'},
            {'codigo': 'insumos', 'nombre': 'Insumos y Materiales'},
            {'codigo': 'servicios', 'nombre': 'Servicios (Luz, Gas, Internet)'},
            {'codigo': 'transporte', 'nombre': 'Transporte y Combustible'},
            {'codigo': 'personal', 'nombre': 'Gastos de Personal'},
            {'codigo': 'mantenimiento', 'nombre': 'Mantenimiento'},
            {'codigo': 'impuestos', 'nombre': 'Impuestos y Tasas'},
            {'codigo': 'otros', 'nombre': 'Otros'}
        ]
    
    @staticmethod
    def calcular_gastos_por_fecha(fecha_desde, fecha_hasta):
        """Calcular gastos por categoría en un rango de fechas"""
        try:
            resultado = db.session.query(
                Gasto.categoria,
                func.sum(Gasto.monto).label('total'),
                func.count(Gasto.id).label('cantidad_gastos')
            ).filter(
                and_(
                    Gasto.fecha >= fecha_desde,
                    Gasto.fecha <= fecha_hasta,
                    Gasto.activo == True
                )
            ).group_by(Gasto.categoria).all()
            
            # Convertir a diccionario
            gastos_por_categoria = {}
            total_general = 0
            
            for categoria, total, cantidad in resultado:
                gastos_por_categoria[categoria] = {
                    'total': float(total),
                    'cantidad_gastos': cantidad
                }
                total_general += float(total)
            
            return {
                'gastos_por_categoria': gastos_por_categoria,
                'total_general': total_general,
                'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
                'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            print(f"Error calculando gastos: {e}")
            return None
    
    @staticmethod
    def obtener_gastos_por_medio_pago(fecha_desde, fecha_hasta):
        """Obtener gastos agrupados por medio de pago"""
        try:
            resultado = db.session.query(
                Gasto.metodo_pago,
                func.sum(Gasto.monto).label('total'),
                func.count(Gasto.id).label('cantidad')
            ).filter(
                and_(
                    Gasto.fecha >= fecha_desde,
                    Gasto.fecha <= fecha_hasta,
                    Gasto.activo == True
                )
            ).group_by(Gasto.metodo_pago).all()
            
            gastos_por_medio = {}
            for medio, total, cantidad in resultado:
                gastos_por_medio[medio] = {
                    'total': float(total),
                    'cantidad': cantidad
                }
            
            return gastos_por_medio
            
        except Exception as e:
            print(f"Error obteniendo gastos por medio: {e}")
            return {}


@app.route('/api/get_cuit')
def get_cuit():
    """Retornar CUIT de la configuración"""
    try:
        return jsonify({
            'success': True,
            'cuit': ARCA_CONFIG.CUIT
        })
    except:
        return jsonify({
            'success': False,
            'cuit': 'N/A'
        })

# ================== RUTAS API PARA REPORTES ==================


@app.route('/api/reporte_medios_pago')
def api_reporte_medios_pago():
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        datos = MedioPago.calcular_recaudacion_por_fecha(desde, hasta)

        medios_pago = []
        if datos and 'recaudacion_por_medio' in datos:
            for medio, valores in datos['recaudacion_por_medio'].items():
                medios_pago.append({
                    'medio_pago': medio,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_operaciones']
                })

        return jsonify({
            'success': True,
            'reporte': {
                'medios_pago': medios_pago,
                'total_general': datos['total_general'] if datos else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/reporte_gastos')
def api_reporte_gastos():
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        datos = Gasto.calcular_gastos_por_fecha(desde, hasta)

        gastos = []
        if datos and 'gastos_por_categoria' in datos:
            for categoria, valores in datos['gastos_por_categoria'].items():
                gastos.append({
                    'categoria': categoria,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_gastos']
                })

        return jsonify({
            'success': True,
            'reporte': {
                'gastos': gastos,
                'total_general': datos['total_general'] if datos else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/reporte_caja_diaria')
def api_reporte_caja_diaria():
    print(">>> API reporte_caja_diaria llamada")
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')

        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})

        desde = datetime.strptime(fecha_desde, "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        ingresos = MedioPago.calcular_recaudacion_por_fecha(desde, hasta)
        gastos = Gasto.calcular_gastos_por_fecha(desde, hasta)

        total_ingresos = ingresos['total_general'] if ingresos else 0
        total_gastos = gastos['total_general'] if gastos else 0
        balance = total_ingresos - total_gastos

        detalle_ingresos = []
        if ingresos and 'recaudacion_por_medio' in ingresos:
            for medio, valores in ingresos['recaudacion_por_medio'].items():
                detalle_ingresos.append({
                    'medio_pago': medio,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_operaciones']
                })

        detalle_gastos = []
        if gastos and 'gastos_por_categoria' in gastos:
            for categoria, valores in gastos['gastos_por_categoria'].items():
                detalle_gastos.append({
                    'categoria': categoria,
                    'total': valores['total'],
                    'cantidad': valores['cantidad_gastos']
                })

        return jsonify({
            'success': True,
            'totalIngresos': total_ingresos,
            'totalGastos': total_gastos,
            'balance': balance,
            'detalleIngresos': detalle_ingresos,
            'detalleGastos': detalle_gastos
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# Clase para manejo de ARCA/AFIP

class CancelacionVenta(db.Model):
    __tablename__ = 'cancelacion_venta'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.now)
    usuario_id = db.Column(db.Integer, nullable=True)
    usuario_nombre = db.Column(db.String(100))
    total = db.Column(Numeric(12, 2))
    cliente_nombre = db.Column(db.String(100))
    detalle = db.Column(db.Text)
    tipo = db.Column(db.String(30), default='cancelacion_medios_pago')

class ARCAClient:
    def __init__(self):
        self.config = ARCA_CONFIG
        self.token = None
        self.sign = None
        self.token_expira = None
        self.cuit = self.config.CUIT
        self.openssl_path = self._buscar_openssl()
        self._cargar_token_cache()
        
        print(f"🔧 AFIP Client inicializado")
        print(f"   CUIT: {self.config.CUIT}")
        print(f"   Ambiente: {'HOMOLOGACIÓN' if self.config.USE_HOMOLOGACION else 'PRODUCCIÓN'}")
    
    def _buscar_openssl(self):
        """Buscar OpenSSL en ubicaciones conocidas"""
        ubicaciones = [
            './openssl.exe',
            'openssl.exe', 
            'openssl',
            r'C:\Program Files\OpenSSL-Win64\bin\openssl.exe',
            r'C:\OpenSSL-Win64\bin\openssl.exe',
            r'C:\Program Files (x86)\OpenSSL-Win32\bin\openssl.exe'
        ]
        
        for ubicacion in ubicaciones:
            try:
                if os.path.exists(ubicacion) or ubicacion in ['openssl.exe', 'openssl']:
                    result = subprocess.run([ubicacion, 'version'], 
                                          capture_output=True, text=True, timeout=5)
                    if result.returncode == 0:
                        print(f"✅ OpenSSL encontrado: {ubicacion}")
                        return ubicacion
            except:
                continue
        
        print("❌ OpenSSL no encontrado, usando 'openssl' por defecto")
        return 'openssl'
    
    def crear_tra(self):
        """Crear Ticket Request Access"""
        now = datetime.now()
        expire = now + timedelta(hours=12)
        unique_id = int(now.timestamp())
        
        tra_xml = f'''<?xml version="1.0" encoding="UTF-8"?>
                    <loginTicketRequest version="1.0">
                        <header>
                            <uniqueId>{unique_id}</uniqueId>
                            <generationTime>{now.strftime('%Y-%m-%dT%H:%M:%S.000-00:00')}</generationTime>
                            <expirationTime>{expire.strftime('%Y-%m-%dT%H:%M:%S.000-00:00')}</expirationTime>
                        </header>
                        <service>wsfe</service>
                    </loginTicketRequest>'''
        
        return tra_xml
    
    def firmar_tra_openssl(self, tra_xml):
        """Firmar TRA usando cryptography (sin OpenSSL externo)"""
        try:
            from cryptography.hazmat.primitives.serialization import pkcs7
            
            print(f"🔐 Firmando TRA con cryptography...")
            
            # Verificar certificados
            if not os.path.exists(self.config.CERT_PATH):
                raise Exception(f"Certificado no encontrado: {self.config.CERT_PATH}")
            if not os.path.exists(self.config.KEY_PATH):
                raise Exception(f"Clave privada no encontrada: {self.config.KEY_PATH}")
            
            # Cargar certificado y clave (ya importados al inicio del archivo)
            with open(self.config.CERT_PATH, "rb") as f:
                cert = x509.load_pem_x509_certificate(f.read())
            with open(self.config.KEY_PATH, "rb") as f:
                key = serialization.load_pem_private_key(f.read(), password=None)
            
            # Firmar el TRA
            cms_data = (
                pkcs7.PKCS7SignatureBuilder()
                .set_data(tra_xml.encode("utf-8"))
                .add_signer(cert, key, hashes.SHA256())
                .sign(serialization.Encoding.DER, [pkcs7.PKCS7Options.Binary])
            )
            
            cms_b64 = base64.b64encode(cms_data).decode("utf-8")
            print("✅ TRA firmado correctamente")
            return cms_b64
                    
        except Exception as e:
            print(f"❌ Error firmando TRA: {e}")
            raise Exception(f"Error firmando TRA: {e}")
    
    def debug_certificados(self):
        """Debug detallado de certificados"""
        try:
            print("🔍 DEBUG: Analizando certificados...")
            
            # Leer certificado
            with open(self.config.CERT_PATH, 'rb') as f:
                cert_data = f.read()
            
            print(f"📄 Certificado: {len(cert_data)} bytes")
            print(f"📄 Primeros 50 caracteres: {cert_data[:50]}")
            
            # Verificar si es PEM o DER
            if b'-----BEGIN CERTIFICATE-----' in cert_data:
                print("✅ Formato: PEM")
            elif cert_data.startswith(b'\x30\x82'):
                print("⚠️ Formato: DER (puede causar problemas)")
            else:
                print("❌ Formato desconocido")
            
            # Leer clave privada
            with open(self.config.KEY_PATH, 'rb') as f:
                key_data = f.read()
            
            print(f"🔑 Clave privada: {len(key_data)} bytes")
            
            if b'-----BEGIN PRIVATE KEY-----' in key_data or b'-----BEGIN RSA PRIVATE KEY-----' in key_data:
                print("✅ Clave formato: PEM")
            else:
                print("⚠️ Clave formato: DER o desconocido")
            
            # Test de firma simple
            print("🧪 Probando firma de prueba...")
            test_data = "test data for signing"
            
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
                f.write(test_data)
                test_file = f.name
            
            with tempfile.NamedTemporaryFile(suffix='.sig', delete=False) as f:
                sig_file = f.name
            
            try:
                cmd = [
                    self.openssl_path, 'smime', '-sign',
                    '-in', test_file,
                    '-out', sig_file,
                    '-signer', self.config.CERT_PATH,
                    '-inkey', self.config.KEY_PATH,
                    '-outform', 'DER',
                    '-nodetach'
                ]
                
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
                
                if result.returncode == 0:
                    print("✅ Test de firma: EXITOSO")
                    return True
                else:
                    print(f"❌ Test de firma FALLÓ: {result.stderr}")
                    return False
                    
            finally:
                try:
                    os.unlink(test_file)
                    os.unlink(sig_file)
                except:
                    pass
            
        except Exception as e:
            print(f"❌ Error en debug certificados: {e}")
            return False


    def _cargar_token_cache(self):
        """Levanta el TA guardado en disco si sigue vigente y es del mismo ambiente."""
        try:
            path = self.config.TOKEN_CACHE_FILE
            if not path or not os.path.exists(path):
                return False
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if bool(data.get('homologacion')) != bool(self.config.USE_HOMOLOGACION):
                return False
            expira = datetime.fromisoformat(data['expira'])
            if expira.tzinfo is not None:
                expira = expira.replace(tzinfo=None)
            self.token = data.get('token')
            self.sign = data.get('sign')
            self.token_expira = expira
            print(f"🎫 TA levantado de cache (vence {expira})")
            return True
        except Exception as e:
            print(f"⚠️ No se pudo leer el TA de cache: {e}")
            return False

    def _guardar_token_cache(self, exp_str):
        """Persiste el TA (token+sign+expiración) en cache/token_arca.json."""
        try:
            expira = None
            if exp_str:
                try:
                    expira = datetime.fromisoformat(exp_str.strip())
                    if expira.tzinfo is not None:
                        expira = expira.replace(tzinfo=None)
                except Exception:
                    expira = None
            if expira is None:
                expira = datetime.now() + timedelta(hours=11)
            self.token_expira = expira
            carpeta = os.path.dirname(self.config.TOKEN_CACHE_FILE) or '.'
            os.makedirs(carpeta, exist_ok=True)
            with open(self.config.TOKEN_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump({
                    'token':        self.token,
                    'sign':         self.sign,
                    'expira':       expira.isoformat(),
                    'homologacion': bool(self.config.USE_HOMOLOGACION),
                }, f)
            print(f"💾 TA guardado en {self.config.TOKEN_CACHE_FILE} (vence {expira})")
        except Exception as e:
            print(f"⚠️ No se pudo guardar el TA en cache: {e}")

    def get_ticket_access(self):
        """Obtener ticket de acceso de WSAA con cache inteligente"""
        try:
            # *** Cache de tokens: memoria o disco, validez por expiración REAL del TA ***
            if not (self.token and self.sign and self.token_expira):
                self._cargar_token_cache()

            if self.token and self.sign and self.token_expira:
                if datetime.now() < (self.token_expira - timedelta(minutes=10)):
                    print(f"🎫 Usando TA vigente (vence {self.token_expira})")
                    return True
                else:
                    print("⏰ TA vencido, obteniendo uno nuevo...")
                    self.token = None
                    self.sign = None
                    self.token_expira = None
            
            print("🎫 Obteniendo nuevo ticket de acceso...")
            
            # Crear y firmar TRA
            tra_xml = self.crear_tra()
            tra_firmado = self.firmar_tra_openssl(tra_xml)
            
            # URL del WSAA
            wsaa_url = self.config.WSAA_URL + '?wsdl' if not self.config.WSAA_URL.endswith('?wsdl') else self.config.WSAA_URL
            
            print(f"🌐 Conectando con WSAA: {wsaa_url}")
            
            # USAR SESIÓN PERSONALIZADA
            session = crear_session_afip()
            
            from zeep.transports import Transport
            
            transport = Transport(session=session, timeout=60)
            client = Client(wsaa_url, transport=transport)
            
            # Enviar solicitud
            response = client.service.loginCms(tra_firmado)
            
            if response:
                # Procesar respuesta XML
                root = ET.fromstring(response)
                
                token_elem = root.find('.//token')
                sign_elem = root.find('.//sign')
                
                if token_elem is None or sign_elem is None:
                    raise Exception("Token o Sign no encontrados en respuesta")
                
                self.token = token_elem.text
                self.sign = sign_elem.text
                
                # *** Persistir el TA en disco ***
                self.token_timestamp = datetime.now()
                _exp = root.find('.//expirationTime')
                self._guardar_token_cache(_exp.text if _exp is not None else None)
                
                print("✅ Ticket de acceso obtenido y guardado en cache")
                return True
            else:
                raise Exception("Respuesta vacía de WSAA")
                
        except Exception as e:
            error_msg = str(e)
            
            # *** NUEVO: Manejo específico del error de token duplicado ***
            if "El CEE ya posee un TA valido" in error_msg:
                print("⚠️ AFIP indica que ya hay un TA válido para este CUIT")
                if self._cargar_token_cache() and self.token and self.sign and \
                   self.token_expira and datetime.now() < (self.token_expira - timedelta(minutes=10)):
                    print("🎫 Reutilizando TA guardado en cache (válido)")
                    return True
                print("❌ Hay un TA válido en AFIP pero NO está guardado localmente.")
                import time
                time.sleep(30)
                
                self.token = None
                self.sign = None
                if hasattr(self, 'token_timestamp'):
                    delattr(self, 'token_timestamp')
                
                # Reintentar una vez
                try:
                    print("🔄 Reintentando obtener token...")
                    tra_xml = self.crear_tra()
                    tra_firmado = self.firmar_tra_openssl(tra_xml)
                    
                    session = crear_session_afip()
                    transport = Transport(session=session, timeout=60)
                    client = Client(wsaa_url, transport=transport)
                    
                    response = client.service.loginCms(tra_firmado)
                    
                    if response:
                        root = ET.fromstring(response)
                        token_elem = root.find('.//token')
                        sign_elem = root.find('.//sign')
                        
                        if token_elem is not None and sign_elem is not None:
                            self.token = token_elem.text
                            self.sign = sign_elem.text
                            self.token_timestamp = datetime.now()
                            _exp2 = root.find('.//expirationTime')
                            self._guardar_token_cache(_exp2.text if _exp2 is not None else None)
                            
                            print("✅ Token obtenido exitosamente en segundo intento")
                            return True
                    
                except Exception as e2:
                    print(f"❌ Segundo intento también falló: {e2}")
            
            print(f"❌ Error obteniendo ticket: {e}")
            return False

    def _cond_iva_codigo(self, valor):
        """Mapea la condición de IVA guardada del cliente al código AFIP. None si no la reconoce."""
        if not valor:
            return None
        v = str(valor).strip().upper().replace(' ', '_')
        try:
            cond = self.config.CONDICIONES_IVA
            if v in cond:
                return int(cond[v])
        except Exception:
            pass
        mapa = {
            'CF': 5, 'CONSUMIDOR_FINAL': 5, 'CONSUMIDORFINAL': 5,
            'RI': 1, 'RESPONSABLE_INSCRIPTO': 1, 'IVA_RESPONSABLE_INSCRIPTO': 1,
            'EXENTO': 4, 'SUJETO_EXENTO': 4, 'IVA_SUJETO_EXENTO': 4,
            'MONOTRIBUTO': 6, 'RESPONSABLE_MONOTRIBUTO': 6, 'MONOTRIBUTISTA': 6,
            'NO_CATEGORIZADO': 7, 'SUJETO_NO_CATEGORIZADO': 7,
            'MONOTRIBUTISTA_SOCIAL': 13, 'NO_ALCANZADO': 15, 'IVA_NO_ALCANZADO': 15,
        }
        if v in mapa:
            return mapa[v]
        if 'EXENTO' in v:        return 4
        if 'MONOTRIB' in v:      return 6
        if 'INSCRIPTO' in v:     return 1
        if 'CONSUMIDOR' in v:    return 5
        if 'NO_CATEGORIZADO' in v: return 7
        return None

    def get_cond_iva_receptor(self, datos_comprobante, tipo_cbte):
        """Condición frente al IVA del receptor (RG 5616)."""
        cond = datos_comprobante.get('condicion_iva_receptor')
        if cond:
            try:
                return int(cond)
            except (TypeError, ValueError):
                pass
        cod = self._cond_iva_codigo(datos_comprobante.get('condicion_iva_str'))
        if cod:
            return cod
        t = int(tipo_cbte)
        doc_tipo = int(datos_comprobante.get('doc_tipo', 99) or 99)
        if t in (1, 2, 3, 51, 52, 53):
            return 1
        if doc_tipo in (96, 99):
            return 5
        if doc_tipo == 80:
            return 6
        return 5

    def autorizar_comprobante(self, datos_comprobante):
        """
        Autorizar comprobante en AFIP usando WSFEv1 - VERSIÓN CORREGIDA CON MÚLTIPLES ALÍCUOTAS IVA
        """
        try:
            print("🎫 Verificando ticket de acceso...")
            
            # Verificar que tenemos ticket válido
            if not self.get_ticket_access():
                raise Exception("No se pudo obtener ticket de acceso")
            
            print("🌐 Conectando con WSFEv1...")
            
            # URL de WSFEv1 según ambiente (homologación o producción) — NO hardcodear.
            wsfev1_url = self.config.WSFEv1_URL
            
            # Crear cliente SOAP con configuración específica
            session = crear_session_afip()
            from zeep.transports import Transport
            from zeep import Settings
            
            # Configuración específica para AFIP
            settings = Settings(strict=False, xml_huge_tree=True)
            transport = Transport(session=session, timeout=60, operation_timeout=60)
            
            try:
                from zeep import Client
                client = Client(wsfev1_url, transport=transport, settings=settings)
                print("✅ Cliente WSFEv1 creado correctamente")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html', 'br line', 'span']):
                    raise Exception("WSFEv1 devolviendo HTML en lugar de XML - Servicio en mantenimiento")
                else:
                    raise Exception(f"Error creando cliente SOAP: {str(e)}")
            
            # Preparar autenticación
            auth = {
                'Token': self.token,
                'Sign': self.sign,
                'Cuit': self.cuit
            }
            
            print("📋 Preparando datos del comprobante...")
            
            # Obtener configuración
            pto_vta = datos_comprobante.get('punto_venta', self.config.PUNTO_VENTA)
            tipo_cbte = datos_comprobante.get('tipo_comprobante', 11)  # 11 = Factura C
            
            # Test rápido con FEDummy para verificar que el servicio funciona
            try:
                print("🧪 Verificando servicio con FEDummy...")
                dummy_response = client.service.FEDummy()
                print(f"✅ FEDummy OK: {dummy_response}")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html']):
                    raise Exception("FEDummy devolviendo HTML - WSFEv1 en mantenimiento")
                else:
                    print(f"⚠️ Warning en FEDummy: {e}")
            
            # Obtener último comprobante autorizado
            try:
                print("📊 Consultando último comprobante autorizado...")
                ultimo_cbte_response = client.service.FECompUltimoAutorizado(
                    Auth=auth,
                    PtoVta=pto_vta,
                    CbteTipo=tipo_cbte
                )
                
                # Verificar errores en la respuesta
                if hasattr(ultimo_cbte_response, 'Errors') and ultimo_cbte_response.Errors:
                    print(f"⚠️ Advertencias al obtener último comprobante:")
                    if hasattr(ultimo_cbte_response.Errors, 'Err'):
                        errors = ultimo_cbte_response.Errors.Err
                        if isinstance(errors, list):
                            for error in errors:
                                print(f"   [{error.Code}] {error.Msg}")
                        else:
                            print(f"   [{errors.Code}] {errors.Msg}")
                
                ultimo_nro = getattr(ultimo_cbte_response, 'CbteNro', 0)
                proximo_nro = ultimo_nro + 1
                
                print(f"📊 Último comprobante AFIP: {ultimo_nro}")
                print(f"📊 Próximo número: {proximo_nro}")
                
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html']):
                    raise Exception("FECompUltimoAutorizado devolviendo HTML")
                else:
                    print(f"⚠️ Error obteniendo último comprobante: {e}")
                    print("🔄 Usando número secuencial local...")
                    proximo_nro = 1
            
            # Preparar datos del comprobante
            fecha_hoy = datetime.now().strftime('%Y%m%d')
            
            # *** NUEVO: CALCULAR ALÍCUOTAS IVA SEPARADAS ***
            # Obtener los items del comprobante (deben venir con detalle por producto)
            items_detalle = datos_comprobante.get('items_detalle', [])
            
            if not items_detalle:
                raise Exception("Se requieren items detallados con alícuotas IVA individuales")
            
            # Agrupar por alícuota de IVA
            alicuotas_iva = {}
            importe_neto_total = 0
            importe_iva_total = 0
            
            print("🧮 Calculando alícuotas de IVA por separado...")
            
            for item in items_detalle:
                subtotal = float(item.get('subtotal', 0))
                iva_porcentaje = float(item.get('iva_porcentaje', 0))
                
                # Calcular IVA del item con redondeo AFIP
                iva_item = round((subtotal * iva_porcentaje / 100), 2)
                
                # Agrupar por alícuota
                if iva_porcentaje not in alicuotas_iva:
                    alicuotas_iva[iva_porcentaje] = {
                        'base_imponible': 0,
                        'iva_total': 0
                    }
                
                alicuotas_iva[iva_porcentaje]['base_imponible'] += subtotal
                alicuotas_iva[iva_porcentaje]['iva_total'] += iva_item
                
                importe_neto_total += subtotal
                importe_iva_total += iva_item
                
                print(f"   📦 Item: ${subtotal:.2f} (IVA {iva_porcentaje}% = ${iva_item:.2f})")
            
            # Redondear totales
            importe_neto_total = round(importe_neto_total, 2)
            importe_iva_total = round(importe_iva_total, 2)
            importe_total = round(importe_neto_total + importe_iva_total, 2)
            
            print(f"💰 Totales calculados: Neto=${importe_neto_total:.2f}, IVA=${importe_iva_total:.2f}, Total=${importe_total:.2f}")

            # ═══ FACTURA C (Monotributo/Exento): AFIP NO admite IVA discriminado ═══
            # Tipos 11/12/13: ImpNeto = Total, ImpIVA = 0 y NO se informa el objeto Iva
            # (errores AFIP 10047 / 10048 / 10071). El bloque de AlicIva de más abajo
            # ya está protegido con 'if importe_iva_total > 0', así que al dejar el IVA
            # en 0 no se agrega el objeto Iva automáticamente.
            es_comprobante_c = int(tipo_cbte) in (11, 12, 13)
            if es_comprobante_c:
                importe_neto_total = importe_total
                importe_iva_total = 0.00
                print("📋 Comprobante C: IVA no discriminado (ImpNeto=Total, ImpIVA=0)")
            
            # Mostrar alícuotas calculadas
            print("📊 Alícuotas de IVA:")
            for porcentaje, datos in alicuotas_iva.items():
                base = round(datos['base_imponible'], 2)
                iva = round(datos['iva_total'], 2)
                print(f"   IVA {porcentaje}%: Base=${base:.2f}, IVA=${iva:.2f}")
            
            # Estructura del comprobante según especificación AFIP
            comprobante = {
                'Concepto': 1,
                'DocTipo': datos_comprobante.get('doc_tipo', 99),
                'DocNro': datos_comprobante.get('doc_nro', 0),
                'CbteDesde': proximo_nro,
                'CbteHasta': proximo_nro,
                'CbteFch': fecha_hoy,
                'ImpTotal': importe_total,
                'ImpTotConc': 0.00,
                'ImpNeto': importe_neto_total,
                'ImpOpEx': 0.00,
                'ImpTrib': 0.00,
                'ImpIVA': importe_iva_total,
                'MonId': 'PES',
                'MonCotiz': 1.00,
                # RG 5616: condición frente al IVA del receptor (obligatorio)
                'CondicionIVAReceptorId': self.get_cond_iva_receptor(datos_comprobante, tipo_cbte),
            }
            
            # *** NUEVO: AGREGAR COMPROBANTES ASOCIADOS (para Notas de Crédito) ***
            cbtes_asoc = datos_comprobante.get('comprobantes_asociados', None)
            if cbtes_asoc:
                print(f"📎 Procesando {len(cbtes_asoc)} comprobante(s) asociado(s)...")
                
                # Formatear comprobantes asociados según estructura AFIP
                cbtes_asoc_afip = []
                
                for idx, cbte in enumerate(cbtes_asoc):
                    cbte_dict = {
                        'Tipo': int(cbte['Tipo']),
                        'PtoVta': int(cbte['PtoVta']),
                        'Nro': int(cbte['Nro'])
                    }
                    
                    # Solo agregar CUIT si existe y no es None
                    if cbte.get('Cuit') and cbte['Cuit'] not in [None, 0, '0', '']:
                        # Limpiar CUIT (quitar guiones)
                        cuit_limpio = str(cbte['Cuit']).replace('-', '')
                        if cuit_limpio and cuit_limpio != '0':
                            cbte_dict['Cuit'] = int(cuit_limpio)
                    
                    cbtes_asoc_afip.append(cbte_dict)
                    
                    print(f"   📄 Cbte {idx+1}: Tipo={cbte['Tipo']}, PtoVta={cbte['PtoVta']}, Nro={cbte['Nro']}")
                
                # Agregar al comprobante
                comprobante['CbtesAsoc'] = {'CbteAsoc': cbtes_asoc_afip}
                print(f"✅ Comprobantes asociados agregados al request AFIP")


            # *** CLAVE: AGREGAR DETALLE DE IVA POR ALÍCUOTA ***
            if importe_iva_total > 0:
                alicuotas_afip = []
                
                for porcentaje, datos in alicuotas_iva.items():
                    if porcentaje > 0:  # Solo agregar si hay IVA
                        # Mapear porcentajes a códigos AFIP
                        codigo_iva = self.get_codigo_iva_afip(porcentaje)
                        
                        if codigo_iva:
                            alicuotas_afip.append({
                                'Id': codigo_iva,
                                'BaseImp': round(datos['base_imponible'], 2),
                                'Importe': round(datos['iva_total'], 2)
                            })
                            
                            print(f"✅ Alícuota AFIP: Código {codigo_iva}, Base=${datos['base_imponible']:.2f}, IVA=${datos['iva_total']:.2f}")
                
                if alicuotas_afip:
                    comprobante['Iva'] = {'AlicIva': alicuotas_afip}
                    print(f"📝 Se agregaron {len(alicuotas_afip)} alícuotas de IVA al comprobante")
                else:
                    print("⚠️ No se pudieron mapear las alícuotas a códigos AFIP")
            
            # Crear request completo
            fe_request = {
                'FeCabReq': {
                    'CantReg': 1,
                    'PtoVta': pto_vta,
                    'CbteTipo': tipo_cbte
                },
                'FeDetReq': {
                    'FECAEDetRequest': [comprobante]
                }
            }
            
            print("📤 Enviando solicitud de autorización a AFIP...")
            print(f"   Tipo comprobante: {tipo_cbte}")
            print(f"   Punto de venta: {pto_vta}")
            print(f"   Número: {proximo_nro}")
            print(f"   Fecha: {fecha_hoy}")
            print(f"   Total: ${importe_total:.2f}")
            print(f"   Alícuotas IVA: {len(alicuotas_iva)} diferentes")
            
            # ENVÍO CRÍTICO
            try:
                response = client.service.FECAESolicitar(Auth=auth, FeCAEReq=fe_request)
                print("✅ Respuesta recibida de AFIP")
            except Exception as e:
                error_str = str(e).lower()
                if any(keyword in error_str for keyword in ['invalid xml', 'mismatch', 'html', 'br line', 'span']):
                    raise Exception("FECAESolicitar devolviendo HTML - WSFEv1 en mantenimiento")
                else:
                    raise Exception(f"Error en FECAESolicitar: {str(e)}")
            
            # Procesar respuesta de AFIP
            print("📋 Procesando respuesta de AFIP...")
            
            # Verificar errores generales
            if hasattr(response, 'Errors') and response.Errors:
                errores = []
                if hasattr(response.Errors, 'Err'):
                    errors = response.Errors.Err
                    if isinstance(errors, list):
                        for error in errors:
                            errores.append(f"[{error.Code}] {error.Msg}")
                    else:
                        errores.append(f"[{errors.Code}] {errors.Msg}")
                
                error_msg = " | ".join(errores)
                raise Exception(f"Errores AFIP: {error_msg}")
            
            # Verificar que hay respuesta de detalle
            if not hasattr(response, 'FeDetResp') or not response.FeDetResp:
                raise Exception("Respuesta de AFIP sin detalles")
            
            # Obtener detalle de respuesta
            if not hasattr(response.FeDetResp, 'FECAEDetResponse'):
                raise Exception("Respuesta de AFIP sin FECAEDetResponse")
            
            detalle_resp = response.FeDetResp.FECAEDetResponse[0]
            
            # Verificar resultado
            resultado = getattr(detalle_resp, 'Resultado', None)
            if resultado != 'A':  # A = Aprobado
                observaciones = []
                if hasattr(detalle_resp, 'Observaciones') and detalle_resp.Observaciones:
                    if hasattr(detalle_resp.Observaciones, 'Obs'):
                        obs_list = detalle_resp.Observaciones.Obs
                        if isinstance(obs_list, list):
                            for obs in obs_list:
                                observaciones.append(f"[{obs.Code}] {obs.Msg}")
                        else:
                            observaciones.append(f"[{obs_list.Code}] {obs_list.Msg}")
                
                obs_msg = " | ".join(observaciones) if observaciones else "Sin observaciones"
                raise Exception(f"Comprobante no autorizado. Resultado: {resultado}. {obs_msg}")
            
            # ÉXITO - Extraer datos
            cae = getattr(detalle_resp, 'CAE', None)
            fecha_vencimiento = getattr(detalle_resp, 'CAEFchVto', None)
            
            if not cae:
                raise Exception("Respuesta sin CAE")
            
            if not fecha_vencimiento:
                raise Exception("Respuesta sin fecha de vencimiento CAE")
            
            numero_completo = f"{pto_vta:04d}-{proximo_nro:08d}"
            
            print(f"🎉 ¡COMPROBANTE AUTORIZADO EXITOSAMENTE!")
            print(f"   Número: {numero_completo}")
            print(f"   CAE: {cae}")
            print(f"   Vencimiento CAE: {fecha_vencimiento}")
            
            return {
                'success': True,
                'cae': cae,
                'numero': numero_completo,
                'punto_venta': pto_vta,
                'numero_comprobante': proximo_nro,
                'fecha_vencimiento': fecha_vencimiento,
                'fecha_proceso': fecha_hoy,
                'importe_total': importe_total,
                'tipo_comprobante': tipo_cbte,
                'estado': 'autorizada',
                'vto_cae': datetime.strptime(fecha_vencimiento, '%Y%m%d').date()
            }
            
        except Exception as e:
            print(f"❌ Error en autorización AFIP: {e}")
            return {
                'success': False,
                'error': str(e),
                'cae': None,
                'vto_cae': None,
                'estado': 'error_afip'
            }

    def get_codigo_iva_afip(self, porcentaje):
        """Mapear porcentajes de IVA a códigos AFIP"""
        mapeo_iva = {
            0: 3,      # Exento
            10.5: 4,   # IVA 10.5%
            21: 5,     # IVA 21%
            27: 6,     # IVA 27%
            2.5: 9     # IVA 2.5%
        }
        
        codigo = mapeo_iva.get(porcentaje, None)
        if codigo is None:
            print(f"⚠️ Porcentaje IVA {porcentaje}% no reconocido, usando código 5 (21%) por defecto")
            return 5  # Por defecto IVA 21%
        
        return codigo
    

    def get_ultimo_comprobante(self, tipo_cbte, punto_venta=None):
        """Obtener último comprobante autorizado.

        punto_venta: si no se pasa, usa el del config (fallback).
        """
        try:
            pto_vta = int(punto_venta) if punto_venta else int(self.config.PUNTO_VENTA)
            print(f"📋 Consultando último comprobante tipo {tipo_cbte} (PV={pto_vta})...")
            
            if not self.get_ticket_access():
                raise Exception("No se pudo obtener acceso a AFIP")
            
            # URL del WSFEv1
            wsfe_url = self.config.WSFEv1_URL
            
            print(f"🌐 Conectando con WSFEv1: {wsfe_url}")
            
            # USAR SESIÓN PERSONALIZADA
            session = crear_session_afip()
            
            from zeep.transports import Transport
            
            transport = Transport(session=session, timeout=60)
            client = Client(wsfe_url, transport=transport)
            
            response = client.service.FECompUltimoAutorizado(
                Auth={
                    'Token': self.token,
                    'Sign': self.sign,
                    'Cuit': self.cuit
                },
                PtoVta=pto_vta,
                CbteTipo=tipo_cbte
            )
            
            if hasattr(response, 'Errors') and response.Errors:
                error_msg = response.Errors.Err[0].Msg
                raise Exception(f"Error AFIP: {error_msg}")
            
            ultimo_num = response.CbteNro
            print(f"✅ Último comprobante: {ultimo_num}")
            return ultimo_num
            
        except Exception as e:
            print(f"❌ Error consultando comprobante: {e}")
            raise Exception(f"Error al obtener último comprobante: {e}")


arca_client = ARCAClient()

# Monitor AFIP simplificado
class AFIPStatusMonitor:
    def __init__(self, arca_config):
        self.config = arca_config
    
    def verificar_rapido(self):
        """Verificación rápida solo de conectividad"""
        try:
            import socket
            from urllib.parse import urlparse
            
            wsaa_host = urlparse(self.config.WSAA_URL).hostname
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((wsaa_host, 443))
            sock.close()
            
            return {
                'conectividad': result == 0,
                'mensaje': '✅ AFIP accesible' if result == 0 else '❌ AFIP no accesible'
            }
        except Exception as e:
            return {
                'conectividad': False,
                'mensaje': f'❌ Error: {str(e)}'
            }

# Crear instancia del monitor
afip_monitor = AFIPStatusMonitor(ARCA_CONFIG)


# DESPUÉS DE DEFINIR LOS MODELOS Y ANTES DE LAS RUTAS:
# Inicializar y registrar el blueprint de estadísticas
estadisticas_bp = init_estadisticas(db, Factura, DetalleFactura, Producto, Cliente)
app.config['MEDIO_PAGO_MODEL'] = MedioPago
app.register_blueprint(estadisticas_bp)

# INICIALIZAR SISTEMA DE CAJA
caja_bp = init_caja_system(db, Factura, DetalleFactura, Producto, Usuario, MedioPago, Gasto)
app.register_blueprint(caja_bp)

# ──────────────────────────────────────────────
# Sistema de caja MULTI-PV (paralelo a caja.py)
# Rutas: /api/caja-pv/*
# ──────────────────────────────────────────────
from caja_multipv import init_caja_multipv_system
caja_multipv_bp = init_caja_multipv_system(db)
app.register_blueprint(caja_multipv_bp)

# INICIALIZAR MODULO DE PRESUPUESTOS
init_presupuestos(app, Cliente, Usuario)

# INICIALIZAR MODULO PUENTE CONTABLE (asientos contables para el contador)
# Le pasamos solo los modelos que estan definidos en este app.py.
# Compras/Pagos/Proveedores se importan desde sus modulos por separado para no romper si faltan.
# OJO con CajaAperturaModel: en caja.py se define DENTRO de init_caja_models(), por eso
# el import directo del principio capturo None. Lo leemos del modulo en runtime, ahora si esta poblado.
import caja as _caja_mod
_CajaApertura = getattr(_caja_mod, 'CajaAperturaModel', None)

try:
    from proveedores import Proveedor, FacturaCompra, PagoProveedor, PagoProveedorMedio
    init_puente_contable(
        app, db,
        Factura=Factura,
        DetalleFactura=DetalleFactura,
        MedioPago=MedioPago,
        NotaCredito=NotaCredito,
        DetalleNotaCredito=DetalleNotaCredito,
        Cliente=Cliente,
        Usuario=Usuario,
        Gasto=Gasto,
        CajaAperturaModel=_CajaApertura,
        Proveedor=Proveedor,
        FacturaCompra=FacturaCompra,
        PagoProveedor=PagoProveedor,
        PagoProveedorMedio=PagoProveedorMedio,
    )
    print("[OK] Puente Contable inicializado con compras y pagos a proveedores")

    # INICIALIZAR MODULO IMPORTADOR
    init_importador(app, db, Cliente=Cliente, Producto=Producto, Proveedor=Proveedor)
    print("[OK] Importador inicializado (con Proveedores)")

    # INICIALIZAR WEBAPP MOVIL
    init_mobile(app, db,
                Usuario=Usuario,
                Factura=Factura,
                DetalleFactura=DetalleFactura,
                Cliente=Cliente,
                Producto=Producto,
                MedioPago=MedioPago,
                Proveedor=Proveedor,
                CajaApertura=_CajaApertura)
    print("[OK] Webapp movil inicializada (con Proveedores)")

except ImportError as e:
    # Fallback: si proveedores.py no esta disponible, igual inicializamos con lo basico
    print(f"[WARN] Puente Contable: proveedores no disponible ({e}), inicializando solo con ventas")
    init_puente_contable(
        app, db,
        Factura=Factura,
        DetalleFactura=DetalleFactura,
        MedioPago=MedioPago,
        NotaCredito=NotaCredito,
        DetalleNotaCredito=DetalleNotaCredito,
        Cliente=Cliente,
        Usuario=Usuario,
        Gasto=Gasto,
        CajaAperturaModel=_CajaApertura,
    )
    # Importador sin Proveedor (las pantallas de Proveedores y Saldos de Proveedores quedaran inhabilitadas)
    init_importador(app, db, Cliente=Cliente, Producto=Producto)
    print("[OK] Importador inicializado (sin Proveedores)")

    # Webapp movil sin Proveedor
    init_mobile(app, db,
                Usuario=Usuario,
                Factura=Factura,
                DetalleFactura=DetalleFactura,
                Cliente=Cliente,
                Producto=Producto,
                MedioPago=MedioPago,
                CajaApertura=_CajaApertura)
    print("[OK] Webapp movil inicializada (sin Proveedores)")

# RUTAS DE LA APLICACION ***  RUTAS DE LA APLICACION *** RUTAS DE LA APLICACION *** RUTAS DE LA APLICACION 

# -------------------------------------------------------------------------
# Helper: devuelve el punto de venta activo de la sesión.
# Si el usuario no tiene PV asignado (p.ej. cron sin sesión), cae al config.
# -------------------------------------------------------------------------
def _pv_activo():
    try:
        pv = session.get('punto_venta')
        if pv:
            return int(pv)
    except Exception:
        pass
    return int(ARCA_CONFIG.PUNTO_VENTA)


# -------------------------------------------------------------------------
# Endpoint: cambiar el PV de la sesión (solo admin).
# No modifica la BD del usuario, solo la sesión actual.
# Útil para que el admin facture puntualmente desde otro PV sin tener
# que hacer logout/login con otro usuario.
# -------------------------------------------------------------------------
@app.route('/api/cambiar_pv_sesion', methods=['POST'])
def api_cambiar_pv_sesion():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401
    if session.get('rol') != 'admin':
        return jsonify({'success': False, 'error': 'Solo el administrador puede cambiar el PV'}), 403

    try:
        data = request.get_json(force=True) or {}
        pv_nuevo = int(data.get('punto_venta') or 0)
        if pv_nuevo < 1 or pv_nuevo > 99999:
            return jsonify({'success': False, 'error': 'PV fuera de rango'}), 400
        session['punto_venta'] = pv_nuevo
        return jsonify({'success': True, 'punto_venta': pv_nuevo})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/')
def index():
    """Router inicial: landing/login para visitantes, dashboard para logueados.
    
    El comportamiento sin sesión depende de config_cliente.MOSTRAR_LANDING:
      - True  → muestra el landing comercial (servidor cloud)
      - False → va directo al login (instalaciones locales)
      - Si la variable no existe → asume False (login directo)
    """
    # Sin sesión → landing (si está activado) o login
    if 'user_id' not in session:
        try:
            from config_cliente import MOSTRAR_LANDING
            mostrar_landing = MOSTRAR_LANDING
        except (ImportError, AttributeError):
            mostrar_landing = False
        
        if mostrar_landing:
            return redirect(url_for('landing'))
        return redirect(url_for('login'))
    
    # Con sesión → chequear licencia y mostrar dashboard
    licencia_info = _licencia_actual()
    if licencia_info.get('tipo_bloqueo', 'error') in ['bloqueado', 'no_encontrada', 'error']:
        return redirect(url_for('licencia_bloqueada'))
    return render_template('dashboard.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        usuario = Usuario.query.filter_by(username=username, activo=True).first()
        
        # Login simple sin encriptación
        if usuario and usuario.password_hash == password:
            session['user_id'] = usuario.id
            session['username'] = usuario.username
            session['nombre'] = usuario.nombre
            session['rol'] = usuario.rol
            session['puede_liquidar'] = bool(getattr(usuario, 'puede_liquidar', False))
            # Punto de venta del usuario. Fallback al PV del config si el usuario no tiene asignado.
            try:
                session['punto_venta'] = int(usuario.punto_venta) if usuario.punto_venta else int(ARCA_CONFIG.PUNTO_VENTA)
            except Exception:
                session['punto_venta'] = int(ARCA_CONFIG.PUNTO_VENTA)
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos')
    
    return render_template('login.html')

    
@app.route('/landing')
def landing():
    return render_template('FactuFacil.html')

@app.route('/usuarios')
def usuarios():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('usuarios.html')

@app.route('/api/usuarios')
def api_usuarios():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        us = Usuario.query.order_by(Usuario.rol.desc(), Usuario.nombre).all()
        return jsonify({'success': True, 'usuarios': [{
            'id': u.id, 'username': u.username, 'nombre': u.nombre,
            'rol': u.rol, 'activo': u.activo,
            'punto_venta': int(u.punto_venta or 1)
        } for u in us]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/usuario/<int:id>')
def api_usuario(id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        u = Usuario.query.get_or_404(id)
        return jsonify({'success': True, 'usuario': {
            'id': u.id, 'username': u.username, 'nombre': u.nombre,
            'rol': u.rol, 'activo': u.activo,
            'punto_venta': int(u.punto_venta or 1),
            'puede_liquidar': bool(getattr(u, 'puede_liquidar', False))
        }})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/guardar_usuario', methods=['POST'])
def api_guardar_usuario():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    # Solo el admin puede crear/editar usuarios (incluye cambiar su PV)
    if session.get('rol') != 'admin':
        return jsonify({'success': False, 'error': 'Solo el administrador puede modificar usuarios'}), 403
    try:
        data = request.get_json()
        uid = data.get('id')
        # Validar PV si vino
        pv_nuevo = None
        if 'punto_venta' in data and data.get('punto_venta') not in (None, ''):
            try:
                pv_nuevo = int(data.get('punto_venta'))
                if pv_nuevo < 1 or pv_nuevo > 99999:
                    return jsonify({'success': False, 'error': 'PV fuera de rango (1-99999)'}), 400
            except (TypeError, ValueError):
                return jsonify({'success': False, 'error': 'PV inválido'}), 400

        if uid:
            u = Usuario.query.get_or_404(uid)
            u.nombre = data.get('nombre', u.nombre)
            u.rol = data.get('rol', u.rol)
            u.activo = data.get('activo', u.activo)
            if pv_nuevo is not None:
                u.punto_venta = pv_nuevo
            if 'puede_liquidar' in data:
                u.puede_liquidar = bool(data.get('puede_liquidar'))
            if data.get('password'):
                u.password_hash = data['password']
        else:
            if Usuario.query.filter_by(username=data['username']).first():
                return jsonify({'success': False, 'error': 'El usuario ya existe'}), 400
            u = Usuario(username=data['username'], nombre=data['nombre'],
                        password_hash=data['password'], rol=data.get('rol','vendedor'),
                        activo=data.get('activo', True),
                        punto_venta=(pv_nuevo if pv_nuevo is not None else 1),
                        puede_liquidar=bool(data.get('puede_liquidar', False)))
            db.session.add(u)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/cambiar_password', methods=['POST'])
def api_cambiar_password():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        data = request.get_json()
        u = Usuario.query.get_or_404(data['id'])
        u.password_hash = data['password']
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/toggle_usuario', methods=['POST'])
def api_toggle_usuario():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        data = request.get_json()
        u = Usuario.query.get_or_404(data['id'])
        u.activo = data['activo']
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/auditoria_cancelaciones')
def auditoria_cancelaciones():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('rol') != 'admin':
        flash('Solo los administradores pueden ver esta sección', 'error')
        return redirect(url_for('index'))
    return render_template('auditoria_cancelaciones.html')

@app.route('/api/cancelaciones')
def api_cancelaciones():
    if 'user_id' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        desde = request.args.get('desde', '')
        hasta = request.args.get('hasta', '')
        tipo = request.args.get('tipo', '')
        filtro = '1=1'
        params = {}
        if desde:
            filtro += ' AND fecha >= :desde'
            params['desde'] = desde
        if hasta:
            filtro += ' AND fecha <= :hasta'
            params['hasta'] = hasta + ' 23:59:59'
        if tipo:
            filtro += ' AND tipo = :tipo'
            params['tipo'] = tipo
        from sqlalchemy import text
        rows = db.session.execute(text(
            'SELECT id, fecha, usuario_nombre, total, cliente_nombre, tipo, detalle '
            'FROM cancelacion_venta WHERE ' + filtro + ' ORDER BY fecha DESC LIMIT 500'
        ), params).fetchall()
        resultado = [{
            'id': r.id,
            'fecha': r.fecha.strftime('%d/%m/%Y %H:%M') if r.fecha else '',
            'usuario': r.usuario_nombre or 'Desconocido',
            'total': float(r.total or 0),
            'cliente': r.cliente_nombre or 'Consumidor Final',
            'tipo': r.tipo or 'cancelacion_medios_pago',
            'detalle': r.detalle
        } for r in rows]
        return jsonify({'success': True, 'cancelaciones': resultado})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/registrar_cancelacion', methods=['POST'])
def registrar_cancelacion():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        import json
        data = request.get_json()
        cancelacion = CancelacionVenta(
            usuario_id=session.get('user_id'),
            usuario_nombre=session.get('nombre', 'Desconocido'),
            total=data.get('total', 0),
            cliente_nombre=data.get('cliente_nombre', 'Consumidor Final'),
            detalle=json.dumps(data.get('detalle', []), ensure_ascii=False),
            tipo=data.get('tipo', 'cancelacion_medios_pago')
        )
        db.session.add(cancelacion)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/cambiar_usuario')
def cambiar_usuario():
    session.clear()
    return redirect(url_for('login'))

    
@app.route('/api/hora_actual')
def hora_actual():
    from datetime import datetime
    import pytz
    
    # Zona horaria de Argentina
    argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
    ahora = datetime.now(argentina_tz)
    
    return jsonify({
        'success': True,
        'timestamp': ahora.isoformat(),
        'fecha_legible': ahora.strftime('%A, %d de %B de %Y'),
        'hora_legible': ahora.strftime('%H:%M:%S'),
        'zona_horaria': 'America/Argentina/Buenos_Aires'
    })

@app.route('/productos')
def productos():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    productos = Producto.query.filter_by(activo=True).all()
    
    # Cargar mapa de colores de categorías para que el render inicial los use
    colores_categoria = {}
    try:
        result = db.session.execute(db.text(
            "SELECT nombre, color FROM categoria WHERE activo = 1"
        ))
        for row in result:
            colores_categoria[(row[0] or '').strip().lower()] = row[1] or '#6c757d'
    except Exception as e:
        print(f"⚠️ No se pudo cargar colores de categorías: {e}")
    
    return render_template('productos.html', productos=productos, colores_categoria=colores_categoria)

@app.route('/combos')
def combos():
    # Obtener solo productos que son combos
    combos = Producto.query.filter_by(es_combo=True).all()
    combos = db.session.query(Producto)\
        .options(
            joinedload(Producto.producto_base),
            joinedload(Producto.producto_base_2),    # NUEVO
            joinedload(Producto.producto_base_3)     # NUEVO
        )\
        .filter(Producto.es_combo == True)\
        .all()

    return render_template('combos.html', combos=combos)

@app.route('/clientes')
def clientes():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Traer clientes con nombre y color de zona (JOIN a mano con raw SQL)
    # c.saldo es el campo VIEJO (cache desactualizado). Calculamos el saldo
    # REAL desde cta_cte_movimiento con la fórmula verificada (misma que
    # saldo_real_cliente en cta_cte.py). Subconsultas: sigue siendo 1 query.
    rows = db.session.execute(text("""
        SELECT c.id, c.nombre, c.documento, c.tipo_documento, c.email, c.telefono,
               c.direccion, c.condicion_iva, c.lista_precio, c.tipo_precio,
               (
                 COALESCE((
                   SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                   WHERE m.cliente_id = c.id AND m.tipo = 'venta_fiada'
                 ), 0)
                 -
                 COALESCE((
                   SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                   WHERE m.cliente_id = c.id AND m.tipo = 'nota_credito'
                 ), 0)
                 -
                 (
                   COALESCE((
                     SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                     WHERE m.cliente_id = c.id AND m.tipo_mov = 'pago'
                       AND m.numero_comprobante = 'A_CUENTA'
                   ), 0)
                   -
                   COALESCE((
                     SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                     WHERE m.cliente_id = c.id AND m.tipo_mov = 'pago'
                       AND m.numero_comprobante = 'SALDO_FAVOR_USADO'
                   ), 0)
                 )
               ) AS saldo,
               c.zona_id, z.nombre AS zona_nombre, z.color AS zona_color
          FROM cliente c
          LEFT JOIN zona z ON z.id = c.zona_id
         ORDER BY c.nombre
    """)).mappings().all()
    clientes = [dict(r) for r in rows]
    return render_template('clientes.html', clientes=clientes)

@app.route('/api/clientes')
def api_clientes():
    """API para obtener lista de todos los clientes (para selects)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        clientes = Cliente.query.order_by(Cliente.nombre).all()
        
        clientes_data = []
        for cliente in clientes:
            clientes_data.append({
                'id': cliente.id,
                'nombre': cliente.nombre,
                'razon_social': cliente.nombre,  # Alias para compatibilidad
                'documento': cliente.documento,
                'tipo_documento': cliente.tipo_documento
            })
        
        return jsonify({
            'success': True,
            'clientes': clientes_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500




# ==================== RUTAS DE CLIENTES ====================

@app.route('/api/cliente/<int:cliente_id>')
def obtener_cliente(cliente_id):
    """Obtener datos de un cliente por ID para edición"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Método compatible con SQLAlchemy antiguo
        cliente = Cliente.query.get_or_404(cliente_id)
        return jsonify({
            'id': cliente.id,
            'nombre': cliente.nombre,
            'documento': cliente.documento or '',
            'tipo_documento': cliente.tipo_documento or 'DNI',
            'email': cliente.email or '',
            'telefono': cliente.telefono or '',
            'direccion': cliente.direccion or '',
            'condicion_iva': cliente.condicion_iva or 'CONSUMIDOR_FINAL',
            'tipo_precio': cliente.tipo_precio or 'venta',
            'lista_precio': cliente.lista_precio or 1,
            'saldo': float(cliente.saldo) if cliente.saldo else 0.0,
            'zona_id': cliente.zona_id,
            'vendedor_id': cliente.vendedor_id,
            'dias_vencimiento_factura': cliente.dias_vencimiento_factura,
            'es_intermediario': bool(getattr(cliente, 'es_intermediario', False))
        })
    except Exception as e:
        return jsonify({'error': f'Error al obtener cliente: {str(e)}'}), 500

@app.route('/guardar_cliente', methods=['POST'])
def guardar_cliente():
    """Crear o actualizar un cliente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validar datos requeridos
        if not data.get('nombre', '').strip():
            return jsonify({'error': 'El nombre es obligatorio'}), 400
        
        # Validar CUIT si es necesario
        if data.get('tipo_documento') == 'CUIT' and data.get('documento'):
            documento = data['documento'].strip()
            if not documento.isdigit() or len(documento) != 11:
                return jsonify({'error': 'El CUIT debe tener 11 dígitos sin guiones'}), 400
        
        cliente_id = data.get('id')
        
        if cliente_id:  # Editar cliente existente
            cliente = Cliente.query.get_or_404(cliente_id)
            accion = 'actualizado'
        else:  # Crear nuevo cliente
            cliente = Cliente()
            accion = 'creado'
        
        # Actualizar datos del cliente
        cliente.nombre = data['nombre'].strip()
        cliente.documento = data.get('documento', '').strip() or None
        cliente.tipo_documento = data.get('tipo_documento', 'DNI')
        cliente.email = data.get('email', '').strip() or None
        cliente.telefono = data.get('telefono', '').strip() or None
        cliente.direccion = data.get('direccion', '').strip() or None
        cliente.condicion_iva = data.get('condicion_iva', 'CONSUMIDOR_FINAL')
        cliente.tipo_precio = data.get('tipo_precio', 'venta')
        cliente.lista_precio = int(data.get('lista_precio', 1) or 1)
        cliente.zona_id = data.get('zona_id') or None
        cliente.vendedor_id = data.get('vendedor_id') or None
        cliente.dias_vencimiento_factura = data.get('dias_vencimiento_factura')
        cliente.es_intermediario = bool(data.get('es_intermediario', False))
        saldo_raw = data.get('saldo', 0)
        cliente.saldo = Decimal(str(round(float(saldo_raw) if saldo_raw is not None else 0, 2)))
        
        # Guardar en base de datos
        if not cliente_id:  # Solo agregar si es nuevo
            db.session.add(cliente)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Cliente {accion} correctamente',
            'cliente_id': cliente.id,
            'cliente_nombre': cliente.nombre
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando cliente: {str(e)}")
        return jsonify({'error': f'Error al guardar cliente: {str(e)}'}), 500

@app.route('/eliminar_cliente/<int:cliente_id>', methods=['DELETE'])
def eliminar_cliente(cliente_id):
    """Eliminar un cliente (marcar como inactivo)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        cliente = Cliente.query.get_or_404(cliente_id)
        
        # Verificar si tiene facturas asociadas
        facturas_count = Factura.query.filter_by(cliente_id=cliente_id).count()
        
        if facturas_count > 0:
            return jsonify({
                'error': f'No se puede eliminar el cliente porque tiene {facturas_count} facturas asociadas'
            }), 400
        
        # Si no tiene facturas, se puede eliminar
        db.session.delete(cliente)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cliente eliminado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error eliminando cliente: {str(e)}")
        return jsonify({'error': f'Error al eliminar cliente: {str(e)}'}), 500

@app.route('/buscar_clientes')
def buscar_clientes():
    """Buscar clientes con filtros (incluye zona)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        buscar = (request.args.get('buscar') or '').strip()
        tipo_doc = (request.args.get('tipo_documento') or '').strip()
        condicion_iva = (request.args.get('condicion_iva') or '').strip()
        zona = (request.args.get('zona') or '').strip()  # "", "SIN_ZONA", o id

        # ── FILTROS NUEVOS POR SALDO (mayo 2026) ──
        filtro_saldo = (request.args.get('filtro_saldo') or '').strip()
        # Valores: 'con_deuda' | 'sin_deuda' | 'a_favor' | ''
        try:
            saldo_min = float(request.args.get('saldo_min') or 0)
        except (ValueError, TypeError):
            saldo_min = 0.0
        orden = (request.args.get('orden') or '').strip()
        # Valores: 'saldo_desc' | '' (default alfabético)

        # Calculamos saldo en subconsulta (mismo cálculo que en /clientes)
        # Luego usamos HAVING para filtrar por saldo si corresponde.

        sql = """
            SELECT c.id, c.nombre, c.documento, c.tipo_documento, c.email, c.telefono,
                   c.direccion, c.condicion_iva, c.lista_precio, c.tipo_precio,
                   c.zona_id, z.nombre AS zona_nombre, z.color AS zona_color,
                   c.dias_vencimiento_factura,
                   (
                     COALESCE((
                       SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                       WHERE m.cliente_id = c.id AND m.tipo = 'venta_fiada'
                     ), 0)
                     -
                     COALESCE((
                       SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                       WHERE m.cliente_id = c.id AND m.tipo = 'nota_credito'
                     ), 0)
                     -
                     (
                       COALESCE((
                         SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                         WHERE m.cliente_id = c.id AND m.tipo_mov = 'pago'
                           AND m.numero_comprobante = 'A_CUENTA'
                       ), 0)
                       -
                       COALESCE((
                         SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                         WHERE m.cliente_id = c.id AND m.tipo_mov = 'pago'
                           AND m.numero_comprobante = 'SALDO_FAVOR_USADO'
                       ), 0)
                     )
                   ) AS saldo
              FROM cliente c
              LEFT JOIN zona z ON z.id = c.zona_id
             WHERE 1=1
        """
        params = {}
        if buscar:
            sql += " AND (c.nombre LIKE :buscar OR c.documento LIKE :buscar OR c.email LIKE :buscar)"
            params['buscar'] = f"%{buscar}%"
        if tipo_doc:
            sql += " AND c.tipo_documento = :tipo_doc"
            params['tipo_doc'] = tipo_doc
        if condicion_iva:
            sql += " AND c.condicion_iva = :condicion_iva"
            params['condicion_iva'] = condicion_iva
        if zona == 'SIN_ZONA':
            sql += " AND c.zona_id IS NULL"
        elif zona:
            try:
                params['zona_id'] = int(zona)
                sql += " AND c.zona_id = :zona_id"
            except ValueError:
                pass

        # ── FILTROS POR SALDO ──
        # Envolvemos la query original en una subconsulta para poder filtrar
        # por el campo calculado `saldo` (es más portable que HAVING sin GROUP BY).
        where_saldo = []
        if filtro_saldo == 'con_deuda':
            where_saldo.append("__sub.saldo > 0")
        elif filtro_saldo == 'sin_deuda':
            where_saldo.append("(__sub.saldo = 0 OR __sub.saldo IS NULL)")
        elif filtro_saldo == 'a_favor':
            where_saldo.append("__sub.saldo < 0")

        if saldo_min and saldo_min > 0:
            where_saldo.append("__sub.saldo >= :saldo_min")
            params['saldo_min'] = float(saldo_min)

        # ── ORDEN ──
        order_sql = " ORDER BY __sub.saldo DESC, __sub.nombre ASC" if orden == 'saldo_desc' else " ORDER BY __sub.nombre"

        if where_saldo:
            # Necesitamos envolver: SELECT * FROM (sql_original) __sub WHERE ...
            sql_final = f"SELECT * FROM ({sql}) AS __sub WHERE " + " AND ".join(where_saldo) + order_sql
        else:
            sql_final = f"SELECT * FROM ({sql}) AS __sub" + order_sql

        rows = db.session.execute(text(sql_final), params).mappings().all()

        resultado = [{
            'id': r['id'],
            'nombre': r['nombre'],
            'documento': r['documento'],
            'tipo_documento': r['tipo_documento'],
            'email': r['email'],
            'telefono': r['telefono'],
            'direccion': r['direccion'],
            'condicion_iva': r['condicion_iva'],
            'lista_precio': r['lista_precio'] or 1,
            'tipo_precio': r['tipo_precio'] or 'venta',
            'saldo': float(r['saldo']) if r['saldo'] else 0.0,
            'zona_id': r['zona_id'],
            'zona_nombre': r['zona_nombre'],
            'zona_color': r['zona_color'],
            'dias_vencimiento_factura': r['dias_vencimiento_factura']
        } for r in rows]

        return jsonify({
            'success': True,
            'clientes': resultado,
            'total': len(resultado)
        })
        
    except Exception as e:
        print(f"Error buscando clientes: {str(e)}")
        return jsonify({'error': f'Error en la búsqueda: {str(e)}'}), 500

# ==================== RUTAS DE PRODUCTOS ====================

# 1. ACTUALIZAR LA RUTA /api/producto_detalle/<int:producto_id>
# RUTA CORREGIDA PARA COMBOS MULTI-PRODUCTO
@app.route('/api/producto_detalle/<int:producto_id>')
def obtener_producto_detalle(producto_id):
    """Obtener datos completos de un producto para edición - CORREGIDO PARA MULTI-PRODUCTO"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        # Usar valores por defecto si costo o margen son 0 o None
        costo = float(producto.costo) if producto.costo and producto.costo > 0 else 0.0
        margen = float(producto.margen) if producto.margen is not None else 30.0
        
        # Si no hay costo pero hay precio, calcular costo aproximado
        if costo == 0.0 and producto.precio > 0:
            costo = float(producto.precio) / (1 + (margen / 100))
        
        # Datos base del producto
        resultado = {
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'descripcion': producto.descripcion or '',
            'precio': float(producto.precio),
            'costo': round(costo, 2),
            'margen': float(margen),
            'stock': producto.stock_dinamico,
            'stock_minimo': float(producto.stock_minimo or 0), 
            'categoria': producto.categoria or '',
            'iva': float(producto.iva),
            'activo': producto.activo,
            'es_pesable': bool(getattr(producto, 'es_pesable', False)),
            'imagen_url': producto.imagen_url or '',
            
            # CAMPOS DE COMBO BÁSICOS
            'es_combo': getattr(producto, 'es_combo', False),
            'cantidad_combo': getattr(producto, 'cantidad_combo', None),
            'producto_base_id': getattr(producto, 'producto_base_id', None),
            
            # NUEVOS CAMPOS PARA MULTI-PRODUCTO
            'producto_base_2_id': getattr(producto, 'producto_base_2_id', None),
            'cantidad_combo_2': getattr(producto, 'cantidad_combo_2', None),
            'producto_base_3_id': getattr(producto, 'producto_base_3_id', None),
            'cantidad_combo_3': getattr(producto, 'cantidad_combo_3', None),
            'codigo_barras': producto.codigo_barras or '',
            'margen2': float(producto.margen2) if producto.margen2 else None,
            'margen3': float(producto.margen3) if producto.margen3 else None,
            'margen4': float(producto.margen4) if producto.margen4 else None,
            'margen5': float(producto.margen5) if producto.margen5 else None,
            'precio2': float(producto.precio2) if producto.precio2 else None,
            'precio3': float(producto.precio3) if producto.precio3 else None,
            'precio4': float(producto.precio4) if producto.precio4 else None,
            'precio5': float(producto.precio5) if producto.precio5 else None,
        }
        
        # SI ES UN COMBO, CARGAR INFORMACIÓN DE TODOS LOS PRODUCTOS BASE
        if resultado['es_combo']:
            print(f"🔍 Cargando detalles de combo multi-producto: {producto.codigo}")
            
            # PRODUCTO BASE 1 (obligatorio)
            if resultado['producto_base_id']:
                try:
                    producto_base = Producto.query.get(resultado['producto_base_id'])
                    if producto_base:
                        resultado['producto_base'] = {
                            'id': producto_base.id,
                            'codigo': producto_base.codigo,
                            'nombre': producto_base.nombre,
                            'precio': float(producto_base.precio)
                        }
                        resultado['precio_unitario_base'] = float(producto_base.precio)
                        print(f"   ✅ Producto base 1: {producto_base.codigo}")
                    else:
                        resultado['producto_base'] = None
                        resultado['precio_unitario_base'] = 0.0
                        print(f"   ❌ Producto base 1 no encontrado: ID {resultado['producto_base_id']}")
                except Exception as e:
                    print(f"   ⚠️ Error cargando producto base 1: {str(e)}")
                    resultado['producto_base'] = None
                    resultado['precio_unitario_base'] = 0.0
            
            # PRODUCTO BASE 2 (opcional)
            if resultado['producto_base_2_id']:
                try:
                    producto_base_2 = Producto.query.get(resultado['producto_base_2_id'])
                    if producto_base_2:
                        resultado['producto_base_2'] = {
                            'id': producto_base_2.id,
                            'codigo': producto_base_2.codigo,
                            'nombre': producto_base_2.nombre,
                            'precio': float(producto_base_2.precio)
                        }
                        print(f"   ✅ Producto base 2: {producto_base_2.codigo}")
                    else:
                        resultado['producto_base_2'] = None
                        print(f"   ❌ Producto base 2 no encontrado: ID {resultado['producto_base_2_id']}")
                except Exception as e:
                    print(f"   ⚠️ Error cargando producto base 2: {str(e)}")
                    resultado['producto_base_2'] = None
            else:
                resultado['producto_base_2'] = None
            
            # PRODUCTO BASE 3 (opcional)
            if resultado['producto_base_3_id']:
                try:
                    producto_base_3 = Producto.query.get(resultado['producto_base_3_id'])
                    if producto_base_3:
                        resultado['producto_base_3'] = {
                            'id': producto_base_3.id,
                            'codigo': producto_base_3.codigo,
                            'nombre': producto_base_3.nombre,
                            'precio': float(producto_base_3.precio)
                        }
                        print(f"   ✅ Producto base 3: {producto_base_3.codigo}")
                    else:
                        resultado['producto_base_3'] = None
                        print(f"   ❌ Producto base 3 no encontrado: ID {resultado['producto_base_3_id']}")
                except Exception as e:
                    print(f"   ⚠️ Error cargando producto base 3: {str(e)}")
                    resultado['producto_base_3'] = None
            else:
                resultado['producto_base_3'] = None
            
            # CALCULAR PRECIO NORMAL TOTAL (TODOS LOS PRODUCTOS)
            precio_normal_total = 0
            
            # Producto 1
            if resultado.get('producto_base') and resultado.get('cantidad_combo'):
                precio_normal_total += resultado['producto_base']['precio'] * float(resultado['cantidad_combo'])
            
            # Producto 2
            if resultado.get('producto_base_2') and resultado.get('cantidad_combo_2'):
                precio_normal_total += resultado['producto_base_2']['precio'] * float(resultado['cantidad_combo_2'])
            
            # Producto 3
            if resultado.get('producto_base_3') and resultado.get('cantidad_combo_3'):
                precio_normal_total += resultado['producto_base_3']['precio'] * float(resultado['cantidad_combo_3'])
            
            # AGREGAR INFORMACIÓN CALCULADA
            resultado['precio_normal_total'] = round(precio_normal_total, 2)
            resultado['ahorro_total'] = round(precio_normal_total - float(producto.precio), 2)
            resultado['descuento_porcentaje_calculado'] = round(
                ((precio_normal_total - float(producto.precio)) / precio_normal_total * 100), 1
            ) if precio_normal_total > 0 else 0
            
            print(f"   💰 Precio normal total: ${precio_normal_total:.2f}")
            print(f"   💰 Precio combo: ${float(producto.precio):.2f}")
            print(f"   💰 Ahorro: ${resultado['ahorro_total']:.2f} ({resultado['descuento_porcentaje_calculado']:.1f}%)")
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Error en obtener_producto_detalle: {str(e)}")
        return jsonify({'error': f'Error al obtener producto: {str(e)}'}), 500


# ============================================================
# SUBIDA DE IMAGEN DE PRODUCTO (para Pedidos Online)
# ============================================================
IMAGENES_PRODUCTOS_DIR = os.path.join(app.static_folder, 'productos_img')
IMAGEN_EXTENSIONES_OK = {'png', 'jpg', 'jpeg', 'webp', 'gif'}
IMAGEN_MAX_BYTES = 3 * 1024 * 1024  # 3 MB

@app.route('/subir_imagen_producto', methods=['POST'])
def subir_imagen_producto():
    """Recibe una imagen, la valida y la guarda en /static/productos_img/.
    Devuelve {success, url} con la ruta pública para guardar en imagen_url."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401

    try:
        if 'imagen' not in request.files:
            return jsonify({'success': False, 'message': 'No se recibió ninguna imagen'}), 400

        archivo = request.files['imagen']
        if not archivo or archivo.filename == '':
            return jsonify({'success': False, 'message': 'Archivo vacío'}), 400

        # Validar extensión
        nombre_original = secure_filename(archivo.filename)
        if '.' not in nombre_original:
            return jsonify({'success': False, 'message': 'Archivo sin extensión'}), 400
        ext = nombre_original.rsplit('.', 1)[1].lower()
        if ext not in IMAGEN_EXTENSIONES_OK:
            return jsonify({'success': False, 'message': 'Formato no permitido (usá JPG, PNG, WebP o GIF)'}), 400

        # Validar tamaño
        archivo.seek(0, os.SEEK_END)
        tam = archivo.tell()
        archivo.seek(0)
        if tam > IMAGEN_MAX_BYTES:
            return jsonify({'success': False, 'message': 'La imagen supera los 3 MB'}), 400

        # Crear carpeta si no existe
        os.makedirs(IMAGENES_PRODUCTOS_DIR, exist_ok=True)

        # Nombre único para evitar pisar y cache
        nombre_final = f"prod_{uuid.uuid4().hex}.{ext}"
        ruta_fisica = os.path.join(IMAGENES_PRODUCTOS_DIR, nombre_final)
        archivo.save(ruta_fisica)

        # URL pública relativa
        url_publica = url_for('static', filename=f'productos_img/{nombre_final}')

        print(f"🖼️  Imagen de producto subida: {nombre_final} ({tam} bytes)")

        return jsonify({'success': True, 'url': url_publica})

    except Exception as e:
        print(f"❌ Error al subir imagen de producto: {e}")
        return jsonify({'success': False, 'message': 'Error interno al guardar la imagen'}), 500


# 2. ACTUALIZAR LA RUTA /guardar_producto
@app.route('/guardar_producto', methods=['POST'])
def guardar_producto():
    """Crear o actualizar un producto con costo y margen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json or {}

        # Helper: trae el valor como str y aplica strip(). Tolera None.
        def s(key, default=''):
            v = data.get(key)
            if v is None:
                return default
            return str(v).strip()

        # Validar datos requeridos
        if not s('codigo'):
            return jsonify({'error': 'El código es obligatorio'}), 400
        
        if not s('nombre'):
            return jsonify({'error': 'El nombre es obligatorio'}), 400
        
        # Validar costo
        try:
            costo = float(data.get('costo', 0) or 0)
            if costo <= 0:
                return jsonify({'error': 'El costo debe ser mayor a 0'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Costo inválido'}), 400
        
        # Validar margen
        try:
            margen = float(data.get('margen', 30) or 30)
            if margen < 0:
                return jsonify({'error': 'El margen no puede ser negativo'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Margen inválido'}), 400
        
        # Calcular precio automáticamente
        precio_calculado = costo * (1 + (margen / 100))
        
        producto_id = data.get('id')
        codigo = s('codigo').upper()
        
        # Verificar que el código no exista (excepto si es el mismo producto)
        producto_existente = Producto.query.filter_by(codigo=codigo).first()
        if producto_existente and (not producto_id or producto_existente.id != int(producto_id)):
            return jsonify({'error': f'Ya existe un producto con el código {codigo}'}), 400
        
        if producto_id:  # Editar producto existente
            producto = Producto.query.get_or_404(producto_id)
            accion = 'actualizado'
        else:  # Crear nuevo producto
            producto = Producto()
            accion = 'creado'
        
        # Actualizar datos del producto
        producto.codigo = codigo
        producto.nombre = s('nombre')
        producto.descripcion = s('descripcion') or None
        precio_anterior = float(producto.precio) if producto_id else None
        producto.precio = aplicar_redondeo_precio(precio_calculado)
        producto.costo = Decimal(str(round(costo, 2)))
        producto.margen = Decimal(str(round(margen, 2)))
        producto.categoria = s('categoria') or None
        producto.iva = Decimal(str(data.get('iva', 21) or 21))
        producto.activo = bool(data.get('activo', True))
        producto.stock_minimo = Decimal(str(data.get('stock_minimo', 0) or 0))
        producto.fecha_modificacion = datetime.now()
        producto.codigo_barras = s('codigo_barras') or None

        # Flag es_pesable (booleano)
        if hasattr(producto, 'es_pesable'):
            producto.es_pesable = bool(data.get('es_pesable', False))

        # Imagen del producto (se muestra en Pedidos Online)
        if hasattr(producto, 'imagen_url'):
            producto.imagen_url = s('imagen_url') or None

        # Actualizar fecha si el precio cambió
        if precio_anterior is None or round(precio_calculado, 2) != round(precio_anterior, 2):
            producto.fecha_actualizacion_precio = datetime.now()
        def parse_precio(key):
            v = data.get(key)
            try:
                return Decimal(str(round(float(v), 2))) if v not in (None, '', 0, '0') else None
            except:
                return None

        # Calcular precio de cada lista a partir del margen y el costo
        for i in range(2, 6):
            margen_lista = parse_precio(f'margen{i}')
            precio_lista = parse_precio(f'precio{i}')
            if margen_lista is not None:
                # Calcular precio desde margen (con redondeo configurable)
                precio_calculado_lista = float(costo) * (1 + float(margen_lista) / 100)
                setattr(producto, f'margen{i}', margen_lista)
                setattr(producto, f'precio{i}', aplicar_redondeo_precio(precio_calculado_lista))
            elif precio_lista is not None:
                # Si vino precio manual sin margen, también lo redondeamos según config
                setattr(producto, f'precio{i}', aplicar_redondeo_precio(precio_lista))
            else:
                setattr(producto, f'margen{i}', None)
                setattr(producto, f'precio{i}', None)
        
        # Solo actualizar stock si es producto nuevo
        if not producto_id:
            producto.stock = int(data.get('stock', 0))
        
        # Guardar en base de datos
        if not producto_id:  # Solo agregar si es nuevo
            db.session.add(producto)
        
        db.session.commit()
        
        print(f"✅ Producto {accion}: {codigo}")
        print(f"   Costo: ${costo:.2f}")
        print(f"   Margen: {margen}%")
        print(f"   Precio: ${precio_calculado:.2f}")
        
        return jsonify({
            'success': True,
            'message': f'Producto {accion} correctamente',
            'producto_id': producto.id,
            'producto_codigo': producto.codigo,
            'precio_calculado': round(precio_calculado, 2),
            'costo': round(costo, 2),
            'margen': round(margen, 2)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error guardando producto: {str(e)}")
        return jsonify({'error': f'Error al guardar producto: {str(e)}'}), 500




@app.route('/ajustar_stock', methods=['POST'])
def ajustar_stock():
    """Ajustar stock de un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        tipo_movimiento = data.get('tipo_movimiento')  # entrada, salida, ajuste
        cantidad = float(data.get('cantidad', 0)) 
        motivo = data.get('motivo', '').strip()
        
        if not producto_id:
            return jsonify({'error': 'ID de producto requerido'}), 400
        
        if cantidad <= 0 and tipo_movimiento != 'ajuste':
            return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
        
        producto = Producto.query.get_or_404(producto_id)
        stock_anterior = producto.stock
        
        # Aplicar movimiento según el tipo
        if tipo_movimiento == 'entrada':
            producto.stock = float(producto.stock) + cantidad
            descripcion = f"Entrada: +{cantidad}"
        elif tipo_movimiento == 'salida':
            if cantidad > producto.stock:
                return jsonify({'error': f'No hay suficiente stock. Stock actual: {producto.stock}'}), 400
            producto.stock = float(producto.stock) - cantidad
            descripcion = f"Salida: -{cantidad}"
        elif tipo_movimiento == 'ajuste':
            if cantidad < 0:
                return jsonify({'error': 'La cantidad para ajuste no puede ser negativa'}), 400
            descripcion = f"Ajuste: {stock_anterior} → {cantidad}"
            producto.stock = cantidad
        else:
            return jsonify({'error': 'Tipo de movimiento inválido'}), 400
        
        # Guardar cambios
        db.session.commit()
        
        # Registrar el movimiento en consola
        print(f"MOVIMIENTO STOCK: Producto {producto.codigo} - {descripcion} - Motivo: {motivo}")
        
        # Registrar en auditoría de stock
        # FIX: para 'ajuste' la cantidad recibida es el STOCK FINAL deseado, no un delta.
        # Calculamos el delta real (stock_nuevo - stock_anterior) y el signo según dirección.
        # Para entrada/salida queda como antes.
        tipo_audit = 'ajuste_entrada' if tipo_movimiento == 'entrada' else ('ajuste_salida' if tipo_movimiento == 'salida' else 'ajuste_manual')
        if tipo_movimiento == 'ajuste':
            delta_ajuste = float(producto.stock) - float(stock_anterior)
            cantidad_audit = abs(delta_ajuste)
            signo_audit = '+' if delta_ajuste >= 0 else '-'
        else:
            cantidad_audit = cantidad
            signo_audit = '+' if tipo_movimiento == 'entrada' else '-'

        registrar_movimiento_stock(
            db=db,
            producto_id=producto_id,
            tipo=tipo_audit,
            cantidad=cantidad_audit,
            signo=signo_audit,
            stock_anterior=float(stock_anterior),
            stock_nuevo=float(producto.stock),
            referencia_tipo='manual',
            motivo=motivo,
            usuario_id=session.get('user_id'),
            usuario_nombre=session.get('nombre', 'Sistema'),
            codigo_producto=producto.codigo,
            nombre_producto=producto.nombre
        )
        
        return jsonify({
            'success': True,
            'message': f'Stock ajustado correctamente',
            'stock_anterior': float(stock_anterior),
            'stock_nuevo': float(producto.stock),
            'movimiento': descripcion
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error ajustando stock: {str(e)}")
        return jsonify({'error': f'Error al ajustar stock: {str(e)}'}), 500

@app.route('/toggle_producto/<int:producto_id>', methods=['POST'])
def toggle_producto(producto_id):
    """Activar/desactivar un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        # Cambiar estado
        producto.activo = not producto.activo
        estado = 'activado' if producto.activo else 'desactivado'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Producto {estado} correctamente',
            'activo': producto.activo
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error cambiando estado del producto: {str(e)}")
        return jsonify({'error': f'Error al cambiar estado: {str(e)}'}), 500

# 3. ACTUALIZAR LA RUTA /buscar_productos_admin
@app.route('/buscar_productos_admin')
def buscar_productos_admin():
    """Buscar productos con filtros para administración - INCLUYE FILTROS PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # ─── Cargar mapa de colores de categorías una sola vez ───
        # Hacemos query directa porque el modelo Categoria está en otro módulo (blueprint)
        try:
            colores_categoria = {}
            result = db.session.execute(db.text(
                "SELECT nombre, color FROM categoria WHERE activo = 1"
            ))
            for row in result:
                # Guardar por nombre (case-insensitive con strip por si hay espacios)
                colores_categoria[(row[0] or '').strip().lower()] = row[1] or '#6c757d'
        except Exception as e:
            print(f"⚠️ No se pudo cargar colores de categorías: {e}")
            colores_categoria = {}
        
        # Obtener parámetros de búsqueda
        buscar = request.args.get('buscar', '').strip()
        categoria = request.args.get('categoria', '').strip()
        filtro_stock = request.args.get('stock', '').strip()
        estado = request.args.get('estado', '').strip()
        
        # ✅ NUEVOS PARÁMETROS PARA COMBOS
        solo_combos = request.args.get('solo_combos', '').strip().lower() == 'true'
        estado = request.args.get('estado', '').strip()  # activo, inactivo
        descuento = request.args.get('descuento', '').strip()  # alto, medio, bajo
        
        print(f"🔍 Búsqueda productos admin:")
        print(f"   Buscar: '{buscar}'")
        print(f"   Solo combos: {solo_combos}")
        print(f"   Estado: '{estado}'")
        print(f"   Descuento: '{descuento}'")
        
        # Construir query base
        query = Producto.query
        
        # ✅ FILTRO PARA SOLO COMBOS
        if solo_combos:
            query = query.filter(Producto.es_combo == True)
            print("   Filtro aplicado: Solo combos")
        
        # Aplicar filtros generales
        if buscar:
            query = query.filter(
                or_(
                    Producto.codigo.ilike(f'%{buscar}%'),
                    Producto.nombre.ilike(f'%{buscar}%'),
                    Producto.descripcion.ilike(f'%{buscar}%')
                )
            )
            print(f"   Filtro aplicado: Búsqueda '{buscar}'")
        
        if categoria:
            query = query.filter(Producto.categoria == categoria)
            print(f"   Filtro aplicado: Categoría '{categoria}'")
        
        if filtro_stock == 'bajo':
            query = query.filter(Producto.stock < 10)
        elif filtro_stock == 'sin_stock':
            query = query.filter(Producto.stock <= 0)
        
        # ✅ FILTRO DE ESTADO (ACTIVO/INACTIVO)
        if estado == 'activo':
            query = query.filter(Producto.activo == True)
            print("   Filtro aplicado: Solo activos")
        elif estado == 'inactivo':
            query = query.filter(Producto.activo == False)
            print("   Filtro aplicado: Solo inactivos")
        
        # Obtener resultados SIN filtro de descuento primero
        productos = query.order_by(Producto.codigo).all()
        print(f"   Productos encontrados (antes filtro descuento): {len(productos)}")
        
        # ✅ APLICAR FILTRO DE DESCUENTO DESPUÉS (solo para combos)
        if descuento and solo_combos:
            productos_filtrados = []
            
            for producto in productos:
                if producto.es_combo and producto.producto_base:
                    # Calcular descuento del combo
                    precio_normal = float(producto.producto_base.precio) * float(producto.cantidad_combo)
                    precio_combo = float(producto.precio)
                    descuento_porcentaje = ((precio_normal - precio_combo) / precio_normal) * 100 if precio_normal > 0 else 0
                    
                    # Aplicar filtro según nivel de descuento
                    if descuento == 'alto' and descuento_porcentaje > 30:
                        productos_filtrados.append(producto)
                    elif descuento == 'medio' and 15 <= descuento_porcentaje <= 30:
                        productos_filtrados.append(producto)
                    elif descuento == 'bajo' and descuento_porcentaje < 15:
                        productos_filtrados.append(producto)
                else:
                    # Si no es combo, incluir sin filtro de descuento
                    productos_filtrados.append(producto)
            
            productos = productos_filtrados
            print(f"   Productos después filtro descuento '{descuento}': {len(productos)}")
        
        # Formatear respuesta
        resultado = []
        for producto in productos:
            # Manejar valores por defecto
            costo = float(producto.costo) if producto.costo else 0.0
            margen = float(producto.margen) if producto.margen is not None else 0.0
            
            # Si no hay costo guardado, calcularlo aproximadamente desde precio
            if costo == 0.0 and producto.precio > 0 and margen > 0:
                costo = float(producto.precio) / (1 + (margen / 100))
            
            producto_dict = {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio': float(producto.precio),
                'costo': round(costo, 2),
                'margen': round(margen, 1),
                'stock': producto.stock_dinamico,
                'categoria': producto.categoria,
                'categoria_color': colores_categoria.get(
                    (producto.categoria or '').strip().lower(),
                    '#6c757d'
                ),
                'iva': float(producto.iva),
                'activo': producto.activo,
                'es_combo': producto.es_combo,
                'acceso_rapido': producto.acceso_rapido if hasattr(producto, 'acceso_rapido') else False,
                'orden_acceso_rapido': producto.orden_acceso_rapido if hasattr(producto, 'orden_acceso_rapido') else 0
            }
            
            # ✅ AGREGAR INFORMACIÓN ESPECÍFICA PARA COMBOS
            if producto.es_combo:
                producto_dict.update({
                    'producto_base_id': producto.producto_base_id,
                    'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
                    'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else 0.0
                })
                
                # Información del producto base
                if producto.producto_base:
                    producto_dict['producto_base'] = {
                        'id': producto.producto_base.id,
                        'codigo': producto.producto_base.codigo,
                        'nombre': producto.producto_base.nombre,
                        'precio': float(producto.producto_base.precio)
                    }
            
            resultado.append(producto_dict)
        
        print(f"✅ Búsqueda completada: {len(resultado)} productos")
        
        return jsonify({
            'success': True,
            'productos': resultado,
            'total': len(resultado),
            'filtros_aplicados': {
                'buscar': buscar,
                'solo_combos': solo_combos,
                'estado': estado,
                'descuento': descuento,
                'categoria': categoria,
                'filtro_stock': filtro_stock
            }
        })
        
    except Exception as e:
        print(f"❌ Error buscando productos: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({'error': f'Error en la búsqueda: {str(e)}'}), 500

# FUNCIÓN PARA ACTUALIZAR PRODUCTOS EXISTENTES CON COSTO CALCULADO
@app.route('/actualizar_costos_productos', methods=['POST'])
def actualizar_costos_productos():
    """Actualizar productos que tienen costo 0 calculándolo desde precio y margen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        productos_sin_costo = Producto.query.filter(
            or_(Producto.costo == 0, Producto.costo.is_(None))
        ).all()
        
        contador_actualizados = 0
        
        for producto in productos_sin_costo:
            if producto.precio > 0:
                margen = float(producto.margen) if producto.margen else 30.0
                # Calcular costo desde precio: costo = precio / (1 + margen/100)
                costo_calculado = float(producto.precio) / (1 + (margen / 100))
                
                producto.costo = Decimal(str(round(costo_calculado, 2)))
                producto.fecha_modificacion = datetime.now()
                
                contador_actualizados += 1
                print(f"📦 Actualizado: {producto.codigo} - Precio=${float(producto.precio):.2f} → Costo=${costo_calculado:.2f}")
        
        if contador_actualizados > 0:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Se actualizaron {contador_actualizados} productos',
            'productos_actualizados': contador_actualizados
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error actualizando costos: {str(e)}")
        return jsonify({'error': f'Error al actualizar costos: {str(e)}'}), 500

@app.route('/obtener_categorias')
def obtener_categorias():
    """Obtener lista de categorías únicas"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener categorías únicas de productos existentes
        categorias = db.session.query(Producto.categoria).filter(
            Producto.categoria.isnot(None),
            Producto.categoria != ''
        ).distinct().all()
        
        categorias_lista = [cat[0] for cat in categorias if cat[0]]
        categorias_lista.sort()
        
        return jsonify({
            'success': True,
            'categorias': categorias_lista
        })
        
    except Exception as e:
        print(f"Error obteniendo categorías: {str(e)}")
        return jsonify({'error': f'Error al obtener categorías: {str(e)}'}), 500


###### RUTAS API PARA MANEJAR COMBOS

@app.route('/api/productos_con_ofertas')
def api_productos_con_ofertas():
    """API para obtener productos con sus ofertas/combos"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        productos = Producto.obtener_productos_con_ofertas()
        return jsonify({
            'success': True,
            'productos': productos
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/crear_combo', methods=['POST'])
def crear_combo():
    try:
        datos = request.get_json()
        
        # Validación básica
        if not datos.get('producto_base_id'):
            return jsonify({'success': False, 'error': 'Producto base 1 es requerido'})
            
        # Si es edición, buscar combo existente
        if datos.get('id'):
            combo = Producto.query.get(datos['id'])
            if not combo:
                return jsonify({'success': False, 'error': 'Combo no encontrado'})
        else:
            combo = Producto()
            combo.es_combo = True
            combo.activo = True
            combo.stock = 0
        
        # Datos básicos del combo
        combo.producto_base_id = datos['producto_base_id']
        combo.cantidad_combo = Decimal(str(float(datos['cantidad_combo'])))  # Conversión segura
        combo.precio = Decimal(str(datos['precio_combo']))
        
        # NUEVOS CAMPOS: Productos adicionales
        combo.producto_base_2_id = datos.get('producto_base_2_id')
        if datos.get('cantidad_combo_2'):
            combo.cantidad_combo_2 = Decimal(str(float(datos['cantidad_combo_2'])))
        else:
            combo.cantidad_combo_2 = Decimal('0')
        
        combo.producto_base_3_id = datos.get('producto_base_3_id')
        if datos.get('cantidad_combo_3'):
            combo.cantidad_combo_3 = Decimal(str(float(datos['cantidad_combo_3'])))
        else:
            combo.cantidad_combo_3 = Decimal('0')
        
        # Generar código automático si no se proporciona
        if not datos.get('codigo_combo'):
            combo.codigo = generar_codigo_combo_multi(combo)
        else:
            combo.codigo = datos['codigo_combo']
            
        # Generar nombre automático si no se proporciona
        if not datos.get('nombre_combo'):
            combo.nombre = generar_nombre_combo_multi(combo)
        else:
            combo.nombre = datos['nombre_combo']
            
        combo.descripcion = datos.get('descripcion_combo', '')

        # IVA: usar el enviado o heredar del producto base
        if datos.get('iva'):
            combo.iva = Decimal(str(datos['iva']))
        else:
            # Heredar IVA del producto base principal
            producto_base = Producto.query.get(combo.producto_base_id)
            if producto_base:
                combo.iva = producto_base.iva
            else:
                combo.iva = Decimal('21')  # Default 21%
        
        # Validar que el precio de oferta sea menor al precio normal
        precio_normal_total = calcular_precio_normal_multi(combo)
        if float(combo.precio) >= precio_normal_total:
            return jsonify({
                'success': False, 
                'error': 'El precio de oferta debe ser menor al precio normal'
            })
        
        db.session.add(combo)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'Combo creado exitosamente',
            'codigo': combo.codigo
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

def calcular_precio_normal_multi(combo):
    """Calcular precio normal total del combo multi-producto"""
    try:
        precio_total = 0.0  # Empezar con float
        
        print(f"🔍 DEBUG calcular_precio_normal_multi:")
        print(f"   Combo ID: {getattr(combo, 'id', 'NUEVO')}")
        
        # Producto 1 (obligatorio)
        if combo.producto_base_id and combo.cantidad_combo:
            producto1 = Producto.query.get(combo.producto_base_id)
            if producto1:
                precio_unit = float(producto1.precio)
                cantidad = float(combo.cantidad_combo)
                subtotal = precio_unit * cantidad
                precio_total += subtotal
                
                print(f"   Producto 1: {precio_unit} x {cantidad} = {subtotal}")
        
        # Producto 2 (opcional)
        if combo.producto_base_2_id and combo.cantidad_combo_2 and float(combo.cantidad_combo_2) > 0:
            producto2 = Producto.query.get(combo.producto_base_2_id)
            if producto2:
                precio_unit = float(producto2.precio)
                cantidad = float(combo.cantidad_combo_2)
                subtotal = precio_unit * cantidad
                precio_total += subtotal
                
                print(f"   Producto 2: {precio_unit} x {cantidad} = {subtotal}")
        
        # Producto 3 (opcional)
        if combo.producto_base_3_id and combo.cantidad_combo_3 and float(combo.cantidad_combo_3) > 0:
            producto3 = Producto.query.get(combo.producto_base_3_id)
            if producto3:
                precio_unit = float(producto3.precio)
                cantidad = float(combo.cantidad_combo_3)
                subtotal = precio_unit * cantidad
                precio_total += subtotal
                
                print(f"   Producto 3: {precio_unit} x {cantidad} = {subtotal}")
        
        print(f"   Total calculado: {precio_total}")
        return precio_total
        
    except Exception as e:
        print(f"❌ Error en calcular_precio_normal_multi: {str(e)}")
        print(f"❌ Tipos de datos:")
        print(f"   combo.cantidad_combo: {type(getattr(combo, 'cantidad_combo', None))}")
        print(f"   combo.cantidad_combo_2: {type(getattr(combo, 'cantidad_combo_2', None))}")
        print(f"   combo.cantidad_combo_3: {type(getattr(combo, 'cantidad_combo_3', None))}")
        raise e

        
def generar_codigo_combo_multi(combo):
    """Generar código automático para combo multi-producto"""
    codigos = []
    
    if combo.producto_base_id:
        producto1 = Producto.query.get(combo.producto_base_id)
        if producto1:
            codigos.append(producto1.codigo)
    
    if combo.producto_base_2_id:
        producto2 = Producto.query.get(combo.producto_base_2_id)
        if producto2:
            codigos.append(producto2.codigo)
    
    if combo.producto_base_3_id:
        producto3 = Producto.query.get(combo.producto_base_3_id)
        if producto3:
            codigos.append(producto3.codigo)
    
    return f"{'_'.join(codigos)}_COMBO"

def generar_nombre_combo_multi(combo):
    """Generar nombre automático para combo multi-producto"""
    nombres = []
    
    if combo.producto_base_id and combo.cantidad_combo:
        producto1 = Producto.query.get(combo.producto_base_id)
        if producto1:
            nombres.append(f"{combo.cantidad_combo}x {producto1.nombre}")
    
    if combo.producto_base_2_id and combo.cantidad_combo_2:
        producto2 = Producto.query.get(combo.producto_base_2_id)
        if producto2:
            nombres.append(f"{combo.cantidad_combo_2}x {producto2.nombre}")
    
    if combo.producto_base_3_id and combo.cantidad_combo_3:
        producto3 = Producto.query.get(combo.producto_base_3_id)
        if producto3:
            nombres.append(f"{combo.cantidad_combo_3}x {producto3.nombre}")
    
    return f"Pack: {' + '.join(nombres)} (Oferta)"

#**************************************************************
@app.route('/api/combos_producto/<int:producto_id>')
def obtener_combos_producto(producto_id):
    """Obtener todos los combos de un producto base"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto_base = Producto.query.get_or_404(producto_id)
        
        # Obtener combos del producto
        combos = Producto.query.filter_by(
            producto_base_id=producto_id,
            es_combo=True,
            activo=True
        ).order_by(Producto.cantidad_combo.asc()).all()
        
        # Preparar respuesta
        combos_data = []
        for combo in combos:
            combo_info = combo.to_dict()
            combos_data.append(combo_info)
        
        return jsonify({
            'success': True,
            'producto_base': producto_base.to_dict(),
            'combos': combos_data,
            'total_combos': len(combos_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# FUNCIÓN PARA MIGRAR PRODUCTOS EXISTENTES
def migrar_productos_para_combos():
    """Migrar productos existentes al nuevo sistema de combos"""
    try:
        print("🔄 Iniciando migración para sistema de combos...")
        
        # Actualizar productos existentes
        productos_sin_migrar = Producto.query.filter(
            Producto.es_combo.is_(None)
        ).all()
        
        contador_migrados = 0
        
        for producto in productos_sin_migrar:
            # Todos los productos existentes son productos base (no combos)
            producto.es_combo = False
            producto.cantidad_combo = Decimal('1.000')
            producto.precio_unitario_base = producto.precio
            producto.descuento_porcentaje = Decimal('0.00')
            
            contador_migrados += 1
        
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} productos actualizados")
        else:
            print("✅ No hay productos para migrar")
            
        return contador_migrados
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en migración: {e}")
        return 0


# EJEMPLO DE USO PARA CREAR COMBOS
def crear_ejemplos_combos():
    """Crear ejemplos de combos para demostración"""
    try:
        # Buscar producto base (milanesa)
        producto_base = Producto.query.filter_by(codigo='MIL001').first()
        
        if not producto_base:
            print("⚠️ Producto base MIL001 no encontrado")
            return
        
        combos_ejemplos = [
            {
                'codigo': 'MIL001-3KG',
                'nombre': '3kg Milanesa de Ternera (Oferta)',
                'cantidad': 3.0,
                'precio': 10000.00,
                'descripcion': 'Oferta especial: 3 kilogramos de milanesa de ternera'
            },
            {
                'codigo': 'MIL001-5KG', 
                'nombre': '5kg Milanesa de Ternera (Super Oferta)',
                'cantidad': 5.0,
                'precio': 18000.00,
                'descripcion': 'Super oferta: 5 kilogramos de milanesa de ternera'
            }
        ]
        
        for combo_data in combos_ejemplos:
            # Verificar si ya existe
            if Producto.query.filter_by(codigo=combo_data['codigo']).first():
                print(f"⚠️ Combo {combo_data['codigo']} ya existe")
                continue
            
            # Calcular valores
            precio_normal = float(producto_base.precio) * combo_data['cantidad']
            descuento_monto = precio_normal - combo_data['precio']
            descuento_porcentaje = (descuento_monto / precio_normal) * 100
            
            # Crear combo
            combo = Producto(
                codigo=combo_data['codigo'],
                nombre=combo_data['nombre'],
                descripcion=combo_data['descripcion'],
                precio=Decimal(str(combo_data['precio'])),
                categoria='OFERTAS',
                iva=producto_base.iva,
                costo=Decimal(str(float(producto_base.costo or 0) * combo_data['cantidad'])),
                stock=int(float(producto_base.stock) / combo_data['cantidad']),
                
                es_combo=True,
                producto_base_id=producto_base.id,
                cantidad_combo=Decimal(str(combo_data['cantidad'])),
                precio_unitario_base=producto_base.precio,
                descuento_porcentaje=Decimal(str(descuento_porcentaje))
            )
            
            db.session.add(combo)
            print(f"✅ Combo creado: {combo_data['codigo']} - Descuento: {descuento_porcentaje:.1f}%")
        
        db.session.commit()
        print("🎉 Ejemplos de combos creados exitosamente")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creando ejemplos: {e}")


# RUTA PARA EJECUTAR MIGRACIÓN
@app.route('/migrar_combos', methods=['POST'])
def ejecutar_migracion_combos():
    """Endpoint para ejecutar migración de combos"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        contador = migrar_productos_para_combos()
        
        return jsonify({
            'success': True,
            'mensaje': f'Migración completada: {contador} productos actualizados',
            'productos_migrados': contador
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500



@app.route('/nueva_venta')
def nueva_venta():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Obtener productos de acceso rápido en lugar de los primeros 8
    productos = Producto.query.filter_by(
        acceso_rapido=True,
        activo=True
    ).order_by(
        Producto.orden_acceso_rapido.asc(),
        Producto.codigo.asc()
    ).limit(8).all()
    
    clientes = Cliente.query.order_by(Cliente.nombre.asc()).all()
    return render_template('nueva_venta.html', productos=productos, clientes=clientes)

# APIs para búsqueda de productos
# 4. ACTUALIZAR LAS APIS DE BÚSQUEDA PARA INCLUIR COSTO
@app.route('/api/buscar_productos/<termino>')
def buscar_productos(termino):
    """Busca productos por código o nombre - CORREGIDO"""
    if not termino or len(termino) < 2:
        return jsonify([])
    
    try:
        # Búsqueda por código exacto primero
        producto_exacto = Producto.query.filter_by(codigo=termino.upper(), activo=True).first()
        if producto_exacto:
            # Obtener costos de productos base si es combo
            costo_unitario_base = 0.0
            costo_unitario_base_2 = 0.0
            costo_unitario_base_3 = 0.0
            cantidad_combo_2 = float(producto_exacto.cantidad_combo_2) if producto_exacto.cantidad_combo_2 else 0.0
            cantidad_combo_3 = float(producto_exacto.cantidad_combo_3) if producto_exacto.cantidad_combo_3 else 0.0
            
            if producto_exacto.es_combo:
                if producto_exacto.producto_base_id:
                    pb = Producto.query.get(producto_exacto.producto_base_id)
                    if pb and pb.costo:
                        costo_unitario_base = float(pb.costo)
                if producto_exacto.producto_base_2_id:
                    pb2 = Producto.query.get(producto_exacto.producto_base_2_id)
                    if pb2 and pb2.costo:
                        costo_unitario_base_2 = float(pb2.costo)
                if producto_exacto.producto_base_3_id:
                    pb3 = Producto.query.get(producto_exacto.producto_base_3_id)
                    if pb3 and pb3.costo:
                        costo_unitario_base_3 = float(pb3.costo)
            
            resultado = {
                'id': producto_exacto.id,
                'codigo': producto_exacto.codigo,
                'nombre': producto_exacto.nombre,
                'precio': float(producto_exacto.precio),
                'precio_base': float(producto_exacto.precio),
                'costo': float(producto_exacto.costo) if producto_exacto.costo else 0.0,
                'margen': float(producto_exacto.margen) if producto_exacto.margen else 0.0,
                'stock': producto_exacto.stock_dinamico,
                'iva': float(producto_exacto.iva),
                'match_tipo': 'codigo_exacto',
                'descripcion': producto_exacto.descripcion or '',
                'es_combo': producto_exacto.es_combo,
                'producto_base_id': producto_exacto.producto_base_id,
                'cantidad_combo': float(producto_exacto.cantidad_combo) if producto_exacto.cantidad_combo else 1.0,
                'cantidad_combo_2': cantidad_combo_2,
                'cantidad_combo_3': cantidad_combo_3,
                'precio_unitario_base': float(producto_exacto.precio_unitario_base) if producto_exacto.precio_unitario_base else float(producto_exacto.precio),
                'costo_unitario_base': costo_unitario_base,
                'costo_unitario_base_2': costo_unitario_base_2,
                'costo_unitario_base_3': costo_unitario_base_3,
                'descuento_porcentaje': float(producto_exacto.descuento_porcentaje) if producto_exacto.descuento_porcentaje else 0.0,
                'ahorro_combo': producto_exacto.calcular_ahorro_combo(),
                'precio_normal': producto_exacto.calcular_precio_normal(),
                'tiene_ofertas': producto_exacto.tiene_ofertas_volumen(),
                'precio2': float(producto_exacto.precio2) if producto_exacto.precio2 else None,
                'precio3': float(producto_exacto.precio3) if producto_exacto.precio3 else None,
                'precio4': float(producto_exacto.precio4) if producto_exacto.precio4 else None,
                'precio5': float(producto_exacto.precio5) if producto_exacto.precio5 else None,
                'margen2': float(producto_exacto.margen2) if producto_exacto.margen2 else None,
                'margen3': float(producto_exacto.margen3) if producto_exacto.margen3 else None,
                'margen4': float(producto_exacto.margen4) if producto_exacto.margen4 else None,
                'margen5': float(producto_exacto.margen5) if producto_exacto.margen5 else None,
            }
            # ✅ ESTA ERA LA LÍNEA QUE FALTABA
            return jsonify([resultado])
        
        # Buscar por codigo_barras exacto
        producto_barras = Producto.query.filter(
            and_(Producto.activo == True, Producto.codigo_barras == termino)
        ).first()
        if not producto_barras:
            producto_barras = Producto.query.filter(
                and_(Producto.activo == True, Producto.codigo_barras == termino.upper())
            ).first()
        if producto_barras:
            res_barras = {
                'id': producto_barras.id, 'codigo': producto_barras.codigo,
                'nombre': producto_barras.nombre, 'precio': float(producto_barras.precio),
                'precio_base': float(producto_barras.precio),
                'costo': float(producto_barras.costo) if producto_barras.costo else 0.0,
                'margen': float(producto_barras.margen) if producto_barras.margen else 0.0,
                'stock': producto_barras.stock_dinamico, 'iva': float(producto_barras.iva),
                'match_tipo': 'codigo_exacto',
                'descripcion': producto_barras.descripcion or '',
                'es_combo': producto_barras.es_combo,
                'producto_base_id': producto_barras.producto_base_id,
                'cantidad_combo': float(producto_barras.cantidad_combo) if producto_barras.cantidad_combo else 1.0,
                'cantidad_combo_2': float(producto_barras.cantidad_combo_2) if producto_barras.cantidad_combo_2 else 0.0,
                'cantidad_combo_3': float(producto_barras.cantidad_combo_3) if producto_barras.cantidad_combo_3 else 0.0,
                'precio_unitario_base': float(producto_barras.precio_unitario_base) if producto_barras.precio_unitario_base else float(producto_barras.precio),
                'costo_unitario_base': 0.0, 'costo_unitario_base_2': 0.0, 'costo_unitario_base_3': 0.0,
                'descuento_porcentaje': float(producto_barras.descuento_porcentaje) if producto_barras.descuento_porcentaje else 0.0,
                'ahorro_combo': producto_barras.calcular_ahorro_combo(),
                'precio_normal': producto_barras.calcular_precio_normal(),
                'tiene_ofertas': producto_barras.tiene_ofertas_volumen(),
                'precio2': float(producto_barras.precio2) if producto_barras.precio2 else None,
                'precio3': float(producto_barras.precio3) if producto_barras.precio3 else None,
                'precio4': float(producto_barras.precio4) if producto_barras.precio4 else None,
                'precio5': float(producto_barras.precio5) if producto_barras.precio5 else None,
                'margen2': float(producto_barras.margen2) if producto_barras.margen2 else None,
                'margen3': float(producto_barras.margen3) if producto_barras.margen3 else None,
                'margen4': float(producto_barras.margen4) if producto_barras.margen4 else None,
                'margen5': float(producto_barras.margen5) if producto_barras.margen5 else None,
            }
            return jsonify([res_barras])
        
        # Búsqueda parcial en código, nombre y codigo_barras
        termino_busqueda = f"%{termino.lower()}%"
        
        productos = Producto.query.filter(
            and_(
                Producto.activo == True,
                or_(
                    Producto.codigo.ilike(termino_busqueda),
                    Producto.nombre.ilike(termino_busqueda),
                    Producto.descripcion.ilike(termino_busqueda),
                    Producto.codigo_barras.ilike(termino_busqueda)
                )
            )
        ).limit(15).all()
        
        resultados = []
        for producto in productos:
            # Determinar tipo de coincidencia para ordenar resultados
            match_tipo = 'nombre'
            if termino.lower() in producto.codigo.lower():
                match_tipo = 'codigo'
            elif termino.lower() in producto.nombre.lower()[:20]:
                match_tipo = 'nombre_inicio'
            
            # Obtener costos de productos base si es combo
            costo_unitario_base = 0.0
            costo_unitario_base_2 = 0.0
            costo_unitario_base_3 = 0.0
            cantidad_combo_2 = float(producto.cantidad_combo_2) if producto.cantidad_combo_2 else 0.0
            cantidad_combo_3 = float(producto.cantidad_combo_3) if producto.cantidad_combo_3 else 0.0
            
            if producto.es_combo:
                if producto.producto_base_id:
                    pb = Producto.query.get(producto.producto_base_id)
                    if pb and pb.costo:
                        costo_unitario_base = float(pb.costo)
                if producto.producto_base_2_id:
                    pb2 = Producto.query.get(producto.producto_base_2_id)
                    if pb2 and pb2.costo:
                        costo_unitario_base_2 = float(pb2.costo)
                if producto.producto_base_3_id:
                    pb3 = Producto.query.get(producto.producto_base_3_id)
                    if pb3 and pb3.costo:
                        costo_unitario_base_3 = float(pb3.costo)
            
            resultado = {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'precio_base': float(producto.precio),
                'costo': float(producto.costo) if producto.costo else 0.0,
                'margen': float(producto.margen) if producto.margen else 0.0,
                'stock': producto.stock_dinamico,
                'iva': float(producto.iva),
                'match_tipo': match_tipo,
                'descripcion': producto.descripcion or '',
                'es_combo': producto.es_combo,
                'producto_base_id': producto.producto_base_id,
                'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
                'cantidad_combo_2': cantidad_combo_2,
                'cantidad_combo_3': cantidad_combo_3,
                'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
                'costo_unitario_base': costo_unitario_base,
                'costo_unitario_base_2': costo_unitario_base_2,
                'costo_unitario_base_3': costo_unitario_base_3,
                'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
                'ahorro_combo': producto.calcular_ahorro_combo(),
                'precio_normal': producto.calcular_precio_normal(),
                'tiene_ofertas': producto.tiene_ofertas_volumen(),
                'precio2': float(producto.precio2) if producto.precio2 else None,
                'precio3': float(producto.precio3) if producto.precio3 else None,
                'precio4': float(producto.precio4) if producto.precio4 else None,
                'precio5': float(producto.precio5) if producto.precio5 else None,
                'margen2': float(producto.margen2) if producto.margen2 else None,
                'margen3': float(producto.margen3) if producto.margen3 else None,
                'margen4': float(producto.margen4) if producto.margen4 else None,
                'margen5': float(producto.margen5) if producto.margen5 else None,
            }
            resultados.append(resultado)
        
        def orden_relevancia(item):
            if item['match_tipo'] == 'codigo_exacto':
                return 0
            elif item['match_tipo'] == 'codigo':
                return 1
            elif item['match_tipo'] == 'nombre_inicio':
                return 2
            else:
                return 3
        
        resultados.sort(key=orden_relevancia)
        return jsonify(resultados)
        
    except Exception as e:
        print(f"❌ Error en buscar_productos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify([])

        

@app.route('/api/producto_por_id/<int:producto_id>')
def get_producto_por_id(producto_id):
    """Obtiene un producto por ID - INCLUYE COSTO"""
    producto = Producto.query.filter_by(id=producto_id, activo=True).first()
    if producto:
        # Obtener costos de productos base si es combo
        costo_unitario_base = 0.0
        costo_unitario_base_2 = 0.0
        costo_unitario_base_3 = 0.0
        cantidad_combo_2 = float(producto.cantidad_combo_2) if producto.cantidad_combo_2 else 0.0
        cantidad_combo_3 = float(producto.cantidad_combo_3) if producto.cantidad_combo_3 else 0.0
        
        if producto.es_combo:
            if producto.producto_base_id:
                pb = Producto.query.get(producto.producto_base_id)
                if pb and pb.costo:
                    costo_unitario_base = float(pb.costo)
            if producto.producto_base_2_id:
                pb2 = Producto.query.get(producto.producto_base_2_id)
                if pb2 and pb2.costo:
                    costo_unitario_base_2 = float(pb2.costo)
            if producto.producto_base_3_id:
                pb3 = Producto.query.get(producto.producto_base_3_id)
                if pb3 and pb3.costo:
                    costo_unitario_base_3 = float(pb3.costo)
        
        return jsonify({
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'precio': float(producto.precio),
            'costo': float(producto.costo) if producto.costo else 0.0,
            'margen': float(producto.margen) if producto.margen else 0.0,
            'stock': producto.stock_dinamico,
            'iva': float(producto.iva),
            'descripcion': producto.descripcion or '',
            'es_combo': producto.es_combo,
            'producto_base_id': producto.producto_base_id,
            'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
            'cantidad_combo_2': cantidad_combo_2,
            'cantidad_combo_3': cantidad_combo_3,
            'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
            'costo_unitario_base': costo_unitario_base,
            'costo_unitario_base_2': costo_unitario_base_2,
            'costo_unitario_base_3': costo_unitario_base_3,
            'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto.calcular_ahorro_combo(),
            'precio_normal': producto.calcular_precio_normal(),
            'precio2': float(producto.precio2) if producto.precio2 else None,
            'precio3': float(producto.precio3) if producto.precio3 else None,
            'precio4': float(producto.precio4) if producto.precio4 else None,
            'precio5': float(producto.precio5) if producto.precio5 else None,
        })
    return jsonify({'error': 'Producto no encontrado'}), 404

@app.route('/api/producto/<codigo>')
def get_producto(codigo):
    """Obtiene un producto por código exacto o codigo_barras - INCLUYE COSTO"""
    producto = Producto.query.filter_by(codigo=codigo.upper(), activo=True).first()
    if not producto:
        # Buscar por codigo_barras si no encuentra por codigo
        producto = Producto.query.filter_by(codigo_barras=codigo, activo=True).first()
    if not producto:
        producto = Producto.query.filter_by(codigo_barras=codigo.upper(), activo=True).first()
    if producto:
        # Obtener costos de productos base si es combo
        costo_unitario_base = 0.0
        costo_unitario_base_2 = 0.0
        costo_unitario_base_3 = 0.0
        cantidad_combo_2 = float(producto.cantidad_combo_2) if producto.cantidad_combo_2 else 0.0
        cantidad_combo_3 = float(producto.cantidad_combo_3) if producto.cantidad_combo_3 else 0.0
        
        if producto.es_combo:
            if producto.producto_base_id:
                pb = Producto.query.get(producto.producto_base_id)
                if pb and pb.costo:
                    costo_unitario_base = float(pb.costo)
            if producto.producto_base_2_id:
                pb2 = Producto.query.get(producto.producto_base_2_id)
                if pb2 and pb2.costo:
                    costo_unitario_base_2 = float(pb2.costo)
            if producto.producto_base_3_id:
                pb3 = Producto.query.get(producto.producto_base_3_id)
                if pb3 and pb3.costo:
                    costo_unitario_base_3 = float(pb3.costo)
        
        return jsonify({
            'id': producto.id,
            'codigo': producto.codigo,
            'nombre': producto.nombre,
            'precio': float(producto.precio),
            'costo': float(producto.costo) if producto.costo else 0.0,
            'margen': float(producto.margen) if producto.margen else 0.0,
            'stock': producto.stock_dinamico,
            'iva': float(producto.iva),
            'descripcion': producto.descripcion or '',
            'es_combo': producto.es_combo,
            'producto_base_id': producto.producto_base_id,
            'cantidad_combo': float(producto.cantidad_combo) if producto.cantidad_combo else 1.0,
            'cantidad_combo_2': cantidad_combo_2,
            'cantidad_combo_3': cantidad_combo_3,
            'precio_unitario_base': float(producto.precio_unitario_base) if producto.precio_unitario_base else float(producto.precio),
            'costo_unitario_base': costo_unitario_base,
            'costo_unitario_base_2': costo_unitario_base_2,
            'costo_unitario_base_3': costo_unitario_base_3,
            'descuento_porcentaje': float(producto.descuento_porcentaje) if producto.descuento_porcentaje else 0.0,
            'ahorro_combo': producto.calcular_ahorro_combo(),
            'precio_normal': producto.calcular_precio_normal(),
            'precio2': float(producto.precio2) if producto.precio2 else None,
            'precio3': float(producto.precio3) if producto.precio3 else None,
            'precio4': float(producto.precio4) if producto.precio4 else None,
            'precio5': float(producto.precio5) if producto.precio5 else None,
        })
    return jsonify({'error': 'Producto no encontrado'}), 404


# 5. FUNCIÓN AUXILIAR PARA MIGRAR PRODUCTOS EXISTENTES
def migrar_productos_sin_costo_margen():
    """Función para migrar productos existentes que no tienen costo ni margen"""
    try:
        productos_sin_costo = Producto.query.filter(
            or_(
                Producto.costo.is_(None),
                Producto.margen.is_(None),
                Producto.costo == 0
            )
        ).all()
        
        contador_migrados = 0
        
        for producto in productos_sin_costo:
            # Si no tiene costo, calcular desde precio con margen del 30%
            if not producto.costo or producto.costo == 0:
                # Asumiendo un margen del 30%, costo = precio / 1.30
                precio_actual = float(producto.precio)
                costo_calculado = precio_actual / 1.30
                margen_calculado = 30.0
                
                producto.costo = Decimal(str(round(costo_calculado, 2)))
                producto.margen = Decimal(str(margen_calculado))
                
                contador_migrados += 1
                print(f"📦 Migrado: {producto.codigo} - Precio=${precio_actual:.2f} → Costo=${costo_calculado:.2f}, Margen=30%")
        
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} productos actualizados")
        else:
            print("✅ No hay productos que migrar")
            
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en migración: {e}")


# REGISTRAR DESCUENTOEN PROCESAR_VENTA 
def registrar_descuento_factura(factura_id, porcentaje, monto, total_original, usuario_id):
    """Registrar descuento en tabla separada - se llama DESPUÉS de crear factura"""
    try:
        if porcentaje > 0 and monto > 0:
            descuento = DescuentoFactura(
                factura_id=factura_id,
                porcentaje_descuento=Decimal(str(porcentaje)),
                monto_descuento=Decimal(str(monto)),
                total_original=Decimal(str(total_original)),
                usuario_id=usuario_id
            )
            db.session.add(descuento)
            db.session.commit()
            print(f"Descuento registrado: {porcentaje}% = ${monto} para factura {factura_id}")
            return True
    except Exception as e:
        print(f"Error registrando descuento: {e}")
        return False


def actualizar_stock_combo(combo, cantidad_vendida, factura_id=None):
    """Actualizar stock de productos base al vender combo"""
    try:
        print(f"Actualizando stock para combo {combo.codigo} - cantidad vendida: {cantidad_vendida}")
        
        # Producto base 1 (obligatorio)
        if combo.producto_base_id and combo.cantidad_combo and float(combo.cantidad_combo) > 0:
            producto_base = Producto.query.get(combo.producto_base_id)
            if producto_base:
                descuento = float(combo.cantidad_combo) * cantidad_vendida
                stock_anterior = float(producto_base.stock)
                producto_base.stock -= Decimal(str(descuento))
                print(f"  Base 1 - {producto_base.codigo}: {stock_anterior} - {descuento} = {float(producto_base.stock)}")
                
                # Auditoría de stock
                registrar_movimiento_stock(
                    db=db,
                    producto_id=producto_base.id,
                    tipo='combo',
                    cantidad=descuento,
                    signo='-',
                    stock_anterior=stock_anterior,
                    stock_nuevo=float(producto_base.stock),
                    referencia_tipo='factura',
                    referencia_id=factura_id,
                    motivo=f'Combo {combo.codigo} x {cantidad_vendida}',
                    usuario_id=session.get('user_id'),
                    usuario_nombre=session.get('nombre', 'Sistema'),
                    codigo_producto=producto_base.codigo,
                    nombre_producto=producto_base.nombre
                )
        
        # Producto base 2 (opcional)
        if combo.producto_base_2_id and combo.cantidad_combo_2 and float(combo.cantidad_combo_2) > 0:
            producto_base_2 = Producto.query.get(combo.producto_base_2_id)
            if producto_base_2:
                descuento = float(combo.cantidad_combo_2) * cantidad_vendida
                stock_anterior = float(producto_base_2.stock)
                producto_base_2.stock -= Decimal(str(descuento))
                print(f"  Base 2 - {producto_base_2.codigo}: {stock_anterior} - {descuento} = {float(producto_base_2.stock)}")
                
                # Auditoría de stock
                registrar_movimiento_stock(
                    db=db,
                    producto_id=producto_base_2.id,
                    tipo='combo',
                    cantidad=descuento,
                    signo='-',
                    stock_anterior=stock_anterior,
                    stock_nuevo=float(producto_base_2.stock),
                    referencia_tipo='factura',
                    referencia_id=factura_id,
                    motivo=f'Combo {combo.codigo} x {cantidad_vendida}',
                    usuario_id=session.get('user_id'),
                    usuario_nombre=session.get('nombre', 'Sistema'),
                    codigo_producto=producto_base_2.codigo,
                    nombre_producto=producto_base_2.nombre
                )
        
        # Producto base 3 (opcional)
        if combo.producto_base_3_id and combo.cantidad_combo_3 and float(combo.cantidad_combo_3) > 0:
            producto_base_3 = Producto.query.get(combo.producto_base_3_id)
            if producto_base_3:
                descuento = float(combo.cantidad_combo_3) * cantidad_vendida
                stock_anterior = float(producto_base_3.stock)
                producto_base_3.stock -= Decimal(str(descuento))
                print(f"  Base 3 - {producto_base_3.codigo}: {stock_anterior} - {descuento} = {float(producto_base_3.stock)}")
                
                # Auditoría de stock
                registrar_movimiento_stock(
                    db=db,
                    producto_id=producto_base_3.id,
                    tipo='combo',
                    cantidad=descuento,
                    signo='-',
                    stock_anterior=stock_anterior,
                    stock_nuevo=float(producto_base_3.stock),
                    referencia_tipo='factura',
                    referencia_id=factura_id,
                    motivo=f'Combo {combo.codigo} x {cantidad_vendida}',
                    usuario_id=session.get('user_id'),
                    usuario_nombre=session.get('nombre', 'Sistema'),
                    codigo_producto=producto_base_3.codigo,
                    nombre_producto=producto_base_3.nombre
                )
        
        return True
        
    except Exception as e:
        print(f"Error actualizando stock de combo {combo.codigo}: {e}")
        return False

# FUNCIÓN PROCESAR_VENTA

@app.route('/procesar_venta', methods=['POST'])
def procesar_venta():
    """Procesar venta con medios de pago y items detallados para AFIP + CTA.CTE"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    # ── Verificar caja abierta ────────────────────────────────────────────────
    caja_abierta = db.session.execute(text("""
        SELECT id FROM cajas WHERE estado = 'abierta' ORDER BY fecha_apertura DESC LIMIT 1
    """)).fetchone()
    if not caja_abierta:
        return jsonify({'success': False, 'error': '❌ No hay una caja abierta. Abrí la caja antes de operar.'}), 400
    # ─────────────────────────────────────────────────────────────────────────
    
    try:
        data = request.json
        
        # Validar datos básicos
        cliente_id = data.get('cliente_id')
        tipo_comprobante = data.get('tipo_comprobante')
        items = data.get('items', [])
        items_detalle = data.get('items_detalle', [])
        medios_pago = data.get('medios_pago', [])
        imprimir_automatico = data.get('imprimir_automatico', True)
        
        # ═══ NUEVO: DATOS PARA CTA.CTE ═══
        productos_cta_cte_ids = data.get('productos_cta_cte_ids', [])  # IDs de cta_cte_detalle
        es_venta_fiada = any(mp.get('medio_pago') == 'CTA.CTE' for mp in medios_pago)
        
        # ═══ CONSIGNACIÓN: factura derivada de un Comprobante Interno ═══
        interno_origen_id = data.get('interno_origen_id')
        interno_origen = None
        if interno_origen_id:
            interno_origen = Factura.query.get(interno_origen_id)
            # Validaciones de seguridad
            if not interno_origen:
                return jsonify({'success': False, 'error': f'Comprobante Interno #{interno_origen_id} no encontrado'}), 400
            if interno_origen.tipo_comprobante != '99':
                return jsonify({'success': False, 'error': 'El comprobante origen no es un Interno'}), 400
            if interno_origen.estado != 'interno':
                return jsonify({'success': False, 'error': f'El interno origen está en estado {interno_origen.estado}'}), 400
            if str(tipo_comprobante) == '99':
                return jsonify({'success': False, 'error': 'No podés tomar un interno para emitir otro interno'}), 400
            # Verificar que no esté ya facturado (chequeo extra a nivel app)
            ya_facturado = Factura.query.filter(
                Factura.interno_origen_id == interno_origen_id,
                Factura.estado != 'anulada'
            ).first()
            if ya_facturado:
                return jsonify({
                    'success': False,
                    'error': f'El interno {interno_origen.numero} ya fue facturado en {ya_facturado.numero}'
                }), 400
            print(f"🔗 Consignación: factura origen es interno {interno_origen.numero} (id={interno_origen_id})")
        
        if not items:
            return jsonify({'success': False, 'error': 'No hay productos en la venta'})
        
        if not medios_pago:
            return jsonify({'success': False, 'error': 'No se especificaron medios de pago'})
        
        if not items_detalle:
            return jsonify({'success': False, 'error': 'No se recibieron items detallados para AFIP'})
        
        # ═══════════════════════════════════════════════════════════════════════════
        # CTA.CTE: siempre genera comprobante (interno o factura).
        # El monto CTA.CTE se registra como pendiente DESPUÉS del commit.
        # ═══════════════════════════════════════════════════════════════════════════
        monto_cta_cte = sum(float(mp.get('importe', 0)) for mp in medios_pago
                            if mp.get('medio_pago') == 'CTA.CTE')
        print(f"🔍 DEBUG CTA.CTE: medios_pago={medios_pago}")
        print(f"🔍 DEBUG CTA.CTE: monto_cta_cte={monto_cta_cte}")

        # ═══════════════════════════════════════════════════════════════════════════
        # CASO 2 Y 3: FACTURACIÓN NORMAL (con o sin productos de CTA.CTE)
        # ═══════════════════════════════════════════════════════════════════════════
        
        # Validar si hay productos de CTA.CTE que NO se puede pagar con CTA.CTE nuevamente
        if len(productos_cta_cte_ids) > 0:
            print("💰 Procesando PAGO de productos que estaban en CTA.CTE...")
            
            # Validar que NO haya medio de pago CTA.CTE
            if any(mp.get('medio_pago') == 'CTA.CTE' for mp in medios_pago):
                return jsonify({
                    'success': False,
                    'error': 'No puede pagar productos fiados con CTA.CTE. Use efectivo, tarjeta o transferencia.'
                }), 400
        
        # ═══ VALIDACIÓN DE MONTOS (tu código original) ═══
        total_medios = sum(float(mp.get('importe', 0)) for mp in medios_pago)
        total_venta = float(data.get('total', 0))
        descuento_monto = float(data.get('descuento_monto', 0))
        descuento_porcentaje = float(data.get('descuento_porcentaje', 0))

        print(f"🔍 DEBUG VALIDACIÓN BACKEND:")
        print(f"   Total medios de pago: ${total_medios:.2f}")
        print(f"   Total venta (con descuento): ${total_venta:.2f}")
        print(f"   Descuento aplicado: ${descuento_monto:.2f}")

        diferencia = total_medios - total_venta
        genera_saldo = data.get('genera_saldo', False)
        saldo_pendiente = float(data.get('saldo_pendiente', 0))

        if diferencia < -0.01:
            faltante = abs(diferencia)
            # Permitir pago parcial si el frontend confirmó genera_saldo
            if genera_saldo and saldo_pendiente > 0 and cliente_id and int(cliente_id) > 1:
                print(f"💳 Pago parcial confirmado: pagó ${total_medios:.2f}, saldo pendiente ${saldo_pendiente:.2f}")
            else:
                print(f"❌ ERROR: Faltan ${faltante:.2f}")
                return jsonify({
                    'success': False, 
                    'error': f'Faltan ${faltante:.2f} para completar el pago'
                })
        elif diferencia > 0.01:
            medios_efectivo = [mp for mp in medios_pago if mp.get('medio_pago') == 'efectivo']
            total_efectivo = sum(float(mp.get('importe', 0)) for mp in medios_efectivo)
            
            if total_efectivo >= total_venta:
                vuelto = diferencia
                print(f"✅ Pago con vuelto: ${vuelto:.2f}")
            else:
                print(f"❌ ERROR: Exceso de ${diferencia:.2f} sin suficiente efectivo")
                return jsonify({
                    'success': False,
                    'error': f'Exceso de ${diferencia:.2f} pero no hay suficiente efectivo para dar vuelto'
                })
        else:
            print(f"✅ Pago exacto")

        print(f"✅ Validación exitosa. Procediendo con la venta.")
        
        # ═══ NUMERACIÓN Y CREACIÓN DE FACTURA (tu código original) ═══
        tipo_comprobante_int = int(tipo_comprobante)
        es_comprobante_interno = (tipo_comprobante_int == 99)
        punto_venta = _pv_activo()   # PV del usuario logueado (con fallback al config)
        
        if es_comprobante_interno:
            # Numeración INDEPENDIENTE con prefijo X (ej: 0001-X0000001)
            ultima_interna = Factura.query.filter_by(
                tipo_comprobante='99',
                punto_venta=punto_venta
            ).order_by(Factura.id.desc()).first()
            
            numero_temporal = 1
            if ultima_interna and ultima_interna.numero:
                try:
                    # Formato esperado: PVPV-X0000001 → tomamos lo que está después de la X
                    parte_numero = ultima_interna.numero.split('-')[1]
                    if parte_numero.startswith('X'):
                        ultimo_numero_local = int(parte_numero[1:])
                    else:
                        ultimo_numero_local = int(parte_numero)
                    numero_temporal = ultimo_numero_local + 1
                except:
                    numero_temporal = 1
            
            numero_factura_temporal = f"{punto_venta:04d}-X{numero_temporal:07d}"
            
            factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
            while factura_existente:
                numero_temporal += 1
                numero_factura_temporal = f"{punto_venta:04d}-X{numero_temporal:07d}"
                factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
            
            print(f"📋 Comprobante Interno: número asignado {numero_factura_temporal}")
        else:
            # Numeración tradicional para facturas A/B/C
            ultima_factura_local = Factura.query.filter_by(
                tipo_comprobante=str(tipo_comprobante_int),
                punto_venta=punto_venta
            ).order_by(Factura.id.desc()).first()
            
            numero_temporal = 1
            if ultima_factura_local and ultima_factura_local.numero:
                try:
                    ultimo_numero_local = int(ultima_factura_local.numero.split('-')[1])
                    numero_temporal = ultimo_numero_local + 1
                except:
                    numero_temporal = 1
            
            numero_factura_temporal = f"{punto_venta:04d}-{numero_temporal:08d}"
            
            factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
            while factura_existente:
                numero_temporal += 1
                numero_factura_temporal = f"{punto_venta:04d}-{numero_temporal:08d}"
                factura_existente = Factura.query.filter_by(numero=numero_factura_temporal).first()
            
            print(f"📝 Número temporal asignado: {numero_factura_temporal}")
        
        total_final = float(data.get('total', total_venta))
        factura = Factura(
            numero=numero_factura_temporal,
            tipo_comprobante=str(tipo_comprobante_int),
            punto_venta=punto_venta,
            cliente_id=cliente_id,
            usuario_id=session['user_id'],
            subtotal=Decimal(str(data['subtotal'])),
            iva=Decimal(str(data['iva'])),
            total=Decimal(str(total_venta)),
            interno_origen_id=interno_origen_id if interno_origen else None
        )
        
        db.session.add(factura)
        db.session.flush()
        
        print(f"✅ Factura creada con ID: {factura.id} y número temporal: {factura.numero}")
        
        # ═══ AGREGAR DETALLES Y DESCONTAR STOCK ═══
        # *** CORREGIDO: Descuento por producto individual ***
        print(f"📦 Procesando {len(items)} productos...")
        if productos_cta_cte_ids:
            print(f"⚠️ {len(productos_cta_cte_ids)} productos vienen de CTA.CTE (ya descontados)")
        
        for i, item in enumerate(items):
            item_detalle = items_detalle[i] if i < len(items_detalle) else {}
            iva_porcentaje = float(item_detalle.get('iva_porcentaje', 21.0))
            
            subtotal = float(item['subtotal'])
            importe_iva = round((subtotal * iva_porcentaje / 100), 2)
            
            # ═══ FIX: congelar el costo del producto al momento de la venta.
            # Antes se guardaba 0 (el front no manda costo_unitario), así que
            # el reporte de margen usaba el costo ACTUAL para ventas viejas y
            # daba mal cuando el costo cambiaba después. Ahora se guarda el
            # Producto.costo vigente en esta venta. Se conserva el valor que
            # mande el front si viniera (>0); si no, se toma de la BD.
            _costo_item = item.get('costo_unitario', 0) or 0
            try:
                _costo_item = float(_costo_item)
            except (TypeError, ValueError):
                _costo_item = 0.0
            if _costo_item <= 0:
                _prod_costo = Producto.query.get(item['producto_id'])
                if _prod_costo and _prod_costo.costo:
                    _costo_item = float(_prod_costo.costo)
            
            detalle = DetalleFactura(
                factura_id=factura.id,
                producto_id=item['producto_id'],
                cantidad=item['cantidad'],
                precio_unitario=Decimal(str(item['precio_unitario'])),
                subtotal=Decimal(str(subtotal)),
                porcentaje_iva=Decimal(str(iva_porcentaje)),
                importe_iva=Decimal(str(importe_iva)),
                costo_unitario=Decimal(str(round(_costo_item, 2)))
            )
            db.session.add(detalle)
            
            # ═══ CONTROL DE DESCUENTO DE STOCK (POR PRODUCTO) ═══
            # Solo descontar si este producto NO viene de CTA.CTE NI de un Comprobante Interno consignado
            # FIX: Usar el flag es_cta_cte del item (antes comparaba producto_id contra IDs de cta_cte_detalle)
            es_producto_cta_cte = item.get('es_cta_cte', False)
            es_consignacion = bool(interno_origen)  # toda la venta proviene de un interno
            
            if es_consignacion:
                # Stock ya fue descontado al emitir el Comprobante Interno
                # Registramos un movimiento informativo para trazabilidad
                producto = Producto.query.get(item['producto_id'])
                if producto:
                    print(f"🔗 {producto.codigo}: stock NO se descuenta (viene del interno {interno_origen.numero})")
                    registrar_movimiento_stock(
                        db=db,
                        producto_id=producto.id,
                        tipo='facturacion_interno',
                        cantidad=item['cantidad'],
                        signo='=',
                        stock_anterior=float(producto.stock),
                        stock_nuevo=float(producto.stock),
                        referencia_tipo='factura',
                        referencia_id=factura.id,
                        motivo=f'Facturación de Interno {interno_origen.numero} → {factura.numero} (sin movimiento de stock)',
                        usuario_id=session.get('user_id'),
                        usuario_nombre=session.get('nombre', 'Sistema'),
                        codigo_producto=producto.codigo,
                        nombre_producto=producto.nombre
                    )
            elif not es_producto_cta_cte:
                producto = Producto.query.get(item['producto_id'])
                if producto:
                    if producto.es_combo:
                        print(f"📦 Combo {producto.codigo}: descontando de productos base...")
                        exito = actualizar_stock_combo(producto, item['cantidad'], factura.id)
                        if exito:
                            print(f"   ✅ Stock actualizado para combo {producto.codigo}")
                        else:
                            print(f"   ❌ Error actualizando stock de combo {producto.codigo}")
                    else:
                        stock_anterior = float(producto.stock)
                        producto.stock -= Decimal(str(item['cantidad']))
                        print(f"📦 {producto.codigo}: {stock_anterior} - {item['cantidad']} = {float(producto.stock)}")
                        
                        # Auditoría de stock
                        registrar_movimiento_stock(
                            db=db,
                            producto_id=producto.id,
                            tipo='venta',
                            cantidad=item['cantidad'],
                            signo='-',
                            stock_anterior=stock_anterior,
                            stock_nuevo=float(producto.stock),
                            referencia_tipo='factura',
                            referencia_id=factura.id,
                            usuario_id=session.get('user_id'),
                            usuario_nombre=session.get('nombre', 'Sistema'),
                            codigo_producto=producto.codigo,
                            nombre_producto=producto.nombre
                        )
            else:
                print(f"⏭️ Producto {item['producto_id']}: ya descontado en CTA.CTE, saltando...")
        
        # ═══ MEDIOS DE PAGO (tu código original) ═══
        print(f"💳 Agregando {len(medios_pago)} medios de pago...")
        for medio_data in medios_pago:
            medio_pago = MedioPago(
                factura_id=factura.id,
                medio_pago=medio_data['medio_pago'],
                importe=Decimal(str(medio_data['importe'])),
                fecha_registro=datetime.now()
            )
            db.session.add(medio_pago)
            print(f"💰 Medio agregado: {medio_data['medio_pago']} ${medio_data['importe']}")
        
        # ═══ AUTORIZACIÓN AFIP (tu código original) ═══
        if es_comprobante_interno:
            # Comprobante Interno: NO se autoriza con AFIP
            factura.estado = 'interno'
            factura.cae = None
            factura.vto_cae = None
            print(f"📋 Comprobante Interno {factura.numero}: NO se envía a AFIP, estado='interno'")
        else:
            try:
                print("📄 Autorizando en AFIP con items detallados...")
                cliente = Cliente.query.get(cliente_id)
                
                datos_comprobante = {
                    'tipo_comprobante': tipo_comprobante_int,
                    'punto_venta': punto_venta,
                    'importe_neto': float(factura.subtotal),
                    'importe_iva': float(factura.iva),
                    'items_detalle': items_detalle,
                    'doc_tipo': 99,
                    'doc_nro': 0
                }
                
                print("🧮 Items detallados enviados a AFIP:")
                for item in items_detalle:
                    print(f"   📦 ${item['subtotal']:.2f} (IVA {item['iva_porcentaje']}%)")
                
                if cliente and cliente.documento:
                    if cliente.tipo_documento == 'CUIT' and len(cliente.documento) == 11:
                        datos_comprobante['doc_tipo'] = 80
                        datos_comprobante['doc_nro'] = int(cliente.documento)
                    elif cliente.tipo_documento == 'DNI' and len(cliente.documento) >= 7:
                        datos_comprobante['doc_tipo'] = 96
                        datos_comprobante['doc_nro'] = int(cliente.documento)
                
                if cliente:
                    datos_comprobante['condicion_iva_str'] = cliente.condicion_iva
                resultado_afip = arca_client.autorizar_comprobante(datos_comprobante)
                
                if resultado_afip['success']:
                    numero_afip = resultado_afip['numero']
                    print(f"✅ AFIP asignó número: {numero_afip}")
                    
                    factura_afip_existente = Factura.query.filter(
                        and_(Factura.numero == numero_afip, Factura.id != factura.id)
                    ).first()
                    
                    if factura_afip_existente:
                        print(f"⚠️ Número AFIP {numero_afip} ya existe, manteniendo temporal")
                        factura.cae = resultado_afip['cae']
                        factura.vto_cae = resultado_afip['vto_cae']
                        factura.estado = 'autorizada'
                    else:
                        factura.numero = numero_afip
                        factura.cae = resultado_afip['cae']
                        factura.vto_cae = resultado_afip['vto_cae']
                        factura.estado = 'autorizada'
                    
                    print(f"✅ Autorización AFIP exitosa. CAE: {factura.cae}")
                    print(f"✅ Número final: {factura.numero}")
                else:
                    factura.estado = 'error_afip'
                    print(f"❌ Error AFIP: {resultado_afip.get('error', 'Error desconocido')}")
                    print(f"📝 Manteniendo número temporal: {factura.numero}")
                
            except Exception as e:
                factura.estado = 'error_afip'
                print(f"❌ Error completo al autorizar en AFIP: {e}")
                print(f"📝 Manteniendo número temporal: {factura.numero}")
        
        # ═══ COMMIT A BASE DE DATOS ═══
        db.session.commit()
        
        print(f"🎉 Venta procesada exitosamente: {factura.numero}")
        
        # ═══ NUEVO: MARCAR PRODUCTOS DE CTA.CTE COMO PAGADOS ═══
        if len(productos_cta_cte_ids) > 0:
            print(f"✅ Marcando {len(productos_cta_cte_ids)} productos de CTA.CTE como pagados...")
            resultado_marca = marcar_productos_como_pagados(
                db=db,
                detalle_ids=productos_cta_cte_ids,
                factura_id=factura.id
            )
            if resultado_marca['success']:
                print("✅ Productos de CTA.CTE marcados como pagados")
            else:
                print(f"⚠️ Error al marcar productos: {resultado_marca['mensaje']}")
        
        # ═══ REGISTRAR EN CTA.CTE (si el pago fue total o parcialmente a CTA.CTE) ═══
        if cliente_id and int(cliente_id) > 1 and len(productos_cta_cte_ids) == 0:
            try:
                saldo_a_registrar = monto_cta_cte if monto_cta_cte > 0 else (round(saldo_pendiente, 2) if genera_saldo and saldo_pendiente > 0 else 0)
                if saldo_a_registrar > 0:
                    registrar_comprobante_en_cta_cte(
                        db=db,
                        factura_id=factura.id,
                        numero_comprobante=factura.numero,
                        monto_total=float(data.get('total', 0)),
                        saldo_pendiente=round(saldo_a_registrar, 2),
                        cliente_id=cliente_id,
                        usuario_id=session['user_id'],
                        observaciones=f'Comprobante {factura.numero}'
                    )
                    # ═══ FIX: el ticket debe mostrar el SALDO REAL acumulado del
                    # cliente en CTA.CTE (no el monto de esta sola factura).
                    # El movimiento de esta venta ya fue insertado por
                    # registrar_comprobante_en_cta_cte() arriba, así que esta
                    # consulta ya lo incluye. Misma fórmula verificada que usa
                    # saldo_real_cliente en cta_cte.py y el panel de Cta.Cte.
                    from sqlalchemy import text as _sql_text
                    _saldo_real_row = db.session.execute(_sql_text("""
                        SELECT (
                          COALESCE((
                            SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                            WHERE m.cliente_id = :cid AND m.tipo = 'venta_fiada'
                          ), 0)
                          -
                          COALESCE((
                            SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                            WHERE m.cliente_id = :cid AND m.tipo = 'nota_credito'
                          ), 0)
                          -
                          (
                            COALESCE((
                              SELECT SUM(m.saldo_pendiente) FROM cta_cte_movimiento m
                              WHERE m.cliente_id = :cid AND m.tipo_mov = 'pago'
                                AND m.numero_comprobante = 'A_CUENTA'
                            ), 0)
                            -
                            COALESCE((
                              SELECT SUM(m.monto_total) FROM cta_cte_movimiento m
                              WHERE m.cliente_id = :cid AND m.tipo_mov = 'pago'
                                AND m.numero_comprobante = 'SALDO_FAVOR_USADO'
                            ), 0)
                          )
                        ) AS saldo
                    """), {'cid': cliente_id}).scalar()
                    saldo_real_cliente_ctacte = round(float(_saldo_real_row or 0), 2)

                    if saldo_real_cliente_ctacte > 0:
                        factura.observaciones = f"Saldo pendiente: ${saldo_real_cliente_ctacte:,.2f}"
                    elif saldo_real_cliente_ctacte < 0:
                        factura.observaciones = f"Saldo a favor: ${abs(saldo_real_cliente_ctacte):,.2f}"
                    else:
                        factura.observaciones = "Saldo pendiente: $0.00"
                    db.session.commit()
                    print(f"💳 Movimiento CTA.CTE registrado — esta factura: ${saldo_a_registrar:.2f} | saldo real cliente: ${saldo_real_cliente_ctacte:.2f}")
            except Exception as e:
                print(f"⚠️ Error registrando en CTA.CTE: {e}")
                import traceback; traceback.print_exc()

        # ═══ GUARDAR DETALLE CTA.CTE EN OBSERVACIONES PARA EL TICKET ═══
        if len(productos_cta_cte_ids) > 0:
            try:
                # Obtener descripción de los productos cobrados de CTA.CTE
                from sqlalchemy import text as sql_text
                result = db.session.execute(sql_text("""
                    SELECT d.descripcion, d.cantidad, d.subtotal, m.fecha
                    FROM cta_cte_detalle d
                    JOIN cta_cte_movimiento m ON m.id = d.movimiento_id
                    WHERE d.id IN :ids
                """), {'ids': tuple(productos_cta_cte_ids)})
                rows = result.fetchall()
                if rows:
                    detalle_obs = "Cobro CTA.CTE:\n"
                    total_ctacte = 0
                    for r in rows:
                        detalle_obs += f"  {r.descripcion} x{float(r.cantidad):.2f} ${float(r.subtotal):,.2f}\n"
                        total_ctacte += float(r.subtotal)
                    detalle_obs += f"Total CTA.CTE: ${total_ctacte:,.2f}"
                    factura.observaciones = detalle_obs
                    db.session.commit()
            except Exception as e:
                print(f"⚠️ Error guardando detalle CTA.CTE en observaciones: {e}")

        # ═══ REGISTRAR DESCUENTO (tu código original) ═══
        if data.get('descuento_monto', 0) > 0:
            total_antes_descuento = float(data.get('subtotal', 0)) + float(data.get('iva', 0))
            registrar_descuento_factura(
                factura.id, 
                data.get('descuento_porcentaje', 0),
                data.get('descuento_monto', 0),
                total_antes_descuento,
                session['user_id']
            )

        # ═══ IMPRESIÓN AUTOMÁTICA (tu código original) ═══
        if imprimir_automatico and IMPRESION_DISPONIBLE:
            try:
                print("🖨️ Imprimiendo factura automáticamente...")
                impresora_termica.imprimir_factura(factura)
            except Exception as e:
                print(f"⚠️ Error en impresión automática: {e}")
        
        return jsonify({
            'success': True,
            'factura_id': factura.id,
            'numero': factura.numero,
            'cae': factura.cae,
            'estado': factura.estado,
            'mensaje': f"Factura {factura.numero} generada correctamente",
            'genera_saldo': genera_saldo,
            'saldo_pendiente': round(saldo_pendiente, 2) if genera_saldo else 0,
            'monto_pagado': round(total_medios, 2),
            'monto_cta_cte': round(monto_cta_cte, 2),
        })
        
    except Exception as e:
        print(f"❌ Error en procesar_venta: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': f'Error al procesar la venta: {str(e)}'}), 500
        

# RUTAS DE IMPRESIÓN
@app.route('/imprimir_factura/<int:factura_id>')
def imprimir_factura(factura_id):
    """Endpoint para imprimir factura específica"""
    try:
        print(f"🖨️ Iniciando impresión de factura ID: {factura_id}")
        
        # Obtener la factura de la base de datos
        factura = Factura.query.get_or_404(factura_id)
        
        # Usar directamente la instancia de la clase
        resultado = impresora_termica.imprimir_factura(factura)
        
        if resultado:
            return jsonify({
                'success': True,
                'mensaje': f'Factura impresa correctamente en {impresora_termica.nombre_impresora}'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Error al imprimir factura'
            })
        
    except Exception as e:
        print(f"❌ Error en endpoint imprimir: {e}")
        return jsonify({
            'success': False,
            'error': f'Error al imprimir factura: {str(e)}'
        }), 500
        
@app.route('/test_impresion')
def test_impresion():
    try:
        print("🧪 Iniciando test de impresión...")
        resultado = impresora_termica.test_impresion()
        
        if resultado:
            return jsonify({
                'success': True,
                'mensaje': f'Test enviado correctamente a {impresora_termica.nombre_impresora}'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Error al enviar test de impresión'
            })
    except Exception as e:
        print(f"❌ ERROR EN TEST: {e}")
        import traceback
        traceback.print_exc()  # ← AGREGAR ESTA LÍNEA
        return jsonify({
            'success': False,
            'error': f'Error en test: {str(e)}'
        }), 500

@app.route('/estado_impresora')
def estado_impresora():
    """Endpoint para verificar estado de impresora"""
    try:
        estado = impresora_termica.verificar_estado()
        return jsonify(estado)
    except Exception as e:
        return jsonify({
            'disponible': False,
            'error': f'Error verificando estado: {str(e)}'
        }), 500

@app.route('/facturas')
def facturas():
    """Página de gestión de facturas (carga dinámica via JavaScript)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # No cargar facturas aquí - se cargan dinámicamente via JavaScript
    # Esto mejora significativamente el tiempo de carga de la página
    
    return render_template('facturas.html', facturas=[])

@app.route('/factura/<int:factura_id>')
def ver_factura(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        return render_template('factura_detalle.html', factura=factura)
        
    except Exception as e:
        app.logger.error(f"Error al cargar factura {factura_id}: {str(e)}")
        flash('Error al cargar la factura', 'error')
        return redirect(url_for('facturas'))

# Funciones de utilidad para limpieza de datos
def limpiar_facturas_duplicadas():
    """Limpiar facturas duplicadas o problemáticas"""
    try:
        # Buscar facturas con números duplicados
        facturas_duplicadas = db.session.query(Factura.numero).group_by(Factura.numero).having(db.func.count(Factura.numero) > 1).all()
        
        if facturas_duplicadas:
            print(f"⚠️ Encontradas {len(facturas_duplicadas)} facturas con números duplicados")
            
            for numero_duplicado in facturas_duplicadas:
                numero = numero_duplicado[0]
                facturas = Factura.query.filter_by(numero=numero).order_by(Factura.id).all()
                
                # Mantener solo la primera, eliminar las demás
                for i, factura in enumerate(facturas):
                    if i > 0:  # Eliminar todas excepto la primera
                        print(f"🗑️ Eliminando factura duplicada: {factura.numero} (ID: {factura.id})")
                        
                        # Eliminar detalles primero
                        DetalleFactura.query.filter_by(factura_id=factura.id).delete()
                        
                        # Eliminar factura
                        db.session.delete(factura)
            
            db.session.commit()
            print("✅ Limpieza de duplicados completada")
        else:
            print("✅ No se encontraron facturas duplicadas")
            
    except Exception as e:
        print(f"❌ Error en limpieza: {e}")
        db.session.rollback()

def verificar_estado_facturas():
    """Verificar el estado actual de las facturas"""
    try:
        total_facturas = Factura.query.count()
        facturas_con_cae = Factura.query.filter(Factura.cae.isnot(None)).count()
        facturas_pendientes = Factura.query.filter_by(estado='pendiente').count()
        facturas_error = Factura.query.filter_by(estado='error_afip').count()
        
        print("\n📊 ESTADO ACTUAL DE FACTURAS:")
        print(f"   Total facturas: {total_facturas}")
        print(f"   Con CAE: {facturas_con_cae}")
        print(f"   Pendientes: {facturas_pendientes}")
        print(f"   Con errores: {facturas_error}")
        
        # Mostrar últimas facturas
        ultimas_facturas = Factura.query.order_by(Factura.id.desc()).limit(5).all()
        print(f"\n📋 ÚLTIMAS 5 FACTURAS:")
        for factura in ultimas_facturas:
            cae_status = f"CAE: {factura.cae[:10]}..." if factura.cae else "Sin CAE"
            print(f"   {factura.numero} - {factura.estado} - {cae_status}")
        
    except Exception as e:
        print(f"❌ Error verificando estado: {e}")

        
@app.route('/api/factura/<int:factura_id>/medios_pago')
def api_medios_pago(factura_id):
    try:
        # Query directa a medios_pago para evitar cache del ORM
        pagos_db = MedioPago.query.filter_by(factura_id=factura_id).order_by(MedioPago.id.asc()).all()

        pagos = [{
            'medio_pago': mp.medio_pago,
            'importe': float(mp.importe)
        } for mp in pagos_db]

        # Obtener observaciones de la factura
        factura = Factura.query.get(factura_id)
        observaciones = factura.observaciones if factura and hasattr(factura, 'observaciones') else None

        from flask import make_response
        resp = make_response(jsonify({'pagos': pagos, 'observaciones': observaciones}))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        return resp

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Funciones de inicialización
def create_tables():
    """Crea las tablas de la base de datos"""
    try:
        db.create_all()
        
        # Crear usuario admin por defecto si no existe
        if not Usuario.query.filter_by(username='admin').first():
            admin = Usuario(
                username='admin',
                password_hash='admin123',  # Sin encriptación para simplicidad
                nombre='Administrador',
                rol='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print("✅ Usuario admin creado (admin/admin123)")
        
        print("✅ Base de datos inicializada correctamente")
        
    except Exception as e:
        print(f"❌ Error al inicializar base de datos: {e}")

@app.route('/qr_afip/<int:factura_id>')
def generar_qr_afip(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        try:
            generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
            
            # Generar imagen QR
            qr_base64 = generador_qr.generar_qr_imagen(factura)
            info_qr = generador_qr.obtener_info_qr(factura)
            
            if qr_base64:
                return jsonify({
                    'success': True,
                    'qr_image': qr_base64,
                    'qr_url': info_qr['url'],
                    'qr_valido': info_qr['valido'],
                    'mensaje': info_qr['mensaje']
                })
            else:
                return jsonify({
                    'success': False,
                    'error': 'Error al generar código QR'
                }), 500
                
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'Módulo QR no disponible'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error: {str(e)}'
        }), 500

# Ruta para mostrar QR en página completa (para escanear con móvil)
@app.route('/mostrar_qr/<int:factura_id>')
def mostrar_qr_completo(factura_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Método compatible con SQLAlchemy antiguo
    factura = Factura.query.get_or_404(factura_id)
    
    try:
        generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
        info_qr = generador_qr.obtener_info_qr(factura)
        qr_base64 = generador_qr.generar_qr_imagen(factura, tamaño=8) if info_qr['valido'] else None
    except Exception as e:
        print(f"⚠️ Error generando QR: {e}")
        info_qr = {'valido': False, 'mensaje': f'Error al generar QR: {str(e)}'}
        qr_base64 = None
    
    return render_template('mostrar_qr.html', 
                         factura=factura, 
                         qr_info=info_qr,
                         qr_image=qr_base64)

# Ruta para validar datos QR
@app.route('/validar_qr/<int:factura_id>')
def validar_qr_factura(factura_id):
    try:
        # Método compatible con SQLAlchemy antiguo
        factura = Factura.query.get_or_404(factura_id)
        
        try:
            generador_qr = crear_generador_qr(ARCA_CONFIG)  # ← Ya está correcto
            info_qr = generador_qr.obtener_info_qr(factura)
            errores = generador_qr.validar_datos_qr(factura)
        except Exception as e:
            print(f"⚠️ Error con módulo QR: {e}")
            info_qr = {'valido': False, 'mensaje': f'Error con módulo QR: {str(e)}'}
            errores = [str(e)]
        
        return jsonify({
            'factura_id': factura_id,
            'numero': factura.numero,
            'qr_valido': info_qr['valido'],
            'qr_url': info_qr['url'] if info_qr['valido'] else None,
            'errores': errores,
            'datos_qr': info_qr.get('datos', {}),
            'mensaje': info_qr['mensaje']
        })
        
    except Exception as e:
        return jsonify({
            'error': f'Error validando QR: {str(e)}'
        }), 500



@app.route('/api/estado_afip_rapido')
def api_estado_afip_rapido():
    """API para verificación rápida de AFIP"""
    try:
        estado = afip_monitor.verificar_rapido()
        return jsonify({
            'success': True,
            'estado': estado
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/test_afip')
def test_afip():
    """Test manual de conexión AFIP con debug detallado"""
    if 'user_id' not in session:
        return jsonify({
            'success': False,
            'error': 'No autorizado'
        }), 401
    
    try:
        print("=" * 50)
        print("TEST AFIP DETALLADO")
        print("=" * 50)
        
        # Verificar archivos
        cert_existe = os.path.exists(ARCA_CONFIG.CERT_PATH)
        key_existe = os.path.exists(ARCA_CONFIG.KEY_PATH)
        
        print(f"Certificado existe: {cert_existe}")
        print(f"Clave privada existe: {key_existe}")
        print(f"Cert path: {ARCA_CONFIG.CERT_PATH}")
        print(f"Key path: {ARCA_CONFIG.KEY_PATH}")
        
        if not cert_existe or not key_existe:
            return jsonify({
                'success': False,
                'mensaje': "Archivos de certificado no encontrados"
            })
        
        # Intentar autenticación
        print("Intentando autenticación...")
        resultado_auth = arca_client.get_ticket_access()
        
        print(f"Resultado autenticación: {resultado_auth}")
        print(f"Tiene token: {bool(arca_client.token)}")
        
        if resultado_auth:
            return jsonify({
                'success': True,
                'mensaje': "Test AFIP exitoso",
                'tiene_token': bool(arca_client.token)
            })
        else:
            return jsonify({
                'success': False,
                'mensaje': "Autenticación falló",
                'tiene_token': False
            })
        
    except Exception as e:
        print(f"EXCEPCIÓN EN TEST: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'mensaje': f"Error: {str(e)}"
        }), 500


# Agregar esta función en app.py para debug de WSFEv1

# Versión simplificada sin verificación de sesión para debug
##### http://localhost:5000/debug_afip_simple LLAMAR ESTA RUTA PARA VERIFICAR ARCA

@app.route('/debug_afip_simple')
def debug_afip_simple():
    """Debug simple de AFIP sin verificación de sesión"""
    try:
        resultado = {
            'timestamp': datetime.now().isoformat(),
            'tests': {}
        }
        
        print("🔍 INICIANDO DEBUG SIMPLE DE AFIP...")
        
        # Test 1: Autenticación WSAA
        print("1️⃣ Test autenticación WSAA...")
        try:
            auth_result = arca_client.get_ticket_access()
            resultado['tests']['wsaa_auth'] = {
                'success': auth_result,
                'token_exists': bool(arca_client.token),
                'message': 'Autenticación exitosa' if auth_result else 'Fallo en autenticación'
            }
        except Exception as e:
            resultado['tests']['wsaa_auth'] = {
                'success': False,
                'error': str(e),
                'message': 'Error en autenticación'
            }
        
        # Test 2: Conectividad WSFEv1
        print("2️⃣ Test conectividad WSFEv1...")
        try:
            import socket
            from urllib.parse import urlparse
            
            wsfe_url = ARCA_CONFIG.WSFEv1_URL
            parsed = urlparse(wsfe_url)
            host = parsed.hostname
            port = 443
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(10)
            connect_result = sock.connect_ex((host, port))
            sock.close()
            
            resultado['tests']['wsfe_connectivity'] = {
                'success': connect_result == 0,
                'host': host,
                'port': port,
                'url': wsfe_url,
                'message': 'Conectividad OK' if connect_result == 0 else f'No se puede conectar (código: {connect_result})'
            }
        except Exception as e:
            resultado['tests']['wsfe_connectivity'] = {
                'success': False,
                'error': str(e),
                'message': 'Error verificando conectividad'
            }
        
        # Test 3: Solo si tenemos autenticación, probar FEDummy
        if resultado['tests']['wsaa_auth']['success']:
            print("3️⃣ Test FEDummy...")
            try:
                from zeep import Client
                from zeep.transports import Transport
                
                session_afip = crear_session_afip()
                transport = Transport(session=session_afip, timeout=30)
                client = Client(ARCA_CONFIG.WSFEv1_URL, transport=transport)
                
                dummy_response = client.service.FEDummy()
                
                resultado['tests']['fe_dummy'] = {
                    'success': True,
                    'response': str(dummy_response),
                    'message': 'FEDummy exitoso'
                }
            except Exception as e:
                resultado['tests']['fe_dummy'] = {
                    'success': False,
                    'error': str(e),
                    'message': 'Error en FEDummy'
                }
        
        # Generar resumen
        total_tests = len(resultado['tests'])
        successful_tests = sum(1 for test in resultado['tests'].values() if test['success'])
        
        resultado['summary'] = {
            'total_tests': total_tests,
            'successful_tests': successful_tests,
            'success_rate': f"{(successful_tests/total_tests*100):.1f}%" if total_tests > 0 else "0%",
            'overall_status': 'OK' if successful_tests == total_tests else 'PROBLEMAS'
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error general: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500



@app.route('/reporte_medios_hoy')
def reporte_medios_hoy():
    """Reporte rápido de medios de pago del día actual"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        fin_hoy = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        
        reporte = MedioPago.calcular_recaudacion_por_fecha(hoy, fin_hoy)
        
        return jsonify({
            'success': True,
            'reporte': reporte
        })
        
    except Exception as e:
        print(f"Error en reporte del día: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/medios_pago_factura/<int:factura_id>')
def medios_pago_factura(factura_id):
    """Obtener los medios de pago de una factura específica"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        medios = [{
            'medio_pago': mp.medio_pago,
            'importe': float(mp.importe),
            'fecha_registro': mp.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
        } for mp in factura.medios_pago]
        
        return jsonify({
            'success': True,
            'factura_numero': factura.numero,
            'total_factura': float(factura.total),
            'medios_pago': medios
        })
        
    except Exception as e:
        print(f"Error obteniendo medios de pago: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/test_stock_dinamico')
def test_stock_dinamico():
    """Ruta temporal para probar cálculo de stock dinámico"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener todos los combos
        combos = Producto.query.filter_by(es_combo=True, activo=True).all()
        
        resultados = []
        
        for combo in combos:
            resultado = {
                'codigo': combo.codigo,
                'nombre': combo.nombre,
                'stock_actual': combo.stock,
                'stock_dinamico': combo.stock_dinamico,
                'diferencia': combo.stock_dinamico - combo.stock,
                'debug': combo.debug_stock_combo()
            }
            resultados.append(resultado)
        
        return jsonify({
            'success': True,
            'combos_analizados': len(resultados),
            'resultados': resultados
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



# AGREGAR esta función a tu app.py (después de la línea que dice @app.route('/medios_pago_factura/<int:factura_id>')):

@app.route('/reporte_ventas')
def reporte_ventas():
    """Página principal del reporte de ventas"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('reporte_ventas.html')


@app.route('/api/reporte_ventas_productos')
def api_reporte_ventas_productos():
    """API para generar reporte de ventas por producto - CORREGIDO PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        categoria = request.args.get('categoria', '').strip()
        orden = request.args.get('orden', 'cantidad_desc')
        solo_con_ventas = request.args.get('solo_con_ventas', 'true').lower() == 'true'
        
        # Validar fechas
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fechas desde y hasta'
            })
        
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido'
            })
        
        # Validar rango de fechas (máximo 2 años)
        if (fecha_hasta_dt - fecha_desde_dt).days > 730:
            return jsonify({
                'success': False,
                'error': 'El rango de fechas no puede ser mayor a 2 años'
            })
        
        print(f"📊 Generando reporte de ventas (CORREGIDO PARA COMBOS):")
        print(f"   Período: {fecha_desde} a {fecha_hasta}")
        print(f"   Categoría: {categoria or 'Todas'}")
        print(f"   Orden: {orden}")
        print(f"   Solo con ventas: {solo_con_ventas}")
        
        # ✅ QUERY CORREGIDA: Incluir cantidad_combo en el cálculo
        query = db.session.query(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo,
            # *** CLAVE: Multiplicar cantidad por cantidad_combo si es combo ***
            func.sum(
                case(
                    (Producto.es_combo == True, DetalleFactura.cantidad * Producto.cantidad_combo),
                    else_=DetalleFactura.cantidad
                )
            ).label('cantidad_real_vendida'),
            func.sum(DetalleFactura.cantidad).label('unidades_combos_vendidas'),  # Para mostrar también combos vendidos
            func.sum(DetalleFactura.subtotal).label('total_vendido'),
            func.avg(DetalleFactura.precio_unitario).label('precio_promedio'),
            func.max(Factura.fecha).label('ultima_venta'),
            func.count(DetalleFactura.id).label('num_transacciones')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt,
                Factura.estado != 'anulada',  # Cuenta autorizadas + error_afip + internos NORMALES. Excluye anuladas (stock reintegrado)
                # ═══ CONSIGNACIÓN: excluir SOLO los internos a CONSIGNATARIO.
                # Un interno (tipo_comprobante='99') cuyo cliente es
                # intermediario NO es una venta real (la venta real es la
                # factura al cliente final); contarlo sería DOBLE conteo.
                # Los internos NORMALES (cliente NO intermediario) SÍ
                # cuentan, como siempre. Ver AYUDA_CIRCUITO_CONSIGNACION.md
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(
                          Factura.tipo_comprobante == '99',
                          Cliente.es_intermediario == True
                      )
                )
            )
        )
        
        print(f"✅ Query configurada con corrección para combos")
        
        # Filtrar por categoría si se especifica
        if categoria:
            query = query.filter(Producto.categoria == categoria)
            print(f"   Filtro categoría aplicado: {categoria}")
        
        # Agrupar por producto
        query = query.group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo
        )
        
        # *** APLICAR ORDENAMIENTO USANDO LA CANTIDAD REAL ***
        if orden == 'cantidad_desc':
            query = query.order_by(desc('cantidad_real_vendida'))
        elif orden == 'cantidad_asc':
            query = query.order_by(asc('cantidad_real_vendida'))
        elif orden == 'total_desc':
            query = query.order_by(desc('total_vendido'))
        elif orden == 'total_asc':
            query = query.order_by(asc('total_vendido'))
        elif orden == 'codigo':
            query = query.order_by(Producto.codigo)
        elif orden == 'nombre':
            query = query.order_by(Producto.nombre)
        
        print(f"   Ordenamiento aplicado: {orden} (usando cantidad real)")
        
        # Ejecutar query
        print(f"🔍 Ejecutando consulta...")
        resultados = query.all()
        print(f"📋 Encontrados {len(resultados)} productos con ventas")
        
        # Consulta adicional: información de estados de facturas para debug
        debug_estados = db.session.query(
            Factura.estado,
            func.count(Factura.id).label('cantidad'),
            func.sum(Factura.total).label('total')
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt
            )
        ).group_by(Factura.estado).all()
        
        estados_info = {}
        for estado, cantidad, total in debug_estados:
            estados_info[estado] = {
                'cantidad': cantidad,
                'total': float(total) if total else 0.0
            }
        
        print(f"📊 Estados de facturas en el período:")
        for estado, info in estados_info.items():
            print(f"   {estado}: {info['cantidad']} facturas (${info['total']:.2f})")
        
        # *** FORMATEAR RESULTADOS CON CANTIDAD REAL ***
        productos = []
        total_unidades_reales = 0
        total_ventas = 0.0
        
        for resultado in resultados:
            # *** USAR CANTIDAD REAL (ya calculada en SQL) ***
            cantidad_real = float(resultado.cantidad_real_vendida) if resultado.cantidad_real_vendida else 0.0
            unidades_combos = int(resultado.unidades_combos_vendidas) if resultado.unidades_combos_vendidas else 0
            total_producto = float(resultado.total_vendido) if resultado.total_vendido else 0.0
            precio_promedio = float(resultado.precio_promedio) if resultado.precio_promedio else 0.0
            
            # *** INFORMACIÓN ADICIONAL PARA COMBOS ***
            info_combo = ""
            unidad_medida = "unidades"
            
            if resultado.es_combo and resultado.cantidad_combo:
                cantidad_combo = float(resultado.cantidad_combo)
                info_combo = f" ({unidades_combos} combos de {cantidad_combo:g} c/u)"
                # Detectar unidad de medida basada en cantidad del combo
                if cantidad_combo >= 1:
                    if cantidad_combo == int(cantidad_combo):
                        unidad_medida = "kg" if cantidad_combo >= 1 else "unidades"
                    else:
                        unidad_medida = "kg"
                
                print(f"📦 {resultado.codigo}: {unidades_combos} combos × {cantidad_combo:g} = {cantidad_real:g} {unidad_medida}")
            
            productos.append({
                'id': resultado.id,
                'codigo': resultado.codigo,
                'nombre': resultado.nombre,
                'descripcion': resultado.descripcion,
                'categoria': resultado.categoria,
                'es_combo': resultado.es_combo,
                'cantidad_combo': float(resultado.cantidad_combo) if resultado.cantidad_combo else 1.0,
                'cantidad_vendida': cantidad_real,  # *** CANTIDAD REAL ***
                'unidades_combos_vendidas': unidades_combos,  # *** COMBOS VENDIDOS ***
                'info_combo': info_combo,  # *** INFORMACIÓN ADICIONAL ***
                'unidad_medida': unidad_medida,  # *** UNIDAD DE MEDIDA ***
                'total_vendido': total_producto,
                'precio_promedio': precio_promedio,
                'ultima_venta': resultado.ultima_venta.isoformat() if resultado.ultima_venta else None,
                'num_transacciones': int(resultado.num_transacciones) if resultado.num_transacciones else 0
            })
            
            total_unidades_reales += cantidad_real
            total_ventas += total_producto
        
        # *** RESUMEN CORREGIDO ***
        resumen = {
            'total_productos': len(productos),
            'total_unidades_reales': total_unidades_reales,  # *** UNIDADES REALES ***
            'total_ventas': total_ventas,
            'promedio_por_producto': total_ventas / len(productos) if len(productos) > 0 else 0,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'categoria_filtro': categoria or 'Todas',
            'incluye_todas_facturas': True,
            'estados_facturas': estados_info,
            'correccion_combos': True  # *** FLAG PARA INDICAR CORRECCIÓN ***
        }
        
        print(f"✅ Reporte generado exitosamente (CON CORRECCIÓN DE COMBOS):")
        print(f"   Productos: {resumen['total_productos']}")
        print(f"   Unidades REALES: {resumen['total_unidades_reales']:g}")
        print(f"   Total: ${resumen['total_ventas']:.2f}")
        print(f"   Incluye corrección de combos: SÍ")
        
        return jsonify({
            'success': True,
            'productos': productos,
            'resumen': resumen,
            'parametros': {
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'categoria': categoria,
                'orden': orden,
                'solo_con_ventas': solo_con_ventas,
                'incluye_todas_facturas': True,
                'correccion_combos_aplicada': True
            }
        })
        
    except Exception as e:
        print(f"❌ Error en reporte de ventas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/exportar_reporte_ventas')
def exportar_reporte_ventas():
    """Exportar reporte de ventas a Excel, CSV o PDF - CORREGIDO PARA COMBOS"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros (mismos que el reporte)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        categoria = request.args.get('categoria', '').strip()
        orden = request.args.get('orden', 'cantidad_desc')
        formato = request.args.get('formato', 'csv')  # csv, excel o pdf
        
        # Validar fechas
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        
        print(f"📤 Exportando reporte a {formato.upper()}: {fecha_desde} a {fecha_hasta}")
        
        # *** QUERY CORREGIDA (MISMA DEL REPORTE) ***
        query = db.session.query(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo,
            # *** CANTIDAD REAL CALCULADA ***
            func.sum(
                case(
                    (Producto.es_combo == True, DetalleFactura.cantidad * Producto.cantidad_combo),
                    else_=DetalleFactura.cantidad
                )
            ).label('cantidad_real_vendida'),
            func.sum(DetalleFactura.cantidad).label('unidades_combos_vendidas'),
            func.sum(DetalleFactura.subtotal).label('total_vendido'),
            func.avg(DetalleFactura.precio_unitario).label('precio_promedio'),
            func.max(Factura.fecha).label('ultima_venta'),
            func.count(DetalleFactura.id).label('num_transacciones')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde_dt,
                Factura.fecha <= fecha_hasta_dt,
                Factura.estado != 'anulada',  # Cuenta autorizadas + error_afip + internos NORMALES. Excluye anuladas
                # CONSIGNACIÓN: excluir internos a consignatario (ver AYUDA)
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(Factura.tipo_comprobante == '99',
                              Cliente.es_intermediario == True)
                )
            )
        )
        
        if categoria:
            query = query.filter(Producto.categoria == categoria)
        
        query = query.group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            Producto.descripcion,
            Producto.categoria,
            Producto.es_combo,
            Producto.cantidad_combo
        )
        
        # *** ORDENAMIENTO USANDO CANTIDAD REAL ***
        if orden == 'cantidad_desc':
            query = query.order_by(desc('cantidad_real_vendida'))
        elif orden == 'cantidad_asc':
            query = query.order_by(asc('cantidad_real_vendida'))
        elif orden == 'total_desc':
            query = query.order_by(desc('total_vendido'))
        elif orden == 'total_asc':
            query = query.order_by(asc('total_vendido'))
        elif orden == 'codigo':
            query = query.order_by(Producto.codigo)
        elif orden == 'nombre':
            query = query.order_by(Producto.nombre)
        
        resultados = query.all()
        print(f"📊 Exportando {len(resultados)} productos")
        
        # Calcular resumen
        total_unidades_reales = 0
        total_ventas = 0.0
        
        productos_formateados = []
        for resultado in resultados:
            cantidad_real = float(resultado.cantidad_real_vendida) if resultado.cantidad_real_vendida else 0.0
            unidades_combos = int(resultado.unidades_combos_vendidas) if resultado.unidades_combos_vendidas else 0
            total_producto = float(resultado.total_vendido) if resultado.total_vendido else 0.0
            
            # Información del tipo de producto
            if resultado.es_combo:
                tipo_producto = "Combo/Oferta"
                cantidad_combo = float(resultado.cantidad_combo) if resultado.cantidad_combo else 1.0
                detalle_unidades = f"{unidades_combos} combos × {cantidad_combo:g} c/u"
            else:
                tipo_producto = "Producto Base"
                detalle_unidades = f"{int(cantidad_real)} unidades"
            
            productos_formateados.append({
                'id': resultado.id,
                'codigo': resultado.codigo,
                'nombre': resultado.nombre,
                'descripcion': resultado.descripcion or '',
                'categoria': resultado.categoria or 'Sin categoría',
                'tipo_producto': tipo_producto,
                'cantidad_real': cantidad_real,
                'detalle_unidades': detalle_unidades,
                'precio_promedio': float(resultado.precio_promedio) if resultado.precio_promedio else 0.0,
                'total_vendido': total_producto,
                'ultima_venta': resultado.ultima_venta,
                'num_transacciones': int(resultado.num_transacciones) if resultado.num_transacciones else 0
            })
            
            total_unidades_reales += cantidad_real
            total_ventas += total_producto
        
        # Crear resumen
        resumen = {
            'total_productos': len(productos_formateados),
            'total_unidades_reales': total_unidades_reales,
            'total_ventas': total_ventas,
            'promedio_por_producto': total_ventas / len(productos_formateados) if len(productos_formateados) > 0 else 0
        }
        
        parametros = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'categoria': categoria,
            'orden': orden
        }
        
        # Generar archivo según formato
        if formato == 'pdf':
            return exportar_pdf_reporte(productos_formateados, resumen, parametros)
        elif formato == 'excel':
            return generar_excel_reporte_mejorado(productos_formateados, resumen, parametros)
        else:  # CSV por defecto
            return generar_csv_reporte_mejorado(productos_formateados, resumen, parametros)
        
    except Exception as e:
        print(f"❌ Error exportando reporte: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500


def exportar_pdf_reporte(productos, resumen, parametros):
    """Generar archivo PDF del reporte usando la función importada"""
    try:
        # Generar PDF usando la función del módulo importado
        # Agrega esto para debugging:
       
        pdf_bytes = generar_pdf_reporte_ventas(productos, resumen, parametros)
        
        # Crear respuesta
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        
        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_ventas_{fecha_desde}_{fecha_hasta}.pdf'
        )
        
    except Exception as e:
        print(f"❌ Error generando PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al generar PDF: {str(e)}'}), 500


def generar_csv_reporte_mejorado(productos, resumen, parametros):
    """Generar archivo CSV del reporte mejorado"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    fecha_desde = parametros['fecha_desde']
    fecha_hasta = parametros['fecha_hasta']
    
    # Encabezado del reporte
    writer.writerow(['Reporte de Ventas por Producto'])
    writer.writerow([f'Período: {fecha_desde} al {fecha_hasta}'])
    writer.writerow([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    writer.writerow([])
    
    # Resumen
    writer.writerow(['RESUMEN DEL PERÍODO'])
    writer.writerow(['Productos Vendidos', resumen['total_productos']])
    writer.writerow(['Total Unidades', f"{resumen['total_unidades_reales']:,.2f}"])
    writer.writerow(['Total Vendido', f"${resumen['total_ventas']:,.2f}"])
    writer.writerow(['Promedio por Producto', f"${resumen['promedio_por_producto']:,.2f}"])
    writer.writerow([])
    
    # Encabezados de datos
    writer.writerow([
        'Código',
        'Producto',
        'Descripción',
        'Categoría',
        'Tipo',
        'Cantidad Real Vendida',
        'Unidades/Combos',
        'Precio Promedio',
        'Total Vendido',
        'Última Venta',
        'Número de Transacciones'
    ])
    
    # Datos
    for producto in productos:
        writer.writerow([
            producto['codigo'],
            producto['nombre'],
            producto['descripcion'],
            producto['categoria'],
            producto['tipo_producto'],
            f"{producto['cantidad_real']:,.2f}",
            producto['detalle_unidades'],
            f"${producto['precio_promedio']:,.2f}",
            f"${producto['total_vendido']:,.2f}",
            producto['ultima_venta'].strftime('%d/%m/%Y') if producto['ultima_venta'] else 'N/A',
            producto['num_transacciones']
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.csv'
    
    return response


def generar_excel_reporte_mejorado(productos, resumen, parametros):
    """Generar archivo Excel del reporte mejorado"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Estilos
        titulo_font = Font(bold=True, size=16, color="FFFFFF")
        titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        encabezado_font = Font(bold=True, size=11, color="FFFFFF")
        encabezado_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:K1')
        celda_titulo = ws['A1']
        celda_titulo.value = 'Reporte de Ventas por Producto'
        celda_titulo.font = titulo_font
        celda_titulo.fill = titulo_fill
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f'Período: {fecha_desde} al {fecha_hasta}'
        ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Resumen
        fila = 5
        ws.merge_cells(f'A{fila}:K{fila}')
        ws[f'A{fila}'] = 'RESUMEN DEL PERÍODO'
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        
        fila += 1
        ws[f'A{fila}'] = 'Productos Vendidos:'
        ws[f'B{fila}'] = resumen['total_productos']
        ws[f'D{fila}'] = 'Total Unidades:'
        ws[f'E{fila}'] = resumen['total_unidades_reales']
        
        fila += 1
        ws[f'A{fila}'] = 'Total Vendido:'
        ws[f'B{fila}'] = resumen['total_ventas']
        ws[f'D{fila}'] = 'Promedio por Producto:'
        ws[f'E{fila}'] = resumen['promedio_por_producto']
        
        # Encabezados de tabla
        fila += 2
        encabezados = [
            'Código', 'Producto', 'Descripción', 'Categoría', 'Tipo',
            'Cantidad Real', 'Unidades/Combos', 'Precio Prom.', 
            'Total Vendido', 'Última Venta', 'Transacciones'
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=fila, column=col, value=encabezado)
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = border
        
        # Datos
        fila += 1
        for producto in productos:
            ws.cell(row=fila, column=1, value=producto['codigo'])
            ws.cell(row=fila, column=2, value=producto['nombre'])
            ws.cell(row=fila, column=3, value=producto['descripcion'])
            ws.cell(row=fila, column=4, value=producto['categoria'])
            ws.cell(row=fila, column=5, value=producto['tipo_producto'])
            ws.cell(row=fila, column=6, value=producto['cantidad_real'])
            ws.cell(row=fila, column=7, value=producto['detalle_unidades'])
            ws.cell(row=fila, column=8, value=producto['precio_promedio'])
            ws.cell(row=fila, column=9, value=producto['total_vendido'])
            ws.cell(row=fila, column=10, value=producto['ultima_venta'].strftime('%d/%m/%Y') if producto['ultima_venta'] else 'N/A')
            ws.cell(row=fila, column=11, value=producto['num_transacciones'])
            
            fila += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.xlsx'
        
        return response
        
    except ImportError:
        return generar_csv_reporte_mejorado(productos, resumen, parametros)        

def generar_csv_reporte(datos, fecha_desde, fecha_hasta):
    """Generar archivo CSV del reporte"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Escribir encabezado del reporte
    writer.writerow([f'Reporte de Ventas por Producto'])
    writer.writerow([f'Período: {fecha_desde} al {fecha_hasta}'])
    writer.writerow([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    writer.writerow([])  # Línea vacía
    
    # Escribir datos
    for fila in datos:
        writer.writerow(fila)
    
    # Crear respuesta
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.csv'
    
    return response

def generar_excel_reporte(datos, fecha_desde, fecha_hasta):
    """Generar archivo Excel del reporte (requiere openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Estilos
        titulo_font = Font(bold=True, size=16)
        encabezado_font = Font(bold=True, size=12, color="FFFFFF")
        encabezado_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Título del reporte
        ws['A1'] = f'Reporte de Ventas por Producto'
        ws['A1'].font = titulo_font
        ws['A2'] = f'Período: {fecha_desde} al {fecha_hasta}'
        ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Fila inicial para datos
        fila_inicio = 5
        
        # Escribir encabezados
        encabezados = datos[0]
        for col, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal='center')
        
        # Escribir datos
        for fila_idx, fila_datos in enumerate(datos[1:], fila_inicio + 1):
            for col_idx, valor in enumerate(fila_datos, 1):
                ws.cell(row=fila_idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_desde}_{fecha_hasta}.xlsx'
        
        return response
        
    except ImportError:
        # Si no está instalado openpyxl, devolver CSV
        return generar_csv_reporte(datos, fecha_desde, fecha_hasta)

# ==================== REPORTE RÁPIDO DE TOP PRODUCTOS ====================

@app.route('/api/top_productos_vendidos')
def api_top_productos_vendidos():
    """API para obtener top 10 productos más vendidos (últimos 30 días)"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Últimos 30 días
        fecha_hasta = datetime.now()
        fecha_desde = fecha_hasta - timedelta(days=30)
        
        # Query para top productos
        resultados = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.sum(DetalleFactura.cantidad).label('cantidad_vendida'),
            func.sum(DetalleFactura.subtotal).label('total_vendido')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                Factura.fecha >= fecha_desde,
                Factura.fecha <= fecha_hasta,
                Factura.estado != 'anulada',  # Cuenta autorizadas, error_afip, internos NORMALES. Excluye anuladas
                # CONSIGNACIÓN: excluir internos a consignatario (ver AYUDA)
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(Factura.tipo_comprobante == '99',
                              Cliente.es_intermediario == True)
                )
            )
        ).group_by(
            Producto.id,
            Producto.codigo,
            Producto.nombre
        ).order_by(
            desc('cantidad_vendida')
        ).limit(10).all()
        
        # Formatear respuesta
        top_productos = []
        for resultado in resultados:
            top_productos.append({
                'codigo': resultado.codigo,
                'nombre': resultado.nombre,
                'cantidad_vendida': int(resultado.cantidad_vendida),
                'total_vendido': float(resultado.total_vendido)
            })
        
        return jsonify({
            'success': True,
            'productos': top_productos,
            'periodo': '30 días'
        })
        
    except Exception as e:
        print(f"Error en top productos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== DASHBOARD DE VENTAS PARA INCLUIR EN MAIN ====================

# REEMPLAZA tu función api_dashboard_ventas() existente con esta versión mejorada:

@app.route('/api/dashboard_ventas')
def api_dashboard_ventas():
    """API para dashboard de ventas (resumen del día) - VERSIÓN CORREGIDA"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        print("🔍 Iniciando consulta dashboard ventas...")
        
        # Obtener fecha actual
        from datetime import date
        hoy = date.today()
        print(f"📅 Consultando ventas para: {hoy}")
        
        # CONSULTA 1: Datos básicos de ventas del día
        # Usar DATE() para comparar solo la fecha, no hora
        # Excluir anuladas (su stock se reintegró, no son venta real)
        consulta_ventas = db.session.query(
            func.count(Factura.id).label('num_facturas'),
            func.coalesce(func.sum(Factura.total), 0).label('total_vendido')
        ).filter(
            and_(
                func.date(Factura.fecha) == hoy,
                Factura.estado != 'anulada',
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(Factura.tipo_comprobante == '99',
                              Cliente.es_intermediario == True)
                )
            )
        ).first()
        
        print(f"📊 Consulta ventas básicas completada")
        print(f"   Facturas: {consulta_ventas.num_facturas}")
        print(f"   Total: ${consulta_ventas.total_vendido}")
        
        # CONSULTA 2: Total de unidades vendidas del día (excluir anuladas)
        consulta_unidades = db.session.query(
            func.coalesce(func.sum(DetalleFactura.cantidad), 0).label('total_unidades')
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                func.date(Factura.fecha) == hoy,
                Factura.estado != 'anulada',
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(Factura.tipo_comprobante == '99',
                              Cliente.es_intermediario == True)
                )
            )
        ).first()
        
        print(f"📦 Unidades vendidas: {consulta_unidades.total_unidades}")
        
        # CONSULTA 3: Producto más vendido del día (excluir anuladas)
        consulta_top_producto = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.sum(DetalleFactura.cantidad).label('cantidad_vendida')
        ).join(
            DetalleFactura, Producto.id == DetalleFactura.producto_id
        ).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                func.date(Factura.fecha) == hoy,
                Factura.estado != 'anulada',
                ~Factura.id.in_(
                    db.session.query(Factura.id)
                      .join(Cliente, Factura.cliente_id == Cliente.id)
                      .filter(Factura.tipo_comprobante == '99',
                              Cliente.es_intermediario == True)
                )
            )
        ).group_by(
            Producto.id,
            Producto.codigo, 
            Producto.nombre
        ).order_by(
            desc('cantidad_vendida')
        ).first()
        
        if consulta_top_producto:
            print(f"👑 Top producto: {consulta_top_producto.codigo} - {consulta_top_producto.nombre} ({consulta_top_producto.cantidad_vendida} unidades)")
        else:
            print("👑 No hay ventas de productos hoy")
        
        # Preparar respuesta
        response_data = {
            'success': True,
            'ventas_hoy': {
                'num_facturas': int(consulta_ventas.num_facturas or 0),
                'total_vendido': float(consulta_ventas.total_vendido or 0),
                'unidades_vendidas': int(consulta_unidades.total_unidades or 0)
            },
            'producto_top_hoy': None
        }
        
        # Agregar producto top si existe
        if consulta_top_producto:
            response_data['producto_top_hoy'] = {
                'codigo': consulta_top_producto.codigo,
                'nombre': consulta_top_producto.nombre,
                'cantidad': int(consulta_top_producto.cantidad_vendida)
            }
        
        print(f"✅ Dashboard data preparada correctamente")
        print(f"📤 Enviando respuesta: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Error en api_dashboard_ventas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        
        # Devolver datos por defecto en caso de error
        return jsonify({
            'success': True,
            'ventas_hoy': {
                'num_facturas': 0,
                'total_vendido': 0.0,
                'unidades_vendidas': 0
            },
            'producto_top_hoy': None,
            'error_debug': str(e)
        })

# AGREGAR TAMBIÉN ESTA RUTA PARA DEBUG:
@app.route('/debug/dashboard_data')
def debug_dashboard_data():
    """Endpoint para debugging del dashboard"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        from datetime import date
        hoy = date.today()
        
        # Información de debug
        debug_info = {
            'fecha_hoy': str(hoy),
            'total_facturas_bd': Factura.query.count(),
            'total_productos_bd': Producto.query.count(),
            'facturas_hoy': Factura.query.filter(func.date(Factura.fecha) == hoy).count(),
            'ultimas_facturas': []
        }
        
        # Últimas 5 facturas para debug
        ultimas_facturas = Factura.query.order_by(Factura.id.desc()).limit(5).all()
        for factura in ultimas_facturas:
            debug_info['ultimas_facturas'].append({
                'id': factura.id,
                'numero': factura.numero,
                'total': float(factura.total),
                'fecha': factura.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                'estado': factura.estado
            })
        
        # Probar la consulta de dashboard
        try:
            dashboard_data = api_dashboard_ventas()
            debug_info['dashboard_response'] = dashboard_data.get_json()
        except Exception as e:
            debug_info['dashboard_error'] = str(e)
        
        return jsonify({
            'success': True,
            'debug_info': debug_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== PASO 3: FUNCIÓN PARA MIGRAR DATOS EXISTENTES ====================

def migrar_detalle_facturas_con_iva():
    """Migrar detalles de facturas existentes para agregar IVA individual"""
    try:
        print("🔄 Iniciando migración de detalles con IVA...")
        
        # Buscar detalles sin porcentaje_iva o importe_iva
        detalles_sin_iva = DetalleFactura.query.filter(
            or_(
                DetalleFactura.porcentaje_iva.is_(None),
                DetalleFactura.importe_iva.is_(None),
                DetalleFactura.porcentaje_iva == 0
            )
        ).all()
        
        print(f"📋 Encontrados {len(detalles_sin_iva)} detalles para migrar")
        
        contador_migrados = 0
        
        for detalle in detalles_sin_iva:
            try:
                # Obtener porcentaje de IVA del producto
                if detalle.producto:
                    porcentaje_iva = float(detalle.producto.iva)
                else:
                    porcentaje_iva = 21.0  # Por defecto
                
                # Calcular importe de IVA
                subtotal = float(detalle.subtotal)
                importe_iva = round((subtotal * porcentaje_iva / 100), 2)
                
                # Actualizar campos
                detalle.porcentaje_iva = Decimal(str(porcentaje_iva))
                detalle.importe_iva = Decimal(str(importe_iva))
                
                contador_migrados += 1
                
                if contador_migrados % 50 == 0:
                    print(f"   📊 Migrados {contador_migrados}/{len(detalles_sin_iva)}")
                
            except Exception as e:
                print(f"⚠️ Error migrando detalle ID {detalle.id}: {e}")
        
        # Guardar cambios
        if contador_migrados > 0:
            db.session.commit()
            print(f"✅ Migración completada: {contador_migrados} detalles actualizados")
        else:
            print("✅ No hay detalles para migrar")
            
        return contador_migrados
        
    except Exception as e:
        print(f"❌ Error en migración: {e}")
        db.session.rollback()
        return 0


# ==================== PASO 4: FUNCIÓN PARA VERIFICAR IVA POR DETALLE ====================

@app.route('/api/verificar_iva_detalle/<int:factura_id>')
def verificar_iva_detalle(factura_id):
    """Verificar que los detalles tengan IVA individual correcto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        detalles_info = []
        total_iva_calculado = 0
        
        for detalle in factura.detalles:
            # IVA guardado en BD
            iva_bd = float(detalle.importe_iva) if detalle.importe_iva else 0
            porcentaje_bd = float(detalle.porcentaje_iva) if detalle.porcentaje_iva else 0
            
            # IVA recalculado
            subtotal = float(detalle.subtotal)
            porcentaje_producto = float(detalle.producto.iva) if detalle.producto else 21.0
            iva_recalculado = round((subtotal * porcentaje_producto / 100), 2)
            
            total_iva_calculado += iva_recalculado
            
            detalle_info = {
                'id': detalle.id,
                'producto': detalle.producto.nombre if detalle.producto else 'Sin producto',
                'subtotal': subtotal,
                'porcentaje_bd': porcentaje_bd,
                'porcentaje_producto': porcentaje_producto,
                'iva_bd': iva_bd,
                'iva_recalculado': iva_recalculado,
                'coincide': abs(iva_bd - iva_recalculado) < 0.01
            }
            
            detalles_info.append(detalle_info)
        
        # Comparar con total de factura
        iva_factura = float(factura.iva)
        
        return jsonify({
            'success': True,
            'factura_numero': factura.numero,
            'iva_factura': iva_factura,
            'iva_calculado_suma': round(total_iva_calculado, 2),
            'total_coincide': abs(iva_factura - total_iva_calculado) < 0.01,
            'detalles': detalles_info
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== PASO 5: RUTA PARA MIGRACIÓN MANUAL ====================

@app.route('/migrar_iva_detalles', methods=['POST'])
def migrar_iva_detalles_endpoint():
    """Endpoint para ejecutar migración de IVA en detalles"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        contador = migrar_detalle_facturas_con_iva()
        
        return jsonify({
            'success': True,
            'mensaje': f'Migración completada: {contador} detalles actualizados',
            'detalles_migrados': contador
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500

@app.route('/importar_productos')
def importar_productos_vista():
    """Vista para importar productos desde Excel"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('importar_productos.html')



@app.route('/api/importar_productos_lote', methods=['POST'])
def importar_productos_lote():
    try:
        data = request.get_json()
        productos = data.get('productos', [])
        opciones = data.get('opciones', {})
        
        solo_actualizar = opciones.get('solo_actualizar', False)
        crear_nuevos = opciones.get('crear_nuevos', True)
        incluir_costo_margen = opciones.get('incluir_costo_margen', True)
        
        if not productos:
            return jsonify({
                'success': False,
                'error': 'No se recibieron productos para importar'
            })
        
        resultados = {
            'nuevos': 0,
            'actualizados': 0,
            'errores': 0,
            'productos_procesados': [],
            'detalles_errores': []
        }
        
        print(f"Procesando {len(productos)} productos...")
        
        for producto_data in productos:
            try:
                codigo = producto_data.get('codigo', '').strip()
                descripcion = producto_data.get('descripcion', '').strip()
                precio = producto_data.get('precio', 0)
                costo = producto_data.get('costo', 0) if incluir_costo_margen else 0
                margen = producto_data.get('margen', 0) if incluir_costo_margen else 0
                
                # Validaciones básicas
                if not codigo or not descripcion or precio <= 0:
                    resultados['errores'] += 1
                    resultados['detalles_errores'].append({
                        'codigo': codigo,
                        'error': 'Datos incompletos o inválidos'
                    })
                    continue
                
                # Verificar si el producto existe (usando SQLAlchemy)
                producto_existente = Producto.query.filter_by(codigo=codigo).first()
                
                if producto_existente:
                    # Producto existe - actualizar
                    if solo_actualizar or crear_nuevos:
                        producto_existente.nombre = descripcion
                        producto_existente.descripcion = descripcion
                        producto_existente.precio = Decimal(str(precio))
                        
                        if incluir_costo_margen:
                            producto_existente.costo = Decimal(str(costo))
                            producto_existente.margen = Decimal(str(margen))
                        
                        producto_existente.fecha_modificacion = datetime.now()
                        
                        resultados['actualizados'] += 1
                        resultados['productos_procesados'].append({
                            'codigo': codigo,
                            'estado': 'actualizado'
                        })
                        print(f"Actualizado: {codigo}")
                    else:
                        resultados['productos_procesados'].append({
                            'codigo': codigo,
                            'estado': 'existente'
                        })
                
                else:
                    # Producto no existe - crear nuevo
                    if crear_nuevos:
                        nuevo_producto = Producto(
                            codigo=codigo,
                            nombre=descripcion,
                            descripcion=descripcion,
                            precio=Decimal(str(precio)),
                            costo=Decimal(str(costo)) if incluir_costo_margen else Decimal('0'),
                            margen=Decimal(str(margen)) if incluir_costo_margen else Decimal('0'),
                            stock=0,
                            categoria='Importado',
                            iva=Decimal('21.00'),
                            activo=True
                        )
                        
                        db.session.add(nuevo_producto)
                        
                        resultados['nuevos'] += 1
                        resultados['productos_procesados'].append({
                            'codigo': codigo,
                            'estado': 'nuevo'
                        })
                        print(f"Creado: {codigo}")
                    else:
                        resultados['productos_procesados'].append({
                            'codigo': codigo,
                            'estado': 'no_creado'
                        })
                
            except Exception as e:
                resultados['errores'] += 1
                resultados['detalles_errores'].append({
                    'codigo': producto_data.get('codigo', 'N/A'),
                    'error': str(e)
                })
                print(f"Error procesando producto {producto_data.get('codigo')}: {e}")
        
        # Guardar todos los cambios
        db.session.commit()
        
        print(f"Importación completada: {resultados['nuevos']} nuevos, "
              f"{resultados['actualizados']} actualizados, {resultados['errores']} errores")
        
        return jsonify({
            'success': True,
            'nuevos': resultados['nuevos'],
            'actualizados': resultados['actualizados'],
            'errores': resultados['errores'],
            'productos_procesados': resultados['productos_procesados'],
            'detalles_errores': resultados['detalles_errores']
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error general en importar_productos_lote: {e}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500
        
                

def detectar_categoria(descripcion):
    """Detectar categoría básica desde la descripción del producto"""
    descripcion_lower = descripcion.lower()
    
    # Mapeo de palabras clave a categorías
    categorias = {
        'POLLO': ['pollo', 'pechuga', 'muslo', 'ala', 'carcasa'],
        'CARNE': ['carne', 'bife', 'asado', 'costilla', 'vacio', 'chorizo'],
        'CERDO': ['cerdo', 'bondiola', 'matambre', 'costilla cerdo'],
        'PESCADO': ['pescado', 'salmon', 'merluza', 'atun'],
        'CHACINADOS': ['salame', 'jamon', 'mortadela', 'chorizo', 'morcilla'],
        'LACTEOS': ['leche', 'queso', 'yogur', 'manteca', 'crema'],
        'CONGELADOS': ['congelado', 'frozen', 'helado'],
        'BEBIDAS': ['gaseosa', 'agua', 'jugo', 'cerveza', 'vino'],
        'PANADERIA': ['pan', 'facturas', 'torta', 'galletas'],
        'LIMPIEZA': ['detergente', 'lavandina', 'jabon', 'shampoo'],
        'VERDURAS': ['verdura', 'lechuga', 'tomate', 'cebolla', 'papa']
    }
    
    for categoria, palabras_clave in categorias.items():
        if any(palabra in descripcion_lower for palabra in palabras_clave):
            return categoria
    
    return 'GENERAL'  # Categoría por defecto

# AGREGAR esta nueva ruta en app.py:

@app.route('/api/eliminar_combo/<int:combo_id>', methods=['DELETE'])
def eliminar_combo(combo_id):
    """Eliminar un combo con validaciones de seguridad"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Verificar que el producto existe y es un combo
        combo = Producto.query.get_or_404(combo_id)
        
        if not combo.es_combo:
            return jsonify({
                'success': False,
                'error': 'El producto no es un combo'
            }), 400
        
        print(f"🗑️ Solicitud de eliminación para combo: {combo.codigo} - {combo.nombre}")
        
        # VALIDACIÓN 1: Verificar si el combo se vendió alguna vez
        ventas_combo = DetalleFactura.query.filter_by(producto_id=combo_id).count()
        
        if ventas_combo > 0:
            print(f"❌ Combo {combo.codigo} tiene {ventas_combo} ventas registradas")
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar el combo porque tiene {ventas_combo} ventas registradas. Solo puedes desactivarlo.',
                'motivo': 'tiene_ventas',
                'ventas_count': ventas_combo,
                'sugerencia': 'Usa el botón de desactivar en lugar de eliminar'
            }), 400
        
        # VALIDACIÓN 2: Verificar si está en facturas pendientes o con errores
        facturas_pendientes = db.session.query(DetalleFactura).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                DetalleFactura.producto_id == combo_id,
                or_(
                    Factura.estado == 'pendiente',
                    Factura.estado == 'error_afip'
                )
            )
        ).count()
        
        if facturas_pendientes > 0:
            print(f"❌ Combo {combo.codigo} está en {facturas_pendientes} facturas pendientes")
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar el combo porque está en {facturas_pendientes} facturas pendientes de autorización.',
                'motivo': 'facturas_pendientes',
                'facturas_count': facturas_pendientes
            }), 400
        
        # VALIDACIÓN 3: Verificar si tiene stock (opcional - puedes eliminarlo igual)
        if combo.stock > 0:
            print(f"⚠️ Advertencia: Combo {combo.codigo} tiene stock {combo.stock}")
            # No bloquear, solo advertir
        
        # SI LLEGAMOS AQUÍ: Es seguro eliminar
        print(f"✅ Combo {combo.codigo} puede eliminarse de forma segura")
        
        # Eliminar el combo de la base de datos
        nombre_combo = combo.nombre
        codigo_combo = combo.codigo
        
        db.session.delete(combo)
        db.session.commit()
        
        print(f"🗑️ Combo eliminado exitosamente: {codigo_combo}")
        
        return jsonify({
            'success': True,
            'message': f'Combo "{nombre_combo}" eliminado correctamente',
            'codigo_eliminado': codigo_combo,
            'motivo': 'eliminacion_segura'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error eliminando combo: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno al eliminar combo: {str(e)}'
        }), 500


@app.route('/api/verificar_eliminacion_combo/<int:combo_id>')
def verificar_eliminacion_combo(combo_id):
    """Verificar si un combo puede ser eliminado de forma segura"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        combo = Producto.query.get_or_404(combo_id)
        
        if not combo.es_combo:
            return jsonify({
                'puede_eliminar': False,
                'motivo': 'no_es_combo'
            })
        
        # Verificar ventas
        ventas_count = DetalleFactura.query.filter_by(producto_id=combo_id).count()
        
        # Verificar facturas pendientes
        facturas_pendientes = db.session.query(DetalleFactura).join(
            Factura, DetalleFactura.factura_id == Factura.id
        ).filter(
            and_(
                DetalleFactura.producto_id == combo_id,
                or_(
                    Factura.estado == 'pendiente',
                    Factura.estado == 'error_afip'
                )
            )
        ).count()
        
        puede_eliminar = (ventas_count == 0 and facturas_pendientes == 0)
        
        resultado = {
            'puede_eliminar': puede_eliminar,
            'ventas_count': ventas_count,
            'facturas_pendientes': facturas_pendientes,
            'stock': combo.stock,
            'codigo': combo.codigo,
            'nombre': combo.nombre
        }
        
        if not puede_eliminar:
            if ventas_count > 0:
                resultado['motivo'] = 'tiene_ventas'
                resultado['mensaje'] = f'El combo tiene {ventas_count} ventas registradas'
            elif facturas_pendientes > 0:
                resultado['motivo'] = 'facturas_pendientes'
                resultado['mensaje'] = f'El combo está en {facturas_pendientes} facturas pendientes'
        else:
            resultado['motivo'] = 'puede_eliminar'
            resultado['mensaje'] = 'El combo puede eliminarse de forma segura'
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'puede_eliminar': False,
            'error': str(e)
        }), 500

# AGREGAR estas rutas en app.py:

@app.route('/api/buscar_facturas')
def buscar_facturas():
    """Buscar facturas con filtros avanzados"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de búsqueda
        numero = request.args.get('numero', '').strip()
        cliente = request.args.get('cliente', '').strip()
        estado = request.args.get('estado', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        limite = int(request.args.get('limite', 100))  # Limitar resultados
        
        print(f"🔍 Búsqueda de facturas:")
        print(f"   Número: '{numero}'")
        print(f"   Cliente: '{cliente}'")
        print(f"   Estado: '{estado}'")
        print(f"   Fecha desde: '{fecha_desde}'")
        print(f"   Fecha hasta: '{fecha_hasta}'")
        print(f"   Límite: {limite}")
        
        # Construir query base con join a cliente
        query = db.session.query(Factura).join(Cliente, Factura.cliente_id == Cliente.id)
        
        # Aplicar filtros
        if numero:
            query = query.filter(Factura.numero.ilike(f'%{numero}%'))
            print(f"   Filtro aplicado: Número contiene '{numero}'")
        
        if cliente:
            query = query.filter(Cliente.nombre.ilike(f'%{cliente}%'))
            print(f"   Filtro aplicado: Cliente contiene '{cliente}'")
        
        if estado:
            query = query.filter(Factura.estado == estado)
            print(f"   Filtro aplicado: Estado = '{estado}'")
        
        # Filtros de fecha
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
                query = query.filter(Factura.fecha >= fecha_desde_dt)
                print(f"   Filtro aplicado: Fecha >= {fecha_desde}")
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha desde inválido. Use YYYY-MM-DD'
                }), 400
        
        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                query = query.filter(Factura.fecha <= fecha_hasta_dt)
                print(f"   Filtro aplicado: Fecha <= {fecha_hasta}")
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha hasta inválido. Use YYYY-MM-DD'
                }), 400
        
        # Ordenar por fecha descendente (más recientes primero)
        query = query.order_by(Factura.fecha.desc())
        
        # Aplicar límite
        facturas = query.limit(limite).all()
        
        print(f"   Facturas encontradas: {len(facturas)}")
        
        # Formatear resultados
        resultado = []
        for factura in facturas:
            # ✅ BUSCAR DESCUENTO APLICADO
            descuento = DescuentoFactura.query.filter_by(factura_id=factura.id).first()

            # Obtener información de medios de pago
            medios_pago = []
            for medio in factura.medios_pago:
                medios_pago.append({
                    'medio_pago': medio.medio_pago,
                    'importe': float(medio.importe)
                })
            
            factura_dict = {
                'id': factura.id,
                'numero': factura.numero,
                'fecha': factura.fecha.strftime('%d/%m/%Y %H:%M'),
                'fecha_iso': factura.fecha.isoformat(),
                'cliente': {
                    'id': factura.cliente.id,
                    'nombre': factura.cliente.nombre,
                    'documento': factura.cliente.documento,
                    'tipo_documento': factura.cliente.tipo_documento
                },
                'tipo_comprobante': factura.tipo_comprobante,
                'tipo_comprobante_nombre': obtener_nombre_comprobante(factura.tipo_comprobante),
                'subtotal': float(factura.subtotal),
                'iva': float(factura.iva),
                'total': float(factura.total),
                'estado': factura.estado,
                'estado_descripcion': obtener_descripcion_estado(factura.estado),
                'cae': factura.cae,
                'vto_cae': factura.vto_cae.strftime('%d/%m/%Y') if factura.vto_cae else None,
                'medios_pago': medios_pago,
                'cantidad_items': len(factura.detalles),
                'usuario': factura.usuario.nombre if factura.usuario else 'Desconocido',
                 'tiene_descuento': bool(descuento),
                'descuento_porcentaje': float(descuento.porcentaje_descuento) if descuento else 0,
                'descuento_monto': float(descuento.monto_descuento) if descuento else 0,
                'total_original': float(descuento.total_original) if descuento else float(factura.total)
            }
            
            resultado.append(factura_dict)
        
        return jsonify({
            'success': True,
            'facturas': resultado,
            'total': len(resultado),
            'limite_aplicado': len(facturas) == limite,
            'filtros_aplicados': {
                'numero': numero,
                'cliente': cliente,
                'estado': estado,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta
            }
        })
        
    except Exception as e:
        print(f"❌ Error buscando facturas: {str(e)}")
        import traceback
        print(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Error en la búsqueda: {str(e)}'
        }), 500


def obtener_nombre_comprobante(tipo):
    """Obtener nombre legible del tipo de comprobante"""
    tipos = {
        '01': 'Factura A',
        '06': 'Factura B', 
        '11': 'Factura C',
        '51': 'Factura M',
        '99': 'Comprobante Interno'
    }
    return tipos.get(str(tipo), f'Tipo {tipo}')


def obtener_descripcion_estado(estado):
    """Obtener descripción legible del estado"""
    estados = {
        'autorizada': 'Autorizada por AFIP',
        'pendiente': 'Pendiente de autorización',
        'error_afip': 'Error en AFIP',
        'anulada': 'Anulada',
        'interno': 'Comprobante Interno (sin AFIP)'
    }
    return estados.get(estado, estado.title())


@app.route('/api/reintentar_afip/<int:factura_id>', methods=['POST'])
def reintentar_afip(factura_id):
    """Reintentar autorización AFIP para una factura pendiente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        if factura.estado not in ['pendiente', 'error_afip']:
            return jsonify({
                'success': False,
                'error': f'No se puede reintentar. Estado actual: {factura.estado}'
            }), 400
        
        print(f"🔄 Reintentando autorización AFIP para factura {factura.numero}")
        
        # Preparar datos para AFIP
        cliente = factura.cliente
        
        # Obtener items con IVA detallado
        items_detalle = []
        for detalle in factura.detalles:
            items_detalle.append({
                'subtotal': float(detalle.subtotal),
                'iva_porcentaje': float(detalle.porcentaje_iva) if detalle.porcentaje_iva else 21.0
            })
        
        datos_comprobante = {
            'tipo_comprobante': int(factura.tipo_comprobante),
            'punto_venta': factura.punto_venta,
            'importe_neto': float(factura.subtotal),
            'importe_iva': float(factura.iva),
            'items_detalle': items_detalle,
            'doc_tipo': 99,  # Sin identificar por defecto
            'doc_nro': 0
        }
        
        # Agregar datos del cliente si existen
        if cliente and cliente.documento:
            if cliente.tipo_documento == 'CUIT' and len(cliente.documento) == 11:
                datos_comprobante['doc_tipo'] = 80  # CUIT
                datos_comprobante['doc_nro'] = int(cliente.documento)
            elif cliente.tipo_documento == 'DNI' and len(cliente.documento) >= 7:
                datos_comprobante['doc_tipo'] = 96  # DNI
                datos_comprobante['doc_nro'] = int(cliente.documento)
        
        if cliente:
            datos_comprobante['condicion_iva_str'] = cliente.condicion_iva
        # Intentar autorizar en AFIP
        resultado_afip = arca_client.autorizar_comprobante(datos_comprobante)
        
        if resultado_afip['success']:
            # Actualizar factura con datos de AFIP
            numero_afip = resultado_afip['numero']
            
            # Verificar si el número de AFIP ya existe
            factura_existente = Factura.query.filter(
                and_(Factura.numero == numero_afip, Factura.id != factura.id)
            ).first()
            
            if not factura_existente:
                factura.numero = numero_afip
            
            factura.cae = resultado_afip['cae']
            factura.vto_cae = resultado_afip['vto_cae']
            factura.estado = 'autorizada'
            
            db.session.commit()
            
            print(f"✅ Reintento exitoso. CAE: {factura.cae}")
            
            return jsonify({
                'success': True,
                'message': f'Factura autorizada exitosamente',
                'numero': factura.numero,
                'cae': factura.cae,
                'estado': factura.estado
            })
        else:
            # Actualizar estado a error
            factura.estado = 'error_afip'
            db.session.commit()
            
            print(f"❌ Reintento falló: {resultado_afip.get('error', 'Error desconocido')}")
            
            return jsonify({
                'success': False,
                'error': f"Error AFIP: {resultado_afip.get('error', 'Error desconocido')}",
                'estado': factura.estado
            })
        
    except Exception as e:
        print(f"❌ Error en reintento AFIP: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500


# ============================================================================
# FUNCIÓN CORREGIDA FINAL: anular_factura CON REINTEGRO DE STOCK
# Reemplazar en app.py desde línea 5645 hasta línea 5684
# ============================================================================

@app.route('/api/anular_factura/<int:factura_id>', methods=['POST'])
def anular_factura(factura_id):
    """Anular una factura y restaurar el stock de los productos"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        factura = Factura.query.get_or_404(factura_id)
        
        if factura.estado == 'anulada':
            return jsonify({
                'success': False,
                'error': 'La factura ya está anulada'
            }), 400
        
        motivo = request.json.get('motivo', '').strip() if request.json else ''
        
        print(f"❌ Anulando factura {factura.numero}. Motivo: {motivo}")
        
        # ═══ VALIDACIÓN CONSIGNACIÓN: si es un Interno con factura derivada, no permitir ═══
        if factura.tipo_comprobante == '99' and factura.estado == 'interno':
            factura_derivada = Factura.query.filter(
                Factura.interno_origen_id == factura.id,
                Factura.estado != 'anulada'
            ).first()
            if factura_derivada:
                return jsonify({
                    'success': False,
                    'error': f'Este Comprobante Interno está vinculado a la factura {factura_derivada.numero}. Anulá primero esa factura.'
                }), 400
        
        # ═══ VALIDACIÓN CONSIGNACIÓN: si es factura derivada, NO reintegrar stock ═══
        es_factura_consignacion = factura.interno_origen_id is not None
        if es_factura_consignacion:
            interno_origen = Factura.query.get(factura.interno_origen_id)
            print(f"🔗 Anulación de factura consignada (origen interno {interno_origen.numero if interno_origen else '?'})")
            print(f"   Stock NO se reintegra (lo reintegrará el interno cuando se anule)")
        
        # ==================== REINTEGRAR STOCK (solo si no es consignación) ====================
        items_factura = DetalleFactura.query.filter_by(factura_id=factura.id).all()
        productos_reintegrados = []
        
        for item in items_factura:
            producto = Producto.query.get(item.producto_id)
            
            if producto:
                if es_factura_consignacion:
                    # Solo registrar movimiento informativo, NO tocar stock
                    registrar_movimiento_stock(
                        db=db,
                        producto_id=producto.id,
                        tipo='facturacion_interno_anulada',
                        cantidad=float(item.cantidad),
                        signo='=',
                        stock_anterior=float(producto.stock),
                        stock_nuevo=float(producto.stock),
                        referencia_tipo='factura',
                        referencia_id=factura.id,
                        motivo=f'Anulación factura consignada {factura.numero} (sin movimiento de stock)' + (f' - {motivo}' if motivo else ''),
                        usuario_id=session.get('user_id'),
                        usuario_nombre=session.get('nombre', 'Sistema'),
                        codigo_producto=producto.codigo,
                        nombre_producto=producto.nombre
                    )
                    productos_reintegrados.append({
                        'codigo': producto.codigo,
                        'nombre': producto.nombre,
                        'cantidad': float(item.cantidad),
                        'stock_anterior': float(producto.stock),
                        'stock_nuevo': float(producto.stock)  # mismo stock
                    })
                else:
                    # Reintegrar stock normal
                    stock_anterior = producto.stock
                    producto.stock += item.cantidad
                    
                    productos_reintegrados.append({
                        'codigo': producto.codigo,
                        'nombre': producto.nombre,
                        'cantidad': float(item.cantidad),
                        'stock_anterior': float(stock_anterior),
                        'stock_nuevo': float(producto.stock)
                    })
                    
                    print(f"   📦 Reintegrando {item.cantidad} unidades de {producto.codigo}")
                    print(f"      Stock: {stock_anterior} → {producto.stock}")
                    
                    # Registrar en auditoría de stock
                    registrar_movimiento_stock(
                        db=db,
                        producto_id=producto.id,
                        tipo='devolucion',
                        cantidad=float(item.cantidad),
                        signo='+',
                        stock_anterior=float(stock_anterior),
                        stock_nuevo=float(producto.stock),
                        referencia_tipo='factura',
                        referencia_id=factura.id,
                        motivo=f'Anulación factura {factura.numero}' + (f' - {motivo}' if motivo else ''),
                        usuario_id=session.get('user_id'),
                        usuario_nombre=session.get('nombre', 'Sistema'),
                        codigo_producto=producto.codigo,
                        nombre_producto=producto.nombre
                    )
            else:
                print(f"   ⚠️ ADVERTENCIA: Producto ID {item.producto_id} no encontrado")
        
        if not es_factura_consignacion:
            print(f"✅ Stock reintegrado: {len(productos_reintegrados)} productos")
        # ==========================================================
        
        # Marcar como anulada
        factura.estado = 'anulada'
        
        # OPCIONAL: Si querés guardar el motivo y fecha
        # (necesitarías agregar estos campos al modelo Factura)
        # factura.fecha_anulacion = datetime.now()
        # factura.motivo_anulacion = motivo
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Factura {factura.numero} anulada correctamente. Stock reintegrado: {len(productos_reintegrados)} productos.',
            'estado': factura.estado,
            'productos_reintegrados': len(productos_reintegrados),
            'detalle_reintegro': productos_reintegrados
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error anulando factura: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Error al anular factura: {str(e)}'
        }), 500


###################################################################################################################
# ============================================================================
# PASO 2: RUTA PARA EMITIR NOTA DE CRÉDITO - DEPRECADA
# Esta función fue movida a notas_credito.py (blueprint)
# COMENTADA para evitar conflictos - la versión correcta está en el blueprint
# ============================================================================
# @app.route('/api/emitir_nota_credito/<int:factura_id>', methods=['POST'])
# def emitir_nota_credito_OLD(factura_id):
    # """
    # Emitir una Nota de Crédito electrónica en AFIP/ARCA para anular una factura.
    
    # Flujo:
    # 1. Validar que la factura esté autorizada
    # 2. Verificar que no tenga NC previa
    # 3. Crear la NC con los datos de la factura
    # 4. Enviar a AFIP para autorización
    # 5. Obtener CAE
    # 6. Reintegrar stock
    # 7. Marcar factura como anulada
    # """
    # if 'user_id' not in session:
        # return jsonify({'error': 'No autorizado'}), 401
    
    # try:
        # # ==================== VALIDACIONES ====================
        # factura = Factura.query.get_or_404(factura_id)
        
        # # 1. Solo se pueden hacer NC de facturas autorizadas
        # if factura.estado != 'autorizada':
            # return jsonify({
                # 'success': False,
                # 'error': f'No se puede emitir NC de una factura en estado: {factura.estado}.\n'
                        # 'Solo se pueden hacer NC de facturas autorizadas.'
            # }), 400
        
        # # 2. Verificar que no tenga ya una NC
        # nc_existente = NotaCredito.query.filter_by(factura_id=factura.id).first()
        # if nc_existente:
            # return jsonify({
                # 'success': False,
                # 'error': f'Esta factura ya tiene una Nota de Crédito: {nc_existente.numero}'
            # }), 400
        
        # # 3. Obtener motivo
        # data = request.get_json()
        # motivo = data.get('motivo', '').strip()
        
        # if not motivo:
            # return jsonify({
                # 'success': False,
                # 'error': 'Debe ingresar un motivo para la Nota de Crédito'
            # }), 400
        
        # print(f"\n{'='*70}")
        # print(f"📝 EMITIENDO NOTA DE CRÉDITO")
        # print(f"{'='*70}")
        # print(f"Factura: {factura.numero}")
        # print(f"Motivo: {motivo}")
        
        # # ==================== DETERMINAR TIPO DE NC ====================
        # # Mapeo: Factura → Nota de Crédito
        # tipo_nc_map = {
            # '01': '03',  # Factura A → NC A
            # '1': '03',
            # '06': '08',  # Factura B → NC B
            # '6': '08',
            # '11': '13',  # Factura C → NC C
            # '51': '53'   # Factura M → NC M
        # }
        
        # tipo_nc = tipo_nc_map.get(str(factura.tipo_comprobante))
        
        # if not tipo_nc:
            # return jsonify({
                # 'success': False,
                # 'error': f'Tipo de factura no soportado para NC: {factura.tipo_comprobante}'
            # }), 400
        
        # print(f"Tipo NC: {tipo_nc}")
        
        # # ==================== OBTENER PRÓXIMO NÚMERO ====================
        # punto_venta = ARCA_CONFIG.PUNTO_VENTA  # Variable global de configuración
        
        # # Buscar último número de NC para este tipo y punto de venta
        # ultima_nc = NotaCredito.query.filter_by(
            # tipo_comprobante=tipo_nc,
            # punto_venta=punto_venta
        # ).order_by(NotaCredito.id.desc()).first()
        
        # if ultima_nc and ultima_nc.numero:
            # # Extraer número de formato 0001-00000123
            # partes = ultima_nc.numero.split('-')
            # proximo_num = int(partes[1]) + 1
        # else:
            # proximo_num = 1
        
        # numero_nc = f"{punto_venta:04d}-{proximo_num:08d}"
        # print(f"Número NC: {numero_nc}")
        
        # # ==================== CREAR NOTA DE CRÉDITO ====================
        # nota_credito = NotaCredito(
            # numero=numero_nc,
            # tipo_comprobante=tipo_nc,
            # punto_venta=punto_venta,
            # fecha=datetime.now(),
            # factura_id=factura.id,
            # factura_numero=factura.numero,
            # cliente_id=factura.cliente_id,
            # usuario_id=session['user_id'],
            # subtotal=factura.subtotal,
            # iva=factura.iva,
            # total=factura.total,
            # estado='pendiente',
            # motivo=motivo
        # )
        
        # db.session.add(nota_credito)
        # db.session.flush()  # Para obtener el ID
        
        # # ==================== COPIAR ITEMS DE LA FACTURA ====================
        # items_factura = DetalleFactura.query.filter_by(factura_id=factura.id).all()
        
        # for item in items_factura:
            # detalle_nc = DetalleNotaCredito(
                # nota_credito_id=nota_credito.id,
                # producto_id=item.producto_id,
                # cantidad=item.cantidad,
                # precio_unitario=item.precio_unitario,
                # subtotal=item.subtotal,
                # porcentaje_iva=item.porcentaje_iva,
                # importe_iva=item.importe_iva
            # )
            # db.session.add(detalle_nc)
        
        # print(f"Items copiados: {len(items_factura)}")
        
        # # ==================== AUTORIZAR EN AFIP ====================
        # try:
            # print("\n📡 Enviando a AFIP/ARCA...")
            
            # # ✅ CAMBIO CRÍTICO: Usar la función mejorada que ya tiene el ID de la NC
            # resultado = autorizar_comprobante_afip(
                # comprobante_id=nota_credito.id,
                # tipo_comprobante='nota_credito'  # ← Especifica que es una NC
            # )
            
            # if resultado['success']:
                # # ✅ AUTORIZADA
                # nota_credito.estado = 'autorizada'
                # nota_credito.cae = resultado['cae']
                # nota_credito.vto_cae = datetime.strptime(resultado['vto_cae'], '%Y%m%d').date()
                # nota_credito.fecha_autorizacion = datetime.now()
                
                # print(f"✅ NC Autorizada - CAE: {resultado['cae']}")
                
            # else:
                # # ❌ ERROR EN AFIP
                # nota_credito.estado = 'error_afip'
                # print(f"❌ Error AFIP: {resultado.get('error', 'Error desconocido')}")
        
        # except Exception as e_afip:
            # print(f"❌ Excepción en AFIP: {str(e_afip)}")
            # nota_credito.estado = 'error_afip'
            # import traceback
            # traceback.print_exc()
        
        # # ==================== REINTEGRAR STOCK ====================
        # if nota_credito.estado == 'autorizada' or nota_credito.estado == 'error_afip':
            # print("\n📦 Reintegrando stock...")
            
            # productos_reintegrados = []
            
            # for item in items_factura:
                # producto = Producto.query.get(item.producto_id)
                
                # if producto:
                    # stock_anterior = producto.stock
                    # producto.stock += item.cantidad
                    
                    # productos_reintegrados.append({
                        # 'codigo': producto.codigo,
                        # 'nombre': producto.nombre,
                        # 'cantidad': float(item.cantidad),
                        # 'stock_anterior': float(stock_anterior),
                        # 'stock_nuevo': float(producto.stock)
                    # })
                    
                    # print(f"   📦 {producto.codigo}: {stock_anterior} → {producto.stock} (+{item.cantidad})")
            
            # print(f"✅ Stock reintegrado: {len(productos_reintegrados)} productos")
            
            # # Marcar factura como anulada
            # factura.estado = 'anulada'
            # factura.tiene_nota_credito = True
            # factura.fecha_anulacion = datetime.now()
            # factura.motivo_anulacion = motivo
        
        # # ==================== GUARDAR TODO ====================
        # db.session.commit()
        
        # print(f"\n{'='*70}")
        # print(f"✅ NOTA DE CRÉDITO COMPLETADA")
        # print(f"{'='*70}\n")
        
        # # ==================== RESPUESTA ====================
        # return jsonify({
            # 'success': True,
            # 'message': f'Nota de Crédito {numero_nc} emitida correctamente',
            # 'nota_credito': {
                # 'id': nota_credito.id,
                # 'numero': nota_credito.numero,
                # 'estado': nota_credito.estado,
                # 'cae': nota_credito.cae,
                # 'vto_cae': nota_credito.vto_cae.isoformat() if nota_credito.vto_cae else None,
                # 'total': float(nota_credito.total)
            # },
            # 'factura_anulada': factura.estado == 'anulada',
            # 'productos_reintegrados': len(productos_reintegrados) if nota_credito.estado == 'autorizada' else 0
        # })
        
    # except Exception as e:
        # db.session.rollback()
        # print(f"\n❌ ERROR GENERAL: {str(e)}")
        # import traceback
        # traceback.print_exc()
        
        # return jsonify({
            # 'success': False,
            # 'error': f'Error al emitir Nota de Crédito: {str(e)}'
        # }), 500


# ============================================================================
# FUNCIÓN MEJORADA: autorizar_comprobante_afip()
# Soporta tanto Facturas como Notas de Crédito
# ============================================================================
# Para FactuFacil - Pablo Gustavo Ré
# Fecha: 2025-11-07
# ============================================================================

def autorizar_comprobante_afip(comprobante_id, tipo_comprobante='factura'):
    """
    Autoriza un comprobante (Factura o Nota de Crédito) en AFIP/ARCA
    
    Args:
        comprobante_id: ID del comprobante (factura_id o nota_credito_id)
        tipo_comprobante: 'factura' o 'nota_credito'
    
    Returns:
        dict: {'success': bool, 'cae': str, 'vto_cae': str, 'error': str}
    """
    
    print("=" * 70)
    print(f"📡 AUTORIZANDO {tipo_comprobante.upper()} EN AFIP")
    print("=" * 70)
    
    try:
        # ====================================================================
        # 1. OBTENER DATOS DEL COMPROBANTE
        # ====================================================================
        
        if tipo_comprobante == 'factura':
            # Obtener factura
            comprobante = Factura.query.get(comprobante_id)
            if not comprobante:
                return {'success': False, 'error': 'Factura no encontrada'}
            
            # Obtener detalles
            detalles = DetalleFactura.query.filter_by(factura_id=comprobante_id).all()
            
            # Determinar tipo de comprobante AFIP
            tipo_cbte_afip = comprobante.tipo_comprobante  # 1=A, 6=B, 11=C, etc.
            
            # NC asociada (no aplica para facturas)
            cbtes_asoc = None
            
        elif tipo_comprobante == 'nota_credito':
            # Obtener NC
            comprobante = NotaCredito.query.get(comprobante_id)
            if not comprobante:
                return {'success': False, 'error': 'Nota de Crédito no encontrada'}
            
            # Obtener detalles
            detalles = DetalleNotaCredito.query.filter_by(nota_credito_id=comprobante_id).all()
            
            # Obtener factura asociada
            factura_original = Factura.query.get(comprobante.factura_id)
            if not factura_original:
                return {'success': False, 'error': 'Factura original no encontrada'}
            
            # Mapear tipo de NC según factura (acepta int y string)
            mapeo_nc = {
                1: 3, '1': 3, '01': 3,    # Factura A → NC A
                6: 8, '6': 8, '06': 8,    # Factura B → NC B
                11: 13, '11': 13,         # Factura C → NC C
                51: 53, '51': 53          # Factura M → NC M
            }

            # Convertir a int si es string
            tipo_factura = factura_original.tipo_comprobante
            if isinstance(tipo_factura, str):
                tipo_factura = int(tipo_factura)

            tipo_cbte_afip = mapeo_nc.get(tipo_factura)
            
            if not tipo_cbte_afip:
                return {
                    'success': False, 
                    'error': f'Tipo de factura no soportado: {factura_original.tipo_comprobante}'
                }
            
            # Preparar comprobante asociado
            punto_vta_str, numero_str = factura_original.numero.split('-')
            cbtes_asoc = [{
                'Tipo': factura_original.tipo_comprobante,
                'PtoVta': int(punto_vta_str),
                'Nro': int(numero_str),
                'Cuit': factura_original.cliente.documento if factura_original.cliente and factura_original.cliente.tipo_documento == 'CUIT' else None
            }]
            
            print(f"📄 NC asociada a Factura: {factura_original.numero}")
            print(f"   Tipo factura: {factura_original.tipo_comprobante}")
            print(f"   Tipo NC: {tipo_cbte_afip}")
        
        else:
            return {'success': False, 'error': 'Tipo de comprobante inválido'}
        
        # ====================================================================
        # 2. VALIDAR DATOS
        # ====================================================================
        
        if not detalles:
            return {'success': False, 'error': 'No hay items en el comprobante'}
        
        # ====================================================================
        # 3. PREPARAR DATOS PARA AFIP
        # ====================================================================
        
        # Cliente
        if comprobante.cliente_id:
            cliente = Cliente.query.get(comprobante.cliente_id)
            if cliente and cliente.documento:
                if cliente.tipo_documento == 'CUIT' and len(cliente.documento) == 11:
                    tipo_doc = 80  # CUIT
                    nro_doc = int(cliente.documento)
                elif cliente.tipo_documento == 'DNI' and len(cliente.documento) >= 7:
                    tipo_doc = 96  # DNI
                    nro_doc = int(cliente.documento)
                else:
                    tipo_doc = 99  # Consumidor Final
                    nro_doc = 0
            else:
                tipo_doc = 99  # Consumidor Final
                nro_doc = 0
        else:
            tipo_doc = 99  # Consumidor Final
            nro_doc = 0
        
        # Obtener número de comprobante
        punto_vta_str, numero_str = comprobante.numero.split('-')
        numero_cbte = int(numero_str)
        
        print(f"📋 Datos preparados:")
        print(f"   Tipo: {tipo_cbte_afip}")
        print(f"   Número: {comprobante.numero}")
        print(f"   Cliente: Tipo Doc {tipo_doc}, Nro {nro_doc}")
        if cbtes_asoc:
            print(f"   Cbte Asociado: {factura_original.numero}")
        
        # ====================================================================
        # 4. PREPARAR DATOS EN FORMATO ARCAClient
        # ====================================================================
        
        # Preparar items_detalle desde los detalles de la BD
        items_detalle = []
        for detalle in detalles:
            items_detalle.append({
                'subtotal': float(detalle.subtotal),
                'iva_porcentaje': float(detalle.porcentaje_iva)
            })
        
        print(f"📦 Items procesados: {len(items_detalle)}")
        
        # Formato para ARCAClient
        datos_arca = {
            'punto_venta': int(punto_vta_str),
            'tipo_comprobante': tipo_cbte_afip,
            'doc_tipo': tipo_doc,
            'doc_nro': nro_doc,
            'items_detalle': items_detalle
        }
        
        # ✅ CRÍTICO: Agregar comprobantes asociados si existen (para NC)
        if cbtes_asoc:
            datos_arca['comprobantes_asociados'] = cbtes_asoc
            print(f"📎 Comprobantes asociados agregados: {len(cbtes_asoc)}")
        
        # ====================================================================
        # 5. LLAMAR A ARCAClient (NO AFIPService)
        # ====================================================================
        
        print(f"📡 Enviando solicitud a AFIP/ARCA...")
        
        # ✅ CORRECCIÓN: Usar ARCAClient en lugar de AFIPService
        afip_service = ARCAClient()
        
        # Autorizar comprobante
        resultado = afip_service.autorizar_comprobante(datos_arca)
        
        # ====================================================================
        # 6. PROCESAR RESPUESTA
        # ====================================================================
        
        if resultado.get('success'):
            cae = resultado['cae']
            vto_cae = resultado.get('vto_cae', resultado.get('fecha_vencimiento', ''))
            
            print(f"✅ AUTORIZADO - CAE: {cae}")
            print(f"   Vencimiento: {vto_cae}")
            
            # Actualizar comprobante
            comprobante.cae = cae
            
            # Convertir fecha de vencimiento (puede venir en formato YYYYMMDD)
            if vto_cae and len(str(vto_cae)) == 8:
                vto_str = str(vto_cae)
                comprobante.vencimiento_cae = datetime.strptime(vto_str, '%Y%m%d').date()
            else:
                # Si viene en otro formato, intentar parsearlo
                try:
                    comprobante.vencimiento_cae = datetime.strptime(str(vto_cae), '%Y%m%d').date()
                except:
                    print(f"⚠️ No se pudo parsear fecha de vencimiento: {vto_cae}")
            
            comprobante.estado = 'autorizada'
            
            db.session.commit()
            
            print("=" * 70)
            print(f"✅ {tipo_comprobante.upper()} AUTORIZADA EXITOSAMENTE")
            print("=" * 70)
            
            return {
                'success': True,
                'cae': cae,
                'vto_cae': vto_cae if isinstance(vto_cae, str) else str(vto_cae)
            }
        else:
            error_msg = resultado.get('error', 'Error desconocido')
            print(f"❌ ERROR: {error_msg}")
            print("=" * 70)
            
            return {
                'success': False,
                'error': error_msg
            }
    
    except Exception as e:
        error_msg = f"Error al autorizar en AFIP: {str(e)}"
        print(f"❌ {error_msg}")
        print("=" * 70)
        
        import traceback
        traceback.print_exc()
        
        return {
            'success': False,
            'error': error_msg
        }

# INICIALIZAR Y REGISTRAR BLUEPRINT DE NOTAS DE CRÉDITO
from notas_credito import init_notas_credito
nc_bp = init_notas_credito(
    db, NotaCredito, DetalleNotaCredito, Factura, DetalleFactura, 
    Cliente, Producto, ARCA_CONFIG, autorizar_comprobante_afip
)
app.register_blueprint(nc_bp)



# ============================================================================
# RUTA ADICIONAL: Ver Nota de Crédito
# ============================================================================

@app.route('/nota_credito/<int:nc_id>')
def ver_nota_credito(nc_id):
    """Vista de detalle de una Nota de Crédito"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    nota_credito = NotaCredito.query.get_or_404(nc_id)
    detalles = DetalleNotaCredito.query.filter_by(nota_credito_id=nc_id).all()
    
    return render_template('nota_credito_detalle.html', 
                         nota_credito=nota_credito,
                         detalles=detalles)


@app.route('/notas_credito')
def notas_credito():
    """Listado de todas las Notas de Crédito"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('notas_credito.html')


@app.route('/api/emitir_nota_credito/<int:factura_id>', methods=['POST'])
def emitir_nota_credito(factura_id):
    """Emite una Nota de Crédito electrónica ante AFIP para anular una factura."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        factura = Factura.query.get_or_404(factura_id)

        if factura.estado == 'anulada':
            return jsonify({'success': False, 'error': 'La factura ya está anulada'}), 400

        # Determinar tipo de NC según tipo de factura
        tipo_nc_map = {
            '01': ('03', 'Nota de Crédito A'),
            '1':  ('03', 'Nota de Crédito A'),
            '06': ('08', 'Nota de Crédito B'),
            '6':  ('08', 'Nota de Crédito B'),
            '11': ('13', 'Nota de Crédito C'),
            '51': ('53', 'Nota de Crédito M'),
        }
        tipo_nc_info = tipo_nc_map.get(str(factura.tipo_comprobante))
        if not tipo_nc_info:
            return jsonify({'success': False, 'error': f'Tipo de factura no soportado para NC: {factura.tipo_comprobante}'}), 400

        tipo_nc, nombre_nc = tipo_nc_info
        motivo = (request.json or {}).get('motivo', 'Anulación de factura').strip()

        print(f"\n{'='*60}")
        print(f"📝 EMITIENDO {nombre_nc}")
        print(f"   Factura: {factura.numero}")
        print(f"   Total: ${factura.total:.2f}")
        print(f"   Motivo: {motivo}")
        print(f"{'='*60}")

        # Armar items con IVA por alícuota (igual que al facturar)
        detalles = DetalleFactura.query.filter_by(factura_id=factura.id).all()
        items_detalle = []
        for d in detalles:
            iva_pct = float(d.porcentaje_iva) if hasattr(d, 'porcentaje_iva') and d.porcentaje_iva is not None else (
                float(d.producto.iva) if d.producto else 21.0
            )
            items_detalle.append({
                'subtotal': float(d.subtotal),
                'iva_porcentaje': iva_pct,
            })

        # Datos del cliente
        cliente = Cliente.query.get(factura.cliente_id)
        doc_tipo = 80 if (cliente and cliente.condicion_iva in ['RI', 'Responsable Inscripto']) else 96
        doc_nro  = int(str(cliente.documento).replace('-','').replace('.','')) if cliente and cliente.documento else 0

        # Número de factura para CbteAsoc (solo el número, sin el punto de venta)
        numero_factura_int = int(factura.numero.split('-')[1])

        datos_nc = {
            'tipo_comprobante':         int(tipo_nc),
            'punto_venta':              factura.punto_venta,
            'doc_tipo':                 doc_tipo,
            'doc_nro':                  doc_nro,
            'condicion_iva_str':        (cliente.condicion_iva if cliente else None),
            'items_detalle':            items_detalle,
            'comprobantes_asociados':   [{
                'Tipo':   int(factura.tipo_comprobante),
                'PtoVta': factura.punto_venta,
                'Nro':    numero_factura_int,
            }],
        }

        # Autorizar ante AFIP
        resultado = arca_client.autorizar_comprobante(datos_nc)

        if not resultado.get('success'):
            error_msg = resultado.get('error', 'Error desconocido de AFIP')
            print(f"❌ Error AFIP: {error_msg}")
            # Guardar NC con error
            nc = NotaCredito(
                tipo_comprobante=tipo_nc,
                punto_venta=factura.punto_venta,
                fecha=datetime.now(),
                factura_id=factura.id,
                factura_numero=factura.numero,
                cliente_id=factura.cliente_id,
                usuario_id=session['user_id'],
                subtotal=factura.subtotal,
                iva=factura.iva,
                total=factura.total,
                estado='error',
                error_afip=error_msg,
                motivo=motivo,
                fecha_creacion=datetime.now(),
            )
            db.session.add(nc)
            db.session.commit()
            return jsonify({'success': False, 'error': f'Error AFIP: {error_msg}'}), 400

        # NC autorizada
        cae         = resultado.get('cae', '')
        vto_cae_str = resultado.get('vto_cae', '')
        numero_nc   = resultado.get('numero_comprobante', '')
        pv          = factura.punto_venta
        numero_nc_fmt = f"{pv:04d}-{numero_nc:08d}" if isinstance(numero_nc, int) else str(numero_nc)

        vto_cae_date = None
        if vto_cae_str:
            try:
                vto_cae_date = datetime.strptime(str(vto_cae_str), '%Y%m%d').date()
            except:
                pass

        nc = NotaCredito(
            numero=numero_nc_fmt,
            tipo_comprobante=tipo_nc,
            punto_venta=pv,
            fecha=datetime.now(),
            factura_id=factura.id,
            factura_numero=factura.numero,
            cliente_id=factura.cliente_id,
            usuario_id=session['user_id'],
            subtotal=factura.subtotal,
            iva=factura.iva,
            total=factura.total,
            estado='autorizada',
            cae=cae,
            vto_cae=vto_cae_date,
            motivo=motivo,
            fecha_creacion=datetime.now(),
            fecha_autorizacion=datetime.now(),
        )
        db.session.add(nc)

        # Copiar detalles de la factura a la NC
        for d in detalles:
            iva_pct = float(d.porcentaje_iva) if hasattr(d, 'porcentaje_iva') and d.porcentaje_iva is not None else (
                float(d.producto.iva) if d.producto else 21.0
            )
            iva_importe = round(float(d.subtotal) * iva_pct / 100, 2)
            det_nc = DetalleNotaCredito(
                nota_credito_id=nc.id,
                producto_id=d.producto_id,
                cantidad=d.cantidad,
                precio_unitario=d.precio_unitario,
                subtotal=d.subtotal,
                porcentaje_iva=Decimal(str(iva_pct)),
                importe_iva=Decimal(str(iva_importe)),
            )
            db.session.add(det_nc)

        # ═══════════════════════════════════════════════════════════════════
        # COMMIT #1 — CRÍTICO: AFIP ya autorizó la NC con CAE (irreversible).
        # Se persiste la NC + sus detalles ANTES de cualquier paso que pueda
        # fallar. Así la NC nunca se pierde aunque falle anular la factura o
        # reintegrar stock. (Parche de seguridad — rediseño completo pendiente.)
        # ═══════════════════════════════════════════════════════════════════
        db.session.commit()
        print(f"✅ NC {numero_nc_fmt} autorizada y guardada. CAE: {cae}")

        # ═══════════════════════════════════════════════════════════════════
        # Pasos secundarios en transacción PROPIA y protegida. Si fallan,
        # la NC ya está salvada (commit #1). Se informa, NO se rompe todo.
        # ═══════════════════════════════════════════════════════════════════
        post_warning = None
        try:
            factura.estado = 'anulada'
            es_consignacion = factura.interno_origen_id is not None
            if not es_consignacion:
                for d in detalles:
                    prod = Producto.query.get(d.producto_id)
                    if prod:
                        stock_ant = float(prod.stock)
                        prod.stock += d.cantidad
                        registrar_movimiento_stock(
                            db=db, producto_id=prod.id, tipo='anulacion_nc',
                            cantidad=float(d.cantidad), signo='+',
                            stock_anterior=stock_ant, stock_nuevo=float(prod.stock),
                            referencia_tipo='nota_credito', referencia_id=None,
                            motivo=f'Reintegro por NC {numero_nc_fmt} - {motivo}',
                            usuario_id=session['user_id'],
                            usuario_nombre=session.get('nombre', 'Sistema'),
                            codigo_producto=prod.codigo, nombre_producto=prod.nombre
                        )

            # ───────────────────────────────────────────────────────────────
            # Detectar si la factura tenía movimiento venta_fiada pendiente
            # en cta_cte_movimiento. NO se modifica ese movimiento — sigue
            # siendo deuda. La NC se inserta como línea aparte (Haber) que
            # compensa el saldo global del cliente vía las fórmulas (que
            # restan tipo='nota_credito').
            #
            # Si la factura era venta al contado (sin movimiento en cta cte),
            # no se inserta línea de NC: no había deuda que compensar.
            # ───────────────────────────────────────────────────────────────
            existe_venta_fiada = db.session.execute(text("""
                SELECT COUNT(*) FROM cta_cte_movimiento
                WHERE factura_id = :factura_id
                  AND tipo = 'venta_fiada'
                  AND estado = 'pendiente'
            """), {'factura_id': factura.id}).scalar() or 0
            print(f"📒 Cta cte cliente: factura tenía {existe_venta_fiada} movimiento(s) venta_fiada pendiente(s)")

            # ───────────────────────────────────────────────────────────────
            # Insertar el movimiento de NC en cta_cte_movimiento.
            # Solo se inserta si la factura tenía movimiento de venta_fiada.
            # Si la venta era al contado, NO hay nada que compensar en cta cte.
            # Valores: tipo='nota_credito' (enum agregado por ALTER TABLE),
            #          tipo_mov='pago' (para que la pantalla la muestre en Haber),
            #          estado='pagado' (la NC no tiene saldo propio a cobrar/pagar),
            #          saldo_pendiente=0 (no entra en la suma de deuda directa,
            #                              el restador de NC en la fórmula lo cuenta).
            # ───────────────────────────────────────────────────────────────
            if existe_venta_fiada > 0:
                db.session.execute(text("""
                    INSERT INTO cta_cte_movimiento
                        (cliente_id, fecha, tipo, tipo_mov, estado, monto_total,
                         saldo_pendiente, factura_id, usuario_id, numero_comprobante,
                         observaciones)
                    VALUES
                        (:cliente_id, NOW(), 'nota_credito', 'pago', 'pagado', :monto,
                         0.00, :factura_id, :usuario_id, :nc_numero,
                         :obs)
                """), {
                    'cliente_id': factura.cliente_id,
                    'monto':      float(factura.total),
                    'factura_id': factura.id,
                    'usuario_id': session['user_id'],
                    'nc_numero':  numero_nc_fmt,
                    'obs':        f'NC {numero_nc_fmt} anula factura {factura.numero} (CAE {cae})',
                })
                print(f"📒 Cta cte cliente: movimiento NC {numero_nc_fmt} insertado por ${float(factura.total):.2f}")

            db.session.commit()
            print(f"✅ Factura {factura.numero} anulada, stock reintegrado, cta cte actualizada.")
        except Exception as e_post:
            db.session.rollback()
            import traceback; traceback.print_exc()
            post_warning = (
                f'La NC {numero_nc_fmt} se emitió y autorizó CORRECTAMENTE ante AFIP '
                f'(CAE {cae}), pero falló un paso posterior (anular factura / stock / cta cte): '
                f'{str(e_post)}. La factura {factura.numero} puede haber quedado sin '
                f'marcar como anulada o con saldo pendiente en cta cte — revisar manualmente. '
                f'La NC NO se perdió.'
            )
            print(f"⚠️ {post_warning}")

        resp = {
            'success':         True,
            'message':         f'Nota de Crédito {numero_nc_fmt} emitida y autorizada por AFIP',
            'numero_nc':       numero_nc_fmt,
            'cae':             cae,
            'vto_cae':         vto_cae_date.strftime('%d/%m/%Y') if vto_cae_date else '',
            'nc_id':           nc.id,
            'factura_anulada': (post_warning is None),
            'productos_reintegrados': (0 if post_warning else len(detalles)),
            'nota_credito': {
                'numero':  numero_nc_fmt,
                'estado':  'Autorizada',
                'cae':     cae,
                'vto_cae': vto_cae_date.strftime('%d/%m/%Y') if vto_cae_date else '',
                'total':   float(factura.total),
            },
        }
        if post_warning:
            resp['warning'] = post_warning
        return jsonify(resp)

    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/factura/<int:factura_id>/lineas_nc', methods=['GET'])
def api_factura_lineas_nc(factura_id):
    """Devuelve las líneas de una factura para el modal de NC parcial,
    marcando las que ya fueron acreditadas por una NC autorizada previa."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        factura = Factura.query.get_or_404(factura_id)

        # Producto_ids ya acreditados por NCs autorizadas de esta factura
        acreditados = set()
        filas = db.session.execute(text("""
            SELECT DISTINCT dnc.producto_id
              FROM detalle_nota_credito dnc
              JOIN notas_credito nc ON nc.id = dnc.nota_credito_id
             WHERE nc.factura_id = :fid AND nc.estado = 'autorizada'
        """), {'fid': factura_id}).fetchall()
        for f in filas:
            acreditados.add(f[0])

        detalles = DetalleFactura.query.filter_by(factura_id=factura_id).all()
        lineas = []
        for d in detalles:
            iva_pct = float(d.porcentaje_iva) if d.porcentaje_iva is not None else (
                float(d.producto.iva) if d.producto else 21.0)
            lineas.append({
                'detalle_id':      d.id,
                'producto_id':     d.producto_id,
                'nombre':          d.producto.nombre if d.producto else f'Producto {d.producto_id}',
                'codigo':          d.producto.codigo if d.producto else '',
                'cantidad':        float(d.cantidad),
                'precio_unitario': float(d.precio_unitario),
                'subtotal':        float(d.subtotal),
                'porcentaje_iva':  iva_pct,
                'ya_acreditada':   d.producto_id in acreditados,
            })

        return jsonify({
            'success': True,
            'factura': {
                'id':               factura.id,
                'numero':           factura.numero,
                'tipo_comprobante': factura.tipo_comprobante,
                'estado':           factura.estado,
            },
            'lineas': lineas,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/emitir_nota_credito_parcial/<int:factura_id>', methods=['POST'])
def emitir_nota_credito_parcial(factura_id):
    """Emite una Nota de Crédito electrónica (AFIP) por LÍNEAS COMPLETAS
    seleccionadas de la factura, SIN anular la factura. Devuelve stock de esas
    líneas y descuenta de la cuenta corriente el monto de la NC."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        factura = Factura.query.get_or_404(factura_id)

        if factura.estado == 'anulada':
            return jsonify({'success': False, 'error': 'La factura está anulada; no se puede emitir NC parcial'}), 400
        if factura.estado != 'autorizada':
            return jsonify({'success': False, 'error': 'Solo se puede emitir NC parcial sobre facturas autorizadas'}), 400

        data = request.json or {}
        detalle_ids = data.get('detalle_ids') or []
        motivo = (data.get('motivo') or 'Nota de crédito parcial').strip()
        if not detalle_ids:
            return jsonify({'success': False, 'error': 'No se seleccionó ninguna línea'}), 400

        tipo_nc_map = {
            '01': ('03', 'Nota de Crédito A'), '1': ('03', 'Nota de Crédito A'),
            '06': ('08', 'Nota de Crédito B'), '6': ('08', 'Nota de Crédito B'),
            '11': ('13', 'Nota de Crédito C'),
            '51': ('53', 'Nota de Crédito M'),
        }
        tipo_nc_info = tipo_nc_map.get(str(factura.tipo_comprobante))
        if not tipo_nc_info:
            return jsonify({'success': False, 'error': f'Tipo de factura no soportado para NC: {factura.tipo_comprobante}'}), 400
        tipo_nc, nombre_nc = tipo_nc_info

        # Líneas seleccionadas (validadas contra la factura)
        detalles_sel = DetalleFactura.query.filter(
            DetalleFactura.id.in_(detalle_ids),
            DetalleFactura.factura_id == factura.id
        ).all()
        if not detalles_sel:
            return jsonify({'success': False, 'error': 'Las líneas seleccionadas no pertenecen a la factura'}), 400

        # Bloquear líneas cuyo producto ya fue acreditado por una NC autorizada
        acreditados = set()
        for f in db.session.execute(text("""
            SELECT DISTINCT dnc.producto_id FROM detalle_nota_credito dnc
              JOIN notas_credito nc ON nc.id = dnc.nota_credito_id
             WHERE nc.factura_id = :fid AND nc.estado = 'autorizada'
        """), {'fid': factura.id}).fetchall():
            acreditados.add(f[0])
        ya = [d for d in detalles_sel if d.producto_id in acreditados]
        if ya:
            nombres = ', '.join((d.producto.nombre if d.producto else str(d.producto_id)) for d in ya)
            return jsonify({'success': False, 'error': f'Estas líneas ya tienen NC: {nombres}'}), 400

        # Totales e items para AFIP (solo líneas seleccionadas)
        items_detalle = []
        subtotal_nc = 0.0
        iva_nc = 0.0
        for d in detalles_sel:
            iva_pct = float(d.porcentaje_iva) if d.porcentaje_iva is not None else (
                float(d.producto.iva) if d.producto else 21.0)
            items_detalle.append({'subtotal': float(d.subtotal), 'iva_porcentaje': iva_pct})
            subtotal_nc += float(d.subtotal)
            iva_nc += round(float(d.subtotal) * iva_pct / 100, 2)
        subtotal_nc = round(subtotal_nc, 2)
        iva_nc = round(iva_nc, 2)
        total_nc = round(subtotal_nc + iva_nc, 2)

        cliente = Cliente.query.get(factura.cliente_id)
        # DocTipo/DocNro del receptor según su documento real.
        doc_tipo = 99
        doc_nro = 0
        if cliente and cliente.documento:
            doc_limpio = str(cliente.documento).replace('-', '').replace('.', '').strip()
            if cliente.tipo_documento == 'CUIT' and len(doc_limpio) == 11:
                doc_tipo, doc_nro = 80, int(doc_limpio)
            elif cliente.tipo_documento == 'DNI' and len(doc_limpio) >= 7:
                doc_tipo, doc_nro = 96, int(doc_limpio)
            elif len(doc_limpio) == 11 and doc_limpio.isdigit():
                doc_tipo, doc_nro = 80, int(doc_limpio)
            elif doc_limpio.isdigit():
                doc_tipo, doc_nro = 96, int(doc_limpio)
        # Comprobantes clase A (01) y M (51) exigen CUIT del receptor (DocTipo 80)
        if str(factura.tipo_comprobante) in ('1', '01', '51') and doc_nro:
            doc_tipo = 80
        numero_factura_int = int(factura.numero.split('-')[1])

        datos_nc = {
            'tipo_comprobante':       int(tipo_nc),
            'punto_venta':            factura.punto_venta,
            'doc_tipo':               doc_tipo,
            'doc_nro':                doc_nro,
            'condicion_iva_str':      (cliente.condicion_iva if cliente else None),
            'items_detalle':          items_detalle,
            'comprobantes_asociados': [{
                'Tipo':   int(factura.tipo_comprobante),
                'PtoVta': factura.punto_venta,
                'Nro':    numero_factura_int,
            }],
        }

        print(f"\n{'='*60}\n📝 EMITIENDO {nombre_nc} PARCIAL  Factura {factura.numero}  Total ${total_nc:.2f}\n{'='*60}")
        resultado = arca_client.autorizar_comprobante(datos_nc)

        if not resultado.get('success'):
            error_msg = resultado.get('error', 'Error desconocido de AFIP')
            nc = NotaCredito(
                tipo_comprobante=tipo_nc, punto_venta=factura.punto_venta, fecha=datetime.now(),
                factura_id=factura.id, factura_numero=factura.numero, cliente_id=factura.cliente_id,
                usuario_id=session['user_id'], subtotal=Decimal(str(subtotal_nc)), iva=Decimal(str(iva_nc)),
                total=Decimal(str(total_nc)), estado='error', error_afip=error_msg,
                motivo=f'[PARCIAL] {motivo}', fecha_creacion=datetime.now(),
            )
            db.session.add(nc); db.session.commit()
            return jsonify({'success': False, 'error': f'Error AFIP: {error_msg}'}), 400

        cae = resultado.get('cae', '')
        vto_cae_str = resultado.get('vto_cae', '')
        numero_nc = resultado.get('numero_comprobante', '')
        pv = factura.punto_venta
        numero_nc_fmt = f"{pv:04d}-{numero_nc:08d}" if isinstance(numero_nc, int) else str(numero_nc)
        vto_cae_date = None
        if vto_cae_str:
            try:
                vto_cae_date = datetime.strptime(str(vto_cae_str), '%Y%m%d').date()
            except Exception:
                pass

        nc = NotaCredito(
            numero=numero_nc_fmt, tipo_comprobante=tipo_nc, punto_venta=pv, fecha=datetime.now(),
            factura_id=factura.id, factura_numero=factura.numero, cliente_id=factura.cliente_id,
            usuario_id=session['user_id'], subtotal=Decimal(str(subtotal_nc)), iva=Decimal(str(iva_nc)),
            total=Decimal(str(total_nc)), estado='autorizada', cae=cae, vto_cae=vto_cae_date,
            motivo=f'[PARCIAL] {motivo}', fecha_creacion=datetime.now(), fecha_autorizacion=datetime.now(),
        )
        db.session.add(nc)
        db.session.flush()  # obtener nc.id

        for d in detalles_sel:
            iva_pct = float(d.porcentaje_iva) if d.porcentaje_iva is not None else (
                float(d.producto.iva) if d.producto else 21.0)
            det_nc = DetalleNotaCredito(
                nota_credito_id=nc.id, producto_id=d.producto_id, cantidad=d.cantidad,
                precio_unitario=d.precio_unitario, subtotal=d.subtotal,
                porcentaje_iva=Decimal(str(iva_pct)),
                importe_iva=Decimal(str(round(float(d.subtotal) * iva_pct / 100, 2))),
            )
            db.session.add(det_nc)

        # COMMIT #1: la NC ya tiene CAE (irreversible). Se persiste antes de los pasos secundarios.
        db.session.commit()
        print(f"✅ NC PARCIAL {numero_nc_fmt} autorizada. CAE: {cae}")

        post_warning = None
        try:
            # Reintegro de stock SOLO de las líneas acreditadas (la factura NO se anula)
            es_consignacion = factura.interno_origen_id is not None
            if not es_consignacion:
                for d in detalles_sel:
                    prod = Producto.query.get(d.producto_id)
                    if prod:
                        aumentar_stock_producto(
                            producto=prod, cantidad=float(d.cantidad), tipo='nota_credito_parcial',
                            referencia_tipo='nota_credito', referencia_id=None,
                            motivo=f'Reintegro por NC parcial {numero_nc_fmt} - {motivo}',
                        )

            # Cta cte: descontar de la deuda el monto de la NC (solo si la factura
            # tenía venta_fiada pendiente). La factura sigue como deuda; la NC se
            # inserta como Haber que compensa vía la fórmula (tipo='nota_credito').
            existe_venta_fiada = db.session.execute(text("""
                SELECT COUNT(*) FROM cta_cte_movimiento
                WHERE factura_id = :fid AND tipo = 'venta_fiada' AND estado = 'pendiente'
            """), {'fid': factura.id}).scalar() or 0
            if existe_venta_fiada > 0:
                db.session.execute(text("""
                    INSERT INTO cta_cte_movimiento
                        (cliente_id, fecha, tipo, tipo_mov, estado, monto_total,
                         saldo_pendiente, factura_id, usuario_id, numero_comprobante, observaciones)
                    VALUES
                        (:cliente_id, NOW(), 'nota_credito', 'pago', 'pagado', :monto,
                         0.00, :factura_id, :usuario_id, :nc_numero, :obs)
                """), {
                    'cliente_id': factura.cliente_id, 'monto': total_nc, 'factura_id': factura.id,
                    'usuario_id': session['user_id'], 'nc_numero': numero_nc_fmt,
                    'obs': f'NC parcial {numero_nc_fmt} s/ factura {factura.numero} (CAE {cae})',
                })
            db.session.commit()
            print(f"✅ NC parcial: stock reintegrado y cta cte actualizada.")
        except Exception as e_post:
            db.session.rollback()
            import traceback; traceback.print_exc()
            post_warning = (
                f'La NC parcial {numero_nc_fmt} se autorizó CORRECTAMENTE ante AFIP (CAE {cae}), '
                f'pero falló un paso posterior (stock / cta cte): {str(e_post)}. Revisar manualmente. '
                f'La NC NO se perdió.'
            )
            print(f"⚠️ {post_warning}")

        resp = {
            'success':   True,
            'message':   f'Nota de Crédito parcial {numero_nc_fmt} emitida y autorizada por AFIP',
            'numero_nc': numero_nc_fmt,
            'cae':       cae,
            'vto_cae':   vto_cae_date.strftime('%d/%m/%Y') if vto_cae_date else '',
            'nc_id':     nc.id,
            'total':     total_nc,
            'lineas_acreditadas': len(detalles_sel),
        }
        if post_warning:
            resp['warning'] = post_warning
        return jsonify(resp)

    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notas_credito/buscar', methods=['GET'])
def api_notas_credito_buscar():
    """Buscar notas de crédito con filtros."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        numero     = request.args.get('numero', '').strip()
        cliente_id = request.args.get('cliente_id', '').strip()
        estado     = request.args.get('estado', '').strip()
        desde      = request.args.get('fecha_desde', '').strip()
        hasta      = request.args.get('fecha_hasta', '').strip()

        q = NotaCredito.query
        if numero:     q = q.filter(NotaCredito.numero.ilike(f'%{numero}%'))
        if cliente_id: q = q.filter(NotaCredito.cliente_id == int(cliente_id))
        if estado:     q = q.filter(NotaCredito.estado == estado)
        if desde:      q = q.filter(NotaCredito.fecha >= datetime.strptime(desde, '%Y-%m-%d'))
        if hasta:      q = q.filter(NotaCredito.fecha <= datetime.strptime(hasta, '%Y-%m-%d').replace(hour=23, minute=59))

        notas = q.order_by(NotaCredito.fecha.desc()).limit(200).all()
        resultado = []
        for nc in notas:
            cliente = Cliente.query.get(nc.cliente_id)
            resultado.append({
                'id':             nc.id,
                'numero':         nc.numero or '—',
                'fecha':          nc.fecha.strftime('%d/%m/%Y') if nc.fecha else '',
                'cliente':        cliente.nombre if cliente else '—',
                'factura_id':     nc.factura_id,
                'factura_numero': nc.factura_numero or '—',
                'tipo_comprobante': nc.tipo_comprobante,
                'total':          float(nc.total or 0),
                'cae':            nc.cae or '',
                'estado':         nc.estado or '',
            })
        return jsonify({'success': True, 'notas': resultado})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notas_credito/estadisticas', methods=['GET'])
def api_notas_credito_estadisticas():
    """Estadísticas de notas de crédito."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        total       = NotaCredito.query.count()
        autorizadas = NotaCredito.query.filter_by(estado='autorizada').count()
        pendientes  = NotaCredito.query.filter_by(estado='pendiente').count()
        con_error   = NotaCredito.query.filter_by(estado='error').count()
        monto = db.session.query(db.func.sum(NotaCredito.total)).filter_by(estado='autorizada').scalar() or 0
        return jsonify({'success': True, 'estadisticas': {
            'total': total, 'autorizadas': autorizadas,
            'pendientes': pendientes, 'con_error': con_error,
            'monto_total': float(monto),
        }})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500





####################################################################################################################

# PASO 2: Nuevas rutas en app.py para gestionar acceso rápido

@app.route('/api/toggle_acceso_rapido/<int:producto_id>', methods=['POST'])
def toggle_acceso_rapido(producto_id):
    """Activar/desactivar producto como acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        if producto.acceso_rapido:
            # Quitar de acceso rápido
            producto.acceso_rapido = False
            producto.orden_acceso_rapido = 0
            accion = 'removido de'
        else:
            # Verificar límite de 8 productos
            productos_acceso_rapido = Producto.query.filter_by(
                acceso_rapido=True, 
                activo=True
            ).count()
            
            if productos_acceso_rapido >= 8:
                return jsonify({
                    'success': False,
                    'error': 'Máximo 8 productos permitidos en acceso rápido. Quita uno primero.'
                }), 400
            
            # Agregar a acceso rápido
            producto.acceso_rapido = True
            # Asignar orden automáticamente
            max_orden = db.session.query(func.max(Producto.orden_acceso_rapido)).scalar() or 0
            producto.orden_acceso_rapido = max_orden + 1
            accion = 'agregado a'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Producto {accion} acceso rápido',
            'acceso_rapido': producto.acceso_rapido,
            'orden': producto.orden_acceso_rapido
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error en toggle_acceso_rapido: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cambiar acceso rápido: {str(e)}'
        }), 500


@app.route('/api/reordenar_acceso_rapido', methods=['POST'])
def reordenar_acceso_rapido():
    """Reordenar productos de acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        productos_orden = data.get('productos_orden', [])
        
        # productos_orden debe ser una lista de IDs en el orden deseado
        for i, producto_id in enumerate(productos_orden):
            producto = Producto.query.get(producto_id)
            if producto and producto.acceso_rapido:
                producto.orden_acceso_rapido = i + 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Orden actualizado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error al reordenar: {str(e)}'
        }), 500


@app.route('/api/productos_acceso_rapido')
def obtener_productos_acceso_rapido():
    """Obtener productos marcados como acceso rápido"""
    try:
        productos = Producto.query.filter_by(
            acceso_rapido=True,
            activo=True
        ).order_by(
            Producto.orden_acceso_rapido.asc(),
            Producto.codigo.asc()
        ).limit(8).all()
        
        productos_data = []
        for producto in productos:
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'stock': producto.stock_dinamico,
                'iva': float(producto.iva),
                'orden': producto.orden_acceso_rapido
            })
        
        return jsonify({
            'success': True,
            'productos': productos_data,
            'total': len(productos_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



def migrar_acceso_rapido():
    """Migrar productos existentes para agregar funcionalidad de acceso rápido"""
    try:
        print("🔄 Migrando productos para acceso rápido...")
        
        # Tomar los primeros 8 productos activos como acceso rápido inicial
        productos_iniciales = Producto.query.filter_by(activo=True).limit(8).all()
        
        for i, producto in enumerate(productos_iniciales):
            producto.acceso_rapido = True
            producto.orden_acceso_rapido = i + 1
            print(f"  ✅ {producto.codigo} marcado como acceso rápido (orden {i + 1})")
        
        # Marcar el resto como NO acceso rápido
        productos_restantes = Producto.query.filter(
            ~Producto.id.in_([p.id for p in productos_iniciales])
        ).all()
        
        for producto in productos_restantes:
            producto.acceso_rapido = False
            producto.orden_acceso_rapido = 0
        
        db.session.commit()
        print(f"✅ Migración completada: {len(productos_iniciales)} productos en acceso rápido")
        
    except Exception as e:
        print(f"❌ Error en migración: {e}")
        db.session.rollback()


# PASO 5: Ruta para ejecutar migración
@app.route('/migrar_acceso_rapido', methods=['POST'])
def ejecutar_migracion_acceso_rapido():
    """Endpoint para ejecutar migración de acceso rápido"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        migrar_acceso_rapido()
        
        return jsonify({
            'success': True,
            'mensaje': 'Migración de acceso rápido completada'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error en migración: {str(e)}'
        }), 500


@app.route('/reportes')
def reportes():
    return render_template('reportes.html')

@app.route('/caja-multipv')
def caja_multipv_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('caja_multipv.html')


from flask import render_template

@app.route("/reporte_ventas_parcial")
def reporte_ventas_parcial():
    return render_template("reporte_ventas_parcial.html")

# ==================== RUTAS DE GASTOS ====================
# Agregar estas rutas al final de tu app.py antes del if __name__ == '__main__':

@app.route('/api/gastos', methods=['GET'])
def obtener_gastos():
    try:
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({'success': False, 'error': 'Fechas requeridas'})
        
        # AGREGA ESTA IMPORTACIÓN
        from sqlalchemy import text
        
        query = """
        SELECT g.*, 
               c.estado as estado_caja,
               CASE WHEN c.estado = 'cerrada' THEN 1 ELSE 0 END as caja_cerrada
        FROM gastos g
        LEFT JOIN cajas c ON g.caja_id = c.id
        WHERE DATE(g.fecha) BETWEEN :fecha_desde AND :fecha_hasta
        ORDER BY g.fecha DESC
        """
        
        # CAMBIA ESTA PARTE
        result = db.session.execute(text(query), {'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta})
        gastos = [dict(row._mapping) for row in result]
        
        # El resto de tu código sigue igual
        return jsonify({
            'success': True,
            'gastos': gastos
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/gastos', methods=['POST'])
def crear_gasto():
    """Crear un nuevo gasto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['descripcion', 'monto', 'fecha']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }), 400
        
        # Validar monto
        try:
            monto = float(data['monto'])
            if monto <= 0:
                return jsonify({
                    'success': False,
                    'error': 'El monto debe ser mayor a 0'
                }), 400
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Monto inválido'
            }), 400
        
        # Validar fecha
        try:
            fecha_obj = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        # Datos del gasto
        descripcion = data['descripcion'].strip()
        categoria = data.get('categoria', 'general')
        metodo_pago = data.get('metodo_pago', 'efectivo')
        notas = data.get('notas', '').strip()
        
        # Validar longitud de campos
        if len(descripcion) > 200:
            return jsonify({
                'success': False,
                'error': 'La descripción no puede exceder 200 caracteres'
            }), 400
        
        # Obtener caja abierta actual - CORREGIDO
        caja_abierta_query = db.session.execute(text("""
            SELECT id FROM cajas 
            WHERE estado = 'abierta' 
            ORDER BY fecha_apertura DESC 
            LIMIT 1
        """)).fetchone()
        
        caja_abierta_id = caja_abierta_query[0] if caja_abierta_query else None

        # Crear nuevo gasto
        gasto = Gasto(
            fecha=fecha_obj,
            descripcion=descripcion,
            monto=Decimal(str(monto)),
            categoria=categoria,
            metodo_pago=metodo_pago,
            notas=notas if notas else None,
            usuario_id=session['user_id'],
            caja_id=caja_abierta_id  # LÍNEA CORREGIDA
        )
        
        db.session.add(gasto)
        db.session.commit()
        
        print(f"✅ Gasto creado: ID {gasto.id} - {descripcion} - ${monto:.2f}")
        
        return jsonify({
            'success': True,
            'message': 'Gasto registrado exitosamente',
            'gasto': gasto.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al crear gasto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/<int:gasto_id>/estado', methods=['GET'])
def verificar_estado_gasto(gasto_id):
    try:
        # Consultar gasto con información de la caja
        query = """
        SELECT g.*, c.estado as estado_caja, c.fecha_cierre 
        FROM gastos g
        LEFT JOIN cajas c ON g.caja_id = c.id
        WHERE g.id = %s
        """
        resultado = db.execute(query, (gasto_id,))
        
        if not resultado:
            return jsonify({'success': False, 'error': 'Gasto no encontrado'})
        
        gasto = resultado[0]
        
        response = {
            'success': True,
            'caja_cerrada': gasto['estado_caja'] == 'cerrada',
            'caja_id': gasto['caja_id'],
            'estado_caja': gasto['estado_caja'],
            'puede_eliminar': gasto['estado_caja'] != 'cerrada',
            'monto': gasto['monto'],
            'fecha': gasto['fecha']
        }
        
        if gasto['estado_caja'] == 'cerrada':
            response['razon'] = 'El gasto pertenece a una caja cerrada'
            
        return jsonify(response)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/gastos/<int:gasto_id>', methods=['DELETE'])
def eliminar_gasto(gasto_id):
    """Eliminar un gasto - Solo si la caja está abierta"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    
    try:
        gasto = Gasto.query.get(gasto_id)
        
        if not gasto:
            return jsonify({'success': False, 'error': 'Gasto no encontrado'}), 404
        
        # ✅ SOLUCIÓN: Consultar directamente con SQL en lugar de usar el modelo importado
        if gasto.caja_id:
            # Consultar el estado de la caja directamente desde la tabla
            from sqlalchemy import text
            
            query = text("SELECT estado FROM cajas WHERE id = :caja_id")
            result = db.session.execute(query, {'caja_id': gasto.caja_id}).fetchone()
            
            if result and result[0] == 'cerrada':
                return jsonify({
                    'success': False, 
                    'error': 'No se puede eliminar gastos de cajas cerradas'
                }), 403
        
        # Eliminar el gasto
        db.session.delete(gasto)
        db.session.commit()
        
        print(f"✅ Gasto {gasto_id} eliminado correctamente")
        
        return jsonify({
            'success': True, 
            'message': 'Gasto eliminado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al eliminar gasto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'error': str(e)
        }), 500
        
                

@app.route('/api/gastos/<int:gasto_id>', methods=['PUT'])
def actualizar_gasto(gasto_id):
    """Actualizar un gasto existente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.get_json()
        gasto = Gasto.query.get_or_404(gasto_id)
        
        if not gasto.activo:
            return jsonify({
                'success': False,
                'error': 'No se puede actualizar un gasto eliminado'
            }), 400
        
        # Actualizar campos proporcionados
        if 'descripcion' in data:
            descripcion = data['descripcion'].strip()
            if not descripcion:
                return jsonify({
                    'success': False,
                    'error': 'La descripción no puede estar vacía'
                }), 400
            if len(descripcion) > 200:
                return jsonify({
                    'success': False,
                    'error': 'La descripción no puede exceder 200 caracteres'
                }), 400
            gasto.descripcion = descripcion
        
        if 'monto' in data:
            try:
                monto = float(data['monto'])
                if monto <= 0:
                    return jsonify({
                        'success': False,
                        'error': 'El monto debe ser mayor a 0'
                    }), 400
                gasto.monto = Decimal(str(monto))
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Monto inválido'
                }), 400
        
        if 'fecha' in data:
            try:
                gasto.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido'
                }), 400
        
        if 'categoria' in data:
            gasto.categoria = data['categoria']
        
        if 'metodo_pago' in data:
            gasto.metodo_pago = data['metodo_pago']
        
        if 'notas' in data:
            notas = data['notas'].strip()
            gasto.notas = notas if notas else None
        
        gasto.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        print(f"✅ Gasto actualizado: ID {gasto_id}")
        
        return jsonify({
            'success': True,
            'message': 'Gasto actualizado exitosamente',
            'gasto': gasto.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al actualizar gasto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/categorias', methods=['GET'])
def obtener_categorias_gastos():
    """Obtener categorías de gastos disponibles y estadísticas de uso"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener categorías disponibles
        categorias_disponibles = Gasto.obtener_categorias_disponibles()
        
        # Obtener estadísticas de uso de categorías
        categorias_usadas = db.session.query(
            Gasto.categoria,
            func.count(Gasto.id).label('cantidad'),
            func.sum(Gasto.monto).label('total')
        ).filter(
            Gasto.activo == True
        ).group_by(Gasto.categoria).order_by(desc('total')).all()
        
        # Combinar información
        categorias_con_stats = []
        for cat_info in categorias_disponibles:
            codigo = cat_info['codigo']
            
            # Buscar estadísticas para esta categoría
            stats = next((cat for cat in categorias_usadas if cat[0] == codigo), None)
            
            categoria_completa = {
                'codigo': codigo,
                'nombre': cat_info['nombre'],
                'cantidad_gastos': stats[1] if stats else 0,
                'total_gastado': float(stats[2]) if stats and stats[2] else 0.0,
                'en_uso': bool(stats)
            }
            
            categorias_con_stats.append(categoria_completa)
        
        return jsonify({
            'success': True,
            'categorias': categorias_con_stats
        })
        
    except Exception as e:
        print(f"❌ Error al obtener categorías: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/medios_pago', methods=['GET'])
def obtener_medios_pago_gastos():
    """Obtener medios de pago disponibles para gastos"""
    try:
        # Usar los mismos medios de pago que las facturas
        medios_disponibles = MedioPago.obtener_medios_disponibles()
        
        return jsonify({
            'success': True,
            'medios_pago': medios_disponibles
        })
        
    except Exception as e:
        print(f"❌ Error al obtener medios de pago: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/resumen_periodo')
def resumen_gastos_periodo():
    """Obtener resumen de gastos para un período específico"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar fechas desde y hasta'
            }), 400
        
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        
        # Obtener resumen usando método del modelo
        resumen = Gasto.calcular_gastos_por_fecha(fecha_desde_dt, fecha_hasta_dt)
        
        if not resumen:
            return jsonify({
                'success': False,
                'error': 'Error al calcular resumen'
            }), 500
        
        # Obtener gastos por medio de pago
        gastos_por_medio = Gasto.obtener_gastos_por_medio_pago(fecha_desde_dt, fecha_hasta_dt)
        
        # Obtener gasto promedio por día
        dias_periodo = (fecha_hasta_dt - fecha_desde_dt).days + 1
        promedio_diario = resumen['total_general'] / dias_periodo if dias_periodo > 0 else 0
        
        return jsonify({
            'success': True,
            'resumen': {
                'total_general': resumen['total_general'],
                'gastos_por_categoria': resumen['gastos_por_categoria'],
                'gastos_por_medio_pago': gastos_por_medio,
                'periodo': {
                    'desde': fecha_desde,
                    'hasta': fecha_hasta,
                    'dias': dias_periodo
                },
                'promedios': {
                    'diario': round(promedio_diario, 2)
                }
            }
        })
        
    except Exception as e:
        print(f"❌ Error en resumen de período: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@app.route('/api/gastos/test_conexion', methods=['GET'])
def test_conexion_gastos():
    """Probar que la tabla de gastos funciona correctamente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Contar gastos activos
        total_gastos = Gasto.query.filter_by(activo=True).count()
        
        # Obtener último gasto
        ultimo_gasto = Gasto.query.filter_by(activo=True).order_by(Gasto.id.desc()).first()
        
        return jsonify({
            'success': True,
            'message': 'Conexión con gastos exitosa',
            'total_gastos': total_gastos,
            'ultimo_gasto': ultimo_gasto.to_dict() if ultimo_gasto else None,
            'categorias_disponibles': len(Gasto.obtener_categorias_disponibles())
        })
        
    except Exception as e:
        print(f"❌ Error test conexión gastos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/producto_precio_oferta/<int:producto_id>/<float:cantidad>')
def obtener_precio_con_oferta_api(producto_id, cantidad):
    """API para obtener precio con oferta aplicada"""
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        precio_final = producto.obtener_precio_con_oferta(cantidad)
        info_oferta = producto.obtener_info_oferta(cantidad)
        
        return jsonify({
            'success': True,
            'precio_base': float(producto.precio),
            'precio_final': precio_final,
            'cantidad': cantidad,
            'info_oferta': info_oferta,
            'producto': {
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'stock': producto.stock
            }
        })
        
    except Exception as e:
        print(f"Error en API precio oferta: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ofertas_producto/<int:producto_id>')
def obtener_ofertas_producto(producto_id):
    """Obtener todas las ofertas de un producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        producto = Producto.query.get_or_404(producto_id)
        
        ofertas = OfertaVolumen.query.filter_by(
            producto_id=producto_id,
            activo=True
        ).order_by(OfertaVolumen.cantidad_minima.asc()).all()
        
        ofertas_data = [oferta.to_dict() for oferta in ofertas]
        
        return jsonify({
            'success': True,
            'producto': {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio_base': float(producto.precio)
            },
            'ofertas': ofertas_data,
            'tiene_ofertas': len(ofertas_data) > 0
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/crear_oferta_volumen', methods=['POST'])
def crear_oferta_volumen():
    """Crear nueva oferta por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        descripcion = data.get('descripcion', '').strip()
        
        # Validaciones
        if not producto_id:
            return jsonify({'error': 'Producto requerido'}), 400
        
        if cantidad_minima <= 0:
            return jsonify({'error': 'La cantidad mínima debe ser mayor a 0'}), 400
        
        if precio_oferta <= 0:
            return jsonify({'error': 'El precio de oferta debe ser mayor a 0'}), 400
        
        # Verificar que el producto existe
        producto = Producto.query.get_or_404(producto_id)
        
        # Verificar que el precio de oferta sea menor al precio normal
        if precio_oferta >= float(producto.precio):
            return jsonify({
                'error': f'El precio de oferta (${precio_oferta}) debe ser menor al precio normal (${producto.precio})'
            }), 400
        
        # Verificar que no exista una oferta igual
        oferta_existente = OfertaVolumen.query.filter_by(
            producto_id=producto_id,
            cantidad_minima=cantidad_minima,
            activo=True
        ).first()
        
        if oferta_existente:
            return jsonify({
                'error': f'Ya existe una oferta para cantidad mínima {cantidad_minima}'
            }), 400
        
        # Crear oferta
        nueva_oferta = OfertaVolumen(
            producto_id=producto_id,
            cantidad_minima=Decimal(str(cantidad_minima)),
            precio_oferta=Decimal(str(precio_oferta)),
            descripcion=descripcion if descripcion else None
        )
        
        db.session.add(nueva_oferta)
        db.session.commit()
        
        print(f"Oferta creada: {producto.codigo} - {cantidad_minima}+ = ${precio_oferta}")
        
        return jsonify({
            'success': True,
            'message': f'Oferta creada: desde {cantidad_minima} unidades a ${precio_oferta}',
            'oferta': nueva_oferta.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error creando oferta: {e}")
        return jsonify({'error': f'Error al crear oferta: {str(e)}'}), 500

############# RUTAS DE OFERTAS
@app.route('/api/eliminar_oferta_volumen/<int:oferta_id>', methods=['DELETE'])
def eliminar_oferta_volumen(oferta_id):
    """Eliminar oferta por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        oferta = OfertaVolumen.query.get_or_404(oferta_id)
        
        # Marcar como inactiva en lugar de eliminar
        oferta.activo = False
        oferta.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Oferta eliminada correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar oferta: {str(e)}'}), 500

# ==================== RUTAS ADICIONALES PARA OFERTAS POR VOLUMEN ====================
# Agregar estas rutas después de las rutas existentes de ofertas en tu app.py

@app.route('/ofertas_volumen')
def ofertas_volumen():
    """Vista principal para gestionar ofertas por volumen"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('ofertas_volumen.html')


@app.route('/api/ofertas_volumen_todas')
def obtener_todas_ofertas_volumen():
    """Obtener todas las ofertas por volumen con información del producto"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Consulta con JOIN para obtener ofertas con información del producto
        ofertas = db.session.query(OfertaVolumen).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            Producto.activo == True  # Solo productos activos
        ).order_by(
            OfertaVolumen.fecha_creacion.desc()
        ).all()
        
        ofertas_data = []
        for oferta in ofertas:
            oferta_dict = oferta.to_dict()
            
            # Agregar información completa del producto
            if oferta.producto:
                oferta_dict['producto'] = {
                    'id': oferta.producto.id,
                    'codigo': oferta.producto.codigo,
                    'nombre': oferta.producto.nombre,
                    'precio': float(oferta.producto.precio),
                    'stock': oferta.producto.stock,
                    'categoria': oferta.producto.categoria,
                    'activo': oferta.producto.activo
                }
            else:
                oferta_dict['producto'] = None
            
            ofertas_data.append(oferta_dict)
        
        print(f"📊 Devolviendo {len(ofertas_data)} ofertas por volumen")
        
        return jsonify({
            'success': True,
            'ofertas': ofertas_data,
            'total': len(ofertas_data)
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo ofertas: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/actualizar_oferta_volumen/<int:oferta_id>', methods=['PUT'])
def actualizar_oferta_volumen(oferta_id):
    """Actualizar una oferta por volumen existente"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        # Validaciones
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        descripcion = data.get('descripcion', '').strip()
        
        if cantidad_minima <= 0:
            return jsonify({'error': 'La cantidad mínima debe ser mayor a 0'}), 400
        
        if precio_oferta <= 0:
            return jsonify({'error': 'El precio de oferta debe ser mayor a 0'}), 400
        
        # Obtener oferta existente
        oferta = OfertaVolumen.query.get_or_404(oferta_id)
        
        # Verificar que el precio de oferta sea menor al precio normal
        if precio_oferta >= float(oferta.producto.precio):
            return jsonify({
                'error': f'El precio de oferta (${precio_oferta}) debe ser menor al precio normal (${oferta.producto.precio})'
            }), 400
        
        # Verificar que no exista otra oferta igual (excluyendo la actual)
        oferta_duplicada = OfertaVolumen.query.filter(
            and_(
                OfertaVolumen.producto_id == oferta.producto_id,
                OfertaVolumen.cantidad_minima == cantidad_minima,
                OfertaVolumen.id != oferta_id,
                OfertaVolumen.activo == True
            )
        ).first()
        
        if oferta_duplicada:
            return jsonify({
                'error': f'Ya existe otra oferta para cantidad mínima {cantidad_minima}'
            }), 400
        
        # Actualizar oferta
        oferta.cantidad_minima = Decimal(str(cantidad_minima))
        oferta.precio_oferta = Decimal(str(precio_oferta))
        oferta.descripcion = descripcion if descripcion else None
        oferta.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        print(f"Oferta actualizada: {oferta.producto.codigo} - {cantidad_minima}+ = ${precio_oferta}")
        
        return jsonify({
            'success': True,
            'message': f'Oferta actualizada: desde {cantidad_minima} unidades a ${precio_oferta}',
            'oferta': oferta.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error actualizando oferta: {e}")
        return jsonify({'error': f'Error al actualizar oferta: {str(e)}'}), 500


@app.route('/api/ofertas_activas_resumen')
def obtener_resumen_ofertas_activas():
    """Obtener resumen de ofertas activas para dashboard"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Contar ofertas activas
        ofertas_activas = OfertaVolumen.query.filter_by(activo=True).count()
        
        # Contar productos con ofertas
        productos_con_ofertas = db.session.query(OfertaVolumen.producto_id).filter_by(activo=True).distinct().count()
        
        # Obtener descuento promedio
        ofertas_con_descuento = db.session.query(
            OfertaVolumen.precio_oferta,
            Producto.precio
        ).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            OfertaVolumen.activo == True
        ).all()
        
        descuento_promedio = 0
        if ofertas_con_descuento:
            descuentos = []
            for precio_oferta, precio_normal in ofertas_con_descuento:
                if float(precio_normal) > 0:
                    descuento = ((float(precio_normal) - float(precio_oferta)) / float(precio_normal)) * 100
                    descuentos.append(descuento)
            
            if descuentos:
                descuento_promedio = sum(descuentos) / len(descuentos)
        
        return jsonify({
            'success': True,
            'ofertas_activas': ofertas_activas,
            'productos_con_ofertas': productos_con_ofertas,
            'descuento_promedio': round(descuento_promedio, 1)
        })
        
    except Exception as e:
        print(f"Error obteniendo resumen de ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/validar_oferta_volumen', methods=['POST'])
def validar_oferta_volumen():
    """Validar datos de una oferta antes de crearla/actualizarla"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        
        producto_id = data.get('producto_id')
        cantidad_minima = float(data.get('cantidad_minima', 0))
        precio_oferta = float(data.get('precio_oferta', 0))
        oferta_id = data.get('oferta_id')  # Para edición
        
        errores = []
        
        # Validar producto
        if not producto_id:
            errores.append('Debe seleccionar un producto')
        else:
            producto = Producto.query.get(producto_id)
            if not producto:
                errores.append('Producto no encontrado')
            elif not producto.activo:
                errores.append('El producto está inactivo')
            else:
                # Validar precio
                if precio_oferta >= float(producto.precio):
                    errores.append(f'El precio de oferta debe ser menor a ${float(producto.precio):.2f}')
                
                # Verificar duplicados
                query = OfertaVolumen.query.filter(
                    and_(
                        OfertaVolumen.producto_id == producto_id,
                        OfertaVolumen.cantidad_minima == cantidad_minima,
                        OfertaVolumen.activo == True
                    )
                )
                
                if oferta_id:  # Excluir la oferta actual si es edición
                    query = query.filter(OfertaVolumen.id != oferta_id)
                
                if query.first():
                    errores.append(f'Ya existe una oferta para cantidad mínima {cantidad_minima}')
        
        # Validaciones básicas
        if cantidad_minima <= 0:
            errores.append('La cantidad mínima debe ser mayor a 0')
        
        if precio_oferta <= 0:
            errores.append('El precio de oferta debe ser mayor a 0')
        
        # Calcular información de la oferta si es válida
        info_oferta = None
        if not errores and producto:
            precio_normal = float(producto.precio)
            descuento_porcentaje = ((precio_normal - precio_oferta) / precio_normal) * 100
            ahorro_total = (precio_normal - precio_oferta) * cantidad_minima
            
            info_oferta = {
                'precio_normal': precio_normal,
                'precio_oferta': precio_oferta,
                'cantidad_minima': cantidad_minima,
                'descuento_porcentaje': round(descuento_porcentaje, 1),
                'ahorro_por_unidad': round(precio_normal - precio_oferta, 2),
                'ahorro_total': round(ahorro_total, 2)
            }
        
        return jsonify({
            'success': len(errores) == 0,
            'errores': errores,
            'info_oferta': info_oferta
        })
        
    except Exception as e:
        print(f"Error validando oferta: {e}")
        return jsonify({
            'success': False,
            'errores': [f'Error de validación: {str(e)}']
        }), 500


@app.route('/api/productos_sin_ofertas')
def obtener_productos_sin_ofertas():
    """Obtener productos que no tienen ofertas por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Productos que NO tienen ofertas activas
        productos_con_ofertas = db.session.query(OfertaVolumen.producto_id).filter_by(activo=True).distinct().subquery()
        
        productos_sin_ofertas = Producto.query.filter(
            and_(
                Producto.activo == True,
                ~Producto.id.in_(productos_con_ofertas)
            )
        ).order_by(Producto.codigo).limit(20).all()
        
        productos_data = []
        for producto in productos_sin_ofertas:
            productos_data.append({
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'stock': producto.stock_dinamico,
                'categoria': producto.categoria
            })
        
        return jsonify({
            'success': True,
            'productos': productos_data,
            'mensaje': f'Se encontraron {len(productos_data)} productos sin ofertas'
        })
        
    except Exception as e:
        print(f"Error obteniendo productos sin ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/estadisticas_ofertas_volumen')
def obtener_estadisticas_ofertas():
    """Obtener estadísticas detalladas de las ofertas por volumen"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Estadísticas básicas
        total_ofertas = OfertaVolumen.query.count()
        ofertas_activas = OfertaVolumen.query.filter_by(activo=True).count()
        ofertas_inactivas = total_ofertas - ofertas_activas
        
        # Productos con ofertas
        productos_con_ofertas = db.session.query(
            OfertaVolumen.producto_id
        ).filter_by(activo=True).distinct().count()
        
        # Ofertas por rango de descuento
        ofertas_con_productos = db.session.query(
            OfertaVolumen.precio_oferta,
            Producto.precio.label('precio_normal')
        ).join(
            Producto, OfertaVolumen.producto_id == Producto.id
        ).filter(
            OfertaVolumen.activo == True
        ).all()
        
        rangos_descuento = {
            'menos_10': 0,      # Menos del 10%
            'entre_10_20': 0,   # Entre 10% y 20%
            'entre_20_30': 0,   # Entre 20% y 30%
            'mas_30': 0         # Más del 30%
        }
        
        descuentos = []
        for precio_oferta, precio_normal in ofertas_con_productos:
            if float(precio_normal) > 0:
                descuento = ((float(precio_normal) - float(precio_oferta)) / float(precio_normal)) * 100
                descuentos.append(descuento)
                
                if descuento < 10:
                    rangos_descuento['menos_10'] += 1
                elif descuento < 20:
                    rangos_descuento['entre_10_20'] += 1
                elif descuento < 30:
                    rangos_descuento['entre_20_30'] += 1
                else:
                    rangos_descuento['mas_30'] += 1
        
        # Calcular estadísticas de descuentos
        descuento_promedio = sum(descuentos) / len(descuentos) if descuentos else 0
        descuento_minimo = min(descuentos) if descuentos else 0
        descuento_maximo = max(descuentos) if descuentos else 0
        
        # Top 5 productos con más ofertas
        top_productos = db.session.query(
            Producto.codigo,
            Producto.nombre,
            func.count(OfertaVolumen.id).label('cantidad_ofertas')
        ).join(
            OfertaVolumen, Producto.id == OfertaVolumen.producto_id
        ).filter(
            OfertaVolumen.activo == True
        ).group_by(
            Producto.id, Producto.codigo, Producto.nombre
        ).order_by(
            func.count(OfertaVolumen.id).desc()
        ).limit(5).all()
        
        top_productos_data = []
        for codigo, nombre, cantidad in top_productos:
            top_productos_data.append({
                'codigo': codigo,
                'nombre': nombre,
                'cantidad_ofertas': cantidad
            })
        
        return jsonify({
            'success': True,
            'estadisticas': {
                'totales': {
                    'total_ofertas': total_ofertas,
                    'ofertas_activas': ofertas_activas,
                    'ofertas_inactivas': ofertas_inactivas,
                    'productos_con_ofertas': productos_con_ofertas
                },
                'descuentos': {
                    'promedio': round(descuento_promedio, 1),
                    'minimo': round(descuento_minimo, 1),
                    'maximo': round(descuento_maximo, 1),
                    'rangos': rangos_descuento
                },
                'top_productos': top_productos_data
            }
        })
        
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/productos_con_ofertas_volumen')
def obtener_productos_con_ofertas_volumen():
    """Obtener todos los productos que tienen ofertas por volumen activas"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener productos con ofertas por volumen
        productos_con_ofertas = db.session.query(Producto).join(
            OfertaVolumen, Producto.id == OfertaVolumen.producto_id
        ).filter(
            and_(
                Producto.activo == True,
                OfertaVolumen.activo == True
            )
        ).distinct().all()
        
        resultado = {}
        
        for producto in productos_con_ofertas:
            ofertas = OfertaVolumen.query.filter_by(
                producto_id=producto.id,
                activo=True
            ).order_by(OfertaVolumen.cantidad_minima.asc()).all()
            
            resultado[str(producto.id)] = {
                'producto': {
                    'id': producto.id,
                    'codigo': producto.codigo,
                    'nombre': producto.nombre,
                    'precio': float(producto.precio),
                    'precio_base': float(producto.precio)
                },
                'ofertas': [oferta.to_dict() for oferta in ofertas]
            }
        
        return jsonify({
            'success': True,
            'productos_ofertas': resultado,
            'total_productos': len(resultado)
        })
        
    except Exception as e:
        print(f"Error obteniendo productos con ofertas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Agregar esta ruta en tu archivo principal de Flask (app.py)

@app.route('/estadisticas')
def estadisticas():
    """Página de estadísticas y reportes avanzados"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        return render_template('estadisticas.html')
    except Exception as e:
        flash(f'Error cargando estadísticas: {str(e)}')
        return redirect(url_for('index'))


@app.route('/api/descuento_factura/<int:factura_id>')
def obtener_descuento_factura(factura_id):
    """Obtener información del descuento aplicado a una factura"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        descuento = DescuentoFactura.query.filter_by(factura_id=factura_id).first()
        
        #print(f"🔍 DEBUG API: Buscando descuento para factura {factura_id}")
        #print(f"🔍 DEBUG API: Descuento encontrado: {bool(descuento)}")
        
        if descuento:
            #print(f"🔍 DEBUG API: Descuento: {descuento.porcentaje_descuento}% = ${descuento.monto_descuento}")
            return jsonify({
                'success': True,
                'tiene_descuento': True,
                'descuento': descuento.to_dict()
            })
        else:
            return jsonify({
                'success': True,
                'tiene_descuento': False
            })
    except Exception as e:
       # print(f"❌ DEBUG API: Error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/test_afip_debug')
def test_afip_debug():
    """Test público para debug AFIP (sin autenticación)"""
    try:
        # CORREGIR LA URL - NO DUPLICAR ?wsdl
        wsaa_base_url = ARCA_CONFIG.WSAA_URL
        if wsaa_base_url.endswith('?wsdl'):
            wsaa_url = wsaa_base_url
        else:
            wsaa_url = wsaa_base_url + '?wsdl'
        
        print(f"URL corregida: {wsaa_url}")
        
        # Crear sesión igual que el diagnóstico
        session_afip = crear_session_afip()
        
        # Hacer petición HTTP directa
        response = session_afip.get(wsaa_url, timeout=15)
        
        print(f"Status Code: {response.status_code}")
        print(f"Content-Type: {response.headers.get('content-type')}")
        print(f"Primeros 200 caracteres:")
        print(response.text[:200])
        
        # Verificar tipo de contenido
        content_lower = response.text.lower()
        is_html = '<h1>' in content_lower or 'axis service' in content_lower
        is_xml = '<?xml' in response.text[:100] or 'wsdl:definitions' in response.text[:500]
        
        return jsonify({
            'success': is_xml and not is_html,
            'status_code': response.status_code,
            'content_type': response.headers.get('content-type'),
            'is_html': is_html,
            'is_xml': is_xml,
            'content_preview': response.text[:300],
            'mensaje': 'AFIP devolviendo HTML - Servicio no disponible' if is_html else ('WSDL XML válido' if is_xml else 'Contenido inesperado'),
            'url_corregida': wsaa_url,
            'problema_url_duplicada': '?wsdl?wsdl' in wsaa_url
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'config_url': ARCA_CONFIG.WSAA_URL
        })


@app.route('/debug_certificados')
def debug_certificados():
    """Debug de certificados - TEMPORAL"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        resultado = arca_client.debug_certificados()
        
        return jsonify({
            'success': resultado,
            'mensaje': 'Debug completado - revisar logs del servidor'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

###### ruta de debug oara comparar stock
@app.route('/api/comparar_stocks')
def comparar_stocks():
    """Comparar stock actual vs stock dinámico"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Solo combos para comparar
        combos = Producto.query.filter_by(es_combo=True, activo=True).all()
        
        comparaciones = []
        for combo in combos:
            comparaciones.append({
                'codigo': combo.codigo,
                'nombre': combo.nombre,
                'stock_actual': float(combo.stock),
                'stock_dinamico': combo.stock_dinamico,
                'diferencia': combo.stock_dinamico - float(combo.stock),
                'necesita_ajuste': combo.stock_dinamico != float(combo.stock)
            })
        
        return jsonify({
            'success': True,
            'comparaciones': comparaciones,
            'combos_con_diferencias': len([c for c in comparaciones if c['necesita_ajuste']])
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/carteles_precios')
def carteles_precios():
    """Vista principal para imprimir carteles de precios"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('carteles_precios.html')

@app.route('/api/productos_para_carteles')
def api_productos_para_carteles():
    """API para obtener productos con información para carteles"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener parámetros de filtro
        buscar = request.args.get('buscar', '').strip()
        categoria = request.args.get('categoria', '').strip()
        estado = request.args.get('estado', 'activo')
        ofertas = request.args.get('ofertas', 'todos')
        desde_fecha = request.args.get('desde_fecha', '').strip()
        
        # Construir query base
        query = Producto.query
        
        # Aplicar filtros
        if buscar:
            query = query.filter(
                or_(
                    Producto.codigo.ilike(f'%{buscar}%'),
                    Producto.nombre.ilike(f'%{buscar}%'),
                    Producto.descripcion.ilike(f'%{buscar}%')
                )
            )
        
        if categoria:
            query = query.filter(Producto.categoria == categoria)
        
        if estado == 'activo':
            query = query.filter(Producto.activo == True)
        
        # Filtro por fecha de actualización de precio
        if desde_fecha:
            try:
                fecha_dt = datetime.strptime(desde_fecha, '%Y-%m-%d')
                query = query.filter(Producto.fecha_actualizacion_precio >= fecha_dt)
            except ValueError:
                pass
        
        # Filtro de ofertas
        if ofertas == 'con_ofertas':
            # Productos con ofertas por volumen O combos
            query = query.filter(
                or_(
                    Producto.es_combo == True,
                    Producto.id.in_(
                        db.session.query(OfertaVolumen.producto_id).filter(
                            OfertaVolumen.activo == True
                        ).distinct()
                    )
                )
            )
        elif ofertas == 'sin_ofertas':
            # Productos SIN ofertas y que NO sean combos
            query = query.filter(
                and_(
                    Producto.es_combo == False,
                    ~Producto.id.in_(
                        db.session.query(OfertaVolumen.producto_id).filter(
                            OfertaVolumen.activo == True
                        ).distinct()
                    )
                )
            )
        
        # Obtener resultados
        productos = query.order_by(Producto.codigo).all()
        
        # Formatear respuesta
        resultado = []
        for producto in productos:
            # Verificar si tiene ofertas
            tiene_ofertas = producto.tiene_ofertas_volumen()
            
            producto_dict = {
                'id': producto.id,
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio': float(producto.precio),
                'stock_dinamico': producto.stock_dinamico,
                'categoria': producto.categoria,
                'activo': producto.activo,
                'es_combo': producto.es_combo,
                'tiene_ofertas': tiene_ofertas,
                'ahorro_combo': producto.calcular_ahorro_combo() if producto.es_combo else 0
            }
            
            resultado.append(producto_dict)
        
        return jsonify({
            'success': True,
            'productos': resultado,
            'total': len(resultado)
        })
        
    except Exception as e:
        print(f"Error en productos_para_carteles: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/imprimir_carteles', methods=['POST'])
def imprimir_carteles():
    """Imprimir carteles de precios en impresora térmica"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        data = request.json
        productos_ids = data.get('productos_ids', [])
        
        if not productos_ids:
            return jsonify({
                'success': False,
                'error': 'No se especificaron productos'
            }), 400
        
        # Obtener productos
        productos = Producto.query.filter(Producto.id.in_(productos_ids)).all()
        
        if not productos:
            return jsonify({
                'success': False,
                'error': 'No se encontraron productos'
            }), 400
        
        # Generar e imprimir carteles
        carteles_impresos = 0
        
        for producto in productos:
            try:
                # Verificar si tiene ofertas
                tiene_ofertas = producto.tiene_ofertas_volumen() or producto.es_combo
                
                # Generar cartel
                resultado = impresora_termica.imprimir_cartel_precio(producto, tiene_ofertas)
                
                if resultado:
                    carteles_impresos += 1
                    print(f"Cartel impreso: {producto.codigo} - {producto.nombre}")
                else:
                    print(f"Error imprimiendo cartel: {producto.codigo}")
                    
            except Exception as e:
                print(f"Error imprimiendo producto {producto.codigo}: {e}")
        
        return jsonify({
            'success': True,
            'carteles_impresos': carteles_impresos,
            'total_solicitados': len(productos),
            'mensaje': f'Se imprimieron {carteles_impresos} de {len(productos)} carteles solicitados'
        })
        
    except Exception as e:
        print(f"Error en imprimir_carteles: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@app.route('/verificar_licencia')
def verificar_licencia_manual():
    '''Verificar estado de licencia manualmente (solo admin)'''
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Solo admin
    usuario = Usuario.query.get(session['user_id'])
    if usuario.rol != 'admin':
        flash('No tiene permisos para ver esta información', 'error')
        return redirect(url_for('index'))
    
    try:
        resultado = verificar_licencia(ARCA_CONFIG.CUIT)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@app.route('/licencia_bloqueada')
def licencia_bloqueada():
    """Página que se muestra cuando el sistema está bloqueado"""
    licencia_info = app.config.get('LICENCIA_INFO', {})
    return render_template('licencia_bloqueada.html', licencia_info=licencia_info)

@app.route('/verificar_licencia_reload')
def verificar_licencia_reload():
    """Re-verifica la licencia y redirige según el resultado"""
    resultado = verificar_licencia(ARCA_CONFIG.CUIT)
    app.config['LICENCIA_INFO'] = resultado
    if resultado['tipo_bloqueo'] in ['sin_bloqueo', 'mora']:
        return redirect(url_for('index'))
    return redirect(url_for('licencia_bloqueada'))



@app.route('/ayuda_ctacte')
def ayuda_ctacte():
    """Ayuda sobre Cuenta Corriente"""
    return render_template('ayuda_ctacte.html')

@app.route('/ayuda_ventas')
def ayuda_ventas():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('ayuda_ventas.html')


@app.route('/api/cta_cte/exportar_pdf')
def exportar_pdf_cta_cte():
    try:
        print("🔍 Iniciando exportación PDF cuentas corrientes")
        
        # Query corregida con los nombres REALES de las columnas
        query = text("""
            SELECT 
                c.id,
                c.nombre,
                c.documento,
                COUNT(DISTINCT m.id) as movimientos_pendientes,
                COALESCE(SUM(CASE WHEN m.tipo = 'venta_fiada' THEN m.monto_total ELSE -m.monto_total END), 0) as saldo_pendiente,
                MAX(m.fecha) as ultima_operacion
            FROM cliente c
            LEFT JOIN cta_cte_movimiento m ON c.id = m.cliente_id AND m.estado = 'pendiente'
            GROUP BY c.id, c.nombre, c.documento
            HAVING saldo_pendiente > 0 OR movimientos_pendientes > 0
            ORDER BY saldo_pendiente DESC
        """)
        
        result = db.session.execute(query)
        clientes = []
        
        total_adeudado = 0
        clientes_con_deuda = 0
        total_movimientos = 0
        
        for row in result:
            saldo = float(row.saldo_pendiente)
            movimientos = int(row.movimientos_pendientes)
            
            # Convertir fecha a string AQUÍ
            ultima_op_str = 'Sin ops.'
            if row.ultima_operacion:
                try:
                    ultima_op_str = row.ultima_operacion.strftime('%d/%m/%Y')
                except:
                    ultima_op_str = str(row.ultima_operacion)[:10]
            
            clientes.append({
                'id': row.id,
                'nombre': row.nombre,
                'documento': row.documento or 'S/D',
                'movimientos_pendientes': movimientos,
                'saldo_pendiente': saldo,
                'ultima_operacion': ultima_op_str  # ✅ YA ES STRING
            })
            
            total_adeudado += saldo
            if saldo > 0:
                clientes_con_deuda += 1
            total_movimientos += movimientos
        
        print(f"📊 Clientes encontrados: {len(clientes)}")
        print(f"💰 Total adeudado: ${total_adeudado:,.2f}")
        print(f"👥 Clientes con deuda: {clientes_con_deuda}")
        
        # Preparar resumen
        resumen = {
            'total_adeudado': total_adeudado,
            'clientes_con_deuda': clientes_con_deuda,
            'total_clientes': len(clientes),
            'total_movimientos': total_movimientos
        }
        
        # Generar el PDF con ambos parámetros
        pdf_bytes = generar_pdf_cuentas_corrientes(clientes, resumen)
        
        # Crear nombre del archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'Cuentas_Corrientes_{fecha_actual}.pdf'
        
        print(f"✅ PDF generado exitosamente: {nombre_archivo}")
        
        # Enviar el archivo
        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        print(f"❌ Error generando PDF de cuentas corrientes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'connection',
            'detail': str(e)
        }), 500


@app.route('/reporte_saldos_clientes')
def reporte_saldos_clientes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    clientes_con_saldo = Cliente.query.filter(
        Cliente.id > 1,
        Cliente.saldo != 0
    ).order_by(desc(Cliente.saldo)).all()
    
    total_a_favor = sum(float(c.saldo) for c in clientes_con_saldo if c.saldo < 0)
    total_deben = sum(float(c.saldo) for c in clientes_con_saldo if c.saldo > 0)
    
    return render_template('reporte_saldos_clientes.html',
                          clientes=clientes_con_saldo,
                          total_a_favor=abs(total_a_favor),
                          total_deben=total_deben)

# ================================================================
# TOTALES POR PERIODO PARA CARDS DEL LISTADO DE FACTURAS
# Devuelve totales agrupados por estado (autorizadas / con error)
# Respeta filtros de numero, cliente, fecha desde/hasta.
# Ignora el filtro Estado a proposito: muestra siempre los 2 totales.
# ================================================================

# ════════════════════════════════════════════════════════════════════════
# CONSIGNACIÓN — Comprobantes Internos pendientes de facturar
# ════════════════════════════════════════════════════════════════════════

@app.route('/api/internos_pendientes_facturar')
def api_internos_pendientes_facturar():
    """Lista los Comprobantes Internos que aún no fueron facturados.
    Filtra opcionalmente por cliente intermediario.
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        intermediario_id = request.args.get('intermediario_id', type=int)

        # Subquery: IDs de internos que YA tienen factura derivada
        ya_facturados = db.session.query(Factura.interno_origen_id).filter(
            Factura.interno_origen_id.isnot(None)
        ).subquery()

        # Query: internos en estado 'interno', NO anulados, NO ya facturados
        query = db.session.query(Factura, Cliente).join(
            Cliente, Factura.cliente_id == Cliente.id
        ).filter(
            Factura.tipo_comprobante == '99',
            Factura.estado == 'interno',
            ~Factura.id.in_(ya_facturados)
        )

        if intermediario_id:
            query = query.filter(Factura.cliente_id == intermediario_id)

        query = query.order_by(Factura.fecha.desc()).limit(100)

        resultado = []
        for factura, cliente in query.all():
            # Contar items
            cant_items = db.session.query(func.count(DetalleFactura.id)).filter(
                DetalleFactura.factura_id == factura.id
            ).scalar() or 0

            resultado.append({
                'id': factura.id,
                'numero': factura.numero,
                'fecha': factura.fecha.strftime('%d/%m/%Y %H:%M') if factura.fecha else '',
                'cliente_id': cliente.id,
                'cliente_nombre': cliente.nombre,
                'cliente_doc': f"{cliente.tipo_documento or ''} {cliente.documento or ''}".strip(),
                'es_intermediario': bool(cliente.es_intermediario),
                'cant_items': int(cant_items),
                'total': float(factura.total or 0),
            })

        return jsonify({'success': True, 'internos': resultado})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/interno/<int:interno_id>/items')
def api_interno_items(interno_id):
    """Devuelve los items de un Comprobante Interno para cargar al carrito."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        interno = Factura.query.get(interno_id)
        if not interno:
            return jsonify({'success': False, 'error': 'Interno no encontrado'}), 404
        if interno.tipo_comprobante != '99':
            return jsonify({'success': False, 'error': 'No es un Comprobante Interno'}), 400
        if interno.estado != 'interno':
            return jsonify({'success': False, 'error': f'Estado inválido: {interno.estado}'}), 400

        # Verificar que NO esté ya facturado
        ya_facturado = Factura.query.filter_by(interno_origen_id=interno_id).first()
        if ya_facturado:
            return jsonify({
                'success': False,
                'error': f'Este interno ya fue facturado en {ya_facturado.numero}'
            }), 400

        items = []
        for det in interno.detalles:
            prod = det.producto
            items.append({
                'producto_id': prod.id,
                'codigo': prod.codigo,
                'nombre': prod.nombre,
                'cantidad': float(det.cantidad),
                'precio_unitario_interno': float(det.precio_unitario),  # precio del interno (lista X)
                'iva': float(det.porcentaje_iva or prod.iva or 21),
                'es_pesable': bool(getattr(prod, 'es_pesable', False)),
                'stock_actual': float(prod.stock or 0),
            })

        return jsonify({
            'success': True,
            'interno': {
                'id': interno.id,
                'numero': interno.numero,
                'fecha': interno.fecha.strftime('%d/%m/%Y %H:%M') if interno.fecha else '',
                'cliente_id': interno.cliente_id,
                'cliente_nombre': interno.cliente.nombre if interno.cliente else '',
                'total': float(interno.total or 0),
            },
            'items': items
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clientes_intermediarios')
def api_clientes_intermediarios():
    """Lista los clientes marcados como intermediarios (consignatarios)."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        clientes = Cliente.query.filter_by(es_intermediario=True).order_by(Cliente.nombre).all()
        return jsonify({
            'success': True,
            'clientes': [{
                'id': c.id,
                'nombre': c.nombre,
                'documento': c.documento or '',
            } for c in clientes]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# LIQUIDACIONES A INTERMEDIARIOS (CONSIGNACIÓN FASE 2)
# ════════════════════════════════════════════════════════════════════════

def _proximo_numero_liquidacion(punto_venta):
    """Genera el próximo número de liquidación con prefijo L"""
    ultima = LiquidacionIntermediario.query.filter(
        LiquidacionIntermediario.numero.like(f'{punto_venta:04d}-L%')
    ).order_by(LiquidacionIntermediario.id.desc()).first()
    
    numero = 1
    if ultima and ultima.numero:
        try:
            parte = ultima.numero.split('-')[1]
            if parte.startswith('L'):
                numero = int(parte[1:]) + 1
        except:
            numero = 1
    
    return f"{punto_venta:04d}-L{numero:07d}"


def _es_admin_o_puede_liquidar(usuario_id):
    """Verifica si el usuario tiene permiso para liquidar."""
    if not usuario_id:
        return False
    u = Usuario.query.get(usuario_id)
    if not u:
        return False
    if u.rol == 'admin':
        return True
    return bool(getattr(u, 'puede_liquidar', False))


@app.route('/liquidaciones')
def vista_liquidaciones():
    """Vista principal del módulo de Liquidaciones a Intermediarios"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not _es_admin_o_puede_liquidar(session.get('user_id')):
        flash('No tenés permiso para acceder a Liquidaciones a Intermediarios', 'danger')
        return redirect(url_for('dashboard'))
    
    return render_template('liquidaciones.html')


@app.route('/liquidacion/<int:liq_id>')
def vista_liquidacion_detalle(liq_id):
    """Vista de detalle de una liquidación"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not _es_admin_o_puede_liquidar(session.get('user_id')):
        flash('No tenés permiso para acceder a Liquidaciones', 'danger')
        return redirect(url_for('dashboard'))
    
    liq = LiquidacionIntermediario.query.get_or_404(liq_id)
    return render_template('liquidacion_detalle.html', liquidacion=liq)


@app.route('/api/liquidaciones/pendientes_por_intermediario')
def api_liquidaciones_pendientes():
    """
    Devuelve resumen por intermediario de internos pendientes de liquidar.
    Un interno está pendiente si:
    - Tiene una factura derivada NO anulada
    - El interno NO está anulado
    - Aún no fue liquidado (no aparece en liquidacion_detalle)
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Subquery: internos ya liquidados
        ya_liquidados = db.session.query(LiquidacionDetalle.interno_id).filter(
            LiquidacionDetalle.interno_id.isnot(None)
        ).subquery()
        
        # Buscar pares interno + factura derivada pendientes de liquidar
        FacturaDerivada = db.aliased(Factura)
        
        pendientes = db.session.query(
            Factura.id.label('interno_id'),
            Factura.numero.label('interno_numero'),
            Factura.fecha.label('interno_fecha'),
            Factura.total.label('interno_total'),
            Factura.subtotal.label('interno_subtotal'),
            Factura.cliente_id,
            Cliente.nombre.label('cliente_nombre'),
            FacturaDerivada.id.label('factura_id'),
            FacturaDerivada.numero.label('factura_numero'),
            FacturaDerivada.total.label('factura_total'),
            FacturaDerivada.subtotal.label('factura_subtotal')
        ).join(
            Cliente, Factura.cliente_id == Cliente.id
        ).join(
            FacturaDerivada, FacturaDerivada.interno_origen_id == Factura.id
        ).filter(
            Factura.tipo_comprobante == '99',
            Factura.estado == 'interno',
            FacturaDerivada.estado != 'anulada',
            ~Factura.id.in_(ya_liquidados),
            Cliente.es_intermediario == True
        ).order_by(Cliente.nombre, Factura.fecha).all()
        
        # Agrupar por intermediario
        resumen = {}
        for r in pendientes:
            cid = r.cliente_id
            if cid not in resumen:
                resumen[cid] = {
                    'intermediario_id': cid,
                    'intermediario_nombre': r.cliente_nombre,
                    'cant_internos': 0,
                    'total_costo_con_iva': 0,         # suma de internos (con IVA)
                    'total_costo_sin_iva': 0,         # suma de internos (sin IVA)
                    'total_vendido_con_iva': 0,       # suma de facturas (con IVA)
                    'total_vendido_sin_iva': 0,       # suma de facturas (sin IVA)
                    'a_liquidar_con_iva': 0,
                    'a_liquidar_sin_iva': 0,
                    'internos': []
                }
            
            interno_total = float(r.interno_total or 0)
            interno_subtotal = float(r.interno_subtotal or 0)
            factura_total = float(r.factura_total or 0)
            factura_subtotal = float(r.factura_subtotal or 0)
            
            dif_con_iva = factura_total - interno_total
            dif_sin_iva = factura_subtotal - interno_subtotal
            
            resumen[cid]['cant_internos'] += 1
            resumen[cid]['total_costo_con_iva'] += interno_total
            resumen[cid]['total_costo_sin_iva'] += interno_subtotal
            resumen[cid]['total_vendido_con_iva'] += factura_total
            resumen[cid]['total_vendido_sin_iva'] += factura_subtotal
            resumen[cid]['a_liquidar_con_iva'] += dif_con_iva
            resumen[cid]['a_liquidar_sin_iva'] += dif_sin_iva
            
            resumen[cid]['internos'].append({
                'interno_id': r.interno_id,
                'interno_numero': r.interno_numero,
                'interno_fecha': r.interno_fecha.strftime('%d/%m/%Y') if r.interno_fecha else '',
                'interno_total': interno_total,
                'interno_subtotal': interno_subtotal,
                'factura_id': r.factura_id,
                'factura_numero': r.factura_numero,
                'factura_total': factura_total,
                'factura_subtotal': factura_subtotal,
                'diferencia_con_iva': dif_con_iva,
                'diferencia_sin_iva': dif_sin_iva,
            })
        
        return jsonify({
            'success': True,
            'intermediarios': list(resumen.values())
        })
        
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidaciones/listar')
def api_liquidaciones_listar():
    """Lista de liquidaciones generadas con filtros."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        intermediario_id = request.args.get('intermediario_id', type=int)
        estado = request.args.get('estado', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        limit = request.args.get('limit', 100, type=int)
        
        query = db.session.query(LiquidacionIntermediario, Cliente).join(
            Cliente, LiquidacionIntermediario.intermediario_id == Cliente.id
        )
        
        if intermediario_id:
            query = query.filter(LiquidacionIntermediario.intermediario_id == intermediario_id)
        if estado:
            query = query.filter(LiquidacionIntermediario.estado == estado)
        if fecha_desde:
            try:
                fd = datetime.strptime(fecha_desde, '%Y-%m-%d')
                query = query.filter(LiquidacionIntermediario.fecha >= fd)
            except:
                pass
        if fecha_hasta:
            try:
                fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                query = query.filter(LiquidacionIntermediario.fecha <= fh)
            except:
                pass
        
        query = query.order_by(LiquidacionIntermediario.fecha.desc()).limit(limit)
        
        resultado = []
        for liq, cli in query.all():
            resultado.append({
                'id': liq.id,
                'numero': liq.numero,
                'fecha': liq.fecha.strftime('%d/%m/%Y %H:%M') if liq.fecha else '',
                'intermediario_id': cli.id,
                'intermediario_nombre': cli.nombre,
                'total_vendido': float(liq.total_vendido or 0),
                'total_costo': float(liq.total_costo or 0),
                'total_liquidar': float(liq.total_liquidar or 0),
                'base_calculo': liq.base_calculo,
                'estado': liq.estado,
                'medio_pago': liq.medio_pago,
                'fecha_pago': liq.fecha_pago.strftime('%d/%m/%Y') if liq.fecha_pago else None,
            })
        
        return jsonify({'success': True, 'liquidaciones': resultado})
        
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidacion/<int:liq_id>')
def api_liquidacion_detalle(liq_id):
    """Detalle completo de una liquidación con sus internos asociados."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        liq = LiquidacionIntermediario.query.get_or_404(liq_id)
        cli = liq.intermediario
        
        # Obtener detalles con info de los comprobantes
        detalles = []
        for d in liq.detalles:
            interno = Factura.query.get(d.interno_id)
            factura = Factura.query.get(d.factura_derivada_id) if d.factura_derivada_id else None
            
            detalles.append({
                'id': d.id,
                'interno_id': d.interno_id,
                'interno_numero': interno.numero if interno else '?',
                'interno_fecha': interno.fecha.strftime('%d/%m/%Y') if interno and interno.fecha else '',
                'factura_id': d.factura_derivada_id,
                'factura_numero': factura.numero if factura else None,
                'factura_fecha': factura.fecha.strftime('%d/%m/%Y') if factura and factura.fecha else None,
                'monto_interno': float(d.monto_interno or 0),
                'monto_factura': float(d.monto_factura or 0),
                'monto_diferencia': float(d.monto_diferencia or 0),
            })
        
        return jsonify({
            'success': True,
            'liquidacion': {
                'id': liq.id,
                'numero': liq.numero,
                'fecha': liq.fecha.strftime('%d/%m/%Y %H:%M') if liq.fecha else '',
                'intermediario_id': cli.id,
                'intermediario_nombre': cli.nombre,
                'intermediario_documento': cli.documento or '',
                'total_vendido': float(liq.total_vendido or 0),
                'total_costo': float(liq.total_costo or 0),
                'total_liquidar': float(liq.total_liquidar or 0),
                'base_calculo': liq.base_calculo,
                'estado': liq.estado,
                'medio_pago': liq.medio_pago,
                'fecha_pago': liq.fecha_pago.strftime('%d/%m/%Y %H:%M') if liq.fecha_pago else None,
                'motivo': liq.motivo or '',
                'usuario_nombre': liq.usuario.nombre if liq.usuario else '',
            },
            'detalles': detalles
        })
        
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidacion/generar', methods=['POST'])
def api_liquidacion_generar():
    """Genera una nueva liquidación a un intermediario.
    
    Body JSON esperado:
    {
      "intermediario_id": 5,
      "internos": [{"interno_id": 12, "factura_derivada_id": 25}, ...],
      "base_calculo": "con_iva" | "sin_iva",
      "forma_pago": "ahora" | "despues",
      "medio_pago": "EFECTIVO" | "TRANSFERENCIA" | "CHEQUE" | null,
      "motivo": "..."
    }
    """
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    if not _es_admin_o_puede_liquidar(session.get('user_id')):
        return jsonify({'success': False, 'error': 'No tenés permiso para liquidar'}), 403
    
    try:
        data = request.json or {}
        intermediario_id = data.get('intermediario_id')
        internos = data.get('internos', [])
        base_calculo = data.get('base_calculo', 'con_iva')
        forma_pago = data.get('forma_pago', 'despues')
        medio_pago = data.get('medio_pago')
        motivo = (data.get('motivo') or '').strip() or None
        
        # Validaciones
        if not intermediario_id:
            return jsonify({'success': False, 'error': 'Falta intermediario_id'}), 400
        if not internos:
            return jsonify({'success': False, 'error': 'No se seleccionaron internos para liquidar'}), 400
        if base_calculo not in ('con_iva', 'sin_iva'):
            return jsonify({'success': False, 'error': 'base_calculo inválido'}), 400
        if forma_pago not in ('ahora', 'despues'):
            return jsonify({'success': False, 'error': 'forma_pago inválido'}), 400
        if forma_pago == 'ahora' and not medio_pago:
            return jsonify({'success': False, 'error': 'Si pagás ahora, especificá medio_pago'}), 400
        
        # Validar intermediario
        cliente = Cliente.query.get(intermediario_id)
        if not cliente:
            return jsonify({'success': False, 'error': 'Intermediario no encontrado'}), 404
        
        # Validar que no haya internos ya liquidados
        interno_ids = [i.get('interno_id') for i in internos]
        ya_liquidados = LiquidacionDetalle.query.filter(
            LiquidacionDetalle.interno_id.in_(interno_ids)
        ).all()
        if ya_liquidados:
            nums_ids = [str(x.interno_id) for x in ya_liquidados]
            return jsonify({
                'success': False,
                'error': f'Estos internos ya fueron liquidados: ID {", ".join(nums_ids)}'
            }), 400
        
        # Calcular totales y validar cada interno
        total_vendido = Decimal('0')
        total_costo = Decimal('0')
        total_liquidar = Decimal('0')
        
        detalles_a_crear = []
        
        for it in internos:
            interno_id = it.get('interno_id')
            factura_id = it.get('factura_derivada_id')
            
            interno = Factura.query.get(interno_id)
            factura = Factura.query.get(factura_id) if factura_id else None
            
            if not interno or interno.tipo_comprobante != '99' or interno.estado != 'interno':
                return jsonify({'success': False, 'error': f'Interno {interno_id} inválido'}), 400
            if not factura or factura.estado == 'anulada':
                return jsonify({'success': False, 'error': f'Factura derivada {factura_id} inválida o anulada'}), 400
            if interno.cliente_id != intermediario_id:
                return jsonify({'success': False, 'error': f'El interno {interno.numero} no pertenece a este intermediario'}), 400
            
            if base_calculo == 'con_iva':
                monto_interno = Decimal(str(interno.total or 0))
                monto_factura = Decimal(str(factura.total or 0))
            else:  # sin_iva
                monto_interno = Decimal(str(interno.subtotal or 0))
                monto_factura = Decimal(str(factura.subtotal or 0))
            
            monto_dif = monto_factura - monto_interno
            
            total_costo += monto_interno
            total_vendido += monto_factura
            total_liquidar += monto_dif
            
            detalles_a_crear.append({
                'interno_id': interno_id,
                'factura_derivada_id': factura_id,
                'monto_interno': monto_interno,
                'monto_factura': monto_factura,
                'monto_diferencia': monto_dif,
            })
        
        # Determinar punto de venta para numerar
        usuario_actual = Usuario.query.get(session.get('user_id'))
        pv = usuario_actual.punto_venta if usuario_actual else 1
        numero = _proximo_numero_liquidacion(pv)
        
        # Crear liquidación
        estado_inicial = 'pagada' if forma_pago == 'ahora' else 'pendiente_pago'
        fecha_pago = datetime.now() if forma_pago == 'ahora' else None
        
        liquidacion = LiquidacionIntermediario(
            numero=numero,
            fecha=datetime.now(),
            intermediario_id=intermediario_id,
            total_vendido=total_vendido,
            total_costo=total_costo,
            total_liquidar=total_liquidar,
            base_calculo=base_calculo,
            estado=estado_inicial,
            medio_pago=medio_pago if forma_pago == 'ahora' else None,
            fecha_pago=fecha_pago,
            usuario_id=session.get('user_id'),
            motivo=motivo
        )
        db.session.add(liquidacion)
        db.session.flush()
        
        # Crear detalles
        internos_liquidados = []
        for d in detalles_a_crear:
            detalle = LiquidacionDetalle(
                liquidacion_id=liquidacion.id,
                interno_id=d['interno_id'],
                factura_derivada_id=d['factura_derivada_id'],
                monto_interno=d['monto_interno'],
                monto_factura=d['monto_factura'],
                monto_diferencia=d['monto_diferencia']
            )
            db.session.add(detalle)
            internos_liquidados.append(d['interno_id'])

        # ═══ CONSIGNACIÓN: cancelar el saldo pendiente del/los comprobante(s)
        # interno(s) en la cta.cte del consignatario. La mercadería ya se
        # rindió y se facturó al cliente final → el interno NO debe seguir
        # figurando como deuda del consignatario (sino se DUPLICA: el
        # interno + la factura al cliente final). Ver
        # AYUDA_CIRCUITO_CONSIGNACION.md (paso 3).
        # Se hace AL GENERAR la liquidación (decisión confirmada con el
        # cliente: el interno se salda al liquidar, no al facturar).
        for interno_id in internos_liquidados:
            # Cancelar los movimientos de cta.cte de ESE interno para ESTE
            # cliente que sigan con saldo pendiente. Identificamos el
            # movimiento del interno por su factura_id (el interno ES una
            # factura type='99'). saldo_pendiente → 0, estado → 'pagado'.
            res = db.session.execute(text("""
                UPDATE cta_cte_movimiento
                   SET saldo_pendiente = 0,
                       estado = 'pagado',
                       observaciones = CONCAT(COALESCE(observaciones,''),
                            ' [Cancelado por liquidación ', :numero, ']')
                 WHERE cliente_id = :cid
                   AND factura_id = :interno_id
                   AND tipo = 'venta_fiada'
                   AND saldo_pendiente > 0
            """), {
                'cid':        cliente.id,
                'interno_id': interno_id,
                'numero':     numero,
            })
            if res.rowcount:
                print(f"🔗 Liquidación {numero}: cancelado saldo pendiente "
                      f"del interno id={interno_id} en cta.cte de {cliente.nombre} "
                      f"({res.rowcount} movimiento/s)")
            else:
                print(f"⚠️ Liquidación {numero}: el interno id={interno_id} no "
                      f"tenía movimiento pendiente en cta.cte (¿ya saldado o "
                      f"nunca fue a cta.cte?) — revisar si es esperado")

        # Si es "pagar después" → registrar como crédito en CTA.CTE del intermediario
        # (saldo negativo = a favor del cliente / Siamotre le debe)
        if forma_pago == 'despues':
            # 1) Movimiento en cta_cte_movimiento (fuente única de saldo).
            #    numero_comprobante='LIQUIDACION' → saldo_real_cliente() lo
            #    suma como saldo a favor del cliente, igual que un A_CUENTA,
            #    pero diferenciado para trazabilidad en la cta.cte.
            db.session.execute(text("""
                INSERT INTO cta_cte_movimiento
                    (cliente_id, tipo, tipo_mov, estado, monto_total, saldo_pendiente,
                     numero_comprobante, usuario_id, observaciones, fecha)
                VALUES
                    (:cid, 'pago', 'pago', 'pagado', :monto, 0,
                     'LIQUIDACION', :uid, :obs, NOW())
            """), {
                'cid':   cliente.id,
                'monto': float(total_liquidar),
                'uid':   session.get('user_id'),
                'obs':   f'Liquidación {numero} (pagar después)',
            })
            # 2) Mantener el campo viejo cliente.saldo sincronizado (todavía
            #    no se elimina; se sincroniza por si algo legacy lo lee).
            saldo_actual = Decimal(str(cliente.saldo or 0))
            cliente.saldo = saldo_actual - total_liquidar  # más negativo
            print(f"💰 Liquidación {numero}: cta.cte +crédito {total_liquidar} a {cliente.nombre} (saldo viejo {saldo_actual} → {cliente.saldo})")
        
        db.session.commit()
        
        print(f"✅ Liquidación {numero} generada. Intermediario: {cliente.nombre}. Total: {total_liquidar}")
        
        return jsonify({
            'success': True,
            'liquidacion': {
                'id': liquidacion.id,
                'numero': liquidacion.numero,
                'total_liquidar': float(total_liquidar),
                'estado': liquidacion.estado,
            }
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidacion/<int:liq_id>/pagar', methods=['POST'])
def api_liquidacion_pagar(liq_id):
    """Marca una liquidación pendiente como pagada."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    if not _es_admin_o_puede_liquidar(session.get('user_id')):
        return jsonify({'success': False, 'error': 'No tenés permiso'}), 403
    
    try:
        liq = LiquidacionIntermediario.query.get_or_404(liq_id)
        
        if liq.estado != 'pendiente_pago':
            return jsonify({'success': False, 'error': f'La liquidación está en estado {liq.estado}, no pendiente'}), 400
        
        data = request.json or {}
        medio_pago = data.get('medio_pago', 'EFECTIVO')
        
        liq.estado = 'pagada'
        liq.medio_pago = medio_pago
        liq.fecha_pago = datetime.now()
        
        # Revertir el crédito en Cta.Cte del intermediario (ya se le pagó)
        cliente = liq.intermediario
        if cliente:
            # 1) Movimiento que CANCELA el crédito de la liquidación.
            #    numero_comprobante='LIQUIDACION_PAGADA' → saldo_real_cliente()
            #    lo resta del saldo a favor (análogo a SALDO_FAVOR_USADO),
            #    dejando el neto en cero para esta liquidación.
            db.session.execute(text("""
                INSERT INTO cta_cte_movimiento
                    (cliente_id, tipo, tipo_mov, estado, monto_total, saldo_pendiente,
                     numero_comprobante, usuario_id, observaciones, fecha)
                VALUES
                    (:cid, 'pago', 'pago', 'pagado', :monto, 0,
                     'LIQUIDACION_PAGADA', :uid, :obs, NOW())
            """), {
                'cid':   cliente.id,
                'monto': float(liq.total_liquidar or 0),
                'uid':   session.get('user_id'),
                'obs':   f'Pago liquidación {liq.numero} ({medio_pago})',
            })
            # 2) Sincronizar campo viejo
            saldo_actual = Decimal(str(cliente.saldo or 0))
            cliente.saldo = saldo_actual + Decimal(str(liq.total_liquidar or 0))  # menos negativo
            print(f"💰 Pago liquidación {liq.numero}: cta.cte cancela crédito {liq.total_liquidar} de {cliente.nombre} (saldo viejo {saldo_actual} → {cliente.saldo})")
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Liquidación {liq.numero} marcada como pagada'})
        
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidacion/<int:liq_id>/anular', methods=['POST'])
def api_liquidacion_anular(liq_id):
    """Anula una liquidación (libera los internos para volver a liquidarse)."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    if not _es_admin_o_puede_liquidar(session.get('user_id')):
        return jsonify({'success': False, 'error': 'No tenés permiso'}), 403
    
    try:
        liq = LiquidacionIntermediario.query.get_or_404(liq_id)
        
        if liq.estado == 'anulada':
            return jsonify({'success': False, 'error': 'Ya está anulada'}), 400
        
        data = request.json or {}
        motivo_anul = (data.get('motivo') or '').strip()
        
        # Si estaba "pendiente de pago" → revertir crédito en Cta.Cte (ya no le debemos)
        if liq.estado == 'pendiente_pago':
            cliente = liq.intermediario
            if cliente:
                # Movimiento que cancela el crédito de la liquidación
                # (mismo efecto que LIQUIDACION_PAGADA: lo resta del saldo
                #  a favor, dejando el neto en cero — la liquidación se anuló).
                db.session.execute(text("""
                    INSERT INTO cta_cte_movimiento
                        (cliente_id, tipo, tipo_mov, estado, monto_total, saldo_pendiente,
                         numero_comprobante, usuario_id, observaciones, fecha)
                    VALUES
                        (:cid, 'pago', 'pago', 'pagado', :monto, 0,
                         'LIQUIDACION_PAGADA', :uid, :obs, NOW())
                """), {
                    'cid':   cliente.id,
                    'monto': float(liq.total_liquidar or 0),
                    'uid':   session.get('user_id'),
                    'obs':   f'Anulación liquidación {liq.numero} (revierte crédito)',
                })
                saldo_actual = Decimal(str(cliente.saldo or 0))
                cliente.saldo = saldo_actual + Decimal(str(liq.total_liquidar or 0))
                print(f"💰 Anulación liquidación {liq.numero}: cta.cte revierte crédito {liq.total_liquidar} (saldo viejo {saldo_actual} → {cliente.saldo})")
        elif liq.estado == 'pagada':
            # Si ya estaba pagada, también revertimos: ahora le debemos el dinero
            # (porque la liquidación ya no existe pero el dinero ya salió de Siamotre)
            # IMPORTANTE: esto puede ser una pifia operativa, registramos pero advertimos
            # NO se toca cta.cte ni saldo: igual que el comportamiento original,
            # el reintegro se maneja manualmente (decisión consciente, no cambiar
            # sin entender el flujo de dinero ya salido).
            print(f"⚠️ Anulación de liquidación YA PAGADA {liq.numero} — revisar reintegro de dinero manualmente (cta.cte NO modificada)")
        
        # ═══ CONSIGNACIÓN: revertir la cancelación de los internos ═══
        # Al generar la liquidación se canceló el saldo pendiente de los
        # internos (saldo_pendiente=0, estado='pagado'). Si se anula la
        # liquidación, esos internos vuelven a ser deuda pendiente del
        # consignatario. IMPORTANTE: capturar los interno_id ANTES de
        # borrar los detalles (abajo se borran por cascade).
        internos_a_revertir = [d.interno_id for d in liq.detalles if d.interno_id]
        cli_liq = liq.intermediario
        for interno_id in internos_a_revertir:
            # Reabrir el movimiento del interno: volver a poner su
            # saldo_pendiente = monto_total y estado='pendiente'.
            # Solo los que fueron cancelados por una liquidación (marca
            # en observaciones) para no tocar otros movimientos.
            res = db.session.execute(text("""
                UPDATE cta_cte_movimiento
                   SET saldo_pendiente = monto_total,
                       estado = 'pendiente',
                       observaciones = CONCAT(COALESCE(observaciones,''),
                            ' [Reabierto por anulación de liquidación ',
                            :numero, ']')
                 WHERE cliente_id = :cid
                   AND factura_id = :interno_id
                   AND tipo = 'venta_fiada'
                   AND estado = 'pagado'
                   AND observaciones LIKE CONCAT('%[Cancelado por liquidación ',
                                                 :numero, ']%')
            """), {
                'cid':        cli_liq.id if cli_liq else None,
                'interno_id': interno_id,
                'numero':     liq.numero,
            })
            if res.rowcount:
                print(f"🔗 Anulación {liq.numero}: reabierto saldo pendiente "
                      f"del interno id={interno_id} ({res.rowcount} mov/s)")

        # Borrar los detalles (gracias al cascade) — esto libera los internos
        # para volver a aparecer como pendientes de liquidar
        for d in list(liq.detalles):
            db.session.delete(d)
        
        liq.estado = 'anulada'
        if motivo_anul:
            liq.motivo = (liq.motivo or '') + f'\n[ANULADA] {motivo_anul}'
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Liquidación {liq.numero} anulada. Los internos volvieron al pool de pendientes.'})
        
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/liquidacion/<int:liq_id>/imprimir_termica', methods=['POST'])
def api_liquidacion_imprimir_termica(liq_id):
    """Imprime el ticket de liquidación por la impresora térmica."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        liq = LiquidacionIntermediario.query.get_or_404(liq_id)
        
        # Resolver detalles aquí (donde sí tenemos acceso a Factura sin import circular)
        detalles_data = []
        for d in liq.detalles:
            interno = Factura.query.get(d.interno_id) if d.interno_id else None
            factura = Factura.query.get(d.factura_derivada_id) if d.factura_derivada_id else None
            detalles_data.append({
                'interno_numero': interno.numero if interno else '?',
                'factura_numero': factura.numero if factura else '?',
                'monto_interno': float(d.monto_interno or 0),
                'monto_factura': float(d.monto_factura or 0),
                'monto_diferencia': float(d.monto_diferencia or 0),
            })
        
        # Importar la impresora térmica
        try:
            from impresora_termica import ImpresoraTermica
        except ImportError as e:
            print(f"❌ ImportError impresora_termica: {e}")
            return jsonify({'success': False, 'error': 'Módulo de impresora no disponible'}), 500
        
        impresora = ImpresoraTermica()
        
        if not hasattr(impresora, 'imprimir_liquidacion'):
            return jsonify({
                'success': False,
                'error': 'La impresora no tiene el método imprimir_liquidacion. Actualizá impresora_termica.py'
            }), 500
        
        ok = impresora.imprimir_liquidacion(liq, detalles_data=detalles_data)
        
        if not ok:
            return jsonify({
                'success': False,
                'error': 'No se pudo imprimir. Revisá la consola NSSM para ver el error específico.'
            }), 500
        
        return jsonify({'success': True, 'message': 'Liquidación impresa'})
        
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/totales_facturas_periodo')
def api_totales_facturas_periodo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        numero      = request.args.get('numero', '').strip()
        cliente     = request.args.get('cliente', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()

        query = db.session.query(Factura).join(Cliente, Factura.cliente_id == Cliente.id)

        if numero:
            query = query.filter(Factura.numero.ilike(f'%{numero}%'))
        if cliente:
            query = query.filter(Cliente.nombre.ilike(f'%{cliente}%'))

        if fecha_desde:
            try:
                fd = datetime.strptime(fecha_desde, '%Y-%m-%d')
                query = query.filter(Factura.fecha >= fd)
            except ValueError:
                return jsonify({'error': 'Formato de fecha desde invalido'}), 400

        if fecha_hasta:
            try:
                fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                query = query.filter(Factura.fecha <= fh)
            except ValueError:
                return jsonify({'error': 'Formato de fecha hasta invalido'}), 400

        subq = query.with_entities(Factura.id, Factura.estado, Factura.total).subquery()

        agregados = db.session.query(
            func.sum(case((subq.c.estado == 'autorizada', subq.c.total), else_=0)).label('tot_aut'),
            func.sum(case((subq.c.estado == 'autorizada', 1),            else_=0)).label('cnt_aut'),
            func.sum(case((subq.c.estado == 'error_afip', subq.c.total), else_=0)).label('tot_err'),
            func.sum(case((subq.c.estado == 'error_afip', 1),            else_=0)).label('cnt_err'),
            func.sum(case((subq.c.estado == 'pendiente',  subq.c.total), else_=0)).label('tot_pen'),
            func.sum(case((subq.c.estado == 'pendiente',  1),            else_=0)).label('cnt_pen'),
            func.sum(case((subq.c.estado == 'anulada',    subq.c.total), else_=0)).label('tot_anu'),
            func.sum(case((subq.c.estado == 'anulada',    1),            else_=0)).label('cnt_anu'),
            func.sum(case((subq.c.estado == 'interno',    subq.c.total), else_=0)).label('tot_int'),
            func.sum(case((subq.c.estado == 'interno',    1),            else_=0)).label('cnt_int'),
        ).one()

        return jsonify({
            'autorizadas': {
                'cantidad': int(agregados.cnt_aut or 0),
                'total':    float(agregados.tot_aut or 0),
            },
            'con_error': {
                'cantidad': int(agregados.cnt_err or 0),
                'total':    float(agregados.tot_err or 0),
            },
            'pendientes': {
                'cantidad': int(agregados.cnt_pen or 0),
                'total':    float(agregados.tot_pen or 0),
            },
            'anuladas': {
                'cantidad': int(agregados.cnt_anu or 0),
                'total':    float(agregados.tot_anu or 0),
            },
            'internos': {
                'cantidad': int(agregados.cnt_int or 0),
                'total':    float(agregados.tot_int or 0),
            },
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500



# ════════════════════════════════════════════════════════════════════════════
# IMPRESIÓN TÉRMICA DE NOTA DE CRÉDITO
# ════════════════════════════════════════════════════════════════════════════

@app.route('/api/nota_credito/<int:nc_id>/imprimir_termica', methods=['POST'])
def api_nota_credito_imprimir_termica(nc_id):
    """Imprime una Nota de Crédito por impresora térmica.
    Reusa imprimir_factura_termica del módulo impresora_termica (formato idéntico
    a factura — el tipo_comprobante '03'/'08'/'13' hace que aparezca como
    'NOTA CRED A/B/C' automáticamente).
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        nc = NotaCredito.query.get(nc_id)
        if not nc:
            return jsonify({'success': False, 'error': 'NC no encontrada'}), 404

        # Armar dict de items (formato esperado por imprimir_factura_termica)
        items = []
        for det in nc.detalles:
            try:
                nombre_prod = det.producto.nombre if det.producto else 'Producto'
            except Exception:
                nombre_prod = 'Producto'
            items.append({
                'nombre':          nombre_prod,
                'cantidad':        float(det.cantidad or 0),
                'precio_unitario': float(det.precio_unitario or 0),
                'subtotal':        float(det.subtotal or 0),
                'porcentaje_iva':  float(det.porcentaje_iva or 21),
                'importe_iva':     float(det.importe_iva or 0),
            })

        # Armar dict del cliente
        cliente_data = None
        if nc.cliente:
            cliente_data = {
                'nombre':         nc.cliente.nombre or 'Consumidor Final',
                'documento':      nc.cliente.documento or None,
                'tipo_documento': getattr(nc.cliente, 'tipo_documento', 'DNI') or 'DNI',
            }

        # Vto CAE: convertir date a string si existe
        vto_cae_str = None
        if nc.vto_cae:
            try:
                vto_cae_str = nc.vto_cae.strftime('%Y-%m-%d') if hasattr(nc.vto_cae, 'strftime') else str(nc.vto_cae)
            except Exception:
                vto_cae_str = str(nc.vto_cae)

        # Observaciones — referencia a factura asociada
        observaciones = f"Nota de Credito sobre Factura {nc.factura_numero}"
        if nc.motivo:
            observaciones += f" — Motivo: {nc.motivo}"

        datos_nc = {
            'numero':           nc.numero,
            'tipo_comprobante': nc.tipo_comprobante,  # '03', '08' o '13' → NC A/B/C
            'subtotal':         float(nc.subtotal or 0),
            'iva':              float(nc.iva or 0),
            'total':            float(nc.total or 0),
            'cae':              nc.cae,
            'vto_cae':          vto_cae_str,
            'cliente':          cliente_data,
            'items':            items,
            'observaciones':    observaciones,
        }

        from impresora_termica import imprimir_nota_credito_termica
        resultado = imprimir_nota_credito_termica(datos_nc)

        if resultado.get('success'):
            return jsonify({
                'success': True,
                'mensaje': resultado.get('mensaje', 'Nota de Crédito impresa correctamente')
            })
        else:
            return jsonify({
                'success': False,
                'error': resultado.get('error', 'No se pudo imprimir la Nota de Crédito')
            }), 500

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════
# ESCÁNER PÚBLICO DE CÓDIGO DE BARRAS — consulta de precios desde celular
# ════════════════════════════════════════════════════════════════════════════

# Rate limit en memoria: { ip -> [timestamps] }
from collections import defaultdict
from threading import Lock as _PrecioLock
_precio_publico_buckets = defaultdict(list)
_precio_publico_lock = _PrecioLock()

def _check_rate_limit_precio(ip, limite=60, ventana_seg=60):
    """True si está dentro del límite, False si lo excedió."""
    import time
    ahora = time.time()
    with _precio_publico_lock:
        _precio_publico_buckets[ip] = [t for t in _precio_publico_buckets[ip] if ahora - t < ventana_seg]
        if len(_precio_publico_buckets[ip]) >= limite:
            return False
        _precio_publico_buckets[ip].append(ahora)
        return True


@app.route('/escaner')
def vista_escaner():
    """Pantalla pública con cámara para escanear códigos de barras.
    Si hay sesión activa, el template muestra info extra (stock, costo)."""
    logueado = 'user_id' in session
    return render_template('escaner.html', logueado=logueado)


@app.route('/api/precio_publico/<path:codigo>')
def api_precio_publico(codigo):
    """Devuelve precio + descripción de un producto buscando por código de barras
    (prioridad) o código interno. PÚBLICO (sin login). Si hay sesión, info extra.
    Rate limit: 60 reqs/min por IP.
    """
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()
    if not _check_rate_limit_precio(ip):
        return jsonify({'success': False, 'error': 'Demasiadas consultas. Esperá un minuto.'}), 429

    try:
        codigo = (codigo or '').strip()
        if not codigo:
            return jsonify({'success': False, 'error': 'Código vacío'}), 400

        # Buscar primero por código interno (igual lógica que /api/producto que SÍ funciona):
        #   1) codigo == upper(buscado)
        #   2) codigo_barras == buscado (sin transformar)
        #   3) codigo_barras == upper(buscado)
        # Sin filtrar por activo=True (la consulta es pública de solo lectura;
        # si el producto existe en la DB, podés ver su precio aunque esté pausado).
        codigo_upper = codigo.upper()
        producto = Producto.query.filter_by(codigo=codigo_upper).first()
        if producto is None:
            producto = Producto.query.filter_by(codigo_barras=codigo).first()
        if producto is None:
            producto = Producto.query.filter_by(codigo_barras=codigo_upper).first()

        if producto is None:
            return jsonify({
                'success': False,
                'error': 'Producto no encontrado',
                'codigo': codigo,
            }), 404

        # Info básica (pública)
        resp = {
            'success':     True,
            'codigo':      producto.codigo,
            'nombre':      producto.nombre,
            'descripcion': producto.descripcion or '',
            'precio':      float(producto.precio or 0),
            'iva':         float(producto.iva or 21),
            'es_pesable':  bool(getattr(producto, 'es_pesable', False)),
        }

        # Info extra (solo logueados)
        if 'user_id' in session:
            try:
                stock = float(producto.stock_dinamico) if hasattr(producto, 'stock_dinamico') and producto.stock_dinamico is not None else float(producto.stock or 0)
            except Exception:
                stock = float(producto.stock or 0)
            resp['stock']        = stock
            resp['costo']        = float(producto.costo or 0)
            resp['margen']       = float(producto.margen or 0)
            resp['stock_minimo'] = float(producto.stock_minimo or 0)
            resp['categoria']    = producto.categoria or ''
            for i in range(2, 6):
                pr = getattr(producto, f'precio{i}', None)
                if pr:
                    resp[f'precio{i}'] = float(pr)

        return jsonify(resp)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════
# QR DEL ESCÁNER PÚBLICO PARA GÓNDOLA
# ════════════════════════════════════════════════════════════════
QR_ESCANER_URL_DEFAULT = 'https://lakapp.dyndns.info/escaner'
QR_LOGO_DIR = os.path.join(app.root_path, 'static', 'images')
QR_LOGO_PATH = os.path.join(QR_LOGO_DIR, 'logo_qr.png')


def _qr_get_url():
    """Devuelve la URL configurada para el QR (o el default)."""
    try:
        cfg = Configuracion.query.filter_by(clave='qr_escaner_url').first()
        if cfg and cfg.valor and cfg.valor.strip():
            return cfg.valor.strip()
    except Exception:
        pass
    return QR_ESCANER_URL_DEFAULT


def _qr_set_url(nueva_url):
    """Guarda la URL del QR en la tabla Configuracion."""
    cfg = Configuracion.query.filter_by(clave='qr_escaner_url').first()
    if cfg is None:
        cfg = Configuracion(clave='qr_escaner_url', valor=nueva_url,
                            descripcion='URL pública del escáner para QR de góndola')
        db.session.add(cfg)
    else:
        cfg.valor = nueva_url
    db.session.commit()


def _qr_generar_imagen(formato='a4', url_qr=None, titulo='Escaneá para ver el precio'):
    """Genera el cartel con QR + logo en memoria.
    formato: 'a4' (1654x2339 hoja A4) | 'termica' (576px ancho rollo 80mm).
    Devuelve BytesIO listo para send_file.
    """
    import qrcode
    from qrcode.constants import ERROR_CORRECT_H
    from PIL import Image, ImageDraw, ImageFont

    if url_qr is None:
        url_qr = _qr_get_url()

    # 1) QR con corrección H
    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_CORRECT_H,
        box_size=(20 if formato == 'a4' else 10),
        border=2,
    )
    qr.add_data(url_qr)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white').convert('RGB')

    # 2) Logo al centro del QR (si existe)
    if os.path.isfile(QR_LOGO_PATH):
        try:
            logo = Image.open(QR_LOGO_PATH).convert('RGBA')
            qw, qh = qr_img.size
            lmax = int(qw * 0.22)
            ratio = min(lmax / logo.width, lmax / logo.height)
            nw = int(logo.width * ratio)
            nh = int(logo.height * ratio)
            logo = logo.resize((nw, nh), Image.LANCZOS)
            pad = 12 if formato == 'a4' else 8
            bg = Image.new('RGB', (nw + 2*pad, nh + 2*pad), 'white')
            qr_img.paste(bg, ((qw - bg.width)//2, (qh - bg.height)//2))
            qr_img.paste(logo, ((qw - nw)//2, (qh - nh)//2), logo)
        except Exception as e:
            print(f"⚠️ Error pegando logo en QR: {e}")

    # 3) Fuente helper
    def get_font(size):
        paths = [
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            'C:/Windows/Fonts/arialbd.ttf',
            'C:/Windows/Fonts/calibrib.ttf',
        ]
        for p in paths:
            if os.path.exists(p):
                return ImageFont.truetype(p, size)
        return ImageFont.load_default()

    # 4) Armado del cartel
    if formato == 'a4':
        W, H = 1654, 2339
        cartel = Image.new('RGB', (W, H), 'white')
        draw = ImageDraw.Draw(cartel)
        draw.rectangle([15, 15, W-15, H-15], outline='#1f2937', width=4)

        # Logo arriba (si existe)
        y_actual = 130
        if os.path.isfile(QR_LOGO_PATH):
            try:
                lb = Image.open(QR_LOGO_PATH).convert('RGBA')
                lb_max = 360
                r = min(lb_max / lb.width, lb_max / lb.height)
                lbw = int(lb.width * r)
                lbh = int(lb.height * r)
                lb = lb.resize((lbw, lbh), Image.LANCZOS)
                cartel.paste(lb, ((W - lbw) // 2, y_actual), lb)
                y_actual += lbh + 60
            except Exception as e:
                print(f"⚠️ Error logo grande: {e}")

        # QR grande
        qr_target = int(W * 0.62)
        qr_resized = qr_img.resize((qr_target, qr_target), Image.LANCZOS)
        cartel.paste(qr_resized, ((W - qr_target) // 2, y_actual))
        y_text = y_actual + qr_target + 60

        ft_title = get_font(110)
        ft_sub = get_font(54)
        ft_foot = get_font(28)

        t1 = 'ESCANEÁ PARA'
        b = draw.textbbox((0,0), t1, font=ft_title)
        draw.text(((W - (b[2]-b[0]))//2, y_text), t1, fill='#1f2937', font=ft_title)
        y_text += (b[3] - b[1]) + 18

        t2 = 'VER EL PRECIO'
        b = draw.textbbox((0,0), t2, font=ft_title)
        draw.text(((W - (b[2]-b[0]))//2, y_text), t2, fill='#1f2937', font=ft_title)
        y_text += (b[3] - b[1]) + 50

        sub = 'Apuntá la cámara del celular al QR'
        b = draw.textbbox((0,0), sub, font=ft_sub)
        draw.text(((W - (b[2]-b[0]))//2, y_text), sub, fill='#6b7280', font=ft_sub)
        y_text += (b[3] - b[1]) + 40

        draw.line([(W*0.25, y_text), (W*0.75, y_text)], fill='#e5e7eb', width=2)

        foot = 'Powered by FactuFácil'
        b = draw.textbbox((0,0), foot, font=ft_foot)
        draw.text(((W - (b[2]-b[0]))//2, H - 80), foot, fill='#9ca3af', font=ft_foot)
    else:
        # Térmica 80mm
        W = 576
        qr_target = int(W * 0.85)
        qr_resized = qr_img.resize((qr_target, qr_target), Image.LANCZOS)

        ft_title = get_font(42)
        ft_sub = get_font(22)
        ft_foot = get_font(16)

        # Logo arriba chico (si existe)
        logo_top = None
        lt_h = 0
        if os.path.isfile(QR_LOGO_PATH):
            try:
                logo_top = Image.open(QR_LOGO_PATH).convert('RGBA')
                lt_max = 130
                r = min(lt_max / logo_top.width, lt_max / logo_top.height)
                ltw = int(logo_top.width * r)
                lt_h = int(logo_top.height * r)
                logo_top = logo_top.resize((ltw, lt_h), Image.LANCZOS)
            except Exception as e:
                print(f"⚠️ Error logo térmica: {e}")
                logo_top = None
                lt_h = 0

        margin_top = 20
        margin_bottom = 30
        spacing = 18

        H = margin_top + lt_h + (spacing if lt_h else 0) + qr_target + spacing + 50 + 50 + spacing + 30 + spacing + 20 + margin_bottom
        cartel = Image.new('RGB', (W, H), 'white')
        draw = ImageDraw.Draw(cartel)

        y = margin_top
        if logo_top:
            cartel.paste(logo_top, ((W - logo_top.width) // 2, y), logo_top)
            y += lt_h + spacing

        cartel.paste(qr_resized, ((W - qr_target) // 2, y))
        y += qr_target + spacing

        t1 = 'ESCANEÁ PARA'
        b = draw.textbbox((0,0), t1, font=ft_title)
        draw.text(((W - (b[2]-b[0]))//2, y), t1, fill='black', font=ft_title)
        y += (b[3] - b[1]) + 8

        t2 = 'VER EL PRECIO'
        b = draw.textbbox((0,0), t2, font=ft_title)
        draw.text(((W - (b[2]-b[0]))//2, y), t2, fill='black', font=ft_title)
        y += (b[3] - b[1]) + spacing

        sub = 'Apuntá la cámara del cel al QR'
        b = draw.textbbox((0,0), sub, font=ft_sub)
        draw.text(((W - (b[2]-b[0]))//2, y), sub, fill='black', font=ft_sub)
        y += (b[3] - b[1]) + spacing

        foot = 'Powered by FactuFácil'
        b = draw.textbbox((0,0), foot, font=ft_foot)
        draw.text(((W - (b[2]-b[0]))//2, y), foot, fill='black', font=ft_foot)

    buf = BytesIO()
    cartel.save(buf, 'PNG', optimize=True)
    buf.seek(0)
    return buf


@app.route('/qr_escaner')
def vista_qr_escaner():
    """Pantalla interna para generar/imprimir el QR del escáner público."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    url_qr = _qr_get_url()
    tiene_logo = os.path.isfile(QR_LOGO_PATH)
    return render_template('qr_escaner.html', url_qr=url_qr, tiene_logo=tiene_logo)


@app.route('/qr_escaner/generar.png')
def qr_escaner_generar():
    """Genera el PNG del cartel (A4 o térmica) con QR + logo embebido."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        formato = request.args.get('formato', 'a4')
        if formato not in ('a4', 'termica'):
            formato = 'a4'
        # Permitir override de URL por querystring (sin guardar)
        url_qr = request.args.get('url') or _qr_get_url()
        buf = _qr_generar_imagen(formato=formato, url_qr=url_qr)
        from flask import send_file
        nombre = f'qr_escaner_{formato}.png'
        return send_file(buf, mimetype='image/png', download_name=nombre)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/qr_escaner/subir_logo', methods=['POST'])
def qr_escaner_subir_logo():
    """Recibe PNG/JPG y lo guarda en /static/images/logo_qr.png normalizado."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        archivo = request.files.get('logo')
        if not archivo or not archivo.filename:
            return jsonify({'success': False, 'error': 'No se recibió ningún archivo'}), 400

        ext = (archivo.filename.rsplit('.', 1)[-1] or '').lower()
        if ext not in ('png', 'jpg', 'jpeg'):
            return jsonify({'success': False, 'error': 'Solo PNG o JPG'}), 400

        os.makedirs(QR_LOGO_DIR, exist_ok=True)

        from PIL import Image
        img = Image.open(archivo.stream)
        max_side = 600
        if max(img.size) > max_side:
            ratio = max_side / max(img.size)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)

        if img.mode not in ('RGB', 'RGBA'):
            img = img.convert('RGBA')
        img.save(QR_LOGO_PATH, format='PNG', optimize=True)

        return jsonify({'success': True, 'mensaje': 'Logo cargado correctamente'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/qr_escaner/quitar_logo', methods=['POST'])
def qr_escaner_quitar_logo():
    """Elimina el logo guardado."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        if os.path.isfile(QR_LOGO_PATH):
            os.remove(QR_LOGO_PATH)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/qr_escaner/guardar_url', methods=['POST'])
def qr_escaner_guardar_url():
    """Persiste la URL del QR para que no haya que escribirla cada vez."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        data = request.get_json(silent=True) or {}
        nueva = (data.get('url') or '').strip()
        if not nueva:
            return jsonify({'success': False, 'error': 'URL vacía'}), 400
        if not (nueva.startswith('http://') or nueva.startswith('https://')):
            return jsonify({'success': False, 'error': 'La URL debe empezar con http:// o https://'}), 400
        _qr_set_url(nueva)
        return jsonify({'success': True, 'url': nueva})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/qr_escaner/imprimir_termica', methods=['POST'])
def qr_escaner_imprimir_termica():
    """Imprime el cartel térmico directo a la impresora vía ESC/POS nativo
    (no usa el PNG sino comandos GS ( k) — más confiable y menos papel."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    try:
        if not IMPRESION_DISPONIBLE:
            return jsonify({'success': False, 'error': 'Impresión térmica no disponible'}), 500
        from impresora_termica import imprimir_qr_escaner_termica
        url_qr = _qr_get_url()
        # Título: nombre de la URL o por defecto
        titulo = 'Escaner'
        resultado = imprimir_qr_escaner_termica(url=url_qr, titulo=titulo)
        return jsonify(resultado)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/factura_pdf/<int:factura_id>')
def factura_pdf(factura_id):
    """Genera y devuelve el PDF A4 de una factura (se abre en el navegador,
    listo para imprimir en hoja A4). Reemplaza/complementa el ticket termico."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        from factura_pdf import generar_pdf_factura

        factura = Factura.query.get_or_404(factura_id)

        # ── Datos del emisor (desde config_cliente, con fallbacks) ──
        emisor = {
            'nombre_comercial': 'COMERCIO',
            'razon_social':     getattr(ARCA_CONFIG, 'RAZON_SOCIAL', '') or '',
            'cuit':             ARCA_CONFIG.CUIT,
            'condicion_iva':    'Responsable Monotributo',
            'direccion':        '',
            'telefono':         '',
            'frase_extra':      '',
        }
        try:
            import config_cliente as _cfg
            emisor['nombre_comercial'] = getattr(_cfg, 'TICKET_NOMBRE_COMERCIAL', emisor['nombre_comercial'])
            emisor['razon_social']     = getattr(_cfg, 'RAZON_SOCIAL', emisor['razon_social'])
            emisor['cuit']             = getattr(_cfg, 'TICKET_CUIT_FORMATO', None) or getattr(_cfg, 'CUIT', emisor['cuit'])
            emisor['condicion_iva']    = getattr(_cfg, 'TICKET_CONDICION_IVA', emisor['condicion_iva'])
            emisor['direccion']        = getattr(_cfg, 'DIRECCION', '') or getattr(_cfg, 'TICKET_DIRECCION', '')
            emisor['telefono']         = getattr(_cfg, 'TELEFONO', '')
            emisor['frase_extra']      = getattr(_cfg, 'TICKET_FRASE_EXTRA', '')
        except Exception as e:
            print(f"WARN: no se pudieron leer datos de emisor de config_cliente: {e}")

        # ── QR de AFIP (solo si la factura tiene CAE) ──
        qr_b64 = None
        if getattr(factura, 'cae', None):
            try:
                generador_qr = crear_generador_qr(ARCA_CONFIG)
                qr_b64 = generador_qr.generar_qr_imagen(factura)
                if qr_b64 and isinstance(qr_b64, str) and qr_b64.startswith('data:'):
                    qr_b64 = qr_b64.split(',', 1)[1]
            except Exception as e:
                print(f"WARN: no se pudo generar QR para PDF: {e}")

        pdf_bytes = generar_pdf_factura(factura, emisor=emisor, qr_base64=qr_b64)

        nombre = f"Factura_{(factura.numero or str(factura.id)).replace('-', '_')}.pdf"
        return send_file(BytesIO(pdf_bytes), mimetype='application/pdf',
                         as_attachment=False, download_name=nombre)

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500



# ================ ARRANQUE DE LA APP ================
if __name__ == '__main__':
    # Crear directorios necesarios
    os.makedirs('cache', exist_ok=True)
    os.makedirs('logs', exist_ok=True)
    os.makedirs('certificados', exist_ok=True)
    
    print("🚀 Iniciando POS Argentina...")
    print(f"📍 URL: http://localhost:5000")
    print(f"🏢 CUIT: {ARCA_CONFIG.CUIT}")
    print(f"🏪 Punto de Venta: {ARCA_CONFIG.PUNTO_VENTA}")
    print(f"🔧 Ambiente: {'HOMOLOGACIÓN' if ARCA_CONFIG.USE_HOMOLOGACION else 'PRODUCCIÓN'}")
    print(f"🖨️ Impresión: {'Disponible' if IMPRESION_DISPONIBLE else 'No disponible'}")
    print(f"👤 Usuario: admin")
    print(f"🔑 Contraseña: admin123")
    print()
    
    with app.app_context():
        create_tables()
        
        migrar_productos_sin_costo_margen()  # ← EJECUTAR UNA SOLA VEZ

        # Limpiar datos problemáticos
        print("🧹 Verificando integridad de datos...")
        limpiar_facturas_duplicadas()
        verificar_estado_facturas()
    
    app.run(debug=True, host='0.0.0.0', port=5075)